"""
AbarVa — Master Dataset Generator
Generates ALL datasets for Arcturus and Meridian in parallel.
Run: python3 generate_datasets.py
Requires: pip install openpyxl
"""

import multiprocessing as mp
import os, sys, time, traceback
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Shared styling helpers ─────────────────────────────────────────────────
def styles():
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "hdr": PatternFill("solid", start_color="1A3A5C"),
        "alt": PatternFill("solid", start_color="F2F7FC"),
        "wht": PatternFill("solid", start_color="FFFFFF"),
        "red": PatternFill("solid", start_color="FDE8E8"),
        "amb": PatternFill("solid", start_color="FFF4E5"),
        "grn": PatternFill("solid", start_color="E8F5E9"),
        "blu": PatternFill("solid", start_color="E8F0F8"),
        "prp": PatternFill("solid", start_color="F0EEFF"),
    }
    return bdr, fills

def h(ws, row, col, val, width=16, fills=None, bdr=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    c.fill = fills["hdr"]
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = bdr
    ws.column_dimensions[get_column_letter(col)].width = width

def cel(ws, row, col, val, fill=None, fmt=None, bold=False, align="left", bdr=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(size=9, name="Arial", bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = bdr
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    return c

def title_row(ws, cols, text, fills):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    ws["A1"].value = text
    ws["A1"].font  = Font(bold=True, size=12, color="1A3A5C", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill  = fills["blu"]
    ws.row_dimensions[1].height = 16

def save(wb, path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

# ══════════════════════════════════════════════════════════════════════════
#  ARCTURUS — PDLC  (Files already generated: F01-F04, F06, F08, F10-F12)
#  Generating: F05 (Data Architecture), F07 (Contract Detail), F09 (MLOps)
# ══════════════════════════════════════════════════════════════════════════

def arcturus_pdlc_f05_data_architecture(base):
    """F05 — Data Architecture & Quality Assessment"""
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Data Architecture"
    title_row(ws, 10, "ARCTURUS FINANCIAL GROUP — Data Architecture & Quality Assessment", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Data Domain",22),("Primary System",20),("Secondary Systems",24),
            ("Record Count",14),("Update Frequency",16),("Data Quality\nScore /100",14),
            ("Completeness %",13),("Accuracy %",13),("Freshness\n(Hours Lag)",13),
            ("Golden Record?",13),("AI Ready?",11),("Blocker",30)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    domains = [
        ("Client Master","Salesforce FSC","Bloomberg AIM, SimCorp, internal CRM",142000,"Daily","38","61%","72%",26,"No — 4 conflicting sources","No","4 systems hold client data with different values. No reconciliation process. FSC master but 44% adoption means 56% of clients have stale or missing data."),
        ("Position Data","Bloomberg AIM","Aladdin, SimCorp, Charles River IMS",2800000,"Real-time (15-45min lag)","52","88%","91%","0.25","No — AIM vs Aladdin conflict","Partial","AIM and Aladdin positions diverge by up to 0.3% daily. Risk team manually reconciles. Real-time AI impossible with manual reconciliation step."),
        ("Trade / Order Data","Bloomberg AIM","Charles River IMS, Murex, Broadridge",4200000,"Real-time","71","94%","96%","0.1","No — settlement lag","Partial","Trade data reliable for execution. Settlement confirmation lags by 4-8 hours. Failed trades not flagged automatically."),
        ("Market Data","Bloomberg Terminal","FactSet, Reuters Eikon",1800000000,"Real-time","88","99%","98%","0","No (not needed)","Yes","Best data domain. External sourcing. High quality. AI-23 and AI-26 rely on this — highest AI readiness."),
        ("Risk Analytics","BlackRock Aladdin","Bloomberg AIM feed",320000,"Daily (monthly stress test)","58","82%","88%",24,"No","No","Daily risk metrics available but stress testing monthly only. Regulatory requirement is daily. Manual process for regulatory reporting."),
        ("Fund Accounting / NAV","SimCorp Dimension","Geneva, SQL Server DW",48000,"Daily T+1","72","91%","94%",24,"No — SimCorp vs Geneva diverge","No","UCITS and alternatives NAV from different systems. Manual reconciliation daily. T+1 lag prevents same-day reporting."),
        ("Client Reporting","SQL Server DW (manual)","SimCorp, Aladdin, FSC, AIM",280000,"Daily T+2 to T+3","28","61%","74%",72,"No","No","Manual aggregation from 14 systems. 3-day lag structural. Data quality degrades with each manual step. Board pack from stale data."),
        ("Compliance / Mandate Data","OpenPages GRC","Bloomberg AIM rules engine",18000,"Weekly","44","71%","68%",168,"No","No","Mandate rules in OpenPages not synchronised with AIM compliance engine. Breaches can go undetected for up to a week."),
        ("ESG Data","Manual (Excel)","FactSet ESG, Bloomberg ESG",9400,"Monthly","32","48%","61%",720,"No","No","ESG data manually compiled from two providers monthly. No automated reconciliation. AI-003 ESG scoring blocked by this."),
        ("Employee / HR Data","Workday","N/A",13000,"Real-time","82","96%","97%",1,"No (not needed)","Partial","HR data clean. Relevant for Maestro team design and knowledge transfer analysis."),
        ("Alternatives / PE Data","Geneva","SimCorp (partial)",4200,"Weekly","51","78%","82%",168,"No","No","Alternatives data in Geneva not connected to main performance system. Valuation lag 1 week."),
        ("Regulatory Reporting","Manual (Excel + OpenPages)","MAS portal, FCA, SEC",82000,"Daily (manual)","21","54%","61%",24,"No","No","All regulatory reporting manual. MAS FEAT breach partly caused by manual process inability to respond at required frequency."),
        ("Treasury / Cash","Murex","Bloomberg AIM",380000,"Intraday","68","89%","91%",4,"No","No","Treasury data reliable. Murex well-integrated. 4-hour lag before available in DW."),
        ("Vendor / Contract Data","Manual (SharePoint)","Various contract management tools",2400,"Ad-hoc","18","42%","55%",720,"No","No","Vendor contracts in SharePoint with no consistent taxonomy. Contract renewal dates tracked in spreadsheets. Highest data quality risk for consulting management."),
    ]

    for r, row in enumerate(domains, 3):
        ws.row_dimensions[r].height = 36
        domain, primary, secondary, records, freq, quality, comp, acc, lag, golden, ai_ready, blocker = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        q_val = int(quality)
        qf = fills["red"] if q_val < 40 else fills["amb"] if q_val < 65 else fills["grn"]
        cel(ws,r,1,domain,sf,bold=True,bdr=bdr)
        cel(ws,r,2,primary,sf,bdr=bdr)
        cel(ws,r,3,secondary,sf,bdr=bdr)
        cel(ws,r,4,records,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,5,freq,sf,align="center",bdr=bdr)
        cel(ws,r,6,q_val,qf,"#,##0",align="center",bold=True,bdr=bdr)
        for col, pct_str in zip([7,8],[comp,acc]):
            pct = int(pct_str.replace("%",""))/100
            pf = fills["red"] if pct < 0.7 else fills["amb"] if pct < 0.85 else fills["grn"]
            cel(ws,r,col,pct,pf,"0%",align="center",bdr=bdr)
        lag_num = float(lag) if isinstance(lag, (int, float)) else 0
        lag_f = fills["red"] if lag_num > 48 else fills["amb"] if lag_num > 4 else fills["grn"]
        cel(ws,r,9,lag,lag_f,"#,##0.0",align="center",bdr=bdr)
        gf = fills["red"] if golden.startswith("No") else fills["grn"]
        cel(ws,r,10,golden,gf,align="center",bdr=bdr)
        af = fills["grn"] if ai_ready=="Yes" else fills["amb"] if ai_ready=="Partial" else fills["red"]
        cel(ws,r,11,ai_ready,af,align="center",bdr=bdr)
        cel(ws,r,12,blocker,sf,bdr=bdr)

    # Sheet 2: Data readiness scorecard
    ws2 = wb.create_sheet("Readiness Scorecard")
    title_row(ws2, 5, "Data Readiness Scorecard — AI Initiative Prerequisites", fills)
    ws2.row_dimensions[2].height = 28
    sc_hdrs = [("Readiness Dimension",28),("Score /100",14),("Peer Benchmark",14),("Gap",12),("Primary Blocker",36)]
    for i,(hd,w) in enumerate(sc_hdrs,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    scorecard = [
        ("Golden Record Availability",4,72,-68,"0 of 14 data domains have a golden record. All domains have 2+ conflicting sources."),
        ("Data Pipeline Automation",12,68,-56,"Manual ETL dominates. 11 of 16 integrations are manual. 3-day lag structural."),
        ("Data Quality (avg across domains)",48,74,-26,"Client, ESG, Regulatory, Vendor data critically low. Market data excellent — skews average upward."),
        ("Real-time Data Availability",18,61,-43,"Only market data and trade execution are real-time. Everything else is T+1 or worse."),
        ("AI Training Data Readiness",11,58,-47,"Insufficient labelled data in most domains. No feature store. No data versioning."),
        ("Data Governance & Ownership",15,62,-47,"CDO role vacant. No data governance board. Data ownership ad-hoc across 14 systems."),
        ("Regulatory Data Compliance",22,71,-49,"MAS FEAT breach. GDPR obligations partially met. Data residency controls incomplete."),
        ("OVERALL DATA READINESS",12,65,-53,"Not ready for production AI deployment. Foundation programmes required first."),
    ]
    for r, (dim, score, bench, gap, blocker) in enumerate(scorecard, 3):
        ws2.row_dimensions[r].height = 28
        sf = fills["alt"] if r%2==0 else fills["wht"]
        sf = fills["prp"] if dim.startswith("OVERALL") else sf
        cel(ws2,r,1,dim,sf,bold=dim.startswith("OVERALL"),bdr=bdr)
        sf_s = fills["red"] if score < 30 else fills["amb"] if score < 60 else fills["grn"]
        cel(ws2,r,2,score,sf_s,"#,##0",align="center",bold=True,bdr=bdr)
        cel(ws2,r,3,bench,sf,"#,##0",align="center",bdr=bdr)
        cel(ws2,r,4,gap,fills["red"],"#,##0",align="center",bold=True,bdr=bdr)
        cel(ws2,r,5,blocker,sf,bdr=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/arcturus/pdlc/F05_Data_Architecture_Assessment.xlsx")
    return "F05 done"


def arcturus_pdlc_f09_mlops(base):
    """F09 — MLOps Infrastructure Assessment"""
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "MLOps Assessment"
    title_row(ws, 8, "ARCTURUS FINANCIAL GROUP — MLOps Infrastructure Assessment", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("MLOps Capability",26),("Current State",28),("Maturity\n(0-5)",11),
            ("Industry\nExpectation",11),("Gap",10),("Initiatives\nBlocked",12),
            ("Effort to Fix",14),("Recommendation",32)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    capabilities = [
        ("Model Registry","No model registry exists. Models tracked in spreadsheets by individual data scientists.",0,4,-4,28,"High — 6 months","Implement MLflow or SageMaker Model Registry. Standardise model versioning immediately."),
        ("Feature Store","No feature store. Each data scientist builds their own feature pipelines. No reuse.",0,3,-3,22,"High — 4 months","Implement Feast or Tecton. Begin with position and client data features — highest demand."),
        ("ML Pipeline / CI-CD","No ML-specific CI/CD. Code deployments use infrastructure pipeline not suitable for ML.",0,4,-4,28,"High — 4 months","Extend existing AWS CodePipeline for ML. Add model validation gates before promotion."),
        ("Model Serving Infrastructure","No serving layer. Models run as scripts on analyst laptops or ad-hoc servers.",0,4,-4,28,"High — 3 months","Deploy SageMaker Endpoints or Seldon. Containerise all models before serving."),
        ("Model Monitoring","No model monitoring. No drift detection. No performance tracking post-deployment.",0,3,-3,18,"Medium — 3 months","Implement Evidently AI or Arize. Start with performance and data drift for first 3 production models."),
        ("Data Versioning","No data versioning. Training data recreated each time. Reproducibility impossible.",0,3,-3,24,"Medium — 2 months","Implement DVC. Version all training datasets. Link to model registry."),
        ("Experiment Tracking","Partial — some teams use MLflow locally. No shared server. No institutional tracking.",1,4,-3,20,"Low — 1 month","Deploy shared MLflow server. Mandate experiment tracking for all initiatives. Quick win."),
        ("Model Validation Framework","No formal model validation. No independent review. CDO approval required but role vacant.",0,4,-4,14,"High — requires CDO","Establish model risk framework. Appoint interim model risk officer. CDO appointment prerequisite for governance."),
        ("Automated Retraining","No automated retraining. Models trained once, never updated. Drift guaranteed.",0,3,-3,18,"Medium — 4 months","Implement scheduled retraining pipelines for top 5 initiatives once serving layer live."),
        ("A/B Testing Framework","No A/B testing capability. No champion/challenger model deployment.",0,2,-2,8,"Medium — 3 months","Implement after serving layer. Not prerequisite for first production deployments."),
        ("Data Pipeline for ML","Manual data extraction. No automated feature pipeline. 3-day lag in source data.",0,4,-4,26,"High — 6 months (depends on golden record)","Blocked by golden record initiative (F05). Data pipeline automation requires clean source data."),
        ("Security & Compliance","No model governance. No audit trail. No GDPR-compliant data handling for ML training.",0,4,-4,12,"High — requires CDO + Legal","GDPR and MAS compliance for AI models undefined. CDO appointment prerequisite."),
    ]

    for r, row in enumerate(capabilities, 3):
        ws.row_dimensions[r].height = 32
        cap, state, maturity, expected, gap, blocked, effort, rec = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws,r,1,cap,sf,bold=True,bdr=bdr)
        cel(ws,r,2,state,fills["red"] if maturity==0 else sf,bdr=bdr)
        mf = fills["red"] if maturity<2 else fills["amb"] if maturity<4 else fills["grn"]
        cel(ws,r,3,maturity,mf,"#,##0",align="center",bold=True,bdr=bdr)
        cel(ws,r,4,expected,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,5,gap,fills["red"],"#,##0",align="center",bdr=bdr)
        cel(ws,r,6,blocked,fills["red"] if blocked>=20 else fills["amb"],"#,##0",align="center",bdr=bdr)
        ef = fills["red"] if "High" in effort else fills["amb"] if "Medium" in effort else fills["grn"]
        cel(ws,r,7,effort,ef,bdr=bdr)
        cel(ws,r,8,rec,sf,bdr=bdr)

    # Initiative blockers sheet
    ws2 = wb.create_sheet("Initiative Blockers by Capability")
    title_row(ws2, 5, "Which MLOps Gaps Block Which Initiatives", fills)
    ws2.row_dimensions[2].height = 28
    ib_hdrs = [("Initiative ID",12),("Initiative Name",32),("Blocking Gap",28),("Unblock Sequence",28),("Earliest Unblock",14)]
    for i,(hd,w) in enumerate(ib_hdrs,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    blockers = [
        ("AI-001","AI Trade Execution Optimisation","No serving infra + no model validation + no data pipeline","1. Experiment tracking → 2. Feature store → 3. Serving → 4. Monitoring","Month 10"),
        ("AI-002","Portfolio Rebalancing","No serving infra + Bloomberg data latency + no retraining","1. Bloomberg API wrapper → 2. Serving infra → 3. Auto-retraining","Month 8"),
        ("AI-005","MAS FEAT Compliance","No model registry + no audit trail + CDO vacant","EMERGENCY: Interim CDO → Compliance model framework → Registry","Month 2 (urgent)"),
        ("AI-006","Risk Model Automation","No model validation framework + CDO vacant","CDO appointment → Model risk framework → Serving infra","Month 4 (CDO prerequisite)"),
        ("AI-011","Predictive Cash Flow","Data pipeline automation only","Golden record → Feature pipeline → Serving","Month 7"),
        ("AI-023","AI Investment Research","Partial MLOps — nearest to production","Serving layer → Monitoring → Production deployment","Month 3 (priority)"),
        ("AI-026","Earnings Call Analysis","External data only — MLOps gap manageable","Serving layer → Lightweight monitoring → Production","Month 3 (priority)"),
        ("AI-019","NL Query Analytics","No golden record → no unified data layer","Golden record first → Then NL layer","Month 10+"),
    ]
    for r, row in enumerate(blockers, 3):
        ws2.row_dimensions[r].height = 28
        id_, name, gap, seq, when = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        uf = fills["red"] if "urgent" in when.lower() or "2" in when.split("Month")[1][:3] else fills["amb"]
        cel(ws2,r,1,id_,sf,align="center",bdr=bdr)
        cel(ws2,r,2,name,sf,bold=True,bdr=bdr)
        cel(ws2,r,3,gap,fills["red"],bdr=bdr)
        cel(ws2,r,4,seq,sf,bdr=bdr)
        cel(ws2,r,5,when,uf,align="center",bold=True,bdr=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/arcturus/pdlc/F09_MLOps_Infrastructure_Assessment.xlsx")
    return "F09 done"


# ══════════════════════════════════════════════════════════════════════════
#  ARCTURUS — AI-POWERED DELIVERY (8 files)
# ══════════════════════════════════════════════════════════════════════════

def arcturus_delivery_files(base):
    results = []

    # D01 — Consulting Audit: Output vs Promise
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Consulting Audit"
    title_row(ws, 9, "ARCTURUS — Consulting Output vs Promise Audit (All Active Engagements)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Vendor",16),("Engagement",28),("Promised Deliverable",28),
            ("Delivered?",12),("Quality\n/10",10),("Knowledge\nTransferred?",14),
            ("Annual\nCost (£M)",12),("Value\nRating",16),("Maestro\nReplacement?",14)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)
    audit = [
        ("Bloomberg LP","OMS Core Maintenance","24/7 support, 99.9% uptime, quarterly feature releases","Partial — SLA met, features delayed","6","No — vendor owns all knowledge",8.4,"Poor — trapped dependency","Yes — API wrapper + internal OMS team"),
        ("Infosys","Risk Tech & Analytics Development","90% sprint velocity, knowledge wiki, internal team upskilling","No — 71% velocity, wiki incomplete, no upskilling","5","No — 22% KT score",3.6,"Below expectations","Yes — 2 Maestros replace 12 Infosys"),
        ("Wipro","Client Portal + Data Platform","85% velocity, full knowledge transfer, golden record delivery","No — 58% velocity, KT 15%, golden record not started","3","No — Wipro own all customisations",4.8,"Poor — knowledge hostage","Yes — exit plan + 2 Maestros"),
        ("Deloitte","MAS FEAT Compliance Programme","Compliance by December 2025, internal team capability","No — deadline missed, internal team not capable","4","Partial — 45% KT",2.6,"Below expectations — advisory only","Partial — internal compliance team with Maestro support"),
        ("TCS","APAC Technology Operations","96% SLA, knowledge transfer, documentation","Yes for SLA, partial for KT","6","Partial — 38% KT",0.9,"Adequate","No — retain, value for APAC ops"),
        ("Google PSO","AI/ML Platform Setup","Full MLOps infrastructure delivered","No — 22% delivered, engagement ended","2","No — 5% KT score",3.5,"Failed","Yes — internal ML engineering with Maestro"),
        ("AWS ProServe","Infrastructure Modernisation","Runbook delivery, internal team training","Mostly — 88% complete","7","Partial — 62% KT",1.4,"Good — retain","No — continue with internal support"),
        ("Salesforce PS","FSC Customisation","Feature delivery, internal admin capability","Partial — 71% features, no admin capability","5","No — 28% KT",1.6,"Below expectations","Yes — internal Salesforce admin team"),
        ("Various Contractors","Enterprise Architecture","Architecture decisions, documentation","No — no documentation, no knowledge retention","1","None — 0% KT",1.8,"Critical risk","Yes — permanent EA hire + Maestro"),
    ]
    for r, row in enumerate(audit, 3):
        ws.row_dimensions[r].height = 28
        vendor, eng, promised, delivered, quality, kt, cost, rating, replace = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws,r,1,vendor,sf,bold=True,bdr=bdr)
        cel(ws,r,2,eng,sf,bdr=bdr); cel(ws,r,3,promised,sf,bdr=bdr)
        df = fills["red"] if "No" in delivered[:3] else fills["amb"] if "Partial" in delivered else fills["grn"]
        cel(ws,r,4,delivered,df,bdr=bdr)
        qv = int(quality); qf = fills["red"] if qv<5 else fills["amb"] if qv<7 else fills["grn"]
        cel(ws,r,5,qv,qf,"#,##0",align="center",bdr=bdr)
        ktf = fills["red"] if "No" in kt else fills["amb"] if "Partial" in kt else fills["grn"]
        cel(ws,r,6,kt,ktf,bdr=bdr)
        cel(ws,r,7,cost,sf,"£#,##0.0",align="center",bdr=bdr)
        rf = fills["red"] if "Poor" in rating or "Failed" in rating or "Critical" in rating else fills["amb"] if "Below" in rating or "Adequate" in rating else fills["grn"]
        cel(ws,r,8,rating,rf,bdr=bdr)
        rpf = fills["grn"] if "Yes" in replace else fills["amb"] if "Partial" in replace else fills["wht"]
        cel(ws,r,9,replace,rpf,bdr=bdr)
    save(wb, f"{base}/arcturus/delivery/D01_Consulting_Audit_Output_vs_Promise.xlsx")
    results.append("D01 done")

    # D02 — Knowledge Retention Risk
    wb2 = Workbook(); bdr, fills = styles()
    ws2 = wb2.active; ws2.title = "Knowledge Risk"
    title_row(ws2, 8, "ARCTURUS — Knowledge Retention Risk Register", fills)
    ws2.row_dimensions[2].height = 32
    hdrs2 = [("Knowledge Domain",26),("Current Owner",22),("Type",14),
              ("If Lost Impact",28),("Retention\nPlan Exists?",14),
              ("Risk Score\n/100",12),("Mitigation",32)]
    for i,(hd,w) in enumerate(hdrs2,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    knowledge = [
        ("Bloomberg AIM customisation logic (14 customisations)","Bloomberg LP engineers","Vendor-owned","Cannot maintain or modify OMS. Entire trading operation at risk.",  "No",98,"AIM API wrapper + internal team capability build. 12-month programme."),
        ("Wipro FSC customisation code","Wipro developers","Vendor-owned","Cannot deploy or maintain Salesforce FSC. Client portal frozen.","No",88,"Code escrow clause + internal Salesforce admin hire immediately."),
        ("Google PSO MLOps design (incomplete)","Google PSO (engagement ended)","Departed vendor","AI/ML platform design knowledge walked out. Starting from zero.","No",95,"Reconstruct from artefacts. Hire ML engineer. Use AbarVa MLOps playbook."),
        ("MAS FEAT compliance process","Deloitte consultants","Vendor-dependent","Regulatory process not documented internally. Next audit will fail.","Partial",72,"Internal compliance team documentation sprint. Deloitte exit plan."),
        ("Enterprise architecture decisions","Contractor (rolling contract)","Contractor-owned","Architecture logic not documented. No permanent owner. Future decisions uninformed.","No",91,"Permanent EA hire. Documentation retrospective. Architecture decision records."),
        ("Data pipeline logic (all 16 integrations)","Wipro, various contractors","Mixed","Manual processes not documented. Loss of any operator causes outage.","No",85,"Runbook creation sprint. All manual processes documented and version-controlled."),
        ("Risk model logic (Aladdin configuration)","Head of Risk + vendor","Internal + vendor","Risk model configuration not fully documented. Vendor changes break models.","Partial",61,"Model documentation programme. Internal model risk function."),
        ("Client mandate rules (93 Bloomberg rules)","Bloomberg LP","Vendor-owned","Compliance rule logic inaccessible without Bloomberg. Mandate breaches undetected.","No",94,"Compliance rule extraction and documentation. Internal rules engine assessment."),
        ("Portfolio Analytics squad IP","Internal team (Rachel Kim)","Internal","Best internal team. Knowledge concentrated in 3 people. Attrition risk high.","Partial",55,"Knowledge sharing programme. Documentation. Succession planning."),
        ("CDO-sponsored initiative context","(CDO VACANT — departed)","Departed employee","14 initiatives orphaned. Context and rationale lost. Decisions without history.","No",89,"Initiative context documentation. CIO assumes interim responsibility with Maestro support."),
    ]
    for r, row in enumerate(knowledge, 3):
        ws2.row_dimensions[r].height = 28
        domain, owner, type_, impact, plan, risk, mit = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws2,r,1,domain,sf,bold=True,bdr=bdr); cel(ws2,r,2,owner,fills["red"] if "Vacant" in owner or "VACANT" in owner or "ended" in owner else sf,bdr=bdr)
        cel(ws2,r,3,type_,sf,bdr=bdr); cel(ws2,r,4,impact,sf,bdr=bdr)
        pf = fills["grn"] if plan=="Yes" else fills["amb"] if plan=="Partial" else fills["red"]
        cel(ws2,r,5,plan,pf,align="center",bdr=bdr)
        rf = fills["red"] if risk>=80 else fills["amb"] if risk>=60 else fills["grn"]
        cel(ws2,r,6,risk,rf,"#,##0",align="center",bold=True,bdr=bdr)
        cel(ws2,r,7,mit,sf,bdr=bdr)
    save(wb2, f"{base}/arcturus/delivery/D02_Knowledge_Retention_Risk.xlsx")
    results.append("D02 done")

    # D03 — Maestro Team Design
    wb3 = Workbook(); bdr, fills = styles()
    ws3 = wb3.active; ws3.title = "Maestro Team Design"
    title_row(ws3, 8, "ARCTURUS — Proposed Maestro Team Design (Replaces 50+ Consultants)", fills)
    ws3.row_dimensions[2].height = 32
    hdrs3 = [("Maestro Role",24),("Scope",32),("Replaces",24),("Annual\nCost Replaced (£M)",16),
             ("Knowledge\nArea",24),("Wave",10),("Engagement\nDuration",14),("Success Metric",24)]
    for i,(hd,w) in enumerate(hdrs3,1): h(ws3,2,i,hd,w,fills=fills,bdr=bdr)
    maestros = [
        ("Delivery Maestro — OMS","Govern Bloomberg AIM relationship. Build internal OMS capability. Reduce vendor dependency.","12-14 Bloomberg LP engineers",7.2,"OMS architecture, Bloomberg AIM API, trade lifecycle","Wave 1","18 months","OMS change capability internal. Bloomberg dependency ratio <40%."),
        ("Delivery Maestro — Data & AI","Build MLOps foundation. Govern AI initiative delivery. Establish golden record.","8-10 Wipro + Google PSO",5.6,"MLOps, data engineering, AI governance, golden record","Wave 1","24 months","3 AI initiatives in production. Data pipeline automated. Reporting lag <4hrs."),
        ("Delivery Maestro — Client Platform","Transfer FSC capability internally. Exit Wipro. Build internal Salesforce admin.","7 Wipro Salesforce team",3.2,"Salesforce FSC, CRM, client data, adoption programmes","Wave 1","12 months","FSC adoption >70%. Internal team owns all deployments. Wipro exited."),
        ("Delivery Maestro — Regulatory Tech","Build internal MAS FEAT compliance capability. Exit Deloitte dependency.","5 Deloitte consultants",2.6,"MAS regulatory, compliance automation, FCA requirements","Wave 1","12 months","MAS FEAT compliant. Internal team capable of next audit without external support."),
        ("Delivery Maestro — Risk Technology","Establish model risk governance. Support CDO appointment. Unblock AI-006.","4 Infosys risk tech",1.8,"Model risk, Aladdin enhancement, regulatory model validation","Wave 2","18 months","Model risk framework established. 5 AI initiatives unblocked."),
        ("Architecture Maestro","Replace contractor EA leadership. Establish permanent architecture governance.","4 contractors (EA function)",1.8,"Enterprise architecture, technology strategy, vendor governance","Wave 1","Permanent knowledge transfer, 12 months active","Architecture decisions documented. No contractor in permanent role."),
    ]
    total_replaced = 0
    for r, row in enumerate(maestros, 3):
        ws3.row_dimensions[r].height = 28
        role, scope, replaces, cost, knowledge, wave, duration, metric = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        wf = fills["grn"] if wave=="Wave 1" else fills["amb"]
        cel(ws3,r,1,role,sf,bold=True,bdr=bdr); cel(ws3,r,2,scope,sf,bdr=bdr)
        cel(ws3,r,3,replaces,sf,bdr=bdr)
        cel(ws3,r,4,cost,fills["grn"],"£#,##0.0",align="center",bold=True,bdr=bdr)
        cel(ws3,r,5,knowledge,sf,bdr=bdr); cel(ws3,r,6,wave,wf,align="center",bdr=bdr)
        cel(ws3,r,7,duration,sf,bdr=bdr); cel(ws3,r,8,metric,sf,bdr=bdr)
        total_replaced += cost
    tr = len(maestros)+3
    cel(ws3,tr,1,"TOTAL CONSULTING SPEND RECOVERABLE",bold=True,bdr=bdr)
    ws3.cell(tr,4,value=f"=SUM(D3:D{tr-1})"); ws3.cell(tr,4).number_format="£#,##0.0"
    ws3.cell(tr,4).font=Font(bold=True,size=11,name="Arial",color="1A3A5C")
    ws3.cell(tr,4).alignment=Alignment(horizontal="center")
    save(wb3, f"{base}/arcturus/delivery/D03_Maestro_Team_Design.xlsx")
    results.append("D03 done")

    return results


# ══════════════════════════════════════════════════════════════════════════
#  ARCTURUS — MARGIN OPTIMIZATION (8 files)
# ══════════════════════════════════════════════════════════════════════════

def arcturus_margin_files(base):
    results = []

    # M01 — P&L by Business Unit
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "P&L by Business Unit"
    title_row(ws, 10, "ARCTURUS FINANCIAL GROUP — P&L by Business Unit FY2025 (£M)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Business Unit",24),("AUM (£B)",12),("Revenue (£M)",14),("Operating\nCosts (£M)",14),
            ("Operating\nProfit (£M)",14),("C/I Ratio",11),("Target C/I",11),
            ("Gap (pp)",10),("Gap (£M)",12),("YoY\nRevenue",11),("Notes",28)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)
    units = [
        ("Global Equities",320,284,238,46,0.838,0.60,-23.8,67.5,0.042,"Largest AUM. Highest C/I. Bloomberg AIM dependency drives cost."),
        ("Fixed Income",210,162,124,38,0.765,0.58,-18.5,29.9,0.021,"Murex and manual processes inflate ops cost."),
        ("Multi-Asset",140,118,92,26,0.780,0.62,-16.0,18.9,0.038,"Cross-asset complexity adds overhead."),
        ("Alternatives / Private Markets",88,96,71,25,0.740,0.60,-14.0,13.4,0.065,"Geneva and manual valuations."),
        ("Asia Pacific (Singapore)",82,74,62,12,0.838,0.62,-21.8,17.2,0.091,"MAS FEAT breach costs emerging."),
        ("Client Solutions / Wealth",0,46,38,8,0.826,0.64,-18.6,8.6,-0.012,"FSC adoption issues reducing cross-sell."),
        ("Corporate / Technology","N/A",0,178,-178,"N/A","N/A","N/A","N/A","N/A","£178M above peer benchmark. IT cost drag on entire business."),
    ]
    for r, row in enumerate(units, 3):
        ws.row_dimensions[r].height = 24
        unit, aum, rev, cost, profit, ci, target, gap_pp, gap_m, yoy, notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws,r,1,unit,sf,bold=True,bdr=bdr)
        cel(ws,r,2,aum,sf,"#,##0" if isinstance(aum,(int,float)) else None,align="center",bdr=bdr)
        cel(ws,r,3,rev,sf,"£#,##0.0" if isinstance(rev,(int,float)) else None,align="center",bdr=bdr)
        cel(ws,r,4,cost,sf,"£#,##0.0" if isinstance(cost,(int,float)) else None,align="center",bdr=bdr)
        pf = fills["grn"] if isinstance(profit,(int,float)) and profit>0 else fills["red"]
        cel(ws,r,5,profit,pf,"£#,##0.0" if isinstance(profit,(int,float)) else None,align="center",bdr=bdr)
        if isinstance(ci,float):
            cif = fills["red"] if ci > 0.75 else fills["amb"] if ci > 0.65 else fills["grn"]
            cel(ws,r,6,ci,cif,"0.0%",align="center",bdr=bdr)
            cel(ws,r,7,target,sf,"0.0%",align="center",bdr=bdr)
            cel(ws,r,8,gap_pp,fills["red"],align="center",bdr=bdr)
            cel(ws,r,9,gap_m,fills["red"],"£#,##0.0",align="center",bdr=bdr)
            cel(ws,r,10,yoy,fills["grn"] if isinstance(yoy,float) and yoy>0 else fills["red"],"0.0%",align="center",bdr=bdr)
        else:
            for c in range(6,11):
                cel(ws,r,c,"N/A",sf,align="center",bdr=bdr)
        cel(ws,r,11,notes,sf,bdr=bdr)

    save(wb, f"{base}/arcturus/margin/M01_PL_by_Business_Unit.xlsx")
    results.append("M01 done")

    # M02 — AI Spend ROI Tracker (the most damning file)
    wb2 = Workbook(); bdr, fills = styles()
    ws2 = wb2.active; ws2.title = "AI Spend ROI"
    title_row(ws2, 8, "ARCTURUS — AI Investment ROI Tracker (£94M Committed · £0 Verified Return)", fills)
    ws2.row_dimensions[2].height = 32
    hdrs2 = [("Initiative",28),("Budget\nCommitted (£M)",14),("Spent\nto Date (£M)",14),
             ("Expected\nROI (£M pa)",14),("Verified\nROI (£M)",14),("ROI\nGap (£M)",14),
             ("Months\nSince Start",12),("Status",14),("Root Cause of Zero ROI",28)]
    for i,(hd,w) in enumerate(hdrs2,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    ai_roi = [
        ("AI Trade Execution Optimisation",4.2,3.8,12.0,0.0,-12.0,25,"Pilot - Stalled","No MLOps. No deployment infrastructure. Model trained, cannot serve."),
        ("Portfolio Rebalancing AI",3.1,2.6,8.0,0.0,-8.0,27,"Pilot - Stalled","Bloomberg data latency. Manual override required. Cannot automate."),
        ("ESG Scoring Automation",2.4,1.8,4.0,0.0,-4.0,19,"In Development","ESG data inconsistent. No golden record. Model not trained on reliable data."),
        ("Client Sentiment Analysis",1.8,1.4,6.0,0.0,-6.0,24,"Pilot - Stalled","CDO vacant. No governance. FSC data sparse. Cannot build model."),
        ("MAS FEAT Compliance AI",3.6,2.9,18.0,0.0,-18.0,21,"BREACH — Overdue","Deployed too late. Manual process continued. Now regulatory breach."),
        ("Risk Model Automation",4.8,4.1,9.0,0.0,-9.0,31,"Pilot - Stalled","CDO vacant for 11 months. No model risk sign-off possible."),
        ("NL Client Reporting",2.1,1.2,5.0,0.0,-5.0,15,"In Development","3-day data lag. Reports on stale data. Useless for clients."),
        ("Fraud Detection AI",2.8,2.3,7.0,0.0,-7.0,26,"Pilot - Stalled","Transaction data in 5 systems. No unified feed. Cannot train reliably."),
        ("AI Trade Surveillance",4.1,3.4,14.0,0.0,-14.0,24,"Pilot - Stalled","FCA model validation required. No framework. CDO vacant."),
        ("Real-Time Performance Attribution",3.8,3.2,8.0,0.0,-8.0,27,"Pilot - Stalled","Bloomberg latency. No streaming infrastructure. T+1 at best."),
        ("Predictive Cash Flow",1.9,1.1,6.0,0.0,-6.0,15,"In Development","3-day data lag makes 30-day forecasts inaccurate."),
        ("AI Regulatory Capital",2.9,2.4,12.0,0.0,-12.0,24,"Pilot - Stalled","Model validation absent. CDO vacant. Cannot deploy."),
        ("AI Investment Research (AI-023)",1.8,1.0,6.0,0.0,-6.0,15,"In Development","Best chance at production. Needs serving infrastructure only."),
        ("AI Trade Matching",3.4,2.9,9.0,0.0,-9.0,26,"Pilot - Stalled","Custodian data fragmented. No unified settlement feed."),
        ("Automated Client Onboarding",3.2,2.7,11.0,0.0,-11.0,22,"Pilot - Stalled","CDO vacant. Regulatory AI framework undefined."),
        ("Earnings Call Analysis (AI-026)",1.4,0.8,4.0,0.0,-4.0,15,"In Development","External data. Near production. Needs serving layer only."),
        ("Automated Compliance Monitoring",2.8,2.3,10.0,0.0,-10.0,22,"Pilot - Stalled","Mandate data in 14 formats. No unified compliance layer."),
        ("Dynamic Asset Allocation",5.6,0.8,22.0,0.0,-22.0,6,"Concept","Foundation not in place. Should not have been started."),
        ("Remaining initiatives (10)",38.4,24.6,63.0,0.0,-63.0,"Various","Various stalled","All blocked by same root causes: MLOps, data quality, CDO vacancy."),
    ]
    for r, row in enumerate(ai_roi, 3):
        ws2.row_dimensions[r].height = 24
        name, budget, spent, expected, verified, gap, months, status, cause = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws2,r,1,name,sf,bold=True,bdr=bdr)
        cel(ws2,r,2,budget,sf,"£#,##0.0",align="center",bdr=bdr)
        cel(ws2,r,3,spent,sf,"£#,##0.0",align="center",bdr=bdr)
        cel(ws2,r,4,expected,fills["grn"],"£#,##0.0",align="center",bdr=bdr)
        cel(ws2,r,5,verified,fills["red"],"£#,##0.0",align="center",bold=True,bdr=bdr)
        cel(ws2,r,6,gap,fills["red"],"£#,##0.0",align="center",bdr=bdr)
        cel(ws2,r,7,months,sf,"#,##0" if isinstance(months,int) else None,align="center",bdr=bdr)
        sf_s = fills["red"] if "Stalled" in str(status) or "BREACH" in str(status) else fills["amb"] if "Development" in str(status) else fills["wht"]
        cel(ws2,r,8,status,sf_s,align="center",bdr=bdr)
        cel(ws2,r,9,cause,sf,bdr=bdr)
    tr2 = len(ai_roi)+3
    cel(ws2,tr2,1,"TOTALS",bold=True,bdr=bdr)
    for col,col_l in zip([2,3,4,5,6],["B","C","D","E","F"]):
        ws2.cell(tr2,col,value=f"=SUM({col_l}3:{col_l}{tr2-1})")
        ws2.cell(tr2,col).number_format="£#,##0.0"
        ws2.cell(tr2,col).font=Font(bold=True,size=10,name="Arial",color="FFFFFF")
        ws2.cell(tr2,col).fill=HDR_FILL=fills["hdr"]
        ws2.cell(tr2,col).alignment=Alignment(horizontal="center")
    save(wb2, f"{base}/arcturus/margin/M02_AI_Spend_ROI_Tracker.xlsx")
    results.append("M02 done")

    return results


# ══════════════════════════════════════════════════════════════════════════
#  MERIDIAN HEALTH — Core datasets (PDLC focus)
# ══════════════════════════════════════════════════════════════════════════

def meridian_files(base):
    results = []

    # MH01 — Clinical Operations Dashboard
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Clinical Operations"
    title_row(ws, 9, "MERIDIAN HEALTH SYSTEM — Clinical Operations Dashboard FY2025", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Metric",28),("Current",14),("Benchmark",14),("Target",14),
            ("Gap",12),("Financial Impact",18),("Trend",12),("Priority",12),("Notes",32)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)
    metrics = [
        ("RCM Denial Rate","18.2%","12.0%","12.0%","+6.2pp","$94M annual write-off","Worsening","Critical","Ensemble Health Partners contract up for renewal. $8M SLA penalties enforceable now."),
        ("Operating Margin","1.8%","4.0%","4.0%","-2.2pp","$246M gap to target","Flat","Critical","Revenue growing 3.8% but costs growing 5.1%. AI investment not delivering."),
        ("Days in AR","52 days","35 days","38 days","+14 days","$42M working capital trapped","Worsening","Critical","Manual prior auth causing downstream AR delay."),
        ("Prior Auth Average Days","4.2 days","1.8 days","2.0 days","+2.2 days","$28M physician attrition risk","Flat","Critical","CMS mandate: electronic prior auth by January 2027. 14 months to compliance."),
        ("MyChart Adoption","34%","60%","65%","-31pp","$18M lost patient engagement revenue","Improving slowly","High","Epic optimization score 58/100 directly correlates. Modules underutilized."),
        ("Epic Optimization Score","58/100","80/100","80/100","-22 points","$12M annual opportunity","Flat","High","22 Epic modules licensed but <30% utilized. Training completion 41%."),
        ("MA Star Rating","3.5","4.0","4.0","-0.5 stars","$24M revenue at risk (CMS bonus threshold)","Flat","High","3.5 = zero CMS quality bonus. 4.0 = 5% revenue bonus on MA premiums."),
        ("Travel Nurse Spend","$48M","$28M","$30M","+$18M","$18M above target","Improving","High","Pandemic peak was $68M. Improving but $18M above benchmark."),
        ("Hospital Occupancy","71%","76%","78%","-7pp","$31M revenue gap","Flat","Medium","Capacity not the constraint — throughput and length of stay are."),
        ("Physician Productivity (RVUs)","4,820/FTE","5,400/FTE","5,200/FTE","-380 RVUs","$14M productivity gap","Worsening","High","Documentation burden 2.1 hrs/day cited as primary cause. GenAI opportunity."),
        ("GenAI Clinical Documentation","Not deployed","Industry: 34% deployed","Deploy by Q2 2026","N/A","$42M productivity opportunity","Not started","High","68% of physicians report burnout partly from documentation. Fastest ROI initiative."),
        ("Claim Automation Rate","31%","65%","60%","-34pp","$22M manual processing cost","Flat","High","Ensemble Health Partners contract tied to automation milestones not being met."),
    ]
    for r, row in enumerate(metrics, 3):
        ws.row_dimensions[r].height = 28
        metric, current, bench, target, gap, impact, trend, priority, notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        pf = fills["red"] if priority=="Critical" else fills["amb"] if priority=="High" else fills["grn"]
        tf = fills["red"] if "Worsening" in trend else fills["amb"] if "Flat" in trend or "slowly" in trend else fills["grn"]
        cel(ws,r,1,metric,sf,bold=True,bdr=bdr); cel(ws,r,2,current,sf,align="center",bdr=bdr)
        cel(ws,r,3,bench,sf,align="center",bdr=bdr); cel(ws,r,4,target,sf,align="center",bdr=bdr)
        cel(ws,r,5,gap,fills["red"],align="center",bdr=bdr)
        cel(ws,r,6,impact,fills["red"],align="center",bdr=bdr)
        cel(ws,r,7,trend,tf,align="center",bdr=bdr)
        cp = cel(ws,r,8,priority,pf,align="center",bold=True,bdr=bdr)
        cp.font=Font(size=9,name="Arial",bold=True)
        cel(ws,r,9,notes,sf,bdr=bdr)
    save(wb, f"{base}/meridian/pdlc/MH01_Clinical_Operations_Dashboard.xlsx")
    results.append("MH01 done")

    # MH02 — AI Initiative Inventory
    wb2 = Workbook(); bdr, fills = styles()
    ws2 = wb2.active; ws2.title = "AI Initiatives"
    title_row(ws2, 8, "MERIDIAN HEALTH — AI Initiative Inventory", fills)
    ws2.row_dimensions[2].height = 32
    hdrs2 = [("Initiative",28),("Sponsor",18),("Status",14),("Budget ($M)",12),
             ("Spent ($M)",12),("ROI Documented?",14),("Genome Pattern",18),
             ("Blocker",28),("Recommendation",28)]
    for i,(hd,w) in enumerate(hdrs2,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    initiatives = [
        ("Prior Auth Automation","Robert Chen (CIO)","Pilot — CMS mandate pressing",8.4,6.2,"No","F006 (79%) — no deployment infra","Epic and Cohere Health integration incomplete. No MLOps.","PRIORITY. $37.6M annual value. 91% Genome confidence. Cohere Health recommended vendor."),
        ("GenAI Clinical Documentation","Dr. Sarah Kim (CMIO)","Pilot — physician demand high",4.2,2.8,"No","F008 (61%) — adoption will fail without workflow","Pilot ran in 2 departments. Physicians positive. No enterprise deployment plan.","High priority. $42M productivity opportunity. Deploy to all physicians by Q2 2026."),
        ("CDO Executive Hire","Emily Rodriguez (CEO)","Search active",0.4,0.1,"N/A","F002 (84%) — no sponsor","Search in month 3. Interim coverage by CIO.","Urgent. $94M in value blocked. Hire enables 12 downstream initiatives."),
        ("Epic Optimization Sprint","Robert Chen (CIO)","Planning",2.1,0.4,"No","F008 (61%) — adoption gap","22 modules underutilized. Training completion 41%.","$12M annual value. Internal capability needed. Maestro to govern sprint."),
        ("RCM Vendor Renegotiation","James Park (CFO)","Active negotiation",0.2,0.2,"N/A","F011 (74%) — vendor misalignment","Ensemble Health Partners. $8M SLA penalties enforceable.","Negotiate now. Leverage SLA breaches. Cohere Health as competitive alternative."),
        ("Predictive Readmission Model","Dr. Sarah Kim (CMIO)","Concept",1.8,0.2,"No","F003 (68%) — data readiness","Epic data quality 58/100. Insufficient clean data for model training.","Block until Epic optimization complete. Data readiness prerequisite."),
        ("Travel Nurse Demand Forecasting","VP HR","In Development",1.2,0.8,"No","F003 (68%)","Workforce data in 3 systems. No unified feed.","Potential $8M value. Unblock with workforce data integration."),
        ("MA Star Rating AI","Emily Rodriguez (CEO)","Concept",3.4,0.4,"No","F002 (84%) — no CDO","CDO hire prerequisite. Multiple quality measures need coordination.","Block until CDO hired. $24M revenue at risk if star rating drops further."),
    ]
    for r, row in enumerate(initiatives, 3):
        ws2.row_dimensions[r].height = 28
        name, sponsor, status, budget, spent, roi, genome, blocker, rec = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        sf_s = fills["red"] if "Pilot" in status and "pressing" in status else fills["amb"] if "Development" in status or "Planning" in status else fills["wht"]
        cel(ws2,r,1,name,sf,bold=True,bdr=bdr); cel(ws2,r,2,sponsor,sf,bdr=bdr)
        cel(ws2,r,3,status,sf_s,bdr=bdr)
        cel(ws2,r,4,budget,sf,"$#,##0.0",align="center",bdr=bdr)
        cel(ws2,r,5,spent,sf,"$#,##0.0",align="center",bdr=bdr)
        rf = fills["grn"] if roi=="Yes" else fills["red"] if roi=="No" else fills["wht"]
        cel(ws2,r,6,roi,rf,align="center",bdr=bdr)
        cel(ws2,r,7,genome,fills["amb"],bdr=bdr)
        cel(ws2,r,8,blocker,fills["red"],bdr=bdr); cel(ws2,r,9,rec,sf,bdr=bdr)
    save(wb2, f"{base}/meridian/pdlc/MH02_AI_Initiative_Inventory.xlsx")
    results.append("MH02 done")

    return results


# ══════════════════════════════════════════════════════════════════════════
#  FIRST CAPITAL — Core datasets
# ══════════════════════════════════════════════════════════════════════════

def firstcapital_files(base):
    results = []

    # FC01 — Core Banking Assessment
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Core Banking Assessment"
    title_row(ws, 8, "FIRST CAPITAL FINANCIAL — Core Banking & Technology Assessment FY2025", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("System / Initiative",26),("Current State",28),("Risk Level",12),
            ("Annual Cost ($M)",14),("Business Impact",24),("EOL / Deadline",14),
            ("Decision Required",24),("Recommendation",28)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)
    systems = [
        ("FIS HORIZON Core Banking","22-year-old system. High customisation. Vendor EOL pathway unclear. 3 major failed upgrade attempts.","Critical",12.4,"Underpins all 84 branches, 4,200 staff, $18B AUM. Failure = operational shutdown.","EOL 2027 (FIS guidance)","Full replacement vs modernisation vs lift-and-shift. Board level decision.","Begin vendor assessment immediately. FIS, Temenos, Thought Machine evaluated. 18-month minimum migration."),
        ("SQL Server 2017 (Data Warehouse)","EOL October 2025 — operating without security patches. 14 data feeds. Manual ETL.","Critical",0.8,"All reporting, compliance, and analytics depend on this. Security breach risk elevated.","EOL PASSED — October 2025","Immediate migration to Azure SQL or Snowflake. Cannot delay.","Emergency remediation. Migrate within 90 days. Azure SQL recommended."),
        ("Q2 Holdings Digital Banking","4 years deployed. 3.2/5 app rating vs 3.8 benchmark. 41% digital adoption vs 67% peers.","High",3.2,"Customer attrition risk. 26pp below peer digital adoption. Lost fee revenue.","Supported — no EOL","Optimise Q2 platform vs replace. Customer journey redesign required.","Q2 optimisation sprint. UX redesign. Target 65% adoption by Q4 2026."),
        ("AML System (manual-heavy)","34% automated. 78% false positive rate. Manual review team 24 FTE. Peers at 72% automation.","High",4.8,"Regulatory risk. Exam findings expected. $3.8M annual excess cost vs peers.","BSA/AML exam due Q2 2026","AML system upgrade or replacement. FedNow dependency resolved first.","Evaluate NICE Actimize, Oracle FCCM. Target 70% automation, <20% false positive."),
        ("FedNow Integration","Not live. 68% of peers already live. Fed expecting compliance.","High",1.2,"Payment competitiveness. Regulatory expectation. Real-time payment capability gap.","Fed guidance Q4 2025 (past)","Integration with FIS HORIZON. Complex given system age.","FedNow integration as part of FIS modernisation programme. Interim workaround via correspondent bank."),
        ("AI/ML Initiatives (6 total)","6 initiatives. 0 in production. $12M committed. $0 verified ROI.","High",12.0,"Digital adoption gap, AML efficiency, fraud reduction all dependent on AI.","Various","Define AI governance framework. Appoint CDO or AI lead.","AI governance framework first. Then sequence: AML AI → Fraud AI → Digital personalisation AI."),
    ]
    for r, row in enumerate(systems, 3):
        ws.row_dimensions[r].height = 32
        name, state, risk, cost, impact, eol, decision, rec = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        rf = fills["red"] if risk=="Critical" else fills["amb"] if risk=="High" else fills["grn"]
        cel(ws,r,1,name,sf,bold=True,bdr=bdr); cel(ws,r,2,state,sf,bdr=bdr)
        cr = cel(ws,r,3,risk,rf,align="center",bold=True,bdr=bdr)
        cel(ws,r,4,cost,sf,"$#,##0.0",align="center",bdr=bdr); cel(ws,r,5,impact,sf,bdr=bdr)
        eol_f = fills["red"] if "PASSED" in eol or "2025" in eol else fills["amb"] if "2026" in eol or "2027" in eol else fills["wht"]
        cel(ws,r,6,eol,eol_f,bdr=bdr); cel(ws,r,7,decision,sf,bdr=bdr); cel(ws,r,8,rec,sf,bdr=bdr)
    save(wb, f"{base}/firstcapital/tech/FC01_Core_Banking_Assessment.xlsx")
    results.append("FC01 done")

    # FC02 — Digital Adoption Analytics
    wb2 = Workbook(); bdr, fills = styles()
    ws2 = wb2.active; ws2.title = "Digital Adoption"
    title_row(ws2, 8, "FIRST CAPITAL FINANCIAL — Digital Adoption Analytics by Product & Channel", fills)
    ws2.row_dimensions[2].height = 32
    hdrs2 = [("Product / Channel",24),("Current\nAdoption",13),("Peer\nBenchmark",13),
             ("Gap",10),("Revenue at\nRisk ($M)",14),("Customer\nSatisfaction",14),
             ("Trend",12),("Root Cause",28),("Initiative",24)]
    for i,(hd,w) in enumerate(hdrs2,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    digital = [
        ("Mobile Banking App","41%","67%","-26pp",18.4,"3.2/5.0 (3.8 peer)","Flat","App rated poorly for UX. Slow load times. Limited features vs peers.","Q2 UX redesign sprint. Target: 4.0 rating, 60% adoption by Q4 2026."),
        ("Online Bill Pay","52%","74%","-22pp",8.2,"3.6/5.0","Slow improvement","FIS integration causes latency. Pending payments show as failed.","FIS API improvement. Real-time payment status. FedNow integration."),
        ("eStatements","61%","82%","-21pp",3.4,"4.1/5.0","Improving","Onboarding process doesn't default to eStatements. Easy fix.","Default new customers to eStatements. Opt-out vs opt-in."),
        ("Digital Account Opening","28%","58%","-30pp",12.6,"2.9/5.0","Flat","Requires branch visit for verification. Peers use biometric verification.","Digital identity verification. Remove branch requirement."),
        ("Person-to-Person Payments","19%","51%","-32pp",6.8,"3.1/5.0","Flat","No Zelle partnership. Competitors all have P2P. Major gap.","Zelle partnership or FedNow P2P capability."),
        ("Business Banking Digital","34%","62%","-28pp",14.2,"3.4/5.0","Flat","Business portal outdated. Manual processes for common requests.","Business portal redesign. API banking for SMB customers."),
        ("OVERALL DIGITAL CHANNEL","41%","67%","-26pp",63.6,"3.2/5.0 avg","Flat","See individual products above.","Digital transformation programme. 18-month roadmap."),
    ]
    for r, row in enumerate(digital, 3):
        ws2.row_dimensions[r].height = 24
        product, current, bench, gap, risk, sat, trend, cause, init = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        sf = fills["prp"] if "OVERALL" in product else sf
        c_val = float(current.replace("%",""))/100 if "%" in current else 0
        b_val = float(bench.replace("%",""))/100 if "%" in bench else 0
        cf = fills["red"] if c_val < 0.45 else fills["amb"] if c_val < 0.60 else fills["grn"]
        cel(ws2,r,1,product,sf,bold="OVERALL" in product,bdr=bdr)
        cel(ws2,r,2,current,cf,align="center",bdr=bdr); cel(ws2,r,3,bench,sf,align="center",bdr=bdr)
        cel(ws2,r,4,gap,fills["red"],align="center",bdr=bdr)
        cel(ws2,r,5,risk,fills["red"],"$#,##0.0",align="center",bdr=bdr)
        sat_f = fills["red"] if float(sat.split("/")[0])<3.5 else fills["amb"] if float(sat.split("/")[0])<4.0 else fills["grn"]
        cel(ws2,r,6,sat,sat_f,align="center",bdr=bdr)
        tf = fills["red"] if "Flat" in trend else fills["amb"] if "slow" in trend.lower() else fills["grn"]
        cel(ws2,r,7,trend,tf,align="center",bdr=bdr)
        cel(ws2,r,8,cause,sf,bdr=bdr); cel(ws2,r,9,init,sf,bdr=bdr)
    save(wb2, f"{base}/firstcapital/tech/FC02_Digital_Adoption_Analytics.xlsx")
    results.append("FC02 done")

    return results


# ══════════════════════════════════════════════════════════════════════════
#  APEX RETAIL — Core datasets
# ══════════════════════════════════════════════════════════════════════════

def apex_files(base):
    results = []

    # AP01 — Operations Dashboard
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Operations Dashboard"
    title_row(ws, 9, "APEX RETAIL GROUP — Operations & Technology Dashboard FY2025", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Metric",26),("Current",13),("Benchmark",13),("Target",13),
            ("Gap",11),("Financial Impact",18),("Technology\nRoot Cause",24),
            ("Priority",12),("Notes",28)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)
    metrics = [
        ("Operating Margin","3.8%","6.0%","6.0%","-2.2pp","$273M gap to target","SAP ECC inefficiency + Salesforce Einstein not activated","Critical","Revenue $12.4B. Target margin requires both cost reduction and AI-driven revenue."),
        ("E-commerce Revenue %","28%","45%","42%","-17pp","$2.1B revenue gap","Digital platform not AI-personalised. Einstein licensed not activated.","Critical","$248M Einstein AI spend. 18 months deployed. 23% adoption. Zero personalisation active."),
        ("Demand Forecast Accuracy","62%","84%","80%","-22pp","$180M excess inventory","o9 demand planning 40% implemented. SAP data quality poor.","Critical","$900M inventory turns at 4.2x vs 6.8x benchmark. $180M trapped capital."),
        ("Shrinkage Rate","2.8%","1.4%","1.6%","+1.4pp","$347M annual loss","No AI-powered loss prevention. Manual exception reporting.","Critical","Industry median 1.4%. $347M = 2.8% of $12.4B. AI loss prevention could recover $180M+."),
        ("Salesforce Einstein Adoption","23%","N/A","80%","-57pp","$248M investment idle","Platform deployed but not configured. No AI models activated.","Critical","18 months since deployment. Personalisation revenue opportunity $248M not captured."),
        ("o9 Demand Planning Implementation","40%","100%","100%","-60pp","$180M forecast cost","SAP data quality prevents o9 from receiving clean data.","High","Only 40% of SKUs in o9. Rest still manual. Forecast accuracy reflects this."),
        ("SAP ECC Version","ECC 6.0 (2011)","S/4HANA (peers)","S/4HANA","14 years old","Migration cost $120-180M if delayed further","EOL 2027. No S/4 migration started.","Critical","EOL 2027 announced. No migration programme. Every quarter of delay adds cost."),
        ("Store Labour Productivity","$284/hr","$312/hr","$300/hr","-$28/hr","$38M annual gap","No AI-assisted scheduling. Manual labour planning.","High","AI scheduling (Einstein) not activated. Manual process."),
        ("Online Order Fulfilment Cost","$8.40","$6.20","$6.80","+$2.20","$44M annual excess","No AI routing optimisation. Einstein not configured for fulfilment.","High","E-commerce margin -2.1% because fulfilment cost exceeds online gross margin."),
        ("Supplier On-Time Delivery","71%","84%","82%","-13pp","$28M stockout cost","SAP supplier portal manual. No AI prediction.","Medium","Stockouts cost $28M in lost sales and emergency sourcing."),
    ]
    for r, row in enumerate(metrics, 3):
        ws.row_dimensions[r].height = 28
        metric, current, bench, target, gap, impact, tech, priority, notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        pf = fills["red"] if priority=="Critical" else fills["amb"] if priority=="High" else fills["grn"]
        cel(ws,r,1,metric,sf,bold=True,bdr=bdr); cel(ws,r,2,current,sf,align="center",bdr=bdr)
        cel(ws,r,3,bench,sf,align="center",bdr=bdr); cel(ws,r,4,target,sf,align="center",bdr=bdr)
        cel(ws,r,5,gap,fills["red"],align="center",bdr=bdr)
        cel(ws,r,6,impact,fills["red"],align="center",bdr=bdr)
        cel(ws,r,7,tech,sf,bdr=bdr)
        cp = cel(ws,r,8,priority,pf,align="center",bold=True,bdr=bdr)
        cel(ws,r,9,notes,sf,bdr=bdr)
    save(wb, f"{base}/apex/margin/AP01_Operations_Dashboard.xlsx")
    results.append("AP01 done")

    # AP02 — Salesforce Einstein Activation Status
    wb2 = Workbook(); bdr, fills = styles()
    ws2 = wb2.active; ws2.title = "Einstein Status"
    title_row(ws2, 7, "APEX RETAIL — Salesforce Einstein Activation Status (18 Months Post-Deployment)", fills)
    ws2.row_dimensions[2].height = 32
    hdrs2 = [("Einstein Module",26),("Licensed?",12),("Activated?",12),("Adoption %",12),
             ("Revenue Opportunity ($M)",18),("Blocker",28),("Priority",12),("Time to Activate",14)]
    for i,(hd,w) in enumerate(hdrs2,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    einstein = [
        ("Einstein Personalisation (Commerce)","Yes","No","0%",142,"No configuration. No product catalogue sync. No customer data model.","Critical","8 weeks with Maestro"),
        ("Einstein Recommendations","Yes","No","0%",48,"Depends on personalisation activation. Sequential.","Critical","10 weeks (after personalisation)"),
        ("Einstein Search","Yes","Partial","18%",22,"Partial activation. No AI ranking. Rules-based only.","High","4 weeks"),
        ("Einstein Next Best Action","Yes","No","0%",18,"No decision logic configured. Sales team not trained.","High","6 weeks"),
        ("Einstein Forecasting","Yes","No","0%",12,"Salesforce data quality insufficient. Clean data required.","High","12 weeks (data prep first)"),
        ("Einstein Bots (Service)","Yes","Partial","12%",8,"2 of 8 intents configured. Bot handles 12% of service queries.","Medium","8 weeks to full activation"),
        ("Einstein Vision (Loss Prevention)","Yes","No","0%",16,"Camera infrastructure not connected. IT project required.","Medium","16 weeks (IT infrastructure)"),
        ("Einstein Analytics (Tableau CRM)","Yes","Partial","34%","N/A","Some dashboards built. Predictive features not enabled.","Medium","4 weeks"),
        ("TOTAL UNACTIVATED OPPORTUNITY","—","—","23% avg","$248M+","See individual modules above","—","10-16 weeks phased"),
    ]
    for r, row in enumerate(einstein, 3):
        ws2.row_dimensions[r].height = 24
        module, licensed, activated, adoption, revenue, blocker, priority, timeline = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        sf = fills["prp"] if "TOTAL" in module else sf
        lf = fills["grn"] if licensed=="Yes" else fills["red"]
        af = fills["red"] if activated=="No" else fills["amb"] if activated=="Partial" else fills["grn"]
        cel(ws2,r,1,module,sf,bold="TOTAL" in module,bdr=bdr)
        cel(ws2,r,2,licensed,lf,align="center",bdr=bdr)
        cel(ws2,r,3,activated,af,align="center",bdr=bdr)
        adp_str = adoption.split("%")[0].strip().split()[0] if "%" in adoption else "0"
        try:
            adp = int(adp_str)/100
        except ValueError:
            adp = 0
        af2 = fills["red"] if adp<0.25 else fills["amb"] if adp<0.5 else fills["grn"]
        cel(ws2,r,4,adoption,af2,align="center",bdr=bdr)
        rev_f = fills["red"] if isinstance(revenue,(int,float)) and revenue>20 else fills["amb"] if isinstance(revenue,(int,float)) else fills["wht"]
        cel(ws2,r,5,revenue,rev_f,"$#,##0" if isinstance(revenue,(int,float)) else None,align="center",bdr=bdr)
        cel(ws2,r,6,blocker,fills["red"] if activated=="No" else sf,bdr=bdr)
        pf = fills["red"] if priority=="Critical" else fills["amb"] if priority=="High" else fills["grn"] if priority=="Medium" else fills["wht"]
        cel(ws2,r,7,priority,pf,align="center",bdr=bdr)
        cel(ws2,r,8,timeline,sf,align="center",bdr=bdr)
    save(wb2, f"{base}/apex/margin/AP02_Salesforce_Einstein_Status.xlsx")
    results.append("AP02 done")

    return results


# ══════════════════════════════════════════════════════════════════════════
#  PROCESS WRAPPER — runs one function safely
# ══════════════════════════════════════════════════════════════════════════

def run_task(args):
    func, base = args
    try:
        result = func(base)
        if isinstance(result, list):
            return result
        return [result]
    except Exception as e:
        return [f"ERROR in {func.__name__}: {str(e)}\n{traceback.format_exc()}"]


# ══════════════════════════════════════════════════════════════════════════
#  MAIN — parallel execution
# ══════════════════════════════════════════════════════════════════════════


import multiprocessing as mp
import os, sys, time, traceback
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def styles():
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "hdr": PatternFill("solid", start_color="1A3A5C"),
        "alt": PatternFill("solid", start_color="F2F7FC"),
        "wht": PatternFill("solid", start_color="FFFFFF"),
        "red": PatternFill("solid", start_color="FDE8E8"),
        "amb": PatternFill("solid", start_color="FFF4E5"),
        "grn": PatternFill("solid", start_color="E8F5E9"),
        "blu": PatternFill("solid", start_color="E8F0F8"),
        "prp": PatternFill("solid", start_color="F0EEFF"),
        "teal":PatternFill("solid", start_color="E0F7F4"),
    }
    return bdr, fills

def h(ws, row, col, val, width=16, fills=None, bdr=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    c.fill = fills["hdr"]
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = bdr
    ws.column_dimensions[get_column_letter(col)].width = width

def cel(ws, row, col, val, fill=None, fmt=None, bold=False, align="left", bdr=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(size=9, name="Arial", bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = bdr
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    return c

def title_row(ws, cols, text, fills):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    ws["A1"].value = text
    ws["A1"].font  = Font(bold=True, size=12, color="1A3A5C", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill  = fills["blu"]
    ws.row_dimensions[1].height = 16

def save(wb, path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

BASE = Path(__file__).parent

# ══════════════════════════════════════════════════════════════════
# MH-P01: Engineering Org & Squad Structure
# ══════════════════════════════════════════════════════════════════
def mh_p01_engineering_org(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Squad Overview"
    title_row(ws, 14, "MERIDIAN HEALTH SYSTEM — Technology Engineering Organisation", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Squad",26),("Domain",18),("Squad Lead",18),("FTE",8),
            ("Contractors",10),("Vendor Staff",10),("Total",8),
            ("Vendor Dep %",12),("Primary Vendor",16),
            ("Mtg Hrs/Wk",10),("Build Hrs/Wk",10),
            ("Sprint Method",13),("Cycle Time (Days)",14),("Status",10)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    squads = [
        ("Epic EHR Core","Clinical Systems","Dr. Sarah Kim (CMIO - interim)",12,2,4,"Epic Systems",18,22,"Scrum",94,"Amber","Epic 80% utilized benchmark. Only 58/100 at Meridian. 22 modules underused."),
        ("Revenue Cycle Technology","Revenue Cycle","James Park (CFO - interim)",6,1,8,"Ensemble Health",24,16,"Scrum",127,"Critical","RCM vendor Ensemble effectively runs this squad. Internal team cannot modify RCM logic."),
        ("Prior Auth & Payer Integration","Revenue Cycle / Clinical","-VACANT-",4,2,6,"Epic / Cohere pilot",26,14,"Kanban",145,"Critical","Lead role vacant 3 months. CMS mandate January 2027. 14 months. No automation started."),
        ("Patient Engagement & MyChart","Clinical / Digital","Linda Chen",5,1,3,"Epic",16,24,"Scrum",88,"Amber","MyChart 34% vs 60% target. Training completion 41%. Module underutilization root cause."),
        ("Clinical Data & Analytics","Data / Clinical","Dr. Marcus Webb",6,2,4,"Epic / Clarity",20,20,"Scrum",102,"Amber","Clarity reporting team. Manual extracts dominate. AI readiness low."),
        ("AI & Innovation Lab","Innovation","(CDO Search - Active)",3,1,6,"Various",22,18,"Agile",0,"Critical","CDO role being recruited. Lab effectively leaderless. $94M AI portfolio has no owner."),
        ("Infrastructure & Cloud","Technology","Amir Singh",5,2,2,"AWS",14,26,"Kanban",45,"Green","Cloud migration 60% complete. AWS well-integrated. Solid internal team."),
        ("Cybersecurity & Compliance","Technology / Compliance","Robert Torres",4,1,1,"Optiv",12,28,"Kanban",0,"Green","HIPAA compliance strong. Security posture solid."),
        ("Integration & Interoperability","Clinical Systems","Priya Nair",5,2,4,"Epic / Rhapsody",22,18,"Scrum",118,"Amber","HL7 and FHIR integration team. 14 payer integrations. Prior auth integration pending."),
        ("Supply Chain Technology","Operations","David Kim",3,1,2,"Infor",18,22,"Scrum",96,"Amber","Infor SCM system. Procurement analytics manual."),
        ("Workforce Technology","HR / Operations","Maria Santos",3,1,1,"Workday",14,26,"Scrum",72,"Green","Workday well-implemented. Travel nurse tracking manual outside system."),
        ("Telehealth & Virtual Care","Clinical / Digital","Dr. James Liu",4,1,3,"Epic / Zoom",20,20,"Scrum",88,"Amber","Telehealth volume 18% of visits. Epic integration partial."),
    ]

    status_fills = {"Critical":fills["red"],"Amber":fills["amb"],"Green":fills["grn"]}
    for r, (sq,domain,lead,fte,contr,consult,vendor,meet,build,method,cycle,status,notes) in enumerate(squads,3):
        ws.row_dimensions[r].height = 28
        sf = fills["alt"] if r%2==0 else fills["wht"]
        total = fte+contr+consult
        vdr = round((contr+consult)/total*100,1)
        cel(ws,r,1,sq,sf,bold=True,bdr=bdr); cel(ws,r,2,domain,sf,bdr=bdr)
        cel(ws,r,3,lead,fills["red"] if "VACANT" in lead else sf,bdr=bdr)
        cel(ws,r,4,fte,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,5,contr,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,6,consult,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,7,f"=D{r}+E{r}+F{r}",sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,8,f"=(E{r}+F{r})/G{r}",sf,"0.0%",align="center",bdr=bdr)
        cel(ws,r,9,vendor,sf,bdr=bdr)
        cel(ws,r,10,meet,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,11,build,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,12,method,sf,align="center",bdr=bdr)
        cel(ws,r,13,cycle if cycle else "Not tracked",sf,align="center",bdr=bdr)
        cs = cel(ws,r,14,status,status_fills.get(status,sf),align="center",bold=True,bdr=bdr)
        cs.font = Font(size=9,name="Arial",bold=True,color="FFFFFF" if status=="Critical" else "333333")

    # DORA sheet
    ws2 = wb.create_sheet("DORA Metrics")
    title_row(ws2, 8, "MERIDIAN HEALTH — DORA Metrics by Squad (FY2025)", fills)
    ws2.row_dimensions[2].height = 32
    dora_hdrs = [("Squad",26),("Deploy Freq",14),("Lead Time (Days)",14),
                 ("Change Fail %",13),("MTTR (Hrs)",12),("DORA Level",12),
                 ("Blocker",28),("Gap vs Elite",14)]
    for i,(hd,w) in enumerate(dora_hdrs,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    dora = [
        ("Epic EHR Core","Monthly","94","4.2%","8","Low","Epic upgrade windows control release schedule","Far below — Elite is daily"),
        ("Revenue Cycle Technology","Bi-monthly","127","8.1%","16","Low","Ensemble Health governs RCM releases","Far below"),
        ("Prior Auth & Payer Integration","Ad hoc","145","12.4%","24","None","No CI/CD. Manual deployments. Lead role vacant.","No comparison — zero deployments"),
        ("Patient Engagement","Monthly","88","3.1%","6","Medium","Some automation. MyChart upgrades Epic-controlled.","Below — target weekly"),
        ("Clinical Data & Analytics","Quarterly","102","5.8%","12","Low","Manual data pipeline. No deployment automation.","Far below"),
        ("AI & Innovation Lab","Never","N/A","N/A","N/A","None","CDO vacant. Zero production AI deployments in 12 months.","No comparison"),
        ("Infrastructure & Cloud","Weekly","45","1.8%","3","Medium","Good infra pipeline. ML pipeline absent.","Near — target daily"),
        ("Cybersecurity","On-demand","28","0.9%","2","High","Strong. Security patches automated.","Near elite"),
        ("Integration & Interoperability","Monthly","118","6.2%","14","Low","Epic governs integration release windows.","Below"),
        ("Supply Chain Technology","Monthly","96","4.4%","8","Low","Infor controls upgrade schedule.","Below"),
        ("Workforce Technology","Bi-weekly","72","2.1%","4","Medium","Workday release cadence manageable internally.","Near"),
        ("Telehealth","Monthly","88","3.8%","7","Medium","Epic integration partial — constrains velocity.","Below"),
    ]
    level_fills = {"High":fills["grn"],"Medium":fills["amb"],"Low":fills["red"],"None":fills["red"]}
    for r,row in enumerate(dora,3):
        ws2.row_dimensions[r].height = 22
        sq,freq,lt,cfr,mttr,lvl,blocker,gap = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws2,r,1,sq,sf,bold=True,bdr=bdr); cel(ws2,r,2,freq,sf,align="center",bdr=bdr)
        cel(ws2,r,3,lt,sf,align="center",bdr=bdr); cel(ws2,r,4,cfr,sf,align="center",bdr=bdr)
        cel(ws2,r,5,mttr,sf,align="center",bdr=bdr)
        cl = cel(ws2,r,6,lvl,level_fills.get(lvl,sf),align="center",bold=True,bdr=bdr)
        cel(ws2,r,7,blocker,sf,bdr=bdr); cel(ws2,r,8,gap,sf,bdr=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/pdlc/MH-P01_Engineering_Organisation.xlsx")
    return "MH-P01 done"


# ══════════════════════════════════════════════════════════════════
# MH-P02: Sprint Velocity & Delivery Metrics
# ══════════════════════════════════════════════════════════════════
def mh_p02_sprint_velocity(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Sprint Velocity"
    title_row(ws, 14, "MERIDIAN HEALTH — Sprint Velocity by Squad (Apr 2025 – Mar 2026)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Squad",26),("Quarter",10),("Planned Pts",11),("Completed Pts",11),
            ("Velocity %",10),("Unplanned %",11),("Carryover %",11),
            ("Deployments",11),("Incidents",10),("Vendor Blocked",12),
            ("AI Initiative Blocked",14),("RAG",8),("Notes",32)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    sprints = [
        # Epic EHR Core — stable but Epic-constrained
        ("Epic EHR Core","Q1 2025",55,48,None,18,14,1,1,2,0,"Amber","Epic upgrade Q1 consumed 3 sprints. Internal feature delivery paused."),
        ("Epic EHR Core","Q2 2025",52,44,None,22,18,1,2,3,0,"Red","Epic upgrade delayed. 3 payer integration fixes required mid-sprint."),
        ("Epic EHR Core","Q3 2025",58,51,None,16,12,2,1,2,0,"Amber","Better quarter. MyChart module work progressing."),
        ("Epic EHR Core","Q4 2025",54,47,None,19,15,1,1,2,0,"Amber","Year-end freeze reduced capacity. Epic controls Q4 window."),
        # Revenue Cycle — Ensemble-dominated
        ("Revenue Cycle Technology","Q1 2025",38,26,None,32,28,0,3,5,0,"Red","Ensemble Health SLA breach Q1. $2.1M in denied claims traced to integration bug. Ensemble fixed after 6 weeks."),
        ("Revenue Cycle Technology","Q2 2025",35,24,None,34,30,0,4,6,0,"Red","Denial rate worsening. Ensemble contract renegotiation underway. Squad morale low."),
        ("Revenue Cycle Technology","Q3 2025",36,26,None,30,26,0,3,5,0,"Red","Ensemble vendor fix deployed. Partial recovery. Internal team still cannot modify RCM logic."),
        ("Revenue Cycle Technology","Q4 2025",34,23,None,35,31,0,4,5,0,"Red","Ensemble SLA penalty $1.8M claimed but disputed. Denial rate 18.2% at year end."),
        # AI Lab — zero deliveries
        ("AI & Innovation Lab","Q1 2025",28,16,None,42,38,0,0,0,6,"Red","6 AI initiatives stalled. CDO search started. No sponsor for AI deployments."),
        ("AI & Innovation Lab","Q2 2025",25,12,None,48,44,0,0,0,8,"Red","Prior auth pilot failed integration testing. CDO search month 4. No production deployments."),
        ("AI & Innovation Lab","Q3 2025",22,10,None,52,48,0,0,0,9,"Red","GenAI documentation pilot positive feedback but no enterprise deployment approved without CDO."),
        ("AI & Innovation Lab","Q4 2025",20,9,None,55,51,0,0,0,10,"Red","10 items blocked for CDO approval. Zero production AI. $94M portfolio idle."),
        # Prior Auth — leaderless
        ("Prior Auth & Payer Integration","Q1 2025",24,14,None,40,36,0,2,0,0,"Red","Lead vacant month 1. CMS prior auth mandate January 2027 — 21 months. No automation started."),
        ("Prior Auth & Payer Integration","Q2 2025",20,11,None,44,40,0,3,0,0,"Red","Cohere Health pilot proposal approved but no project lead to run it."),
        ("Prior Auth & Payer Integration","Q3 2025",18,10,None,46,42,0,2,0,0,"Red","Interim contractor assigned. 18 months to CMS deadline. Electronic prior auth design not started."),
        ("Prior Auth & Payer Integration","Q4 2025",16,9,None,48,44,0,3,0,0,"Red","14 months to CMS mandate. Electronic prior auth — zero progress. $37.6M value idle."),
        # Patient Engagement — improving slowly
        ("Patient Engagement & MyChart","Q1 2025",42,36,None,14,10,3,1,0,0,"Amber","MyChart adoption 34%. Module rollout programme started Q1."),
        ("Patient Engagement & MyChart","Q2 2025",44,38,None,12,8,4,1,0,1,"Amber","MyChart 35%. 1 AI feature blocked (sentiment analysis — no CDO)."),
        ("Patient Engagement & MyChart","Q3 2025",46,41,None,10,6,4,0,0,0,"Green","Best quarter. Patient portal redesign launched. Adoption 36%."),
        ("Patient Engagement & MyChart","Q4 2025",44,39,None,12,8,3,1,0,1,"Amber","Adoption 37%. Progress slow — target 60%. 1 AI feature still blocked."),
        # Infrastructure — best performing
        ("Infrastructure & Cloud","Q1 2025",60,56,None,8,4,8,1,0,0,"Green","Strong. Cloud migration progressing. AWS Bedrock POC started."),
        ("Infrastructure & Cloud","Q2 2025",62,58,None,7,3,9,0,0,0,"Green","Excellent. No ML pipeline yet but infra for it being designed."),
        ("Infrastructure & Cloud","Q3 2025",64,60,None,6,3,10,1,0,0,"Green","Best squad. ML pipeline scoped — waiting on AI Lab direction."),
        ("Infrastructure & Cloud","Q4 2025",62,57,None,8,4,8,1,0,0,"Green","Solid. Year-end good. AWS Bedrock access configured for GenAI."),
    ]

    rag_fills = {"Green":fills["grn"],"Amber":fills["amb"],"Red":fills["red"]}
    for r,row in enumerate(sprints,3):
        ws.row_dimensions[r].height = 24
        sq,qtr,planned,completed,_,unpl,carry,deploys,incidents,vblocked,aiblocked,rag,notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws,r,1,sq,sf,bold=True,bdr=bdr); cel(ws,r,2,qtr,sf,align="center",bdr=bdr)
        cel(ws,r,3,planned,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,4,completed,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,5,f"=D{r}/C{r}",sf,"0.0%",align="center",bdr=bdr)
        cel(ws,r,6,unpl/100,sf,"0.0%",align="center",bdr=bdr)
        cel(ws,r,7,carry/100,sf,"0.0%",align="center",bdr=bdr)
        cel(ws,r,8,deploys,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,9,incidents,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,10,vblocked,fills["red"] if vblocked>=4 else fills["amb"] if vblocked>0 else sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,11,aiblocked,fills["red"] if aiblocked>=6 else fills["amb"] if aiblocked>0 else sf,"#,##0",align="center",bdr=bdr)
        cs = cel(ws,r,12,rag,rag_fills.get(rag,sf),align="center",bold=True,bdr=bdr)
        cs.font = Font(size=9,name="Arial",bold=True,color="FFFFFF" if rag=="Red" else "333333")
        cel(ws,r,13,notes,sf,bdr=bdr)

    ws.freeze_panes = "C3"
    save(wb, f"{base}/meridian/pdlc/MH-P02_Sprint_Velocity.xlsx")
    return "MH-P02 done"


# ══════════════════════════════════════════════════════════════════
# MH-P03: AI Initiative Inventory (comprehensive)
# ══════════════════════════════════════════════════════════════════
def mh_p03_ai_initiatives(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "AI Initiatives"
    title_row(ws, 10, "MERIDIAN HEALTH SYSTEM — AI Initiative Inventory (FY2026)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("ID",7),("Initiative",30),("Sponsor",18),("Status",14),
            ("Budget ($M)",12),("Spent ($M)",12),("ROI Verified ($M)",14),
            ("Data Readiness",13),("Genome Pattern",18),("Blocker",24),
            ("Annual Value ($M)",14),("Recommendation",30)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    initiatives = [
        ("AI-001","Prior Auth Automation (Cohere Health)","Robert Chen (CIO)","Pilot — CMS mandate pressing",8.4,6.2,0.0,"62%","F006 (79%) — no MLOps infra","Epic integration 60% complete. No deployment pipeline.",37.6,"PRIORITY 1. $37.6M value. 91% Genome confidence. Cohere Health scored 94/100. CMS mandate January 2027."),
        ("AI-002","GenAI Clinical Documentation","Dr. Sarah Kim (CMIO)","Pilot — physician demand high",4.2,2.8,0.0,"58%","F008 (61%) — adoption risk","Pilot positive. 2 departments. No enterprise deployment plan. CDO needed.",42.0,"PRIORITY 2. $42M productivity. 68% of physicians cite documentation burden. AWS Bedrock ready."),
        ("AI-003","CDO Executive Hire","Emily Rodriguez (CEO)","Search active — month 4",0.4,0.2,0.0,"N/A","F002 (84%) — no sponsor","Search ongoing. Interim coverage by CIO.",94.0,"CRITICAL enabler. $94M in blocked value. Every month of delay costs the portfolio."),
        ("AI-004","Epic Optimization Sprint","Robert Chen (CIO)","Planning — not started",2.1,0.4,0.0,"68%","F008 (61%) — training gap","22 modules underutilized. Training completion 41%. No MLOps needed.",12.0,"HIGH. $12M annual value. 8-week sprint. Internal capability sufficient with Maestro."),
        ("AI-005","RCM Vendor Renegotiation (Ensemble)","James Park (CFO)","Active negotiation",0.2,0.2,0.0,"N/A","F011 (74%) — vendor misalignment","$8M SLA penalties enforceable. Ensemble underperforming on denial rate.",8.0,"ACT NOW. Leverage SLA breach. Cohere Health as competitive threat in negotiation."),
        ("AI-006","Predictive Readmission Model","Dr. Sarah Kim (CMIO)","Concept — data not ready",1.8,0.2,0.0,"32%","F003 (68%) — data readiness","Epic data quality 58/100. Insufficient clean data for model training. Block until Epic optimized.",6.0,"BLOCK. Prerequisite: Epic Optimization Sprint. Then fast to deploy with GenAI."),
        ("AI-007","Travel Nurse Demand Forecasting","VP HR","In Development",1.2,0.8,0.0,"52%","F003 (68%)","Workforce data in 3 systems. Workday, manual spreadsheets, agency portal.",8.0,"MEDIUM. $8M savings. Workforce data consolidation required first."),
        ("AI-008","MA Star Rating AI","Emily Rodriguez (CEO)","Concept",3.4,0.4,0.0,"44%","F002 (84%) — no CDO","CDO hire prerequisite. Multiple quality measures need coordinated AI overlay.",24.0,"BLOCK until CDO. $24M revenue at risk if star rating drops below 3.5."),
        ("AI-009","Clinical Coding Automation","James Park (CFO)","Pilot",2.8,1.9,0.0,"61%","F006 (79%)","Coding AI pilot with 3M HIS. Accuracy 84% vs 96% required. Not production-ready.",11.0,"CONTINUE PILOT. Target 96% accuracy before enterprise deployment."),
        ("AI-010","Patient No-Show Prediction","Robert Chen (CIO)","In Development",0.9,0.6,0.0,"71%","F008 (61%)","Model trained on Epic scheduling data. Deployment blocked — no serving infrastructure.",4.0,"QUICK WIN once MLOps infra live. Model ready. Infrastructure gap only."),
        ("AI-011","Supply Chain AI (Stockout Prediction)","COO","Concept",1.4,0.2,0.0,"38%","F003 (68%)","Supply chain data in Infor and manual. No unified feed.",3.0,"LOW priority. Data consolidation required. Defer to Wave 2."),
        ("AI-012","Telehealth Triage AI","Dr. Sarah Kim (CMIO)","Concept",2.2,0.3,0.0,"42%","F002 (84%)","CDO required for clinical AI governance. No framework.",8.0,"BLOCK until CDO and clinical AI governance established."),
    ]

    status_fills = {
        "Pilot — CMS mandate pressing":fills["red"],
        "Pilot — physician demand high":fills["red"],
        "Search active — month 4":fills["amb"],
        "Active negotiation":fills["amb"],
        "Planning — not started":fills["amb"],
        "In Development":fills["amb"],
        "Concept":fills["wht"],
        "Pilot":fills["amb"],
    }

    for r,row in enumerate(initiatives,3):
        ws.row_dimensions[r].height = 36
        id_,name,sponsor,status,budget,spent,roi,data_r,genome,blocker,value,rec = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws,r,1,id_,sf,align="center",bdr=bdr)
        cel(ws,r,2,name,sf,bold=True,bdr=bdr)
        cel(ws,r,3,sponsor,sf,bdr=bdr)
        cs = cel(ws,r,4,status,status_fills.get(status,sf),bdr=bdr)
        cel(ws,r,5,budget,sf,"$#,##0.0",align="center",bdr=bdr)
        cel(ws,r,6,spent,sf,"$#,##0.0",align="center",bdr=bdr)
        cel(ws,r,7,roi,fills["red"],"$#,##0.0",align="center",bdr=bdr)
        dr_val = int(data_r.replace("%","")) if "%" in data_r else 0
        dr_f = fills["red"] if dr_val<40 else fills["amb"] if dr_val<65 else fills["grn"]
        cel(ws,r,8,data_r,dr_f,align="center",bdr=bdr)
        cel(ws,r,9,genome,fills["amb"],bdr=bdr)
        cel(ws,r,10,blocker,fills["red"] if "BLOCK" in rec or "no" in blocker.lower() else sf,bdr=bdr)
        cel(ws,r,11,value,fills["grn"],"$#,##0.0",align="center",bold=True,bdr=bdr)
        cel(ws,r,12,rec,sf,bdr=bdr)

    # Summary
    tr = len(initiatives)+3
    cel(ws,tr,1,"TOTALS",bold=True,bdr=bdr)
    ws.cell(tr,5,value=f"=SUM(E3:E{tr-1})"); ws.cell(tr,5).number_format="$#,##0.0"
    ws.cell(tr,6,value=f"=SUM(F3:F{tr-1})"); ws.cell(tr,6).number_format="$#,##0.0"
    ws.cell(tr,7,value=0); ws.cell(tr,7).number_format="$#,##0.0"
    ws.cell(tr,11,value=f"=SUM(K3:K{tr-1})"); ws.cell(tr,11).number_format="$#,##0.0"
    for c in [5,6,7,11]:
        ws.cell(tr,c).font=Font(bold=True,size=10,name="Arial"); ws.cell(tr,c).alignment=Alignment(horizontal="center")

    ws.freeze_panes = "C3"
    save(wb, f"{base}/meridian/pdlc/MH-P03_AI_Initiative_Inventory.xlsx")
    return "MH-P03 done"


# ══════════════════════════════════════════════════════════════════
# MH-P04: Technology Landscape
# ══════════════════════════════════════════════════════════════════
def mh_p04_technology_landscape(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "System Inventory"
    title_row(ws, 11, "MERIDIAN HEALTH SYSTEM — Technology Landscape & System Inventory", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("System",24),("Vendor",18),("Category",16),("Age (Yrs)",9),
            ("Annual Cost ($M)",14),("Data Domain",18),("Integrations",10),
            ("EOL Status",16),("AI Ready?",10),("Internal Capability",14),("Notes",32)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    systems = [
        ("Epic EHR","Epic Systems","Electronic Health Record",11,18.4,"Clinical, Orders, Notes, Scheduling, Billing",28,"Supported — quarterly upgrades","Partial","High — large team","Core clinical platform. 58/100 optimization vs 80 benchmark. 22 modules underutilized. FHIR R4 supported."),
        ("Ensemble Health Partners","Ensemble Health","Revenue Cycle Management",4,14.2,"Claims, Denials, Collections, AR",12,"Supported","No","Low — vendor-dependent","RCM outsourced. Denial rate 18.2% vs 12% benchmark. $8M SLA penalties enforceable. Contract renegotiation active."),
        ("Epic Clarity / Caboodle","Epic Systems","Clinical Data Warehouse",11,2.1,"Clinical reporting, analytics, population health",8,"Supported — linked to Epic","Partial","Medium","Clarity reporting team. Manual extracts dominate. AI training data readiness 58/100."),
        ("Cerner (legacy, 2 hospitals)","Oracle Health","Electronic Health Record (Legacy)",14,3.8,"Clinical — legacy hospitals",6,"EOL — migration to Epic by Q4 2026","No","Low — being retired","2 community hospitals still on Cerner. Epic migration in progress. Data migration complexity high."),
        ("Infor Cloverleaf","Infor","Integration Engine",8,1.2,"HL7 messaging, ADT, Lab, Radiology",42,"Supported","Partial","Medium","42 HL7 interfaces. Prior auth integration requires Cloverleaf enhancement. Well-maintained."),
        ("Cohere Health (Pilot)","Cohere Health","Prior Auth AI",1,0.4,"Prior authorization, payer communication",3,"Supported — pilot","Yes","Low — new vendor","Pilot with 3 payer contracts. 94/100 Genome vendor score. CMS mandate January 2027 makes this urgent."),
        ("3M HIS (Coding AI Pilot)","3M / Solventum","Clinical Coding AI",1,0.6,"Clinical coding, DRG assignment",2,"Supported — pilot","Yes","Low — new","Coding AI pilot. 84% accuracy vs 96% needed. Ongoing training with Meridian-specific data."),
        ("AWS (Cloud)","Amazon Web Services","Cloud Infrastructure",3,4.2,"Infrastructure, storage, compute",18,"Supported — strategic","Yes","High — growing team","AWS Bedrock configured. GenAI infrastructure ready. ML pipeline not yet built."),
        ("Workday","Workday","HRIS / Workforce",5,1.8,"HR, Payroll, Workforce planning",6,"Supported","Partial","High","Well-implemented. Travel nurse tracking manual outside system. Workforce analytics limited."),
        ("Strata Decision Technology","Strata","Financial Planning & Analytics",6,0.9,"Financial planning, cost accounting, budgeting",4,"Supported","No","Medium","CFO uses extensively. Not AI-integrated."),
        ("Press Ganey","Press Ganey","Patient Experience",8,0.6,"Patient satisfaction, HCAHPS, CG-CAHPS",2,"Supported","No","High — self-service","HCAHPS data feed into Epic. MA Star Rating dependency."),
        ("Nuance DAX (Pilot)","Nuance / Microsoft","AI Ambient Documentation",0.5,0.3,"Clinical documentation, physician notes",1,"Supported — pilot","Yes","Low — new","10 physician pilot. Excellent reception. Reduces documentation time 40%. Competing with GenAI option."),
        ("Salesforce Health Cloud","Salesforce","CRM / Care Coordination",2,1.4,"Patient outreach, care management, referrals",4,"Supported","Partial","Low — Wipro-dependent","Care coordination team. Wipro administers. Internal team limited. 38% adoption vs 70% target."),
    ]

    for r,row in enumerate(systems,3):
        ws.row_dimensions[r].height = 32
        name,vendor,cat,age,cost,domain,ints,eol,ai_r,cap,notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws,r,1,name,sf,bold=True,bdr=bdr); cel(ws,r,2,vendor,sf,bdr=bdr)
        cel(ws,r,3,cat,sf,bdr=bdr)
        af = fills["red"] if age>12 else fills["amb"] if age>8 else sf
        cel(ws,r,4,age,af,"#,##0",align="center",bdr=bdr)
        cel(ws,r,5,cost,sf,"$#,##0.0",align="center",bdr=bdr)
        cel(ws,r,6,domain,sf,bdr=bdr); cel(ws,r,7,ints,sf,"#,##0",align="center",bdr=bdr)
        eol_f = fills["red"] if "EOL" in eol else fills["amb"] if "legacy" in eol.lower() else sf
        cel(ws,r,8,eol,eol_f,bdr=bdr)
        ai_f = fills["grn"] if ai_r=="Yes" else fills["amb"] if ai_r=="Partial" else fills["red"]
        cel(ws,r,9,ai_r,ai_f,align="center",bdr=bdr)
        cap_f = fills["grn"] if "High" in cap else fills["amb"] if "Medium" in cap else fills["red"]
        cel(ws,r,10,cap,cap_f,bdr=bdr); cel(ws,r,11,notes,sf,bdr=bdr)

    # Epic Module Utilization sheet
    ws2 = wb.create_sheet("Epic Module Utilization")
    title_row(ws2, 6, "MERIDIAN HEALTH — Epic Module Utilization (22 Modules Assessed)", fills)
    ws2.row_dimensions[2].height = 32
    epic_hdrs = [("Module",26),("Licensed?",11),("Go-Live Date",13),("Utilization %",13),
                 ("Benchmark %",13),("Gap",10),("Annual Value Unrealized ($M)",18),("Notes",30)]
    for i,(hd,w) in enumerate(epic_hdrs,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    epic_modules = [
        ("EHR Core (Inpatient)","Yes","2014-06-01",94,95,-1,0.0,"Core well-utilized. Upgrade to latest version pending."),
        ("MyChart Patient Portal","Yes","2016-03-01",34,60,-26,4.2,"Major gap. 60% benchmark. Patient engagement revenue tied to adoption."),
        ("Epic Ambulatory","Yes","2015-09-01",78,88,-10,2.1,"Ambulatory workflows partially adopted. Templates not standardized."),
        ("Epic Beacon (Oncology)","Yes","2018-04-01",52,82,-30,1.4,"Oncology workflow adoption poor. Manual workarounds common."),
        ("Epic Beaker (Lab)","Yes","2016-01-01",88,92,-4,0.3,"Lab well-integrated. Minor optimization gaps."),
        ("Epic Radiant (Radiology)","Yes","2016-01-01",82,90,-8,0.6,"Radiology good. PACS integration could improve."),
        ("Epic Willow (Pharmacy)","Yes","2015-03-01",76,88,-12,0.8,"Pharmacy dispensing good. Clinical decision support underused."),
        ("Epic Cupid (Cardiology)","Yes","2019-07-01",44,78,-34,1.2,"Cardiology workflows manual. Significant optimization opportunity."),
        ("Epic Stork (OB/GYN)","Yes","2017-01-01",68,85,-17,0.7,"OB module partial. Labor and delivery well-adopted. Prenatal gaps."),
        ("Epic Kaleidoscope (Ophthalmology)","Yes","2020-03-01",38,72,-34,0.5,"Ophthalmology adoption poor. Staff prefer legacy workflows."),
        ("Epic Healthy Planet (Pop Health)","Yes","2021-01-01",22,65,-43,2.8,"Population health massively underutilized. MA Star Rating tied to this. Priority."),
        ("Epic Care Everywhere (HIE)","Yes","2018-06-01",61,80,-19,0.9,"Care Everywhere sharing good. Receiving external records less utilized."),
        ("Epic Tapestry (Health Plan)","Yes","2019-01-01",44,72,-28,3.1,"Health plan module for 187k MA lives. Underutilization directly impacts MA Star Rating."),
        ("Epic Cogito (Analytics)","Yes","2020-06-01",28,68,-40,1.8,"Analytics platform severely underused. Most analytics done in Clarity manually."),
        ("Epic Welcome (Check-in)","Yes","2021-03-01",42,75,-33,0.6,"Self-service check-in low adoption. Staff habit and patient literacy cited."),
        ("Epic Caboodle (Data Warehouse)","Yes","2021-01-01",38,70,-32,0.4,"Data warehouse underutilized. Manual Clarity queries dominate."),
        ("Epic Rover (Mobile)","Yes","2022-01-01",52,80,-28,0.7,"Mobile nursing workflows partial. Device availability cited."),
        ("Epic MyChart Bedside","Yes","2022-06-01",24,65,-41,0.3,"Bedside patient tablet low adoption. Workflow training gap."),
        ("Epic Cheers (CRM)","Yes","2023-01-01",18,60,-42,0.8,"Patient outreach CRM barely deployed. Marketing and care gap outreach opportunity."),
        ("Epic GPT Integration (AI)","Yes","2024-01-01",12,45,-33,2.4,"AI integration (ambient documentation) pilot only. Enterprise deployment opportunity."),
        ("Epic Prior Auth Automation","Yes","2024-06-01",8,50,-42,6.2,"Prior auth automation available in Epic. Not activated. CMS mandate January 2027. CRITICAL."),
        ("Epic Scheduling AI","Licensed","2024-03-01",0,40,-40,1.1,"Scheduling AI licensed but not activated. No-show prediction sits here."),
    ]
    for r,row in enumerate(epic_modules,3):
        ws2.row_dimensions[r].height = 22
        module,lic,golive,util,bench,gap,value,notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        uf = fills["red"] if util<40 else fills["amb"] if util<65 else fills["grn"]
        gf = fills["red"] if gap<-25 else fills["amb"] if gap<-10 else fills["grn"]
        vf = fills["red"] if value>2 else fills["amb"] if value>0.5 else sf
        cel(ws2,r,1,module,sf,bold=True,bdr=bdr); cel(ws2,r,2,lic,sf,align="center",bdr=bdr)
        cel(ws2,r,3,golive,sf,align="center",bdr=bdr)
        cel(ws2,r,4,util/100,uf,"0%",align="center",bdr=bdr)
        cel(ws2,r,5,bench/100,sf,"0%",align="center",bdr=bdr)
        cel(ws2,r,6,gap,gf,"#,##0",align="center",bdr=bdr)
        cel(ws2,r,7,value,vf,"$#,##0.0",align="center",bdr=bdr)
        cel(ws2,r,8,notes,sf,bdr=bdr)

    tr2 = len(epic_modules)+3
    cel(ws2,tr2,1,"TOTAL UNREALIZED VALUE",bold=True,bdr=bdr)
    ws2.cell(tr2,7,value=f"=SUM(G3:G{tr2-1})"); ws2.cell(tr2,7).number_format="$#,##0.0"
    ws2.cell(tr2,7).font=Font(bold=True,size=11,name="Arial",color="CC0000")
    ws2.cell(tr2,7).alignment=Alignment(horizontal="center")

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/pdlc/MH-P04_Technology_Landscape.xlsx")
    return "MH-P04 done"


# ══════════════════════════════════════════════════════════════════
# MH-P05: MLOps Infrastructure Assessment
# ══════════════════════════════════════════════════════════════════
def mh_p05_mlops(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "MLOps Assessment"
    title_row(ws, 7, "MERIDIAN HEALTH — MLOps Infrastructure Assessment", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("MLOps Capability",26),("Current State",28),("Maturity (0-5)",12),
            ("Initiatives Blocked",13),("Effort to Fix",14),
            ("AWS Bedrock Status",18),("Recommendation",32)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    caps = [
        ("Model Registry","No registry. Models tracked in spreadsheets by data scientists.",0,8,"Medium — 2 months","Not configured","Implement MLflow on AWS. Link to Bedrock model catalog."),
        ("Feature Store","No feature store. Each model rebuilds features independently.",0,6,"High — 3 months","Not configured","Start with Epic clinical features — highest demand across initiatives."),
        ("ML Pipeline / CI-CD","No ML CI-CD. Infrastructure pipeline exists (AWS CodePipeline) but not extended to ML.",0,8,"Medium — 2 months","Partially available","Extend existing CodePipeline. Add model validation gate. 2-month effort."),
        ("Model Serving","No serving layer. Models run ad-hoc on analyst workstations.",0,10,"Medium — 2 months","SageMaker available","Deploy SageMaker Endpoints. Containerise prior auth and GenAI models first."),
        ("Model Monitoring","No monitoring. No drift detection. No performance tracking.",0,6,"Medium — 2 months","CloudWatch available","Implement using CloudWatch + custom metrics. Prior auth model first."),
        ("Experiment Tracking","Partial — 2 data scientists use MLflow locally. No shared server.",1,5,"Low — 3 weeks","Not configured","Deploy shared MLflow server on EC2. Quick win — 3 weeks."),
        ("AWS Bedrock / GenAI","AWS Bedrock access configured. Foundation models available. No clinical fine-tuning.",2,3,"Low — in progress","Active — POC stage","BEST ASSET. GenAI documentation and prior auth both viable. Prioritise."),
        ("Data Pipeline for ML","Manual Epic Clarity extracts. No automated clinical data pipeline for ML.",0,9,"High — 4 months (Epic dependency)","Not configured","Epic Clarity automation + feature pipeline. Requires Epic cooperation."),
        ("HIPAA Compliance for AI","No AI-specific HIPAA framework. PHI in training data not governed.",0,6,"High — requires CDO + Legal","Not assessed","CDO appointment prerequisite. Legal review of AI data handling required."),
        ("Model Validation (Clinical)","No clinical AI validation framework. FDA/ONC requirements undefined.",0,5,"High — requires CDO + CMIO","Not assessed","Clinical AI governance framework required before production deployment."),
        ("Automated Retraining","No automated retraining. All models static after initial training.",0,4,"Medium — 3 months","Not configured","Implement after serving layer. Prior auth model will drift as payer rules change."),
    ]

    for r,row in enumerate(caps,3):
        ws.row_dimensions[r].height = 28
        cap,state,maturity,blocked,effort,aws,rec = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws,r,1,cap,sf,bold=True,bdr=bdr); cel(ws,r,2,state,fills["red"] if maturity<2 else sf,bdr=bdr)
        mf = fills["red"] if maturity<2 else fills["amb"] if maturity<4 else fills["grn"]
        cel(ws,r,3,maturity,mf,"#,##0",align="center",bold=True,bdr=bdr)
        cel(ws,r,4,blocked,fills["red"] if blocked>=6 else fills["amb"],"#,##0",align="center",bdr=bdr)
        ef = fills["red"] if "High" in effort else fills["amb"] if "Medium" in effort else fills["grn"]
        cel(ws,r,5,effort,ef,bdr=bdr)
        awsf = fills["grn"] if "Active" in aws else fills["amb"] if "available" in aws.lower() else fills["red"]
        cel(ws,r,6,aws,awsf,bdr=bdr); cel(ws,r,7,rec,sf,bdr=bdr)

    # Priority deployment plan
    ws2 = wb.create_sheet("Priority Deployment Plan")
    title_row(ws2, 5, "MERIDIAN — AI Deployment Priority Sequence", fills)
    ws2.row_dimensions[2].height = 28
    plan_hdrs = [("Priority",10),("Initiative",28),("Prerequisite MLOps",24),
                 ("Earliest Deploy",14),("Annual Value ($M)",14),("Notes",30)]
    for i,(hd,w) in enumerate(plan_hdrs,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    plan = [
        (1,"GenAI Clinical Documentation (AWS Bedrock)","Serving layer only — Bedrock handles model","Month 2","$42M","Fastest path. Bedrock ready. No training data required. Physician-facing."),
        (2,"Prior Auth Automation (Cohere Health)","Serving layer + Epic integration","Month 3","$37.6M","CMS mandate January 2027. Cohere Health manages the model. Epic API integration is the blocker."),
        (3,"Patient No-Show Prediction","Serving layer only — model trained","Month 3","$4M","Model already trained on Epic scheduling data. Serving layer is only blocker."),
        (4,"Clinical Coding AI (3M HIS)","Serving layer + accuracy improvement","Month 5","$11M","Pilot at 84% accuracy. Target 96%. Continue training while infrastructure built."),
        (5,"Epic Optimization Sprint (all modules)","No MLOps needed — configuration only","Month 1","$12M","First initiative to start. No AI/MLOps required. Pure Epic configuration and training."),
        (6,"Predictive Readmission","Feature store + Epic data pipeline","Month 8","$6M","Requires clean Epic data first. Epic Optimization Sprint prerequisite."),
        (7,"Travel Nurse Forecasting","Feature store + workforce data pipeline","Month 6","$8M","Workforce data consolidation required. Then standard ML pipeline."),
        (8,"MA Star Rating AI","Full MLOps + CDO approval","Month 9","$24M","CDO appointment and clinical AI governance prerequisite."),
    ]
    for r,row in enumerate(plan,3):
        ws2.row_dimensions[r].height = 24
        pri,name,prereq,when,value,notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        pf = fills["red"] if pri<=2 else fills["amb"] if pri<=4 else fills["grn"]
        cel(ws2,r,1,pri,pf,"#,##0",align="center",bold=True,bdr=bdr)
        cel(ws2,r,2,name,sf,bold=True,bdr=bdr); cel(ws2,r,3,prereq,sf,bdr=bdr)
        cel(ws2,r,4,when,sf,align="center",bdr=bdr)
        cel(ws2,r,5,float(value.replace("$","").replace("M","")),fills["grn"],"$#,##0.0",align="center",bold=True,bdr=bdr)
        cel(ws2,r,6,notes,sf,bdr=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/pdlc/MH-P05_MLOps_Assessment.xlsx")
    return "MH-P05 done"


# ══════════════════════════════════════════════════════════════════
# MH-M01: Revenue Cycle P&L by Payer
# ══════════════════════════════════════════════════════════════════
def mh_m01_revenue_cycle(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "RCM by Payer"
    title_row(ws, 10, "MERIDIAN HEALTH — Revenue Cycle P&L by Payer (FY2025, $M)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Payer",20),("Revenue Mix %",13),("Gross Revenue ($M)",16),
            ("Net Revenue ($M)",14),("Denial Rate",12),("Denial $ Impact ($M)",16),
            ("Days in AR",12),("Collection Rate",13),("vs Benchmark",14),("Notes",28)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    payers = [
        ("Medicare FFS","28%",3136,2620,0.142,445,54,0.871,"Benchmark 12% denial","Highest denial rate. Prior auth rules most complex. CMS mandate directly impacts this."),
        ("Medicare Advantage","18%",2016,1512,0.224,452,61,0.812,"Benchmark 10% denial","MA denial rate critical — tied to Star Rating. 3.5 stars = zero quality bonus."),
        ("Medicaid","22%",2464,1651,0.168,414,58,0.842,"Benchmark 14% denial","Medicaid managed care plans increasingly complex. Prior auth expanding."),
        ("Commercial — Blue Cross","12%",1344,1210,0.088,118,42,0.924,"At benchmark","Best-performing payer relationship. Prior auth electronic where possible."),
        ("Commercial — Aetna","8%",896,806,0.096,86,44,0.912,"Slight above benchmark","Good relationship. Prior auth partially automated."),
        ("Commercial — Cigna","6%",672,598,0.112,75,48,0.904,"Above benchmark","Manual prior auth for most specialties. Automation opportunity."),
        ("Self-Pay / Uninsured","4%",448,201,0.0,247,78,0.448,"Expected low","High charity care and bad debt. 78 days AR reflects collection difficulty."),
        ("Other / Managed Care","2%",224,188,0.094,21,45,0.894,"At benchmark","Various managed care plans. Reasonable performance."),
        ("TOTAL","100%",11200,8786,0.182,1858,52,0.876,"Benchmark 12% denial","$1.858B annual denial impact. $94M write-off after recovery. $37.6M recoverable via prior auth automation."),
    ]

    for r,row in enumerate(payers,3):
        ws.row_dimensions[r].height = 24
        payer,mix,gross,net,denial_r,denial_d,ar,coll,vs_bench,notes = row
        sf = fills["prp"] if payer=="TOTAL" else fills["alt"] if r%2==0 else fills["wht"]
        bold = payer=="TOTAL"
        cel(ws,r,1,payer,sf,bold=bold,bdr=bdr)
        cel(ws,r,2,mix,sf,align="center",bdr=bdr)
        cel(ws,r,3,gross,sf,"$#,##0",align="center",bdr=bdr)
        cel(ws,r,4,net,sf,"$#,##0",align="center",bdr=bdr)
        dr_f = fills["red"] if isinstance(denial_r,float) and denial_r>0.15 else fills["amb"] if isinstance(denial_r,float) and denial_r>0.10 else fills["grn"] if isinstance(denial_r,float) else sf
        cel(ws,r,5,denial_r,dr_f,"0.0%" if isinstance(denial_r,float) else None,align="center",bdr=bdr)
        cel(ws,r,6,denial_d,fills["red"] if isinstance(denial_d,(int,float)) and denial_d>100 else sf,"$#,##0",align="center",bdr=bdr)
        ar_f = fills["red"] if isinstance(ar,int) and ar>55 else fills["amb"] if isinstance(ar,int) and ar>40 else fills["grn"] if isinstance(ar,int) else sf
        cel(ws,r,7,ar,ar_f,"#,##0",align="center",bdr=bdr)
        cel(ws,r,8,coll,sf,"0.0%",align="center",bdr=bdr)
        bench_f = fills["red"] if "above" in vs_bench.lower() and "slight" not in vs_bench.lower() else fills["amb"] if "slight" in vs_bench.lower() or "At" in vs_bench else fills["grn"]
        cel(ws,r,9,vs_bench,bench_f,bdr=bdr); cel(ws,r,10,notes,sf,bdr=bdr)

    # Denial analysis by denial reason
    ws2 = wb.create_sheet("Denial Root Cause")
    title_row(ws2, 6, "MERIDIAN HEALTH — Denial Root Cause Analysis (FY2025)", fills)
    ws2.row_dimensions[2].height = 28
    dc_hdrs = [("Denial Reason",28),("Volume (Claims)",14),("Dollar Value ($M)",14),
               ("% of Total Denials",14),("Recovery Rate",12),("Root Cause",28),("Fix",28)]
    for i,(hd,w) in enumerate(dc_hdrs,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    denials = [
        ("Prior authorization not obtained",48200,412,0.222,0.18,"Manual prior auth process. 4.2-day average vs 1.8 peer. CMS mandate January 2027.","Prior auth automation (Cohere Health). Epic prior auth module activation."),
        ("Authorization — wrong procedure/site",18400,156,0.084,0.32,"Manual mapping of procedure to auth. Errors common in specialties.","Automated procedure-to-auth mapping in Epic."),
        ("Medical necessity not established",22800,194,0.104,0.24,"Clinical documentation insufficient. Epic note templates not standardized.","GenAI clinical documentation. Epic template standardization."),
        ("Duplicate claim submission",8400,42,0.023,0.71,"Billing system reconciliation gap. Manual resubmission creates duplicates.","Epic billing automation enhancement."),
        ("Timely filing — missed deadline",12600,108,0.058,0.12,"Manual workflow. Claims held in work queue. Ensemble SLA breach caused delays.","Ensemble SLA enforcement. Automated timely filing alerts."),
        ("Incorrect patient information",6200,28,0.015,0.68,"Registration errors. Manual data entry. No real-time eligibility verification.","Real-time eligibility verification at registration."),
        ("Non-covered service",14800,124,0.067,0.08,"Benefit verification manual. High-cost procedures checked, low-cost often not.","Automated benefit verification for all scheduled procedures."),
        ("Out-of-network provider",4800,38,0.020,0.14,"Provider credentialing data lags. Staff unaware of network status changes.","Real-time provider network status in Epic scheduling."),
        ("Coding errors (ICD/CPT)",16400,142,0.077,0.44,"Manual coding. 3M HIS pilot only 84% accurate. Human coder variability.","3M coding AI to 96%+ accuracy. CDI programme."),
        ("Other / misc",24400,186,0.100,0.28,"Various. Payer-specific rules. Appeals in progress.","Payer-specific rule automation. Appeal workflow."),
        ("COORDINATION OF BENEFITS","Ensemble-managed portion",428,None,0.22,"Above are primary reasons. Ensemble manages appeals. Internal visibility limited.","Negotiate transparency clause with Ensemble in contract renewal."),
    ]
    for r,row in enumerate(denials,3):
        ws2.row_dimensions[r].height = 24
        reason,vol,dollar,pct,recovery,cause,fix = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws2,r,1,reason,sf,bold=True,bdr=bdr)
        cel(ws2,r,2,vol,sf,"#,##0" if isinstance(vol,int) else None,align="center",bdr=bdr)
        cel(ws2,r,3,dollar,sf,"$#,##0",align="center",bdr=bdr)
        if pct: cel(ws2,r,4,pct,sf,"0.0%",align="center",bdr=bdr)
        else: cel(ws2,r,4,"N/A",sf,align="center",bdr=bdr)
        if recovery:
            rf = fills["grn"] if recovery>0.5 else fills["amb"] if recovery>0.25 else fills["red"]
            cel(ws2,r,5,recovery,rf,"0.0%",align="center",bdr=bdr)
        else: cel(ws2,r,5,"N/A",sf,align="center",bdr=bdr)
        cel(ws2,r,6,cause,sf,bdr=bdr); cel(ws2,r,7,fix,sf,bdr=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/margin/MH-M01_Revenue_Cycle_PL_by_Payer.xlsx")
    return "MH-M01 done"


# ══════════════════════════════════════════════════════════════════
# MH-M02: Operating Margin by Service Line
# ══════════════════════════════════════════════════════════════════
def mh_m02_margin_service_line(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Margin by Service Line"
    title_row(ws, 9, "MERIDIAN HEALTH — Operating Margin by Service Line FY2025 ($M)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Service Line",24),("Net Revenue ($M)",14),("Direct Cost ($M)",14),
            ("Indirect Cost ($M)",14),("Operating Profit ($M)",14),("Margin %",11),
            ("Target Margin",11),("Gap (pp)",10),("AI Opportunity ($M)",14)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    service_lines = [
        ("Inpatient Medicine",2184,1614,436,134,0.061,0.090,-2.9,18.4),
        ("Surgical Services",1848,1219,314,315,0.170,0.200,-3.0,12.1),
        ("Cardiovascular",924,692,175,57,0.062,0.100,-3.8,8.4),
        ("Oncology",756,604,166,-14,-0.019,0.050,-6.9,6.2),
        ("Emergency Services",672,596,134,-58,-0.086,0.020,-10.6,4.8),
        ("Orthopaedics",588,412,94,82,0.139,0.160,-2.1,3.1),
        ("Women's Health",504,378,101,25,0.050,0.080,-3.0,2.8),
        ("Neurology / Neurosurgery",420,336,101,-17,-0.040,0.040,-8.0,4.4),
        ("Behavioural Health",336,302,101,-67,-0.199,0.010,-20.9,1.2),
        ("Outpatient / Ambulatory",1428,1000,200,228,0.160,0.180,-2.0,6.8),
        ("Post-Acute / Rehab",336,278,67,-9,-0.027,0.030,-5.7,1.8),
        ("TOTAL / SYSTEM",11200,8431,1889,-58,-0.005,0.040,-4.5,70.0),
    ]

    for r,row in enumerate(service_lines,3):
        ws.row_dimensions[r].height = 22
        sl,rev,direct,indirect,profit,margin,target,gap,ai_opp = row
        sf = fills["prp"] if "TOTAL" in sl else fills["alt"] if r%2==0 else fills["wht"]
        bold = "TOTAL" in sl
        cel(ws,r,1,sl,sf,bold=bold,bdr=bdr)
        cel(ws,r,2,rev,sf,"$#,##0",align="center",bdr=bdr)
        cel(ws,r,3,direct,sf,"$#,##0",align="center",bdr=bdr)
        cel(ws,r,4,indirect,sf,"$#,##0",align="center",bdr=bdr)
        pf = fills["grn"] if profit>100 else fills["amb"] if profit>0 else fills["red"]
        cel(ws,r,5,profit,pf,"$#,##0",align="center",bold=bold,bdr=bdr)
        mf = fills["red"] if margin<0 else fills["amb"] if margin<0.05 else fills["grn"]
        cel(ws,r,6,margin,mf,"0.0%",align="center",bdr=bdr)
        cel(ws,r,7,target,sf,"0.0%",align="center",bdr=bdr)
        gf = fills["red"] if gap<-5 else fills["amb"] if gap<-2 else sf
        cel(ws,r,8,gap,gf,"0.0",align="center",bdr=bdr)
        cel(ws,r,9,ai_opp,fills["teal"],"$#,##0.0",align="center",bold=bold,bdr=bdr)

    # Cost breakdown sheet
    ws2 = wb.create_sheet("Cost Structure")
    title_row(ws2, 6, "MERIDIAN HEALTH — Cost Structure & AI Impact Opportunity", fills)
    ws2.row_dimensions[2].height = 28
    cs_hdrs = [("Cost Category",24),("Annual Cost ($M)",14),("% of Net Revenue",14),
               ("Benchmark %",13),("Gap ($M)",12),("AI Recovery ($M)",13),("Initiative",24)]
    for i,(hd,w) in enumerate(cs_hdrs,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    costs = [
        ("Labour — Clinical (FTE)",2480,0.221,0.195,289,0.0,"Workforce AI — travel nurse reduction $18M"),
        ("Labour — Travel Nurses / Agency",480,0.043,0.025,202,18.0,"Travel nurse demand forecasting AI"),
        ("Labour — Non-Clinical",892,0.080,0.072,90,0.0,"Admin automation partial opportunity"),
        ("Physician Fees",1120,0.100,0.095,56,0.0,"Physician productivity — GenAI documentation"),
        ("Supplies & Implants",1344,0.120,0.112,90,0.0,"Supply chain AI — demand forecasting"),
        ("Purchased Services (RCM)",682,0.061,0.048,146,24.0,"RCM vendor renegotiation + prior auth automation"),
        ("IT & Technology",448,0.040,0.028,135,12.0,"Epic optimization — reduce manual workarounds"),
        ("Facilities & Depreciation",672,0.060,0.058,22,0.0,"Limited AI opportunity"),
        ("AI Portfolio (no documented ROI)",94,0.008,0.000,94,94.0,"Activate portfolio — $148M recoverable value"),
        ("Other Operating Costs",448,0.040,0.036,45,0.0,"Various"),
        ("TOTAL OPERATING COSTS",8660,0.773,0.729,631,148.0,"Total recoverable via AI + vendor + Epic"),
    ]
    for r,row in enumerate(costs,3):
        ws2.row_dimensions[r].height = 22
        cat,cost,pct,bench,gap,ai_r,init = row
        sf = fills["prp"] if "TOTAL" in cat else fills["alt"] if r%2==0 else fills["wht"]
        bold = "TOTAL" in cat
        cel(ws2,r,1,cat,sf,bold=bold,bdr=bdr)
        cel(ws2,r,2,cost,sf,"$#,##0",align="center",bdr=bdr)
        cel(ws2,r,3,pct,sf,"0.0%",align="center",bdr=bdr)
        cel(ws2,r,4,bench,sf,"0.0%",align="center",bdr=bdr)
        gf = fills["red"] if gap>80 else fills["amb"] if gap>20 else sf
        cel(ws2,r,5,gap,gf,"$#,##0",align="center",bdr=bdr)
        af = fills["grn"] if ai_r>10 else fills["teal"] if ai_r>0 else sf
        cel(ws2,r,6,ai_r,af,"$#,##0.0",align="center",bold=bold,bdr=bdr)
        cel(ws2,r,7,init,sf,bdr=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/margin/MH-M02_Operating_Margin_by_Service_Line.xlsx")
    return "MH-M02 done"


# ══════════════════════════════════════════════════════════════════
# MH-M03: AI Spend ROI Tracker
# ══════════════════════════════════════════════════════════════════
def mh_m03_ai_roi(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "AI Spend ROI"
    title_row(ws, 8, "MERIDIAN HEALTH — AI Spend ROI Tracker ($94M Committed · $0 Verified)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Initiative",28),("Budget ($M)",12),("Spent ($M)",12),
            ("Expected ROI ($M pa)",14),("Verified ROI ($M)",13),
            ("Months Active",12),("Status",14),("Root Cause / Fix",32)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    ai_spend = [
        ("Prior Auth Automation",8.4,6.2,37.6,0.0,18,"Pilot - Stalled","Epic integration incomplete. No MLOps. Cohere Health pilot approved but not deployed. Fix: Epic API + serving layer."),
        ("GenAI Clinical Documentation",4.2,2.8,42.0,0.0,12,"Pilot - Limited","AWS Bedrock ready. 10-physician pilot positive. No enterprise deployment plan. CDO needed for clinical AI governance. Fix: CDO + deployment plan."),
        ("Predictive Readmission",1.8,0.2,6.0,0.0,6,"Concept - Blocked","Epic data quality insufficient. Fix: Epic optimization sprint first."),
        ("Clinical Coding AI (3M HIS)",2.8,1.9,11.0,0.0,14,"Pilot - Improving","84% accuracy vs 96% needed. Continuing. Fix: More Meridian-specific training data."),
        ("Patient No-Show Prediction",0.9,0.6,4.0,0.0,10,"Pilot - Ready","Model trained. No serving infrastructure. Fix: SageMaker endpoint — 4 weeks."),
        ("Supply Chain Stockout AI",1.4,0.2,3.0,0.0,6,"Concept","Data in Infor not accessible to ML. Fix: Infor API integration."),
        ("Telehealth Triage AI",2.2,0.3,8.0,0.0,4,"Concept - Blocked","CDO required for clinical AI governance. Fix: CDO appointment."),
        ("MA Star Rating AI",3.4,0.4,24.0,0.0,6,"Concept - Blocked","CDO required. Multiple Epic modules need coordination. Fix: CDO + Epic Healthy Planet activation."),
        ("Travel Nurse Forecasting",1.2,0.8,8.0,0.0,10,"In Development","Workforce data in 3 systems. Fix: Data consolidation + ML pipeline."),
        ("Ambient Documentation (Nuance DAX)",0.3,0.3,3.0,0.0,6,"Pilot - Limited","10-physician pilot. Budget small. Competing with GenAI option. Fix: Choose GenAI or DAX — don't run both."),
        ("Salesforce Health Cloud AI",1.4,0.4,5.0,0.0,8,"Pilot - Low adoption","38% adoption limits AI value. Fix: Adoption programme first."),
        ("All other / exploration",66.0,48.7,0.0,0.0,"Various","Various","Various exploratory spending. Poorly tracked. CDO needed to govern."),
    ]

    for r,row in enumerate(ai_spend,3):
        ws.row_dimensions[r].height = 28
        name,budget,spent,expected,verified,months,status,cause = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws,r,1,name,sf,bold=True,bdr=bdr)
        cel(ws,r,2,budget,sf,"$#,##0.0",align="center",bdr=bdr)
        cel(ws,r,3,spent,sf,"$#,##0.0",align="center",bdr=bdr)
        cel(ws,r,4,expected,fills["grn"],"$#,##0.0",align="center",bdr=bdr)
        cel(ws,r,5,verified,fills["red"],"$#,##0.0",align="center",bold=True,bdr=bdr)
        cel(ws,r,6,months,sf,"#,##0" if isinstance(months,int) else None,align="center",bdr=bdr)
        sf_s = fills["red"] if "Stalled" in status or "Blocked" in status else fills["amb"] if "Development" in status or "Pilot" in status else fills["wht"]
        cel(ws,r,7,status,sf_s,bdr=bdr); cel(ws,r,8,cause,sf,bdr=bdr)

    tr = len(ai_spend)+3
    cel(ws,tr,1,"TOTALS",bold=True,bdr=bdr)
    for col,col_l in zip([2,3,4,5],["B","C","D","E"]):
        ws.cell(tr,col,value=f"=SUM({col_l}3:{col_l}{tr-1})")
        ws.cell(tr,col).number_format="$#,##0.0"
        ws.cell(tr,col).font=Font(bold=True,size=10,name="Arial")
        ws.cell(tr,col).alignment=Alignment(horizontal="center")

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/margin/MH-M03_AI_Spend_ROI_Tracker.xlsx")
    return "MH-M03 done"


# ══════════════════════════════════════════════════════════════════
# MH-L01: Leadership & Governance
# ══════════════════════════════════════════════════════════════════
def mh_l01_leadership(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Leadership Structure"
    title_row(ws, 8, "MERIDIAN HEALTH SYSTEM — Technology & AI Leadership Register", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Role",24),("Name",20),("Status",14),("Reports To",18),
            ("AI Initiatives Sponsored",12),("Decision Authority",20),("Risk",12),("Notes",32)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    leaders = [
        ("CEO","Emily Rodriguez","Active","Board of Directors",3,"Final approval >$5M",  "Low","AI-native commitment public. Aware of portfolio gap. Driving CDO hire."),
        ("CFO","James Park","Active","CEO",2,"Budget approval, P&L accountability","Low","Focused on RCM improvement. Key for prior auth investment approval. Vendour renegotiation sponsor."),
        ("CIO","Robert Chen","Active","CEO",4,"Technology investment <$5M","Medium","Strong Epic background. Frustrated by AI Lab stall. Supporting CDO search."),
        ("CDO","VACANT — Search Active","Month 4 of search","CIO (interim coverage)",8,"None — role vacant","Critical","$94M portfolio blocked. Clinical AI governance absent. 8 initiatives require CDO sign-off."),
        ("CMIO","Dr. Sarah Kim","Active","CEO",4,"Clinical workflow and AI clinical decisions","Medium","GenAI documentation champion. Prior auth clinical lead. Key demo stakeholder."),
        ("COO","Patricia Walsh","Active","CEO",1,"Operational decisions","Low","Supply chain and operations. AI interest limited but open."),
        ("Chief Nursing Officer","Dr. Angela Torres","Active","COO",0,"Nursing workflow decisions","Low","Travel nurse reduction programme owner. Workforce analytics key for her."),
        ("VP Revenue Cycle","Michael O'Brien","Active","CFO",1,"RCM operational decisions","High","Ensemble Health relationship owner. Denial rate accountability. SLA renegotiation."),
        ("Director AI & Innovation","(Reporting to CDO — vacant)","Leaderless","CDO (VACANT)",0,"None — reports to vacant role","Critical","Small team. Good skills. No direction or authority without CDO."),
        ("Epic System Administrator","Linda Chen (also squad lead)","Dual role — strained","CIO",0,"Epic configuration below $100k","Medium","Running epic squad AND system admin. Capacity strained. Epic optimization needs dedicated resource."),
    ]

    stat_fills={"Active":fills["grn"],"Month 4 of search":fills["red"],"Leaderless":fills["red"],"Dual role — strained":fills["amb"]}
    risk_fills={"Critical":fills["red"],"High":fills["amb"],"Medium":fills["amb"],"Low":fills["grn"]}
    for r,row in enumerate(leaders,3):
        ws.row_dimensions[r].height = 28
        role,name,status,reports,ai_c,auth,risk,notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws,r,1,role,sf,bold=True,bdr=bdr)
        cel(ws,r,2,name,fills["red"] if "VACANT" in name else sf,bold="VACANT" in name,bdr=bdr)
        cs = cel(ws,r,3,status,stat_fills.get(status,sf),align="center",bdr=bdr)
        cs.font=Font(size=9,name="Arial",bold=True)
        cel(ws,r,4,reports,sf,bdr=bdr)
        cel(ws,r,5,ai_c,fills["red"] if isinstance(ai_c,int) and ai_c>=6 else fills["amb"] if isinstance(ai_c,int) and ai_c>=3 else sf,"#,##0" if isinstance(ai_c,int) else None,align="center",bdr=bdr)
        cel(ws,r,6,auth,sf,bdr=bdr)
        cr = cel(ws,r,7,risk,risk_fills.get(risk,sf),align="center",bold=True,bdr=bdr)
        cel(ws,r,8,notes,sf,bdr=bdr)

    # Strategic commitments vs reality
    ws2 = wb.create_sheet("Strategic Commitments")
    title_row(ws2, 6, "MERIDIAN HEALTH — Strategic Commitments vs Delivery Reality", fills)
    ws2.row_dimensions[2].height = 28
    sc_hdrs = [("Commitment",32),("Source",18),("Date",11),("Status",28),("Gap",28),("Risk",10)]
    for i,(hd,w) in enumerate(sc_hdrs,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    commitments = [
        ("'We will be AI-native in clinical operations by 2025'","Annual Report 2023","2023-03-15","$94M committed. Zero AI in clinical production. CDO search month 4.","AI-native requires production AI. Not one model deployed at scale in clinical settings. 2025 passed.","Critical"),
        ("'Prior auth automation live by Q2 2025'","Board Technology Committee","2024-01-20","Q2 2025 passed. No prior auth automation. Cohere Health pilot not deployed.","CMS mandate January 2027. 14 months. Zero progress in 12 months of approved pilot.","Critical"),
        ("'Denial rate to 14% by year-end 2025'","CFO Improvement Plan","2024-04-01","Denial rate 18.2% at year-end 2025. Target 14% missed by 4.2pp.","$45M annual impact of missing target. Ensemble Health SLA breach contributed.","Critical"),
        ("'MyChart adoption to 55% by Q4 2025'","CIO Technology Roadmap","2024-02-01","MyChart adoption 37% at Q4 2025. Target 55% missed by 18pp.","Patient engagement revenue, MA Star Rating, and Epic optimization score all tied to this.","High"),
        ("'Epic optimization score to 75/100 by Q3 2025'","Epic Governance Committee","2024-06-01","Epic score 58/100 at Q4 2025. Target 75 missed. 17-point gap.","$12M annual optimization opportunity unrealized. 22 modules still underused.","High"),
        ("'CDO appointed by Q2 2025'","CEO commitment to Board","2024-09-01","CDO search active month 4. Target Q2 2025 missed. Now Q2 2026 estimate.","Every month of vacancy = $7.8M in blocked AI value. 8 initiatives waiting.","Critical"),
        ("'Travel nurse spend below $35M by year-end 2025'","COO Cost Reduction Plan","2024-03-01","Travel nurse spend $48M at year-end 2025. Target $35M missed by $13M.","Demand forecasting AI would recover $8M. Workforce optimization not deployed.","High"),
        ("'MA Star Rating maintained at 4.0'","CEO investor commitment","2024-01-01","MA Star Rating 3.5. CMS quality bonus threshold MISSED. $24M revenue at risk.","3.5 = zero quality bonus. 4.0 = 5% premium on MA revenue. Epic Tapestry and Healthy Planet underused.","Critical"),
    ]
    for r,row in enumerate(commitments,3):
        ws2.row_dimensions[r].height = 36
        commit,source,date,status,gap,risk = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        rf = risk_fills.get(risk,sf)
        cel(ws2,r,1,commit,sf,bold=True,bdr=bdr); cel(ws2,r,2,source,sf,bdr=bdr)
        cel(ws2,r,3,date,sf,align="center",bdr=bdr)
        cel(ws2,r,4,status,fills["red"],bdr=bdr); cel(ws2,r,5,gap,sf,bdr=bdr)
        cr = cel(ws2,r,6,risk,rf,align="center",bold=True,bdr=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/pdlc/MH-L01_Leadership_Governance.xlsx")
    return "MH-L01 done"


# ══════════════════════════════════════════════════════════════════
# MH-T01: Prior Auth Workflow Analysis
# ══════════════════════════════════════════════════════════════════
def mh_t01_prior_auth_workflow(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Prior Auth Workflow"
    title_row(ws, 8, "MERIDIAN HEALTH — Prior Auth Workflow Analysis (Current State vs Target)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Step",26),("Owner",16),("Current Avg Time",14),("Target Time",12),
            ("Volume/Day",11),("Error Rate",11),("Cost per Step",12),("Automation Possible?",16),("Notes",28)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    steps = [
        ("1. Order entered in Epic by physician","Physician / NP",0.25,0.10,847,0.042,"$8.40","Yes — Epic SmartForms","Physician enters procedure. Incomplete documentation causes 22% of denials downstream."),
        ("2. Insurance eligibility verification","Revenue Cycle Staff",0.50,0.05,847,0.068,"$12.80","Yes — real-time API","Manual payer website check. Should be real-time via API. Current: staff manually check 3 systems."),
        ("3. Prior auth requirement check","Revenue Cycle Staff",0.75,0.05,847,0.084,"$18.20","Yes — payer API","Manual lookup of payer auth requirements. Each payer has different rules. Errors common."),
        ("4. Clinical documentation compilation","Clinical Documentation Specialist",1.50,0.20,847,0.124,"$38.40","Partial — GenAI assist","Compile physician notes, labs, imaging into auth package. Most time-consuming step. GenAI can reduce 70%."),
        ("5. Auth submission to payer","Auth Coordinator",0.50,0.10,847,0.056,"$12.80","Yes — X12 278 standard","Manual fax or payer portal. X12 278 electronic standard available for 68% of payers. Not activated."),
        ("6. Payer review period","Payer (external)",48.0,4.0,847,0.0,"$0","No — payer controlled","Average 48-hour payer review. Cohere Health reduces to 2-4 hours via AI review with payers."),
        ("7. Auth decision monitoring","Auth Coordinator",8.0,0.50,847,0.0,"$18.20","Yes — automated alerts","Staff manually check payer portals for decisions. Automated notification available but not configured."),
        ("8. Auth denial — initial review","Clinical RN / Auth Coordinator",2.0,0.50,254,0.0,"$51.20","Partial","30% of auths denied initially. Manual clinical review. Cohere Health peer-to-peer automation."),
        ("9. Peer-to-peer request (if needed)","Physician",4.0,0.50,127,0.0,"$128.00","Partial — AI prep","15% require physician peer-to-peer. Scheduling takes days. Physician time cost significant."),
        ("10. Auth approval — notification","Revenue Cycle Staff",0.25,0.05,593,0.0,"$6.40","Yes — automated","Auth approval manual notification to scheduling and clinical teams. Should be automated."),
        ("11. Auth entered into Epic","Revenue Cycle Staff",0.50,0.05,593,0.044,"$12.80","Yes — API","Manual entry of auth number into Epic. Errors here cause downstream denials. Should be automated."),
        ("TOTAL CURRENT STATE","Mixed",65.75,6.10,847,0.18,"$307 avg","Target: $48 avg via automation","65-hour average end-to-end. Target with Cohere Health + Epic automation: 6 hours. $37.6M annual value."),
    ]

    for r,row in enumerate(steps,3):
        ws.row_dimensions[r].height = 28
        step,owner,curr_time,target,vol,error,cost,auto,notes = row
        sf = fills["prp"] if "TOTAL" in step else fills["alt"] if r%2==0 else fills["wht"]
        bold = "TOTAL" in step
        cel(ws,r,1,step,sf,bold=bold,bdr=bdr); cel(ws,r,2,owner,sf,bdr=bdr)
        tf = fills["red"] if isinstance(curr_time,(int,float)) and curr_time>10 else fills["amb"] if isinstance(curr_time,(int,float)) and curr_time>1 else sf
        cel(ws,r,3,curr_time,tf,align="center",bdr=bdr)
        cel(ws,r,4,target,fills["grn"],align="center",bdr=bdr)
        cel(ws,r,5,vol,sf,"#,##0",align="center",bdr=bdr)
        if isinstance(error,float) and error>0:
            ef = fills["red"] if error>0.08 else fills["amb"]
            cel(ws,r,6,error,ef,"0.0%",align="center",bdr=bdr)
        else:
            cel(ws,r,6,error or "N/A",sf,align="center",bdr=bdr)
        cel(ws,r,7,cost,sf,align="center",bdr=bdr)
        af = fills["grn"] if "Yes" in auto else fills["amb"] if "Partial" in auto else fills["red"]
        cel(ws,r,8,auto,af,bdr=bdr); cel(ws,r,9,notes,sf,bdr=bdr)

    # Payer complexity matrix
    ws2 = wb.create_sheet("Payer Auth Complexity")
    title_row(ws2, 7, "MERIDIAN — Prior Auth Complexity by Payer (Top 8 Payers)", fills)
    ws2.row_dimensions[2].height = 28
    pay_hdrs = [("Payer",20),("Auth Volume/Month",14),("Electronic Auth %",14),
                ("Avg Decision Time",14),("Denial Rate %",12),("Appeal Win Rate",12),
                ("Cohere Compatible?",14),("Priority for Automation",14)]
    for i,(hd,w) in enumerate(pay_hdrs,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    payer_complexity = [
        ("CMS Medicare FFS",2840,0.22,72,0.168,0.44,"Yes — CMS X12 standard","CRITICAL — CMS mandate January 2027"),
        ("UnitedHealthcare (MA)",1620,0.18,96,0.242,0.38,"Yes — pilot done","CRITICAL — highest MA denial rate"),
        ("Humana (MA)",980,0.24,84,0.198,0.41,"Yes","HIGH — significant MA volume"),
        ("Aetna (Commercial + MA)",720,0.31,60,0.142,0.52,"Yes","HIGH"),
        ("BCBS North Carolina",840,0.42,48,0.118,0.58,"Yes","MEDIUM — already partial electronic"),
        ("Cigna",480,0.28,72,0.128,0.51,"In evaluation","MEDIUM"),
        ("Medicaid NC (State)",1440,0.12,120,0.198,0.34,"Yes — state MMIS","HIGH — high volume, complex rules"),
        ("Various Commercial",680,0.35,60,0.094,0.61,"Varies","LOW — near benchmark already"),
    ]
    for r,row in enumerate(payer_complexity,3):
        ws2.row_dimensions[r].height = 22
        payer,vol,elec_pct,decision,denial,appeal,cohere,priority = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cel(ws2,r,1,payer,sf,bold=True,bdr=bdr)
        cel(ws2,r,2,vol,sf,"#,##0",align="center",bdr=bdr)
        ef = fills["grn"] if elec_pct>0.40 else fills["amb"] if elec_pct>0.25 else fills["red"]
        cel(ws2,r,3,elec_pct,ef,"0%",align="center",bdr=bdr)
        df = fills["red"] if decision>90 else fills["amb"] if decision>60 else fills["grn"]
        cel(ws2,r,4,f"{decision}hrs",df,align="center",bdr=bdr)
        drf = fills["red"] if denial>0.18 else fills["amb"] if denial>0.12 else fills["grn"]
        cel(ws2,r,5,denial,drf,"0.0%",align="center",bdr=bdr)
        cel(ws2,r,6,appeal,sf,"0.0%",align="center",bdr=bdr)
        cf = fills["grn"] if "Yes" in cohere else fills["amb"] if "evaluat" in cohere.lower() else fills["red"]
        cel(ws2,r,7,cohere,cf,bdr=bdr)
        pf = fills["red"] if "CRITICAL" in priority else fills["amb"] if "HIGH" in priority else fills["grn"]
        cel(ws2,r,8,priority,pf,bold=True,bdr=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/tech/MH-T01_Prior_Auth_Workflow_Analysis.xlsx")
    return "MH-T01 done"


# ══════════════════════════════════════════════════════════════════
# MH-T02: Vendor Assessment (RCM)
# ══════════════════════════════════════════════════════════════════
def mh_t02_vendor_assessment(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "RCM Vendor Assessment"
    title_row(ws, 9, "MERIDIAN HEALTH — RCM Vendor Assessment & Cohere Health Scoring", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Vendor",18),("Category",16),("Overall Score /100",14),
            ("Genome Outcome Rate",14),("Reference Match %",13),("Annual Cost ($M)",12),
            ("Prior Auth Capability",16),("Epic Integration",14),
            ("Recommendation",24)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    vendors = [
        ("Cohere Health","Prior Auth AI",94,0.84,0.91,2.8,"Full — AI-powered, payer-integrated","Native Epic FHIR R4","RECOMMENDED — 94 Genome score. $37.6M annual value. CMS-compliant electronic prior auth."),
        ("Waystar","RCM + Prior Auth",81,0.72,0.78,4.2,"Good — rule-based + some AI","Epic certified","STRONG ALTERNATIVE — broader RCM scope but prior auth less sophisticated than Cohere."),
        ("Olive AI (now Availity)","Prior Auth AI",74,0.61,0.68,3.1,"Moderate — AI but less payer coverage","Epic compatible","Consider for secondary payers where Cohere coverage limited."),
        ("Ensemble Health Partners (current)","Full RCM Outsource",58,0.44,0.52,14.2,"Manual-heavy, limited automation","Deep Epic integration","CURRENT VENDOR. SLA breach. $8M penalties enforceable. Denial rate 18.2% vs 12% benchmark. Renegotiate."),
        ("Change Healthcare / Optum","RCM + Clearinghouse",71,0.64,0.72,5.8,"Good clearinghouse, limited AI","Epic certified","Broad but not specialized. Better as clearinghouse than prior auth AI."),
        ("Experian Health","RCM Analytics",66,0.58,0.64,2.4,"Limited prior auth","Epic compatible","Analytics strength. Prior auth not primary capability."),
        ("nThrive (now Kodiak Solutions)","RCM",62,0.54,0.61,3.8,"Rule-based only","Epic compatible","Legacy RCM. No meaningful AI. Do not recommend for prior auth."),
        ("3M/Solventum (Clinical Coding)","Coding AI",82,0.76,0.84,0.6,"Not prior auth — coding focus","Epic API","RECOMMEND for coding automation. Separate from prior auth decision."),
    ]

    for r,row in enumerate(vendors,3):
        ws.row_dimensions[r].height = 28
        vendor,cat,score,genome,ref_match,cost,pa_cap,epic_int,rec = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        sf_v = fills["grn"] if "RECOMMENDED" in rec else fills["amb"] if "STRONG" in rec or "RECOMMEND" in rec else fills["red"] if "Do not" in rec else fills["wht"]
        cel(ws,r,1,vendor,sf,bold=True,bdr=bdr); cel(ws,r,2,cat,sf,bdr=bdr)
        sf_s = fills["grn"] if score>=85 else fills["amb"] if score>=70 else fills["red"]
        cel(ws,r,3,score,sf_s,"#,##0",align="center",bold=True,bdr=bdr)
        cel(ws,r,4,genome,sf,"0.0%",align="center",bdr=bdr)
        cel(ws,r,5,ref_match,sf,"0.0%",align="center",bdr=bdr)
        cf = fills["red"] if cost>10 else fills["amb"] if cost>3 else fills["grn"]
        cel(ws,r,6,cost,cf,"$#,##0.0",align="center",bdr=bdr)
        cel(ws,r,7,pa_cap,sf,bdr=bdr)
        ei_f = fills["grn"] if "Native" in epic_int or "certified" in epic_int else fills["amb"]
        cel(ws,r,8,epic_int,ei_f,bdr=bdr)
        cel(ws,r,9,rec,sf_v,bdr=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/tech/MH-T02_Vendor_Assessment_RCM.xlsx")
    return "MH-T02 done"


# ══════════════════════════════════════════════════════════════════
# MH-M04: Physician Productivity & Documentation Burden
# ══════════════════════════════════════════════════════════════════
def mh_m04_physician_productivity(base):
    wb = Workbook(); bdr, fills = styles()
    ws = wb.active; ws.title = "Physician Productivity"
    title_row(ws, 8, "MERIDIAN HEALTH — Physician Productivity & Documentation Burden Analysis", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Specialty",22),("FTE Count",10),("RVUs/FTE (Actual)",14),("RVUs/FTE (Benchmark)",14),
            ("RVU Gap",10),("Doc Hours/Day (Actual)",14),("Doc Hours/Day (Target)",13),
            ("Burnout Rate %",12),("AI Documentation\nOpportunity ($M pa)",16)]
    for i,(hd,w) in enumerate(hdrs,1): h(ws,2,i,hd,w,fills=fills,bdr=bdr)

    specialties = [
        ("Internal Medicine",124,4420,5200,-780,2.8,1.0,0.72,6.2),
        ("Emergency Medicine",88,4180,4800,-620,3.1,1.0,0.81,4.8),
        ("Hospitalist",96,4620,5400,-780,2.6,1.0,0.68,5.1),
        ("Surgery (General)",42,5840,6400,-560,1.8,0.8,0.58,2.4),
        ("Cardiovascular",38,5240,5800,-560,2.2,0.8,0.62,2.1),
        ("Oncology",28,4840,5200,-360,2.9,1.0,0.76,1.8),
        ("Orthopaedics",34,6120,6800,-680,1.6,0.6,0.48,1.4),
        ("Neurology",22,4280,4800,-520,2.4,0.8,0.64,1.2),
        ("OB/GYN",32,4620,5200,-580,2.1,0.8,0.61,1.6),
        ("Psychiatry / Behavioural",18,3840,4400,-560,3.4,1.2,0.84,1.1),
        ("Radiology",24,5640,6000,-360,0.6,0.3,0.38,0.4),
        ("Outpatient / Primary Care",186,4180,4800,-620,3.2,1.0,0.74,8.4),
        ("TOTAL / SYSTEM",732,4820,5400,-580,2.6,0.9,0.68,36.5),
    ]

    for r,row in enumerate(specialties,3):
        ws.row_dimensions[r].height = 22
        spec,fte,rvu_a,rvu_b,rvu_gap,doc_a,doc_t,burnout,ai_opp = row
        sf = fills["prp"] if "TOTAL" in spec else fills["alt"] if r%2==0 else fills["wht"]
        bold = "TOTAL" in spec
        cel(ws,r,1,spec,sf,bold=bold,bdr=bdr)
        cel(ws,r,2,fte,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,3,rvu_a,fills["amb"],"#,##0",align="center",bdr=bdr)
        cel(ws,r,4,rvu_b,sf,"#,##0",align="center",bdr=bdr)
        cel(ws,r,5,rvu_gap,fills["red"],"#,##0",align="center",bdr=bdr)
        df = fills["red"] if doc_a>2.5 else fills["amb"] if doc_a>1.5 else fills["grn"]
        cel(ws,r,6,doc_a,df,"0.0",align="center",bdr=bdr)
        cel(ws,r,7,doc_t,fills["grn"],"0.0",align="center",bdr=bdr)
        bf = fills["red"] if burnout>0.70 else fills["amb"] if burnout>0.55 else fills["grn"]
        cel(ws,r,8,burnout,bf,"0%",align="center",bdr=bdr)
        cel(ws,r,9,ai_opp,fills["teal"],"$#,##0.0",align="center",bold=bold,bdr=bdr)

    # GenAI documentation ROI model
    ws2 = wb.create_sheet("GenAI Documentation ROI")
    title_row(ws2, 5, "MERIDIAN — GenAI Clinical Documentation ROI Model", fills)
    ws2.row_dimensions[2].height = 28
    roi_hdrs = [("Assumption",28),("Current State",18),("With GenAI",16),("Impact",14),("Confidence",12)]
    for i,(hd,w) in enumerate(roi_hdrs,1): h(ws2,2,i,hd,w,fills=fills,bdr=bdr)
    assumptions = [
        ("Physician count (FTE)","732","732","Unchanged","100%"),
        ("Avg documentation hours per day","2.6 hrs","0.8 hrs","1.8 hrs saved","88% — Nuance DAX validated"),
        ("Documentation time value ($/hr)","$186","$186","Same","Physician avg fully-loaded"),
        ("Annual documentation cost ($M)","$104.2M","$32.1M","-$72.1M","High confidence"),
        ("Productivity gain — RVUs/FTE","-580 vs benchmark","Recover 400","$28M revenue","Medium — adoption dependent"),
        ("Reduction in burnout-related attrition","18% annual turnover","12% target","$14M recruitment saving","Medium"),
        ("Documentation error reduction","22% prior auth denial linked","Target 14%","$18M denial reduction","High — directly measurable"),
        ("TOTAL ANNUAL VALUE ($M)","—","—","$42M","Genome-validated: 91% confidence"),
    ]
    for r,row in enumerate(assumptions,3):
        ws2.row_dimensions[r].height = 22
        asmp,curr,genai,impact,conf = row
        sf = fills["prp"] if "TOTAL" in asmp else fills["alt"] if r%2==0 else fills["wht"]
        bold = "TOTAL" in asmp
        cel(ws2,r,1,asmp,sf,bold=bold,bdr=bdr); cel(ws2,r,2,curr,sf,bdr=bdr)
        cel(ws2,r,3,genai,fills["grn"],bdr=bdr)
        cel(ws2,r,4,impact,fills["teal"] if "$" in impact else sf,bold=bold,bdr=bdr)
        cel(ws2,r,5,conf,sf,bdr=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/margin/MH-M04_Physician_Productivity.xlsx")
    return "MH-M04 done"


# ══════════════════════════════════════════════════════════════════
# PARALLEL RUNNER
# ══════════════════════════════════════════════════════════════════
def run_task(args):
    func, base = args
    try:
        result = func(base)
        return [result] if isinstance(result, str) else result
    except Exception as e:
        return [f"ERROR in {func.__name__}: {e}\n{traceback.format_exc()}"]

if __name__ == "__main__":
    import time, sys
    from pathlib import Path
    BASE = Path(__file__).parent

    tasks_arcturus = [
        (arcturus_pdlc_f05_data_architecture, BASE),
        (arcturus_pdlc_f09_mlops, BASE),
        (arcturus_delivery_files, BASE),
        (arcturus_margin_files, BASE),
        (firstcapital_files, BASE),
        (apex_files, BASE),
    ]

    tasks_meridian = [
        (mh_p01_engineering_org, BASE),
        (mh_p02_sprint_velocity, BASE),
        (mh_p03_ai_initiatives, BASE),
        (mh_p04_technology_landscape, BASE),
        (mh_p05_mlops, BASE),
        (mh_m01_revenue_cycle, BASE),
        (mh_m02_margin_service_line, BASE),
        (mh_m03_ai_roi, BASE),
        (mh_l01_leadership, BASE),
        (mh_t01_prior_auth_workflow, BASE),
        (mh_t02_vendor_assessment, BASE),
        (mh_m04_physician_productivity, BASE),
    ]

    all_tasks = tasks_arcturus + tasks_meridian

    print(f"\n{'='*60}")
    print(f"AbarVa Dataset Generator — {len(all_tasks)} task groups")
    print(f"Generating Arcturus + Meridian datasets...")
    print(f"{'='*60}\n")

    start = time.time()
    with mp.Pool(processes=min(len(all_tasks), mp.cpu_count())) as pool:
        results = pool.map(run_task, all_tasks)

    elapsed = time.time() - start
    all_results = [r for group in results for r in group]
    errors = [r for r in all_results if "ERROR" in r]
    successes = [r for r in all_results if "ERROR" not in r]

    print("Results:")
    for r in all_results:
        print(f"  {'OK' if 'ERROR' not in r else 'FAIL'} {r}")

    print(f"\n{'='*60}")
    print(f"Complete: {len(successes)} files generated, {len(errors)} errors")
    print(f"Time: {elapsed:.1f}s")

    all_files = [f for f in BASE.rglob("*.xlsx") if "scripts" not in str(f)]
    print(f"\nAll files ({len(all_files)}):")
    for f in sorted(all_files):
        print(f"  {str(f.relative_to(BASE))}")

    if errors:
        for e in errors: print(e)
        sys.exit(1)
    else:
        print(f"\nDone. Upload datasets/ folder to your repo.")
