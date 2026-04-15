"""
AbarVa — Margin Optimization Extended Datasets
Covers: Back Office BPO, Supply Chain, HR Operations, IT Ops
For: Meridian Health System + Arcturus Financial Group
Run: python3 generate_margin_extended.py
"""
import multiprocessing as mp
import sys, time, traceback
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).parent))
from constants import MER, ARC, GENOME

BASE = Path(__file__).parent

def S():
    t = Side(style="thin", color="CCCCCC")
    b = Border(left=t, right=t, top=t, bottom=t)
    f = {
        "hdr":  PatternFill("solid", start_color="1A3A5C"),
        "alt":  PatternFill("solid", start_color="F2F7FC"),
        "wht":  PatternFill("solid", start_color="FFFFFF"),
        "red":  PatternFill("solid", start_color="FDE8E8"),
        "amb":  PatternFill("solid", start_color="FFF4E5"),
        "grn":  PatternFill("solid", start_color="E8F5E9"),
        "blu":  PatternFill("solid", start_color="E8F0F8"),
        "prp":  PatternFill("solid", start_color="F0EEFF"),
        "teal": PatternFill("solid", start_color="E0F7F4"),
    }
    return b, f

def H(ws, row, col, val, w=16, f=None, b=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    c.fill = f["hdr"]; c.border = b
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(col)].width = w

def C(ws, row, col, val, fill=None, fmt=None, bold=False, align="left", b=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(size=9, name="Arial", bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = b
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    return c

def T(ws, cols, text, f):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    ws["A1"].value = text
    ws["A1"].font = Font(bold=True, size=12, color="1A3A5C", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = f["blu"]; ws.row_dimensions[1].height = 16

def save(wb, path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════
# MERIDIAN — BACK OFFICE / BPO OPPORTUNITY
# ══════════════════════════════════════════════════════════════════════════

def mer_bo01_back_office(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Back Office Assessment"
    T(ws, 9, f"{MER['name']} — Back Office Operations & BPO Opportunity Assessment", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Function",22),("Current FTE",12),("Annual Cost $M",13),
            ("Cost per FTE $K",13),("Industry Benchmark $M",14),("Overspend $M",12),
            ("BPO / Offshore Opportunity",20),("Est Savings $M",12),("Notes",28)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    functions = [
        # Function, FTE, Cost $M, Cost/FTE $K, Benchmark $M, Overspend, BPO opp, Savings $M, Notes
        ("Finance & Accounting — AP/AR",          124, 14.9, 120, 9.3,  5.6, "High — 60-70% of AP/AR processable offshore",  4.2,  "Invoice processing $18 per invoice vs $8 benchmark. Manual matching dominates."),
        ("Finance — Month-end Close",              48,  6.8, 142, 4.1,  2.7, "Medium — structured close process amenable to offshore", 1.8, "18-day close vs 8-day benchmark. Manual journal entries. No automation."),
        ("Finance — Management Reporting",         36,  5.4, 150, 3.2,  2.2, "Medium — templated reporting offshortable", 1.2, "Board pack manual. 3-day lag from source data. Report factory model applicable."),
        ("Revenue Cycle — Coding",                 284, 28.4, 100, 18.4, 10.0, "High — clinical coding 70% offshortable to certified coders", 7.2, "84% coding accuracy in-house vs 96% benchmark. 3M HIS pilot shows AI + offshore hybrid."),
        ("Revenue Cycle — Billing & Collections",  196, 19.6, 100, 12.8,  6.8, "High — billing ops standard BPO candidate", 4.8, "First-pass resolution 72% vs 85% benchmark. Manual follow-up dominates."),
        ("Revenue Cycle — Prior Auth Admin",       142, 14.2, 100,  6.8,  7.4, "Very High — prior auth admin 80% automatable/offshortable", 8.4, "4.2 day avg vs 1.8 day peer. CMS mandate Jan 2027. Cohere Health replaces most of this."),
        ("Supply Chain — Procurement Admin",        68,  8.2, 121,  5.1,  3.1, "High — purchase order and vendor management offshortable", 2.4, "Non-clinical procurement particularly amenable. GPO compliance tracking manual."),
        ("HR — Benefits Administration",            42,  5.0, 119,  2.8,  2.2, "High — benefits admin fully BPO-able",   1.8, "Open enrollment, COBRA, benefits query handling all standard BPO."),
        ("HR — Payroll Operations",                 38,  4.6, 121,  2.4,  2.2, "High — payroll processing standard BPO",  1.6, "42,000 employees. Bi-weekly payroll. High volume, low complexity."),
        ("HR — Recruitment Administration",         28,  3.4, 121,  2.0,  1.4, "Medium — sourcing admin offshortable, judgment stays",  0.8, "Job posting, screening scheduling, onboarding admin offshortable."),
        ("IT — Service Desk / Helpdesk",            84,  9.2, 110,  5.6,  3.6, "Very High — L1/L2 support standard offshore", 3.2, "42,000 staff. 1.8 tickets per user per month. L1/L2 = 80% of volume."),
        ("IT — Application Support",                62,  7.4, 119,  4.8,  2.6, "Medium — depends on system complexity", 1.6, "Epic application support partially offshortable. Clinical systems need onshore judgment."),
        ("Compliance — Audit & Documentation",      48,  6.2, 129,  3.8,  2.4, "Medium — documentation and evidence gathering offshortable", 1.2, "HIPAA audit prep, policy documentation, evidence collection."),
        ("Patient Access — Scheduling Admin",       186, 16.8,  90, 11.2,  5.6, "High — scheduling and registration admin offshortable", 4.0, "Appointment scheduling, insurance verification, registration. High volume, standardised."),
    ]

    total_cost = total_bench = total_savings = 0
    for r, row in enumerate(functions, 3):
        ws.row_dimensions[r].height = 28
        func, fte, cost, cpf, bench, over, bpo, savings, notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        total_cost += cost; total_bench += bench; total_savings += savings
        C(ws,r,1,func,sf,bold=True,b=bdr)
        C(ws,r,2,fte,sf,"#,##0",align="center",b=bdr)
        C(ws,r,3,cost,sf,"$#,##0.0",align="center",b=bdr)
        C(ws,r,4,cpf,sf,"$#,##0",align="center",b=bdr)
        C(ws,r,5,bench,sf,"$#,##0.0",align="center",b=bdr)
        C(ws,r,6,over,fills["red"] if over>3 else fills["amb"],"$#,##0.0",align="center",b=bdr)
        bpo_f = fills["red"] if "Very High" in bpo else fills["amb"] if "High" in bpo else fills["wht"]
        C(ws,r,7,bpo,bpo_f,b=bdr)
        C(ws,r,8,savings,fills["grn"],"$#,##0.0",align="center",bold=True,b=bdr)
        C(ws,r,9,notes,sf,b=bdr)

    tr = len(functions)+3
    C(ws,tr,1,f"TOTAL — {sum(r[1] for r in functions):,} FTE — BPO/offshore opportunity",fills["prp"],bold=True,b=bdr)
    for col, val, fmt in [(3,total_cost,"$#,##0.0"),(5,total_bench,"$#,##0.0"),(8,total_savings,"$#,##0.0")]:
        ws.cell(tr,col,value=val); ws.cell(tr,col).number_format=fmt
        ws.cell(tr,col).font=Font(bold=True,size=10,name="Arial"); ws.cell(tr,col).alignment=Alignment(horizontal="center")

    # BPO approach comparison sheet
    ws2 = wb.create_sheet("BPO Approach Comparison")
    T(ws2, 7, "Meridian — BPO Delivery Model Options", fills)
    ws2.row_dimensions[2].height = 28
    hdrs2 = [("Approach",22),("Description",28),("Savings %",12),("Savings $M",12),
             ("Implementation Cost $M",16),("Payback Months",13),("Risk",12),("Recommendation",24)]
    for i,(hd,w) in enumerate(hdrs2,1): H(ws2,2,i,hd,w,fills,bdr)
    approaches = [
        ("Full BPO Outsource","Hand all back-office functions to single vendor (e.g. Accenture, Conduent, Optum)",0.45,25.4,4.2,24,"High","Not recommended — single vendor dependency, knowledge loss, contract lock-in"),
        ("Selective BPO — High Volume","Offshore only high-volume, low-judgment functions: AP/AR, payroll, scheduling admin, helpdesk",0.38,18.6,2.4,18,"Medium","RECOMMENDED Wave 1 — preserves clinical and strategic functions internally"),
        ("AI + Offshore Hybrid","Automate with AI first, offshore residual. Prior auth: Cohere Health. Coding: 3M HIS. Helpdesk: ServiceNow AI.",0.52,22.8,3.8,20,"Medium","RECOMMENDED Wave 2 — higher savings, lower offshore headcount needed"),
        ("Nearshore / Domestic","Keep offshore to low-risk functions. Use nearshore (Mexico, Canada) for clinical-adjacent.",0.28,12.4,1.8,24,"Low","Conservative option — lower savings, lower risk"),
        ("Internal Automation Only","No offshore. Automate with AI and RPA only.",0.22,9.8,2.8,28,"Low","Leaves 50%+ of savings on table. Not recommended as primary strategy."),
    ]
    for r, row in enumerate(approaches, 3):
        ws2.row_dimensions[r].height = 28
        approach, desc, pct, savings, impl, payback, risk, rec = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        C(ws2,r,1,approach,sf,bold=True,b=bdr); C(ws2,r,2,desc,sf,b=bdr)
        C(ws2,r,3,pct,sf,"0%",align="center",b=bdr)
        C(ws2,r,4,savings,fills["grn"],"$#,##0.0",align="center",b=bdr)
        C(ws2,r,5,impl,sf,"$#,##0.0",align="center",b=bdr)
        C(ws2,r,6,payback,sf,"#,##0",align="center",b=bdr)
        rf = fills["red"] if risk=="High" else fills["amb"] if risk=="Medium" else fills["grn"]
        C(ws2,r,7,risk,rf,align="center",b=bdr)
        rec_f = fills["grn"] if "RECOMMENDED" in rec else fills["red"] if "Not recommended" in rec else fills["wht"]
        C(ws2,r,8,rec,rec_f,b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/margin/MER-BO01_Back_Office_BPO_Opportunity.xlsx")
    return "MER-BO01 done"


def mer_sc01_supply_chain(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Supply Chain Analysis"
    T(ws, 9, f"{MER['name']} — Supply Chain & Procurement Analysis", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Category",22),("Annual Spend $M",14),("GPO Contract %",13),
            ("Actual GPO Compliance %",16),("Compliance Gap %",12),("Savings Opportunity $M",16),
            ("Price vs GPO Benchmark",16),("AI Opportunity",20),("Priority",10)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    categories = [
        # Cat, Spend, GPO %, Compliance %, Gap %, Savings, vs benchmark, AI opp, Priority
        ("Medical/Surgical Supplies",    284, 0.92, 0.74, -0.18,  8.4, "+12% vs GPO contract price", "Demand forecasting AI — reduce stockouts, eliminate emergency buys","Critical"),
        ("Pharmaceuticals",              196, 0.98, 0.91, -0.07,  4.2, "+4% vs GPO (340B compliance gap)", "340B compliance AI — automated split billing","High"),
        ("Capital Equipment",            142, 0.68, 0.52, -0.16,  6.8, "+18% vs benchmark on non-GPO purchases","Equipment utilization AI — reduce unnecessary purchases","Critical"),
        ("Laboratory Supplies",          84,  0.88, 0.79, -0.09,  2.4, "+8% vs GPO contract","Lab supply demand AI — reduce waste and expiry","High"),
        ("Food & Nutrition Services",    68,  0.72, 0.61, -0.11,  2.1, "+15% vs benchmark on non-contract spend","Menu planning AI — reduce food waste (currently 22%)","Medium"),
        ("Linen & Laundry",              42,  0.84, 0.76, -0.08,  0.9, "At benchmark for contracted items","Linen utilization tracking — reduce loss","Low"),
        ("Facility Maintenance",         58,  0.62, 0.48, -0.14,  2.6, "+22% vs benchmark on reactive maintenance","Predictive maintenance AI — reduce reactive call-outs","High"),
        ("IT Hardware & Software",       84,  0.78, 0.64, -0.14,  3.2, "+16% vs benchmark. Unmanaged spend high.","License optimization AI — identify unused licenses","High"),
        ("Professional Services",        126, 0.24, 0.18, -0.06,  4.8, "Largely uncontracted. High spend variability.","Spend analytics AI — identify consolidation opportunities","Critical"),
        ("Non-Clinical Supplies",        38,  0.82, 0.71, -0.11,  1.2, "+9% vs GPO on office/admin supplies","None required — process improvement sufficient","Low"),
    ]

    total_spend = total_savings = 0
    for r, row in enumerate(categories, 3):
        ws.row_dimensions[r].height = 28
        cat, spend, gpo_pct, compliance, gap, savings, vs_bench, ai_opp, priority = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        total_spend += spend; total_savings += savings
        pf = fills["red"] if priority=="Critical" else fills["amb"] if priority=="High" else fills["grn"]
        C(ws,r,1,cat,sf,bold=True,b=bdr)
        C(ws,r,2,spend,sf,"$#,##0",align="center",b=bdr)
        C(ws,r,3,gpo_pct,sf,"0%",align="center",b=bdr)
        cf = fills["red"] if compliance < 0.65 else fills["amb"] if compliance < 0.80 else fills["grn"]
        C(ws,r,4,compliance,cf,"0%",align="center",b=bdr)
        C(ws,r,5,gap,fills["red"],"0%",align="center",b=bdr)
        C(ws,r,6,savings,fills["teal"],"$#,##0.0",align="center",bold=True,b=bdr)
        C(ws,r,7,vs_bench,sf,b=bdr)
        C(ws,r,8,ai_opp,fills["teal"] if "AI" in ai_opp else sf,b=bdr)
        C(ws,r,9,priority,pf,align="center",bold=True,b=bdr)

    tr = len(categories)+3
    C(ws,tr,1,f"TOTAL — ${total_spend:.0f}M spend — ${total_savings:.1f}M recoverable",fills["prp"],bold=True,b=bdr)
    C(ws,tr,2,total_spend,fills["prp"],"$#,##0",align="center",bold=True,b=bdr)
    C(ws,tr,6,total_savings,fills["grn"],"$#,##0.0",align="center",bold=True,b=bdr)

    # Inventory analysis sheet
    ws2 = wb.create_sheet("Inventory Performance")
    T(ws2, 7, "Meridian — Inventory Performance vs Benchmark", fills)
    ws2.row_dimensions[2].height = 28
    inv_hdrs = [("Category",22),("Inventory Value $M",14),("Turns (Actual)",12),
                ("Turns (Benchmark)",14),("Days on Hand",12),("Days on Hand Benchmark",16),
                ("Excess Inventory $M",14),("Stockout Events/Yr",13),("Stockout Cost $M",12)]
    for i,(hd,w) in enumerate(inv_hdrs,1): H(ws2,2,i,hd,w,fills,bdr)
    inventory = [
        ("Medical/Surgical",  42.0, 8.2,  12.4, 44, 29, 8.4,  124, 2.2),
        ("Pharmaceuticals",   28.0, 18.4, 22.0, 20, 17, 4.8,  48,  0.8),
        ("Lab Supplies",      12.0, 10.2, 14.8, 36, 25, 3.2,  28,  0.4),
        ("Capital Equipment",  8.0,  2.1,  3.2, 174,114, 2.8,  8,  0.2),
        ("Food & Nutrition",   4.2, 24.0, 28.0, 15, 13, 0.4,  12,  0.1),
        ("Linen & Laundry",    2.8, 12.0, 16.0, 30, 23, 0.6,  8,   0.1),
    ]
    for r, row in enumerate(inventory, 3):
        ws2.row_dimensions[r].height = 22
        cat, inv_val, turns_a, turns_b, days_a, days_b, excess, stockouts, stockout_cost = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        tf = fills["red"] if turns_a < turns_b*0.75 else fills["amb"] if turns_a < turns_b else fills["grn"]
        C(ws2,r,1,cat,sf,bold=True,b=bdr)
        C(ws2,r,2,inv_val,sf,"$#,##0.0",align="center",b=bdr)
        C(ws2,r,3,turns_a,tf,"#,##0.0",align="center",b=bdr)
        C(ws2,r,4,turns_b,fills["grn"],"#,##0.0",align="center",b=bdr)
        C(ws2,r,5,days_a,fills["red"] if days_a>days_b*1.3 else fills["amb"],"#,##0",align="center",b=bdr)
        C(ws2,r,6,days_b,fills["grn"],"#,##0",align="center",b=bdr)
        C(ws2,r,7,excess,fills["red"],"$#,##0.0",align="center",b=bdr)
        C(ws2,r,8,stockouts,sf,"#,##0",align="center",b=bdr)
        C(ws2,r,9,stockout_cost,fills["amb"],"$#,##0.0",align="center",b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/margin/MER-SC01_Supply_Chain_Procurement.xlsx")
    return "MER-SC01 done"


def mer_hr01_workforce_operations(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "HR Operations Cost"
    T(ws, 8, f"{MER['name']} — HR Operations Cost & Optimization", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("HR Function",22),("FTE",10),("Annual Cost $M",13),("Benchmark $M",13),
            ("Gap $M",11),("Automation Opportunity",22),("Est Savings $M",12),("Notes",28)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    hr_functions = [
        # Function, FTE, Cost, Benchmark, Gap, Automation, Savings, Notes
        ("Travel Nurse Procurement",       24, 48.0, 28.0, -20.0, "Demand forecasting AI + preferred agency programme", 8.0, "Travel nurse $48M vs $28M benchmark. Largest single HR cost opportunity. Demand AI reduces need by 30%."),
        ("Permanent Nurse Recruitment",    18,  8.4,  5.2,  -3.2, "AI sourcing + video screening + predictive fit scoring", 1.8, "Cost per hire $18,400 vs $12,000 benchmark. 62-day time-to-fill vs 38-day benchmark."),
        ("Physician Recruitment",          12,  6.2,  4.1,  -2.1, "AI job matching + relationship management platform", 0.8, "Physician recruitment $142,000 per hire. Retention programmes reduce rehire frequency."),
        ("Benefits Administration",        28,  3.4,  1.8,  -1.6, "Full BPO to Aon/Mercer or AI self-service portal", 1.2, "42,000 employees. High query volume. Benefits AI handles 80% of queries."),
        ("Payroll Processing",             22,  2.6,  1.4,  -1.2, "Automated payroll + BPO residual", 0.8, "Bi-weekly payroll for 42,000. High automation potential. Error rate 0.8% vs 0.2% benchmark."),
        ("Credentialing & Privileging",    18,  2.2,  1.6,  -0.6, "AI-assisted credentialing + primary source verification", 0.4, "72-day credentialing process vs 45-day benchmark. Delays revenue start."),
        ("Employee Relations & HR BP",     42,  5.8,  4.8,  -1.0, "AI case management + self-service HR portal", 0.6, "HR business partner ratio 1:620 vs 1:400 benchmark. Case management manual."),
        ("Learning & Development",         24,  3.2,  2.4,  -0.8, "AI personalized learning + virtual simulation", 0.4, "Epic training completion 41% — L&D effectiveness problem."),
        ("Workforce Analytics",            8,   1.2,  0.8,  -0.4, "Integrated workforce AI platform", 0.2, "Analytics currently manual. No predictive capability."),
        ("Workforce Planning",             6,   0.9,  0.6,  -0.3, "AI demand forecasting + scenario planning", 0.2, "Annual planning only. No real-time workforce adjustment."),
    ]

    for r, row in enumerate(hr_functions, 3):
        ws.row_dimensions[r].height = 28
        func, fte, cost, bench, gap, auto_opp, savings, notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,func,sf,bold=True,b=bdr)
        C(ws,r,2,fte,sf,"#,##0",align="center",b=bdr)
        C(ws,r,3,cost,sf,"$#,##0.0",align="center",b=bdr)
        C(ws,r,4,bench,sf,"$#,##0.0",align="center",b=bdr)
        C(ws,r,5,gap,fills["red"] if gap<-2 else fills["amb"],"$#,##0.0",align="center",b=bdr)
        C(ws,r,6,auto_opp,fills["teal"],b=bdr)
        C(ws,r,7,savings,fills["grn"],"$#,##0.0",align="center",bold=True,b=bdr)
        C(ws,r,8,notes,sf,b=bdr)

    tr = len(hr_functions)+3
    C(ws,tr,1,"TOTAL HR OPERATIONS OPPORTUNITY",fills["prp"],bold=True,b=bdr)
    C(ws,tr,3,sum(r[2] for r in hr_functions),fills["prp"],"$#,##0.0",align="center",bold=True,b=bdr)
    C(ws,tr,7,sum(r[6] for r in hr_functions),fills["grn"],"$#,##0.0",align="center",bold=True,b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/margin/MER-HR01_Workforce_Operations.xlsx")
    return "MER-HR01 done"


def mer_it01_it_operations(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "IT Operations"
    T(ws, 8, f"{MER['name']} — IT Operations Cost & Managed Services Opportunity", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("IT Function",22),("FTE",10),("Annual Cost $M",13),("Benchmark $M",13),
            ("Gap $M",11),("Managed Services / AI Opportunity",24),("Savings $M",11),("Notes",28)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    it_functions = [
        ("Service Desk (L1/L2)",          84, 9.2, 5.6, -3.6, "L1/L2 offshore + AI chatbot handles 60% of tickets", 3.2, "1.8 tickets/user/month × 42,000 staff. $38 per ticket vs $18 benchmark. AI + offshore = $12."),
        ("Epic Application Support",       62, 7.4, 4.8, -2.6, "Offshore Epic support center (India, Philippines) for L2", 1.6, "Epic support 24/7. Most issues pattern-based. 40% offshortable safely."),
        ("Infrastructure & Cloud Ops",     48, 5.8, 4.2, -1.6, "Managed cloud services (AWS) for infrastructure", 1.2, "Cloud migration 60% complete. AWS managed services reduce ops FTE need."),
        ("Cybersecurity Operations",       28, 4.2, 3.2, -1.0, "MSSP for SOC monitoring — 24/7 coverage cheaper external", 0.8, "SOC monitoring $4.2M internal vs $2.8M MSSP benchmark. Response quality equivalent."),
        ("Data & Analytics Ops",           24, 3.2, 2.4, -0.8, "Offshore analytics operations + AI-automated reporting", 0.6, "Manual Clarity extracts. Automated pipeline replaces 40% of FTE need."),
        ("End User Computing",             18, 2.4, 1.8, -0.6, "Device-as-a-service + remote management platform", 0.4, "Device procurement and management. DaaS reduces capital and ops cost."),
        ("Project Management Office",      14, 2.1, 1.6, -0.5, "AI project tracking + offshore PMO coordination", 0.3, "IT project management. AI tools reduce manual tracking overhead."),
        ("Vendor Management",              8,  1.2, 0.8, -0.4, "AI contract management + consolidated vendor portal", 0.2, "42 active IT vendors. Manual contract tracking. Consolidation opportunity."),
        ("IT Finance & Governance",        6,  0.9, 0.6, -0.3, "Automated IT financial management", 0.2, "IT budget tracking manual. Shadow IT untracked. AI identifies waste."),
    ]

    for r, row in enumerate(it_functions, 3):
        ws.row_dimensions[r].height = 28
        func, fte, cost, bench, gap, mssp_opp, savings, notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,func,sf,bold=True,b=bdr)
        C(ws,r,2,fte,sf,"#,##0",align="center",b=bdr)
        C(ws,r,3,cost,sf,"$#,##0.0",align="center",b=bdr)
        C(ws,r,4,bench,fills["grn"],"$#,##0.0",align="center",b=bdr)
        C(ws,r,5,gap,fills["red"] if gap<-2 else fills["amb"],"$#,##0.0",align="center",b=bdr)
        C(ws,r,6,mssp_opp,fills["teal"],b=bdr)
        C(ws,r,7,savings,fills["grn"],"$#,##0.0",align="center",bold=True,b=bdr)
        C(ws,r,8,notes,sf,b=bdr)

    tr = len(it_functions)+3
    C(ws,tr,1,"TOTAL IT OPERATIONS OPPORTUNITY",fills["prp"],bold=True,b=bdr)
    C(ws,tr,3,sum(r[2] for r in it_functions),fills["prp"],"$#,##0.0",align="center",bold=True,b=bdr)
    C(ws,tr,7,sum(r[6] for r in it_functions),fills["grn"],"$#,##0.0",align="center",bold=True,b=bdr)
    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/margin/MER-IT01_IT_Operations.xlsx")
    return "MER-IT01 done"


# ══════════════════════════════════════════════════════════════════════════
# MERIDIAN — MARGIN OPPORTUNITY MAP (the full universe)
# ══════════════════════════════════════════════════════════════════════════

def mer_m05_opportunity_map(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Margin Opportunity Map"
    T(ws, 9, f"{MER['name']} — Full Margin Opportunity Map ($11.2B Health System)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Category",22),("Lever",26),("Est Annual Opportunity $M",18),
            ("Genome Confidence",14),("Data Available?",14),("Data Required to Unlock",24),
            ("Wave",10),("Status",14)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    opportunities = [
        # Revenue side — already have data
        ("Revenue — RCM", "Prior auth automation (Cohere Health)", 37.6, "91%", "Yes — MER-M01", "Already loaded", 1, "Ready to analyse"),
        ("Revenue — RCM", "Denial reduction — coding accuracy", 11.0, "82%", "Yes — MER-M01", "Already loaded", 1, "Ready to analyse"),
        ("Revenue — RCM", "Ensemble renegotiation (SLA breach)", 8.0, "95%", "Yes — MER-C02", "Already loaded", 1, "Ready to analyse"),
        ("Revenue — RCM", "Timely filing — workflow automation", 6.4, "74%", "Partial", "Claims aging report needed", 1, "Upload: claims aging by payer"),
        ("Revenue — Clinical", "GenAI physician documentation", 42.0, "88%", "Yes — MER-M04", "Already loaded", 1, "Ready to analyse"),
        ("Revenue — Clinical", "MA Star Rating 3.5 → 4.0", 24.0, "78%", "Yes — MER-C01", "Already loaded", 1, "Ready to analyse"),
        ("Revenue — AI Portfolio", "AI portfolio reorientation ($28M idle)", 28.0, "89%", "Yes — MER-M03", "Already loaded", 1, "Ready to analyse"),
        # Back office — need uploads
        ("Back Office — Finance", "AP/AR automation + offshore", 6.0, "74%", "No", "Finance ops data: AP volume, cost/invoice, FTE count", 2, "Upload: MER-BO01 or provide 3 numbers"),
        ("Back Office — Finance", "Month-end close cycle reduction", 1.8, "68%", "No", "Close timeline report, journal entry count", 2, "Upload: close cycle data"),
        ("Back Office — RCM Admin", "Billing & collections BPO/offshore", 4.8, "72%", "No", "Collections FTE count, first-pass resolution rate", 2, "Upload: RCM ops data"),
        ("Back Office — Prior Auth Admin", "Prior auth admin BPO (pre-Cohere)", 8.4, "84%", "No", "Prior auth admin FTE count, volume by payer", 2, "Upload: prior auth staffing data"),
        # Supply chain
        ("Supply Chain", "GPO compliance improvement", 8.4, "78%", "No", "Spend by category, GPO compliance rate", 2, "Upload: MER-SC01 or spend by category"),
        ("Supply Chain", "Inventory optimization (excess + stockouts)", 4.8, "72%", "No", "Inventory turns by category, stockout log", 2, "Upload: inventory data"),
        ("Supply Chain", "Non-clinical procurement consolidation", 2.4, "68%", "No", "Top 50 suppliers by spend", 2, "Upload: spend analytics"),
        # Workforce
        ("Workforce", "Travel nurse demand forecasting AI", 8.0, "84%", "Yes — MER-C03", "Already loaded", 1, "Ready to analyse"),
        ("Workforce", "Permanent recruitment cost reduction", 1.8, "72%", "No", "Cost per hire, time-to-fill, turnover by department", 2, "Upload: MER-HR01 or HR metrics"),
        ("Workforce", "Benefits admin BPO", 1.2, "74%", "No", "Benefits admin FTE, query volume", 3, "Upload: HR ops data"),
        # IT operations
        ("IT Operations", "Service desk offshore + AI chatbot", 3.2, "82%", "No", "Ticket volume, cost per ticket, resolution rates", 2, "Upload: IT service desk data"),
        ("IT Operations", "Managed security services (SOC)", 0.8, "78%", "No", "SOC staffing, incident volume, current vendor", 2, "Upload: IT ops data"),
        ("IT Operations", "Epic support offshore center", 1.6, "74%", "No", "Epic support ticket volume, FTE count", 2, "Upload: IT ops data"),
        # Clinical operations
        ("Clinical Ops", "OR utilization improvement", 12.0, "72%", "No", "OR utilization %, block time data, case mix", 2, "Upload: OR operations data"),
        ("Clinical Ops", "Length of stay reduction (AI discharge)", 8.4, "74%", "No", "ALOS by DRG vs benchmark, discharge planning data", 2, "Upload: MER-M02 already partial"),
        ("Clinical Ops", "Readmission reduction (AI prediction)", 6.2, "78%", "No", "30-day readmission rate by DRG, discharge disposition", 2, "Upload: quality metrics data"),
    ]

    status_fills = {"Ready to analyse": fills["grn"], "Upload:": fills["amb"]}
    for r, row in enumerate(opportunities, 3):
        ws.row_dimensions[r].height = 24
        cat, lever, opp, confidence, data_avail, data_req, wave, status = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        data_f = fills["grn"] if data_avail.startswith("Yes") else fills["amb"] if data_avail=="Partial" else fills["red"]
        wave_f = fills["grn"] if wave==1 else fills["amb"] if wave==2 else fills["wht"]
        status_f = fills["grn"] if "Ready" in status else fills["amb"] if "Upload" in status else fills["wht"]
        C(ws,r,1,cat,sf,bold=True,b=bdr); C(ws,r,2,lever,sf,b=bdr)
        C(ws,r,3,opp,fills["teal"],"$#,##0.0",align="center",bold=True,b=bdr)
        C(ws,r,4,confidence,sf,align="center",b=bdr)
        C(ws,r,5,data_avail,data_f,b=bdr); C(ws,r,6,data_req,sf,b=bdr)
        C(ws,r,7,wave,wave_f,"#,##0",align="center",bold=True,b=bdr)
        C(ws,r,8,status,status_f,b=bdr)

    tr = len(opportunities)+3
    C(ws,tr,1,f"TOTAL OPPORTUNITY — {len(opportunities)} levers identified",fills["prp"],bold=True,b=bdr)
    C(ws,tr,3,sum(r[2] for r in opportunities),fills["prp"],"$#,##0.0",align="center",bold=True,b=bdr)

    wave1 = sum(r[2] for r in opportunities if r[6]==1)
    wave2 = sum(r[2] for r in opportunities if r[6]==2)
    ready = sum(r[2] for r in opportunities if "Ready" in r[7])

    tr2 = tr+2
    for label, val in [
        ("Wave 1 (data already loaded): $M", wave1),
        ("Wave 2 (upload to unlock): $M", wave2),
        ("Already analysable without uploads: $M", ready),
        ("Total addressable margin opportunity: $M", sum(r[2] for r in opportunities))
    ]:
        ws.cell(tr2,1,value=label).font = Font(bold=True,size=9,name="Arial")
        ws.cell(tr2,3,value=val).number_format = "$#,##0.0"
        ws.cell(tr2,3).font = Font(bold=True,size=10,name="Arial",color="1A3A5C")
        ws.cell(tr2,3).alignment = Alignment(horizontal="center")
        tr2 += 1

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/margin/MER-M05_Margin_Opportunity_Map.xlsx")
    return "MER-M05 done"


# ══════════════════════════════════════════════════════════════════════════
# ARCTURUS — BACK OFFICE / MIDDLE OFFICE
# ══════════════════════════════════════════════════════════════════════════

def arc_bo01_middle_office(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Middle Office Operations"
    T(ws, 8, f"{ARC['name']} — Middle Office Operations Cost & Automation Opportunity", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Function",24),("FTE",10),("Annual Cost £M",13),("Benchmark £M",13),
            ("Gap £M",11),("Automation / Offshore Opportunity",24),("Savings £M",11),("Notes",28)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    functions = [
        ("Trade Operations — Settlement",        42, 4.8, 3.2, -1.6, "STP automation — target 98% straight-through", 1.2, "Settlement failure rate 2.1% vs 0.8% benchmark. Manual matching team 42 FTE. STP AI."),
        ("Trade Operations — Reconciliation",    38, 4.2, 2.8, -1.4, "Automated reconciliation — eliminate manual breaks", 1.0, "Daily position reconciliation AIM vs Aladdin manual. 3-day lag perpetuates breaks."),
        ("Trade Operations — Corporate Actions", 24, 2.8, 1.8, -1.0, "AI corporate actions processing — complex event handling", 0.6, "Corporate actions processing manual for complex events. AI can handle 70% of standard events."),
        ("Client Reporting — Production",        36, 4.4, 2.6, -1.8, "Report factory automation — offshore + AI generation", 1.4, "Monthly client reports manual. 3-day data lag = reports always stale. Offshore + GenAI."),
        ("Client Reporting — Data Aggregation",  28, 3.2, 1.8, -1.4, "Automated data pipeline replaces manual aggregation", 1.0, "Manual extraction from 14 systems. This IS the 3-day lag. Pipeline automation solves this."),
        ("Finance — Fund Accounting Ops",        32, 3.8, 2.4, -1.4, "NAV automation + offshore fund ops", 0.8, "SimCorp Dimension partially automated. Manual reconciliation steps remain."),
        ("Finance — Performance Attribution",    18, 2.4, 1.4, -1.0, "Automated performance attribution — real-time capable", 0.6, "T+1 performance. Manual attribution for complex strategies. AI can automate 80%."),
        ("Finance — Regulatory Reporting",       24, 3.2, 1.8, -1.4, "Automated regulatory filing — MAS, FCA, SEC", 0.8, "Manual regulatory reporting. MAS FEAT breach partly from manual process."),
        ("Risk — Model Validation Admin",        14, 1.8, 1.0, -0.8, "AI model validation support + offshore validation ops", 0.4, "Model validation backlog. Manual process. CDO vacancy compounds."),
        ("Compliance — KYC/AML Operations",      22, 2.6, 1.6, -1.0, "AI-assisted KYC + transaction monitoring automation", 0.6, "KYC refresh manual. AML monitoring 34% automated vs 72% peer."),
        ("IT Operations — Market Data",          12, 2.8, 1.8, -1.0, "Market data spend rationalisation + vendor consolidation", 0.4, "Bloomberg Terminal + FactSet + Reuters = overlapping data. Consolidation saves 15%."),
        ("IT Operations — Support Desk",         18, 1.8, 1.2, -0.6, "IT support offshore centre — standard financial services model", 0.4, "1,200 support tickets/month. 70% L1/L2 resolvable offshore."),
    ]

    for r, row in enumerate(functions, 3):
        ws.row_dimensions[r].height = 28
        func, fte, cost, bench, gap, auto_opp, savings, notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,func,sf,bold=True,b=bdr)
        C(ws,r,2,fte,sf,"#,##0",align="center",b=bdr)
        C(ws,r,3,cost,sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,4,bench,fills["grn"],"£#,##0.0",align="center",b=bdr)
        C(ws,r,5,gap,fills["red"] if gap<-1 else fills["amb"],"£#,##0.0",align="center",b=bdr)
        C(ws,r,6,auto_opp,fills["teal"],b=bdr)
        C(ws,r,7,savings,fills["grn"],"£#,##0.0",align="center",bold=True,b=bdr)
        C(ws,r,8,notes,sf,b=bdr)

    tr = len(functions)+3
    C(ws,tr,1,"TOTAL MIDDLE OFFICE OPPORTUNITY",fills["prp"],bold=True,b=bdr)
    C(ws,tr,3,sum(r[2] for r in functions),fills["prp"],"£#,##0.0",align="center",bold=True,b=bdr)
    C(ws,tr,7,sum(r[6] for r in functions),fills["grn"],"£#,##0.0",align="center",bold=True,b=bdr)
    ws.freeze_panes = "B3"
    save(wb, f"{base}/arcturus/margin/ARC-BO01_Middle_Office_Operations.xlsx")
    return "ARC-BO01 done"


def arc_m04_opportunity_map(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Margin Opportunity Map"
    T(ws, 9, f"{ARC['name']} — Full Margin Opportunity Map (£16.2B Asset Manager)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Category",22),("Lever",28),("Est Annual Opportunity £M",18),
            ("Genome Confidence",14),("Data Available?",14),("Data Required",22),
            ("Wave",10),("Status",14)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    opportunities = [
        # Already have data
        ("AI Portfolio", "AI portfolio reorientation — £94M idle", 94.0, "89%", "Yes — ARC-M02", "Already loaded", 1, "Ready to analyse"),
        ("AI Portfolio", "CDO hire — unblocks 14 initiatives", "£94M unlocked", "84%", "Yes — ARC-C03", "Already loaded", 1, "Ready to analyse"),
        ("IT Cost", "Consulting spend reduction — £42M, 24% KT", 22.4, "84%", "Yes — ARC-D01", "Already loaded", 1, "Ready to analyse"),
        ("IT Cost", "Bloomberg AIM overpay — £8.4M vs £2.1M peer", 6.3, "78%", "Yes — ARC-C04", "Already loaded", 1, "Ready to analyse"),
        ("IT Cost", "IT overspend £178M above peer benchmark", 12.0, "72%", "Yes — ARC-C02", "Already loaded", 1, "Ready to analyse"),
        ("Revenue", "FSC adoption 44% → 75% — cross-sell unlock", 8.4, "68%", "Yes — ARC-C04", "Already loaded", 2, "Ready to analyse"),
        ("Revenue", "Performance fee recovery — AI attribution", 6.2, "72%", "Partial", "Performance attribution data by strategy needed", 2, "Upload: ARC-FO01 or P&L by strategy"),
        # Need uploads
        ("Middle Office", "Settlement STP — failure rate 2.1% → 0.8%", 4.8, "78%", "No", "Settlement failure log, ops FTE count", 2, "Upload: ARC-BO01 or settlement data"),
        ("Middle Office", "Client reporting automation + offshore", 2.8, "74%", "No", "Report production FTE, report count, current process", 2, "Upload: middle office ops data"),
        ("Middle Office", "Reconciliation automation", 2.4, "72%", "No", "Break count by type, reconciliation FTE", 2, "Upload: operations data"),
        ("Revenue", "Fee yield analysis — AUM mix optimization", 12.0, "68%", "No", "Revenue by strategy, AUM by client type, fee schedule", 2, "Upload: revenue by strategy"),
        ("Revenue", "Client retention — churn analysis", 8.4, "72%", "No", "Client attrition data, AUM movement, NPS by segment", 2, "Upload: client analytics"),
        ("Finance", "Fund accounting automation — NAV ops", 1.6, "68%", "No", "Fund accounting FTE, NAV process timeline", 2, "Upload: finance ops data"),
        ("Finance", "Regulatory reporting automation", 1.6, "74%", "No", "Regulatory filing volume, manual FTE count", 2, "Upload: compliance ops data"),
        ("Compliance", "KYC/AML automation", 1.2, "72%", "No", "KYC refresh volume, AML alert volume, false positive rate", 2, "Upload: compliance data"),
        ("People", "Investment team productivity — AI research tools", 6.2, "68%", "No", "Research process, analyst hours by activity", 3, "Upload: front office time data"),
    ]

    for r, row in enumerate(opportunities, 3):
        ws.row_dimensions[r].height = 24
        cat, lever, opp, confidence, data_avail, data_req, wave, status = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        data_f = fills["grn"] if data_avail.startswith("Yes") else fills["amb"] if "Partial" in data_avail else fills["red"]
        wave_f = fills["grn"] if wave==1 else fills["amb"] if wave==2 else fills["wht"]
        status_f = fills["grn"] if "Ready" in status else fills["amb"]
        C(ws,r,1,cat,sf,bold=True,b=bdr); C(ws,r,2,lever,sf,b=bdr)
        opp_val = opp if isinstance(opp,(int,float)) else 0
        C(ws,r,3,opp,fills["teal"],"£#,##0.0" if isinstance(opp,(int,float)) else None,align="center",bold=True,b=bdr)
        C(ws,r,4,confidence,sf,align="center",b=bdr)
        C(ws,r,5,data_avail,data_f,b=bdr); C(ws,r,6,data_req,sf,b=bdr)
        C(ws,r,7,wave,wave_f,"#,##0",align="center",bold=True,b=bdr)
        C(ws,r,8,status,status_f,b=bdr)

    tr = len(opportunities)+3
    numeric_opps = [r[2] for r in opportunities if isinstance(r[2],(int,float))]
    C(ws,tr,1,f"TOTAL QUANTIFIED OPPORTUNITY",fills["prp"],bold=True,b=bdr)
    C(ws,tr,3,sum(numeric_opps),fills["prp"],"£#,##0.0",align="center",bold=True,b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/arcturus/margin/ARC-M04_Margin_Opportunity_Map.xlsx")
    return "ARC-M04 done"


# ══════════════════════════════════════════════════════════════════════════
# DYNAMIC DATA REQUEST TEMPLATES
# ══════════════════════════════════════════════════════════════════════════

def data_request_templates(base):
    """
    Generate the data request template library.
    When Maestro selects a focus area, platform pulls from this library
    to generate a client-facing upload request list.
    """
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Data Request Library"
    T(ws, 8, "AbarVa — Dynamic Data Request Template Library (All Solutions × All Focus Areas)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Solution",16),("Focus Area",20),("File / Data Requested",26),
            ("Format",12),("Why Needed",28),("What It Unlocks",24),
            ("Priority",10),("3-Number Alternative",28)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    templates = [
        # MARGIN — HEALTHCARE — RCM
        ("Margin","RCM — Denial Analysis","Denial rate report by payer (12 months)","Excel/CSV","Root cause denial analysis by payer and denial reason","Payer-specific recovery strategy. Prior auth automation ROI per payer.","Critical","1) Overall denial rate % 2) Top denial reason % 3) Days in AR"),
        ("Margin","RCM — Prior Auth","Prior auth denial log by procedure code","Excel/CSV","Procedure-level prior auth denial rate","Cohere Health compatibility assessment. Electronic prior auth ROI.","Critical","1) Prior auth denial rate % 2) Top denied procedure 3) Avg days to decision"),
        ("Margin","RCM — AR Analysis","Claims aging report (30/60/90/120 day buckets)","Excel/CSV","Days in AR diagnosis by payer and service line","Working capital opportunity. Payer performance ranking.","High","1) Total AR $M 2) >90 day AR % 3) Days in AR"),
        ("Margin","RCM — Payer Contracts","Payer contract summary or fee schedule","PDF/Excel","SLA penalty identification. Renegotiation leverage.","Enforceable penalty calculations. Ensemble renegotiation ammunition.","High","1) Ensemble annual fee $M 2) SLA breach count 3) Contract end date"),
        # MARGIN — HEALTHCARE — BACK OFFICE
        ("Margin","Back Office — Finance","Finance department org chart + headcount by function","Excel/PDF","FTE ratio vs benchmark calculation","BPO opportunity by function. Offshore savings estimate.","High","1) Finance FTE count 2) Annual finance dept cost $M 3) AP invoices processed/month"),
        ("Margin","Back Office — Finance","AP invoice processing volume and cost","Excel/CSV","Cost per invoice vs $8 benchmark","AP automation ROI. Offshore processing opportunity.","High","1) Monthly invoice count 2) Cost per invoice $ 3) % matched automatically"),
        ("Margin","Back Office — Finance","Month-end close timeline (last 3 cycles)","Excel","Close cycle benchmarking vs 8-day peer","Close automation opportunity. CFO credibility for accelerated close.","Medium","1) Close cycle days 2) Journal entry count 3) Manual adjustment count"),
        # MARGIN — HEALTHCARE — SUPPLY CHAIN
        ("Margin","Supply Chain","Annual spend by category with GPO vs non-GPO split","Excel","GPO compliance rate and off-contract spend identification","Compliance improvement opportunity. Price benchmark gap.","High","1) Total supply spend $M 2) GPO compliance % 3) Non-contract spend $M"),
        ("Margin","Supply Chain","Top 50 suppliers by spend","Excel/CSV","Consolidation opportunity. Vendor rationalization.","Vendor reduction opportunity. Negotiation leverage.","Medium","1) Total suppliers # 2) Top 10 suppliers % of spend 3) Sole-source contracts #"),
        ("Margin","Supply Chain","Inventory turns by category","Excel","Excess inventory and stockout analysis","Working capital improvement. Demand forecasting AI ROI.","Medium","1) Avg inventory turns 2) Stockout events/year 3) Expired inventory write-off $M"),
        # MARGIN — HEALTHCARE — WORKFORCE
        ("Margin","Workforce","Travel nurse spend by unit and agency (12 months)","Excel","Unit-level demand pattern analysis","AI demand forecasting savings estimate. Preferred agency programme ROI.","Critical","1) Travel nurse spend $M 2) Travel nurse FTE % of total 3) Top 3 units by TN usage"),
        ("Margin","Workforce","Turnover rate by department and role","Excel","Retention programme ROI. Recruitment cost analysis.","Turnover cost quantification. AI retention prediction value.","High","1) Overall turnover % 2) Nursing turnover % 3) Cost per hire $"),
        # MARGIN — ASSET MANAGEMENT — FEE/REVENUE
        ("Margin","Revenue — Fee Analysis","Revenue by strategy, AUM by strategy (12 months)","Excel","Fee yield analysis. Revenue per AUM by strategy.","Fee compression diagnosis. Highest-yield strategy identification.","High","1) Total AUM £B 2) Total management fee revenue £M 3) Performance fee £M"),
        ("Margin","Revenue — Client Analytics","Client attrition data (gained/lost AUM by client)","Excel/CSV","Churn analysis. At-risk client identification.","Retention programme ROI. Client lifetime value model.","High","1) Clients lost last 12mo # 2) AUM lost £M 3) Main reason for leaving"),
        # MARGIN — ASSET MANAGEMENT — OPERATIONS
        ("Margin","Middle Office","Settlement failure log (12 months)","Excel/CSV","Settlement failure rate vs 0.8% benchmark","STP automation ROI. Failed trade cost quantification.","High","1) Settlement fail rate % 2) Failed trade cost £M pa 3) Manual resolution FTE #"),
        ("Margin","Middle Office","Reconciliation breaks by type and age","Excel","Break resolution time and cost","Reconciliation automation opportunity. Offshore ops model.","Medium","1) Daily break count avg 2) >5 day breaks % 3) Reconciliation FTE count"),
        # TECH — HEALTHCARE
        ("Tech","Epic Optimization","Epic optimization score by module (Epic Compass report)","PDF/Excel","Module-level utilization vs 80/100 benchmark","Module activation roadmap. $36.5M unrealized value prioritization.","Critical","1) Epic optimization score /100 2) MyChart adoption % 3) Training completion %"),
        ("Tech","Cerner Migration","Cerner configuration documentation for 2 hospitals","PDF","Migration complexity assessment. Data migration scope.","Migration timeline and cost estimate. Risk identification.","High","1) Hospital beds # 2) Years on Cerner 3) Customizations count"),
        # TECH — ASSET MANAGEMENT
        ("Tech","Bloomberg AIM","Bloomberg AIM customization register (if documented)","Excel/PDF","Customization complexity scoring. Migration risk assessment.","Which modernization approach is viable. Why prior attempts failed.","Critical","1) AIM customizations count 2) Bloomberg annual cost £M 3) Internal AIM team FTE"),
        ("Tech","Data Architecture","Data dictionary or system inventory","Excel/PDF","Golden record gap analysis. Integration map completion.","Data pipeline automation scope. ML data readiness score.","High","1) Total systems count 2) Manual ETL processes count 3) Reporting lag days"),
        # PDLC
        ("PDLC","Delivery Velocity","Sprint velocity report by squad (12 months)","Excel/CSV","Actual cycle time vs benchmark. Delivery predictability.","Cycle time baseline for Phase 3. DORA metric calculation.","Critical","1) Avg sprint velocity % 2) Avg cycle time days 3) Deployments per month"),
        ("PDLC","AI Portfolio","AI initiative inventory with status, budget, blockers","Excel","28-initiative Genome pattern matching. Blocker analysis.","Which initiatives to accelerate. MLOps priority sequencing.","Critical","1) AI initiatives count 2) In production count 3) Total budget £M"),
        ("PDLC","MLOps","MLOps capability assessment or tool inventory","Excel/PDF","MLOps maturity scoring. Deployment infrastructure gap.","Infrastructure build sequence. Time to first AI in production.","High","1) Model registry? Y/N 2) CI/CD for ML? Y/N 3) Serving infrastructure? Y/N"),
    ]

    for r, row in enumerate(templates, 3):
        ws.row_dimensions[r].height = 28
        solution, focus, file_req, fmt, why, unlocks, priority, alt_3num = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        pf = fills["red"] if priority=="Critical" else fills["amb"] if priority=="High" else fills["grn"]
        sol_f = {"Margin":fills["teal"],"Tech":fills["prp"],"PDLC":fills["blu"]}.get(solution,sf)
        C(ws,r,1,solution,sol_f,align="center",bold=True,b=bdr)
        C(ws,r,2,focus,sf,bold=True,b=bdr); C(ws,r,3,file_req,sf,b=bdr)
        C(ws,r,4,fmt,sf,align="center",b=bdr); C(ws,r,5,why,sf,b=bdr)
        C(ws,r,6,unlocks,fills["grn"],b=bdr)
        C(ws,r,7,priority,pf,align="center",bold=True,b=bdr)
        C(ws,r,8,alt_3num,fills["amb"],b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/data_request_templates.xlsx")
    return "Data Request Templates done"


# ══════════════════════════════════════════════════════════════════════════
# RUNNER
# ══════════════════════════════════════════════════════════════════════════

def run_task(args):
    func, base = args
    try:
        result = func(base)
        return [result] if isinstance(result, str) else result
    except Exception as e:
        return [f"ERROR in {func.__name__}: {e}\n{traceback.format_exc()}"]

if __name__ == "__main__":
    tasks = [
        (mer_bo01_back_office, BASE),
        (mer_sc01_supply_chain, BASE),
        (mer_hr01_workforce_operations, BASE),
        (mer_it01_it_operations, BASE),
        (mer_m05_opportunity_map, BASE),
        (arc_bo01_middle_office, BASE),
        (arc_m04_opportunity_map, BASE),
        (data_request_templates, BASE),
    ]

    print(f"\n{'='*60}")
    print(f"AbarVa — Margin Extended Datasets + Data Request Library")
    print(f"{len(tasks)} files — Meridian + Arcturus — BPO/Back Office/Supply Chain")
    print(f"{'='*60}\n")

    start = time.time()
    with mp.Pool(processes=min(len(tasks), mp.cpu_count())) as pool:
        results = pool.map(run_task, tasks)
    elapsed = time.time() - start

    all_results = [r for group in results for r in group]
    errors = [r for r in all_results if "ERROR" in r]
    successes = [r for r in all_results if "ERROR" not in r]

    print("Results:")
    for r in all_results:
        print(f"  {'OK' if 'ERROR' not in r else 'FAIL'} {r}")

    print(f"\n{'='*60}")
    print(f"Complete: {len(successes)} files, {len(errors)} errors, {elapsed:.1f}s")

    all_files = list(BASE.rglob("*.xlsx"))
    all_files = [f for f in all_files if "scripts" not in str(f)]
    print(f"\nNew files:")
    for f in sorted(all_files):
        if any(x in str(f) for x in ["BO01","SC01","HR01","IT01","M04","M05","data_request"]):
            print(f"  {str(f.relative_to(BASE))}")

    if errors:
        for e in errors: print(e)
        sys.exit(1)
    else:
        print(f"\nRun: git add . && git commit -m 'datasets: margin extended + data request library'")
