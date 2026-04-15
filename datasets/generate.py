"""
AbarVa — Complete Dataset Generator v2
70 files across Arcturus + Meridian, all cross-referenced.
Run: python3 generate.py
Requires: pip install openpyxl
"""
import sys, time, traceback, multiprocessing as mp
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).parent))
from constants import ARC, MER, GENOME

BASE = Path(__file__).parent

# ── Styling helpers ────────────────────────────────────────────────────────
def S():
    t = Side(style="thin", color="CCCCCC")
    b = Border(left=t, right=t, top=t, bottom=t)
    f = {
        "hdr":  PatternFill("solid", start_color="1A3A5C"),
        "alt":  PatternFill("solid", start_color="F2F7FC"),
        "wht":  PatternFill("solid", start_color="FFFFFF"),
        "red":  PatternFill("solid", start_color="FDE8E8"),
        "amb":  PatternFill("solid", start_color="FFF4E5"),
        "grn":  PatternFill("solid", start_color="E8F5E9"),
        "blu":  PatternFill("solid", start_color="E8F0F8"),
        "prp":  PatternFill("solid", start_color="F0EEFF"),
        "teal": PatternFill("solid", start_color="E0F7F4"),
        "dkred":PatternFill("solid", start_color="FF6B6B"),
    }
    return b, f

def H(ws, row, col, val, w=16, f=None, b=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    c.fill = f["hdr"]; c.border = b
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(col)].width = w

def C(ws, row, col, val, fill=None, fmt=None, bold=False, align="left", b=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(size=9, name="Arial", bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = b
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    return c

def T(ws, cols, text, f):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    ws["A1"].value = text
    ws["A1"].font = Font(bold=True, size=12, color="1A3A5C", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = f["blu"]; ws.row_dimensions[1].height = 16

def save(wb, path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

def rf(f, val, hi_bad=True):
    """Return fill based on value — red=bad/high, grn=good/low"""
    if not isinstance(val, (int, float)): return None
    if hi_bad:
        return f["red"] if val > 0.75 else f["amb"] if val > 0.50 else f["grn"]
    else:
        return f["grn"] if val > 0.75 else f["amb"] if val > 0.50 else f["red"]

# ══════════════════════════════════════════════════════════════════════════
# ARCTURUS — CORE
# ══════════════════════════════════════════════════════════════════════════

def arc_c01_engineering_org(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Squad Overview"
    T(ws, 14, f"{ARC['name']} — Engineering Organisation & Squad Structure {ARC['fy']}", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Squad",26),("Domain",18),("Squad Lead",18),("FTE",8),("Contractors",10),
            ("Vendor Staff",10),("Total",8),("Vendor Dep %",11),("Primary Vendor",16),
            ("Mtg Hrs/Wk",10),("Build Hrs/Wk",10),("Sprint Method",13),
            ("Cycle Time Days",13),("DORA Level",10),("Status",10)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    squad_leads = {
        "OMS Core Engineering":     ARC["people"]["VP_OMS"],
        "OMS Integration & APIs":   "Priya Nambiar",
        "Risk Technology":          "Marcus Webb",
        "Client Data Platform":     "-VACANT-",
        "Portfolio Analytics":      ARC["people"]["Head_PA"],
        "Compliance & Regulatory":  "Thomas Brennan (CCO-interim)",
        "Client Portal (FSC)":      "Anna Johansson (MD)",
        "Data Engineering":         "Raj Sharma",
        "AI/ML Platform":           ARC["people"]["Head_AI"],
        "Singapore Technology":     ARC["people"]["SG_Tech"],
        "New York Markets":         ARC["people"]["NY_Tech"],
        "Enterprise Architecture":  ARC["people"]["EA_Lead"],
        "DevOps & Infrastructure":  "Amir Patel",
        "Digital Innovation Lab":   f"({ARC['people']['CDO']})",
    }
    squad_domains = {
        "OMS Core Engineering":    "Technology - Trading Systems",
        "OMS Integration & APIs":  "Technology - Trading Systems",
        "Risk Technology":         "Technology - Risk",
        "Client Data Platform":    "Technology - Data",
        "Portfolio Analytics":     "Technology - Investments",
        "Compliance & Regulatory": "Technology - Compliance",
        "Client Portal (FSC)":     "Technology - Client",
        "Data Engineering":        "Technology - Data",
        "AI/ML Platform":          "Technology - Innovation",
        "Singapore Technology":    "Technology - APAC",
        "New York Markets":        "Technology - Americas",
        "Enterprise Architecture": "Technology - Architecture",
        "DevOps & Infrastructure": "Technology - Platform",
        "Digital Innovation Lab":  "Technology - Innovation",
    }
    squad_methods = {
        "OMS Core Engineering":    "Scrum",  "OMS Integration & APIs":  "Scrum",
        "Risk Technology":         "Scrum",  "Client Data Platform":    "Kanban",
        "Portfolio Analytics":     "Scrum",  "Compliance & Regulatory": "Waterfall",
        "Client Portal (FSC)":     "Scrum",  "Data Engineering":        "Kanban",
        "AI/ML Platform":          "Scrum",  "Singapore Technology":    "Scrum",
        "New York Markets":        "Scrum",  "Enterprise Architecture": "N/A",
        "DevOps & Infrastructure": "Kanban", "Digital Innovation Lab":  "Agile",
    }
    squad_mtg = {
        "OMS Core Engineering":22,"OMS Integration & APIs":24,"Risk Technology":18,
        "Client Data Platform":26,"Portfolio Analytics":19,"Compliance & Regulatory":28,
        "Client Portal (FSC)":25,"Data Engineering":22,"AI/ML Platform":20,
        "Singapore Technology":18,"New York Markets":21,"Enterprise Architecture":30,
        "DevOps & Infrastructure":16,"Digital Innovation Lab":24,
    }

    status_fills = {"Critical":fills["red"],"Amber":fills["amb"],"Green":fills["grn"],"None":fills["red"]}
    squad_statuses = {
        "OMS Core Engineering":"Critical","OMS Integration & APIs":"Critical",
        "Risk Technology":"Amber","Client Data Platform":"Critical",
        "Portfolio Analytics":"Green","Compliance & Regulatory":"Critical",
        "Client Portal (FSC)":"Amber","Data Engineering":"Critical",
        "AI/ML Platform":"Critical","Singapore Technology":"Amber",
        "New York Markets":"Amber","Enterprise Architecture":"Critical",
        "DevOps & Infrastructure":"Amber","Digital Innovation Lab":"Critical",
    }

    total_fte=total_contr=total_consult=0
    for r, sq in enumerate(ARC["squads"], 3):
        ws.row_dimensions[r].height = 22
        name = sq["name"]; fte=sq["fte"]; contr=sq["contractors"]; consult=sq["consulting"]
        total = fte+contr+consult
        total_fte+=fte; total_contr+=contr; total_consult+=consult
        build = 40-squad_mtg.get(name,20)
        sf = fills["alt"] if r%2==0 else fills["wht"]
        status = squad_statuses.get(name,"Amber")
        C(ws,r,1,name,sf,bold=True,b=bdr); C(ws,r,2,squad_domains.get(name,""),sf,b=bdr)
        lead = squad_leads.get(name,"")
        lf = fills["red"] if "VACANT" in lead or "vacant" in lead.lower() else sf
        C(ws,r,3,lead,lf,b=bdr)
        C(ws,r,4,fte,sf,"#,##0",align="center",b=bdr)
        C(ws,r,5,contr,sf,"#,##0",align="center",b=bdr)
        C(ws,r,6,consult,sf,"#,##0",align="center",b=bdr)
        C(ws,r,7,f"=D{r}+E{r}+F{r}",sf,"#,##0",align="center",b=bdr)
        C(ws,r,8,f"=(E{r}+F{r})/G{r}",fills["red"] if (contr+consult)/total>0.55 else fills["amb"],"0.0%",align="center",b=bdr)
        C(ws,r,9,sq["vendor"],sf,b=bdr)
        mtg=squad_mtg.get(name,20)
        C(ws,r,10,mtg,fills["red"] if mtg>=24 else fills["amb"] if mtg>=18 else fills["grn"],"#,##0",align="center",b=bdr)
        C(ws,r,11,build,sf,"#,##0",align="center",b=bdr)
        C(ws,r,12,squad_methods.get(name,"Scrum"),sf,align="center",b=bdr)
        cycle=sq["cycle_days"]
        C(ws,r,13,cycle if cycle else "Not tracked",fills["red"] if cycle>120 else fills["amb"] if cycle>80 else fills["grn"] if cycle>0 else fills["red"],align="center",b=bdr)
        C(ws,r,14,sq["dora"],fills["grn"] if sq["dora"]=="High" else fills["amb"] if sq["dora"]=="Medium" else fills["red"],align="center",b=bdr)
        cs=C(ws,r,15,status,status_fills.get(status,sf),align="center",bold=True,b=bdr)
        cs.font=Font(size=9,name="Arial",bold=True,color="FFFFFF" if status=="Critical" else "333333")

    # Totals
    tr=len(ARC["squads"])+3
    C(ws,tr,1,f"TOTALS — {total_fte+total_contr+total_consult} headcount",fills["hdr"],bold=True,b=bdr)
    ws.cell(tr,1).font=Font(bold=True,color="FFFFFF",size=9,name="Arial")
    for col,val in [(4,total_fte),(5,total_contr),(6,total_consult)]:
        ws.cell(tr,col,value=val); ws.cell(tr,col).number_format="#,##0"
        ws.cell(tr,col).font=Font(bold=True,color="FFFFFF",size=9,name="Arial")
        ws.cell(tr,col).fill=fills["hdr"]; ws.cell(tr,col).alignment=Alignment(horizontal="center")
        ws.cell(tr,col).border=bdr
    tot=total_fte+total_contr+total_consult
    ws.cell(tr,7,value=tot); ws.cell(tr,7).number_format="#,##0"
    ws.cell(tr,7).font=Font(bold=True,color="FFFFFF",size=9,name="Arial")
    ws.cell(tr,7).fill=fills["hdr"]; ws.cell(tr,7).alignment=Alignment(horizontal="center"); ws.cell(tr,7).border=bdr
    vdr_pct=(total_contr+total_consult)/tot
    ws.cell(tr,8,value=vdr_pct); ws.cell(tr,8).number_format="0.0%"
    ws.cell(tr,8).font=Font(bold=True,color="FFFFFF",size=9,name="Arial")
    ws.cell(tr,8).fill=fills["hdr"]; ws.cell(tr,8).alignment=Alignment(horizontal="center"); ws.cell(tr,8).border=bdr

    # Key insight banner
    tr2=tr+2
    ws.merge_cells(f"A{tr2}:O{tr2}")
    ws.cell(tr2,1,value=f"KEY INSIGHT: Vendor dependency ratio {vdr_pct:.0%}. {ARC['people']['CDO']}. {ARC['cdo_vacancy']['initiatives_blocked']} of {ARC['ai_initiatives']} AI initiatives blocked. Annual consulting spend £{ARC['total_consulting_m']:.0f}M.")
    ws.cell(tr2,1).font=Font(bold=True,size=9,name="Arial",color="CC0000")
    ws.cell(tr2,1).fill=fills["red"]; ws.row_dimensions[tr2].height=18

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/core/ARC-C01_Engineering_Organisation.xlsx")
    return "ARC-C01 done"


def arc_c02_financial_statements(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "P&L Summary"
    T(ws, 6, f"{ARC['name']} — Financial Statements Summary {ARC['fy']} (£M)", fills)
    ws.row_dimensions[2].height = 28

    # P&L
    hdrs=[("Line Item",32),("FY2023 (£M)",14),("FY2024 (£M)",14),("FY2025 (£M)",14),("Budget (£M)",14),("Variance (£M)",14)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    pl_lines=[
        ("REVENUE","","","","",""),
        ("Management fees",          610,  648,  672,  680,   -8),
        ("Performance fees",          82,   61,   48,   90,  -42),
        ("Advisory & other fees",     52,   54,   60,   58,    2),
        ("TOTAL REVENUE",            744,  763,  780,  828,  -48),
        ("","","","","",""),
        ("OPERATING COSTS","","","","",""),
        ("Employee compensation",    312,  328,  348,  340,   -8),
        ("Technology & systems",      28,   32,   37,   30,   -7),
        ("Bloomberg AIM licence",      7,    8,    8,    7,   -1),
        ("Consulting & contractors",  38,   40,   42,   35,   -7),
        ("Data & market data",        18,   19,   21,   19,   -2),
        ("Premises & occupancy",      24,   25,   26,   25,   -1),
        ("Regulatory & compliance",   12,   14,   16,   14,   -2),
        ("AI portfolio (no ROI)",      0,   42,   94,   40,  -54),
        ("Other operating costs",     48,   52,   55,   50,   -5),
        ("TOTAL OPERATING COSTS",    487,  560,  647,  560,  -87),
        ("","","","","",""),
        ("OPERATING PROFIT",         257,  203,  133,  268, -135),
        ("OPERATING MARGIN %",      0.345,0.266,0.171,0.324, None),
        ("","","","","",""),
        ("CORPORATE / TECHNOLOGY OVERHEAD","","","","",""),
        ("Corporate IT infrastructure",28,   30,   37,   28,   -9),
        ("Digital transformation prog.",12,  18,   24,   15,   -9),
        ("AI governance & CDO function", 2,   3,    0,    4,    4),
        ("TOTAL CORP/TECH OVERHEAD",   42,   51,   61,   47,  -14),
        ("","","","","",""),
        ("PROFIT BEFORE TAX",        215,  152,   72,  221, -149),
        ("EFFECTIVE TAX RATE",       0.21, 0.21,  0.21, 0.21, None),
        ("PROFIT AFTER TAX",         170,  120,   57,  175, -118),
        ("","","","","",""),
        ("KEY RATIOS","","","","",""),
        ("Cost-to-Income Ratio",    0.655,0.734,0.710,0.580,-0.130),
        ("Tech Cost / Revenue",     0.038,0.042,0.047,0.037,-0.010),
        ("AI Spend / Revenue",      0.000,0.055,0.121,0.048,-0.073),
        ("AI Verified ROI / AI Spend",0.0, 0.0,  0.0,  0.50, -0.50),
    ]

    section_rows={"REVENUE","OPERATING COSTS","CORPORATE / TECHNOLOGY OVERHEAD","KEY RATIOS"}
    total_rows={"TOTAL REVENUE","TOTAL OPERATING COSTS","OPERATING PROFIT","TOTAL CORP/TECH OVERHEAD","PROFIT BEFORE TAX","PROFIT AFTER TAX"}

    for r,row in enumerate(pl_lines,3):
        ws.row_dimensions[r].height=18
        label,fy23,fy24,fy25,budget,var=row
        if not label: continue
        is_section=label in section_rows; is_total=label in total_rows
        sf=fills["prp"] if is_total else fills["blu"] if is_section else fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,label,sf,bold=is_total or is_section,b=bdr)
        for col,val in [(2,fy23),(3,fy24),(4,fy25),(5,budget),(6,var)]:
            if val=="": C(ws,r,col,None,sf,b=bdr); continue
            if isinstance(val,float) and val<2:
                fmt="0.0%" if abs(val)<2 else "£#,##0.0"
                vf=fills["red"] if "Ratio" in label and val>0.65 else fills["grn"] if "Ratio" in label and val<0.60 else sf
            else:
                fmt="£#,##0.0"
                vf=fills["red"] if col==6 and isinstance(val,(int,float)) and val<-10 else fills["grn"] if col==6 and isinstance(val,(int,float)) and val>5 else sf
            C(ws,r,col,val,vf if col in [4,6] else sf,fmt,bold=is_total,align="center",b=bdr)

    ws.freeze_panes="B3"

    # Sheet 2 — IT cost breakdown (ties to F08)
    ws2=wb.create_sheet("IT Cost Breakdown")
    T(ws2,7,f"{ARC['name']} — IT Cost Detail & Peer Benchmark Comparison {ARC['fy']} (£M)",fills)
    ws2.row_dimensions[2].height=28
    it_hdrs=[("Category",24),("Arcturus (£M)",13),("% of Revenue",13),
             ("Peer Median (£M)",14),("Peer %",10),("Overspend (£M)",13),("Notes",28)]
    for i,(hd,w) in enumerate(it_hdrs,1): H(ws2,2,i,hd,w,fills,bdr)
    it_costs=[
        ("Bloomberg AIM — licence + maintenance",ARC["bloomberg"]["annual_cost_m"],None,2.1,None,None,"28-year-old OMS. Peers pay £2.1M for equivalent. Vendor lock-in premium."),
        ("Other trading systems (Aladdin, etc.)",7.1,None,6.8,None,None,"BlackRock Aladdin + MTS + Bloomberg Terminal. Reasonable."),
        ("Client technology (Salesforce FSC)",3.1,None,2.4,None,None,"44% adoption vs 78% target. Cost per active user excessive."),
        ("Data & analytics platforms",4.2,None,3.8,None,None,"FactSet, Bloomberg data, internal DW. Reasonable."),
        ("Infrastructure & cloud",3.8,None,4.1,None,None,"AWS + on-premise. At benchmark. ML infrastructure absent."),
        ("AI portfolio (no documented ROI)",ARC["ai_budget_committed_m"],None,8.0,None,None,f"£{ARC['ai_budget_committed_m']:.0f}M committed. £{ARC['ai_verified_roi_m']:.0f}M verified ROI. Genome pattern F010."),
        ("Consulting (tech-related)",ARC["total_consulting_m"],None,18.0,None,None,f"£{ARC['total_consulting_m']:.0f}M annual. £{ARC['total_consulting_m']-18:.0f}M above peer median. Vendor dependency embedded."),
        ("Internal technology staff",24.0,None,22.0,None,None,"67 FTEs + 37 contractors. Contractor ratio 47% vs 28% peer."),
        ("Other IT costs",4.2,None,3.8,None,None,"Licences, support, misc."),
    ]
    total_arc=sum(r[1] for r in it_costs if isinstance(r[1],(int,float)))
    total_peer=sum(r[3] for r in it_costs if isinstance(r[3],(int,float)))
    for r,row in enumerate(it_costs,3):
        ws2.row_dimensions[r].height=22
        cat,arc_v,_,peer_v,__,___,notes=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws2,r,1,cat,sf,bold=True,b=bdr)
        C(ws2,r,2,arc_v,sf,"£#,##0.0",align="center",b=bdr)
        C(ws2,r,3,arc_v/ARC["total_revenue_m"] if isinstance(arc_v,(int,float)) else None,sf,"0.00%",align="center",b=bdr)
        C(ws2,r,4,peer_v,sf,"£#,##0.0",align="center",b=bdr)
        C(ws2,r,5,peer_v/ARC["total_revenue_m"] if isinstance(peer_v,(int,float)) else None,sf,"0.00%",align="center",b=bdr)
        overspend=arc_v-peer_v if isinstance(arc_v,(int,float)) and isinstance(peer_v,(int,float)) else None
        C(ws2,r,6,overspend,fills["red"] if overspend and overspend>2 else fills["grn"] if overspend and overspend<0 else sf,"£#,##0.0",align="center",b=bdr)
        C(ws2,r,7,notes,sf,b=bdr)
    tr2=len(it_costs)+3
    C(ws2,tr2,1,f"TOTAL — {ARC['it_pct_revenue']:.1%} of revenue vs {ARC['it_peer_benchmark_pct']:.1%} peer",fills["prp"],bold=True,b=bdr)
    C(ws2,tr2,2,total_arc,fills["prp"],"£#,##0.0",align="center",bold=True,b=bdr)
    C(ws2,tr2,3,total_arc/ARC["total_revenue_m"],fills["prp"],"0.00%",align="center",bold=True,b=bdr)
    C(ws2,tr2,4,total_peer,fills["prp"],"£#,##0.0",align="center",bold=True,b=bdr)
    C(ws2,tr2,5,total_peer/ARC["total_revenue_m"],fills["prp"],"0.00%",align="center",bold=True,b=bdr)
    C(ws2,tr2,6,total_arc-total_peer,fills["red"],"£#,##0.0",align="center",bold=True,b=bdr)

    save(wb, f"{base}/arcturus/core/ARC-C02_Financial_Statements.xlsx")
    return "ARC-C02 done"


def arc_c03_leadership(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Leadership Register"
    T(ws, 8, f"{ARC['name']} — Technology & AI Leadership Register {ARC['fy']}", fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Role",24),("Name",22),("Status",16),("Reports To",18),
          ("AI Initiatives Sponsored",13),("Decision Authority",22),("Risk",10),("Notes",34)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    leaders=[
        ("CEO",ARC["people"]["CEO"],"Active","Board of Directors",6,"Final approval >£5M","Low","AI-native commitment stated. Aware of £94M portfolio gap. Driving CDO hire."),
        ("CFO",ARC["people"]["CFO"],"Active","CEO",3,"Budget approval, fee authorisation","Low","Conservative on AI. Focused on C/I reduction. Key AbarVa engagement approver."),
        ("CIO",ARC["people"]["CIO"],"Active","CEO",8,"Technology investment <£5M","Medium","Strong tech background. Frustrated by delivery velocity. 14 squads reporting."),
        ("CDO",ARC["people"]["CDO"],"VACANT — 11 months","CIO (interim)",14,"None — role vacant","Critical",f"{ARC['cdo_vacancy']['initiatives_blocked']} of {ARC['ai_initiatives']} AI initiatives blocked. AI governance council cannot convene. MAS FEAT oversight absent."),
        ("CCO",ARC["people"]["CCO"],"Active","CEO",2,"Regulatory decisions","High","MAS FEAT breach on his watch. Under regulatory scrutiny. MAS engagement ongoing."),
        ("MD Client Solutions",ARC["people"]["MD_Client"],"Active","CEO",1,"Client platform decisions","Medium",f"Salesforce FSC adoption accountability. 44% vs 78% target. £38M invested."),
        ("Head of Risk",ARC["people"]["Head_Risk"],"Active","CIO",2,"Risk model approval","Medium","CDO vacancy means model risk governance falls here. Capacity strained."),
        ("Head of Portfolio Analytics",ARC["people"]["Head_PA"],"Active","CIO",0,"Squad delivery","Low","Best performing squad lead. Advocate for AI. Frustrated by cross-squad blockers."),
        ("VP Engineering — OMS",ARC["people"]["VP_OMS"],"Active","CIO",1,"OMS squad delivery","High","Bloomberg dependency blocker. Aware but lacks authority to resolve."),
        ("VP Engineering — Data",ARC["people"]["VP_Data"],"Contractor (interim)","CIO",0,"Data squad delivery","Critical","Contractor in permanent role. No knowledge retention obligation. Wipro relationship managed by contractor."),
        ("Head of AI/ML",ARC["people"]["Head_AI"],"Active","CDO (VACANT)","—","None — reports to vacant role","Critical","Talented. Frustrated. 3 engineers interviewing elsewhere. Zero production deliveries in 12 months."),
        ("Singapore Technology Lead",ARC["people"]["SG_Tech"],"Active","CIO",0,"APAC tech delivery","Medium","MAS FEAT compliance gap primary concern. Singapore AUM data segregation required."),
        ("AI Governance Council","Not constituted","Requires CDO","CDO (VACANT)","N/A","AI initiative approval","Critical","Cannot convene without CDO. 14 initiatives awaiting governance sign-off."),
    ]

    stat_fills={"Active":fills["grn"],"VACANT — 11 months":fills["red"],"Contractor (interim)":fills["amb"],"Not constituted":fills["red"]}
    risk_fills={"Critical":fills["red"],"High":fills["amb"],"Medium":fills["amb"],"Low":fills["grn"]}
    for r,row in enumerate(leaders,3):
        ws.row_dimensions[r].height=28
        role,name,status,reports,ai_c,auth,risk,notes=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,role,sf,bold=True,b=bdr)
        C(ws,r,2,name,fills["red"] if "VACANT" in name or "Contractor" in name else sf,bold="VACANT" in name,b=bdr)
        cs=C(ws,r,3,status,stat_fills.get(status,sf),align="center",bold=True,b=bdr)
        cs.font=Font(size=9,name="Arial",bold=True)
        C(ws,r,4,reports,sf,b=bdr)
        C(ws,r,5,ai_c,fills["red"] if isinstance(ai_c,int) and ai_c>=8 else fills["amb"] if isinstance(ai_c,int) and ai_c>=3 else sf,"#,##0" if isinstance(ai_c,int) else None,align="center",b=bdr)
        C(ws,r,6,auth,sf,b=bdr)
        cr=C(ws,r,7,risk,risk_fills.get(risk,sf),align="center",bold=True,b=bdr)
        C(ws,r,8,notes,sf,b=bdr)

    # Strategic commitments sheet
    ws2=wb.create_sheet("Strategic Commitments vs Reality")
    T(ws2,6,f"{ARC['name']} — Board Commitments vs Delivery Reality",fills)
    ws2.row_dimensions[2].height=28
    sc_hdrs=[("Commitment",34),("Source",20),("Date Made",12),("Current Status",30),("Gap",30),("Risk",10)]
    for i,(hd,w) in enumerate(sc_hdrs,1): H(ws2,2,i,hd,w,fills,bdr)
    commitments=[
        (f"'We will be AI-native by 2026'","CEO Annual Report 2024","2024-03-15",
         f"0 of {ARC['ai_initiatives']} AI initiatives in production. No MLOps. CDO vacant {ARC['cdo_vacancy']['months_vacant']} months.",
         f"AI-native requires production AI. Not one model deployed. 2026 target impossible without immediate structural change.","Critical"),
        (f"'AI will reduce C/I ratio to {ARC['ci_target']:.0%} by 2027'","Investor Day 2024","2024-06-10",
         f"C/I ratio {ARC['ci_ratio']:.0%}. No AI-driven cost reduction with documented baseline.",
         f"{ARC['ci_ratio']-ARC['ci_target']:.0%} gap to target. No programme exists to deliver it. £{ARC['ai_budget_committed_m']:.0f}M AI spend has zero verified ROI.","Critical"),
        ("'Bloomberg AIM modernisation begins 2025'","Board Technology Committee","2024-09-18",
         "No modernisation programme started. No vendor selected. No business case approved.",
         f"3rd failed commitment. Same statement made in 2009 and 2016. Root cause — F001+F002 — unaddressed.","Critical"),
        (f"'MAS FEAT compliance by {ARC['mas_feat']['deadline']}'",f"Regulatory Commitment to MAS","2024-01-15",
         f"MAS FEAT {ARC['mas_feat']['status']}. Regulatory notice received. £{ARC['mas_feat']['singapore_aum_b']:.1f}B Singapore AUM at risk.",
         "Regulatory breach. Not a strategy gap — a compliance failure with material financial consequences.","Critical"),
        ("'Golden record for client data by Q3 2025'","CDO Appointment Brief","2024-02-01",
         "CDO appointed then left after 4 months. Role vacant 11 months. Golden record not started.",
         "Initiative orphaned. No owner. 14 siloed systems unchanged. 3-day reporting lag persists.","Critical"),
        ("'AI Governance Council operational Q1 2025'","Board Risk Committee","2023-12-05",
         "AI Governance Council not constituted. No CDO to chair.",
         f"{ARC['cdo_vacancy']['governance_decisions_blocked']} governance decisions blocked. FCA model risk examination risk elevated.","Critical"),
        (f"'Salesforce FSC adoption 80% by Q2 2025'","MD Client Solutions plan","2024-01-20",
         "Adoption 44% vs 78% target. Q2 2025 deadline missed by 34pp.",
         "£38M invested. 44% adopted. Cost per active user ~£2,100/month.","High"),
        ("'Consulting spend to reduce 20% by year-end 2025'","CFO cost reduction plan","2024-04-01",
         f"Consulting spend unchanged at £{ARC['total_consulting_m']:.0f}M. Google PSO ended but Wipro expanded.",
         "Net consulting spend flat. No structural reduction programme.","High"),
    ]
    for r,row in enumerate(commitments,3):
        ws2.row_dimensions[r].height=40
        commit,source,date,status,gap,risk=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        rf2=risk_fills.get(risk,sf)
        C(ws2,r,1,commit,sf,bold=True,b=bdr); C(ws2,r,2,source,sf,b=bdr)
        C(ws2,r,3,date,sf,align="center",b=bdr)
        C(ws2,r,4,status,fills["red"],b=bdr); C(ws2,r,5,gap,sf,b=bdr)
        C(ws2,r,6,risk,rf2,align="center",bold=True,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/core/ARC-C03_Leadership_Governance.xlsx")
    return "ARC-C03 done"


def arc_c04_technology_landscape(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "System Inventory"
    T(ws,11,f"{ARC['name']} — Technology Landscape {ARC['fy']}",fills)
    ws.row_dimensions[2].height=32
    hdrs=[("System",24),("Vendor",18),("Category",16),("Age Yrs",9),
          ("Annual Cost £M",13),("Data Domain",18),("Integrations",10),
          ("EOL Status",16),("Internal Capability",14),("AI Ready?",10),("Notes",32)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    systems=[
        ("Bloomberg AIM","Bloomberg LP","Order Management",28,ARC["bloomberg"]["annual_cost_m"],"Orders, positions, executions",14,"Supported (vendor controlled)","Low","No",f"28yr OMS. {ARC['bloomberg']['customisations']} customisations. {ARC['bloomberg']['failed_modernisations']} failed migrations. £{ARC['bloomberg']['annual_cost_m']:.1f}M annual."),
        ("BlackRock Aladdin","BlackRock","Risk Management",11,4.2,"Portfolio risk, stress, analytics",8,"Supported","Medium","Partial","Monthly stress testing. Regulatory daily requirement unmet. Internal team capable."),
        ("Salesforce FSC","Salesforce","CRM / Client Platform",4,3.1,"Client profiles, AUM, interactions",6,"Supported","Low","Yes","44% adoption vs 78% target. £38M invested. Wipro owns customisations. Internal team cannot deploy."),
        ("SimCorp Dimension","SimCorp","Portfolio Accounting",9,3.8,"Fund accounting, NAV, positions",7,"Supported","Medium","Partial","UCITS and alternatives accounting. Stable."),
        ("Charles River IMS","Charles River (SS&T)","Investment Mgmt System",6,2.9,"Orders, compliance, analytics",5,"Supported","High","Partial","Best internal capability. Portfolio Analytics squad."),
        ("FactSet Research","FactSet","Market Data / Research",8,2.4,"Market data, research",4,"Supported","High","Yes","Highest AI readiness. AI-023 and AI-026 rely on this."),
        ("Bloomberg Terminal","Bloomberg LP","Market Data",15,1.8,"Prices, news, analytics",12,"Supported","High","Partial","Separate from AIM. Standard tool. Different risk profile."),
        ("SQL Server 2017 DW","Microsoft","Data Warehouse",7,0.4,"Aggregated reporting, MIS",9,"EOL Oct 2025 — PAST","Low","No","EOL passed. Operating without security patches. 3-day reporting lag source."),
        ("Murex","Murex","Treasury Management",12,2.6,"FX, derivatives, collateral",8,"Supported","Medium","No","Treasury booking. No AI capability. Manual reporting."),
        ("Geneva (SS&C)","SS&C Advent","Alternatives Accounting",8,1.9,"Alternatives, limited partnership",4,"Supported","Medium","No","Separate from SimCorp. Data duplication."),
        ("Linedata Longview","Linedata","OMS (Alternatives)",7,1.2,"Alternatives orders, compliance",3,"Under review","Low","No","Vendor acquired twice. Future uncertain."),
        ("OpenPages (IBM)","IBM","Risk & Compliance GRC",5,0.8,"Operational risk, audit",4,"Supported","Low","No","Low internal capability. IBM PS dependent."),
        ("Broadridge","Broadridge","Post-Trade Processing",11,2.2,"Settlement, corporate actions",6,"Supported","High","No","Solid. Settlement failure rate elevated — not platform cause."),
    ]

    for r,row in enumerate(systems,3):
        ws.row_dimensions[r].height=32
        name,vendor,cat,age,cost,domain,ints,eol,cap,ai_r,notes=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,name,sf,bold=True,b=bdr); C(ws,r,2,vendor,sf,b=bdr); C(ws,r,3,cat,sf,b=bdr)
        C(ws,r,4,age,fills["red"] if age>20 else fills["amb"] if age>10 else sf,"#,##0",align="center",b=bdr)
        C(ws,r,5,cost,sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,6,domain,sf,b=bdr); C(ws,r,7,ints,sf,"#,##0",align="center",b=bdr)
        ef=fills["red"] if "EOL" in eol and "PAST" in eol else fills["amb"] if "EOL" in eol or "review" in eol.lower() else sf
        C(ws,r,8,eol,ef,b=bdr)
        cf=fills["grn"] if "High" in cap else fills["amb"] if "Medium" in cap else fills["red"]
        C(ws,r,9,cap,cf,b=bdr)
        af=fills["grn"] if ai_r=="Yes" else fills["amb"] if ai_r=="Partial" else fills["red"]
        C(ws,r,10,ai_r,af,align="center",b=bdr)
        C(ws,r,11,notes,sf,b=bdr)

    # Integration map
    ws2=wb.create_sheet("Integration Map")
    T(ws2,7,f"{ARC['name']} — System Integration Map ({len(ARC['squads'])*2} connections)",fills)
    ws2.row_dimensions[2].height=28
    i_hdrs=[("Source",20),("Target",20),("Data Flow",22),("Protocol",12),("Frequency",12),("Automated?",11),("Risk",32)]
    for i,(hd,w) in enumerate(i_hdrs,1): H(ws2,2,i,hd,w,fills,bdr)
    integrations=[
        ("Bloomberg AIM","BlackRock Aladdin","Positions, P&L","FIX/SFTP","Real-time (15-45min lag)","Partial","AIM governs feed. Latency means risk models run on stale positions."),
        ("Bloomberg AIM","SQL Server DW","EOD positions, P&L","SFTP flat file","Daily T+1","No — manual","MANUAL. Source of 3-day reporting lag. Operator-dependent."),
        ("Bloomberg AIM","SimCorp Dimension","Trade instructions","FIX","Real-time","Yes","Bloomberg controls format. Breaking changes need vendor."),
        ("Bloomberg AIM","Charles River IMS","Order routing","FIX","Real-time","Yes","Stable. Internal team understands."),
        ("Bloomberg AIM","Broadridge","Settlement instructions","SWIFT MT","Daily","Yes","Standard SWIFT. Low risk."),
        ("BlackRock Aladdin","SQL Server DW","Risk metrics","API/SFTP","Daily","No — manual","Manual extract. Risk team exports manually. T+1 at earliest."),
        ("Salesforce FSC","SQL Server DW","Client AUM, contacts","SFTP","Daily","No — manual","Wipro runs extract. Inconsistent. Missing records common."),
        ("FactSet","Bloomberg Terminal","Security master, prices","API","Real-time","Yes","Reliable. Well-integrated."),
        ("SimCorp","SQL Server DW","NAV, fund accounting","SFTP flat file","Daily T+1","No — manual","Manual. Fund ops team runs nightly. Format varies."),
        ("Murex","Bloomberg AIM","FX rates, derivative vals","API","Intraday","Partial","Partial automation. Manual check for complex structures."),
        ("Geneva","SQL Server DW","Alternatives NAV","SFTP","Weekly","No — manual","Weekly manual extract. Alternatives always 3-7 days stale."),
        ("OpenPages","SQL Server DW","Operational risk events","Manual entry","Weekly","No","Manual entry. Completeness not guaranteed."),
        ("SQL Server DW","Reporting Layer","Aggregated MIS","SQL queries","Daily T+2","Partial","Some automated. Board pack manual. 3-day lag baked in."),
        ("Bloomberg AIM","MAS MASNET Portal","Regulatory trade reporting","Custom FTP","Daily","Yes","MAS FEAT reporting. Compliance gap — not all trades reported correctly."),
    ]
    for r,row in enumerate(integrations,3):
        ws2.row_dimensions[r].height=24
        src,tgt,flow,proto,freq,auto,risk=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        af=fills["grn"] if auto=="Yes" else fills["amb"] if auto=="Partial" else fills["red"]
        C(ws2,r,1,src,sf,bold=True,b=bdr); C(ws2,r,2,tgt,sf,b=bdr); C(ws2,r,3,flow,sf,b=bdr)
        C(ws2,r,4,proto,sf,align="center",b=bdr); C(ws2,r,5,freq,sf,align="center",b=bdr)
        ca=C(ws2,r,6,auto,af,align="center",bold=True,b=bdr)
        C(ws2,r,7,risk,fills["red"] if "MANUAL" in risk or "lag" in risk.lower() else sf,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/core/ARC-C04_Technology_Landscape.xlsx")
    return "ARC-C04 done"


def arc_c05_sprint_velocity(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Sprint Velocity"
    T(ws,13,f"{ARC['name']} — Sprint Velocity & DORA Metrics (Apr 2025–Mar 2026)",fills)
    ws.row_dimensions[2].height=32
    hdrs=[("Squad",26),("Quarter",10),("Planned Pts",11),("Completed Pts",11),
          ("Velocity %",10),("Unplanned %",11),("Carryover %",10),
          ("Deployments",11),("Incidents",10),("Vendor Blocked",12),
          ("CDO Blocked",11),("RAG",8),("Notes",32)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    sprints=[
        # OMS Core — Bloomberg-dominated
        ("OMS Core Engineering","Q1 2025",42,31,None,28,22,0,3,4,0,"Red","2 sprints blocked by Bloomberg release freeze. Emergency AIM-OMS sync patch."),
        ("OMS Core Engineering","Q2 2025",40,28,None,31,28,0,4,5,0,"Red","Bloomberg upgrade required 6-week regression. Zero feature delivery."),
        ("OMS Core Engineering","Q3 2025",44,32,None,26,24,1,2,4,0,"Amber","1 deployment — hotfix only. Vendor controls all change windows."),
        ("OMS Core Engineering","Q4 2025",42,29,None,30,25,0,3,5,0,"Red","Year-end freeze. Bloomberg governs all windows. Squad in maintenance mode."),
        # Portfolio Analytics — best
        ("Portfolio Analytics","Q1 2025",65,61,None,8,5,8,1,0,0,"Green","Strong. 8 production deployments. Factor model updated."),
        ("Portfolio Analytics","Q2 2025",68,64,None,7,4,9,1,0,1,"Green","CDO vacancy blocked 1 AI feature. Otherwise excellent."),
        ("Portfolio Analytics","Q3 2025",66,63,None,9,6,10,0,0,0,"Green","Best quarter. ESG data integration delivered."),
        ("Portfolio Analytics","Q4 2025",64,60,None,10,7,8,1,0,2,"Green","2 items awaiting AI governance (CDO vacant)."),
        # AI/ML Platform — zero deliveries
        ("AI/ML Platform","Q1 2025",30,18,None,45,40,0,0,0,6,"Red","6 features blocked by absent CDO sponsor. Zero production deployments."),
        ("AI/ML Platform","Q2 2025",28,14,None,48,44,0,0,0,8,"Red","Google PSO engagement ended. Knowledge transfer incomplete."),
        ("AI/ML Platform","Q3 2025",25,12,None,52,48,0,0,0,9,"Red","9 items require CDO approval. Interim CDO refuses to sign off."),
        ("AI/ML Platform","Q4 2025",22,10,None,55,50,0,0,0,11,"Red","11 CDO-blocked items. 0 deliveries in 12 months. 3 engineers interviewing elsewhere."),
        # Client Data Platform — Wipro-dominated
        ("Client Data Platform","Q1 2025",35,22,None,38,34,0,2,0,0,"Red","Team lead vacancy month 2. Wipro PM running squad."),
        ("Client Data Platform","Q2 2025",32,19,None,40,36,0,3,0,0,"Red","Golden record stalled. Wipro scope change added cost without delivery."),
        ("Client Data Platform","Q3 2025",30,18,None,42,38,0,2,0,0,"Red","Data pipeline POC attempted. Wipro delivered incomplete."),
        ("Client Data Platform","Q4 2025",28,16,None,44,40,0,4,0,0,"Red","3-day reporting lag unresolved. Same root cause as Q1. No progress."),
        # Compliance — MAS breach
        ("Compliance & Regulatory","Q1 2025",20,14,None,22,18,0,1,0,0,"Amber","MAS FEAT gap analysis delivered (Deloitte). No implementation started."),
        ("Compliance & Regulatory","Q2 2025",18,12,None,24,20,0,1,0,0,"Red","MAS deadline Dec 2025 confirmed. Deloitte 180-page requirements doc."),
        ("Compliance & Regulatory","Q3 2025",22,15,None,20,16,0,0,0,0,"Red","Implementation sprint started. 8 weeks behind plan."),
        ("Compliance & Regulatory","Q4 2025",20,13,None,26,22,0,1,0,0,"Red","MAS deadline MISSED. Dec 2025. Regulatory notice received."),
        # Risk Technology — moderate
        ("Risk Technology","Q1 2025",48,42,None,14,10,4,1,0,2,"Green","Solid. 2 CDO-dependent AI features blocked."),
        ("Risk Technology","Q2 2025",50,44,None,12,8,5,1,0,3,"Amber","Model validation backlog growing."),
        ("Risk Technology","Q3 2025",46,40,None,15,12,4,2,0,4,"Amber","4 AI risk overlays blocked pending CDO."),
        ("Risk Technology","Q4 2025",48,41,None,14,11,5,1,0,5,"Amber","5 features awaiting CDO. Regulatory model validation deadline Q1 2026 at risk."),
    ]

    rag_fills={"Green":fills["grn"],"Amber":fills["amb"],"Red":fills["red"]}
    for r,row in enumerate(sprints,3):
        ws.row_dimensions[r].height=22
        sq,qtr,planned,completed,_,unpl,carry,deploys,incidents,vblocked,cdoblocked,rag,notes=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,sq,sf,bold=True,b=bdr); C(ws,r,2,qtr,sf,align="center",b=bdr)
        C(ws,r,3,planned,sf,"#,##0",align="center",b=bdr)
        C(ws,r,4,completed,sf,"#,##0",align="center",b=bdr)
        C(ws,r,5,f"=D{r}/C{r}",sf,"0.0%",align="center",b=bdr)
        C(ws,r,6,unpl/100,sf,"0.0%",align="center",b=bdr)
        C(ws,r,7,carry/100,sf,"0.0%",align="center",b=bdr)
        C(ws,r,8,deploys,sf,"#,##0",align="center",b=bdr)
        C(ws,r,9,incidents,sf,"#,##0",align="center",b=bdr)
        C(ws,r,10,vblocked,fills["red"] if vblocked>=4 else fills["amb"] if vblocked>0 else sf,"#,##0",align="center",b=bdr)
        C(ws,r,11,cdoblocked,fills["red"] if cdoblocked>=6 else fills["amb"] if cdoblocked>0 else sf,"#,##0",align="center",b=bdr)
        cs=C(ws,r,12,rag,rag_fills.get(rag,sf),align="center",bold=True,b=bdr)
        cs.font=Font(size=9,name="Arial",bold=True,color="FFFFFF" if rag=="Red" else "333333")
        C(ws,r,13,notes,sf,b=bdr)

    ws.freeze_panes="C3"
    save(wb, f"{base}/arcturus/core/ARC-C05_Sprint_Velocity.xlsx")
    return "ARC-C05 done"


def arc_p01_ai_initiatives(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "AI Initiative Inventory"
    T(ws,12,f"{ARC['name']} — AI Initiative Inventory ({ARC['ai_initiatives']} Initiatives · {ARC['fy']})",fills)
    ws.row_dimensions[2].height=35
    hdrs=[("ID",7),("Initiative",30),("Division",18),("Sponsor",18),
          ("Status",14),("Budget £M",11),("Spent £M",11),("ROI Verified £M",13),
          ("Data Readiness",12),("MLOps Infra",10),("Genome Pattern",18),
          ("Blocker",24),("Value if Delivered £M pa",14)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    p=ARC["people"]
    initiatives=[
        ("AI-001","AI Trade Execution Optimisation","Trading/Technology",p["CIO"],"Pilot—Stalled",4.2,3.8,0.0,"38%","None","F006+F002","No MLOps. No CDO sign-off. Data from 14 systems inconsistent.",12.0),
        ("AI-002","Automated Portfolio Rebalancing","Investments/Technology",p["CEO"],"Pilot—Stalled",3.1,2.6,0.0,"44%","None","F001+F003","No AIM automated feed. Manual override required.",8.0),
        ("AI-003","ESG Scoring Automation","Investments/Data",p["CEO"],"In Development",2.4,1.8,0.0,"62%","Partial","F003","ESG data inconsistent across regions.",4.0),
        ("AI-004","Client Sentiment (CRM)","Client/Technology",f"({p['CDO']})","Pilot—Stalled",1.8,1.4,0.0,"28%","None","F002+F003","CDO vacant. FSC adoption 44%. Sparse training data.",6.0),
        ("AI-005","MAS FEAT Compliance AI","Compliance/Technology",p["CCO"],"BREACH—Overdue",3.6,2.9,0.0,"52%","None","F006+F008","MAS deadline MISSED Dec 2025. Manual process ongoing. Regulatory notice.",18.0),
        ("AI-006","Risk Model Automation","Risk/Technology",f"({p['CDO']})","Pilot—Stalled",4.8,4.1,0.0,"71%","None","F002","CDO sign-off required. Vacant 11 months. Monthly runs remain manual.",9.0),
        ("AI-007","NL Client Reporting","Client/Data",p["MD_Client"],"In Development",2.1,1.2,0.0,"41%","None","F003","3-day reporting lag. Reports from stale data.",5.0),
        ("AI-008","Fraud Detection AI","Operations/Technology",p["CFO"],"Pilot—Stalled",2.8,2.3,0.0,"55%","None","F001+F006","Transaction data across 5 systems. No unified feed.",7.0),
        ("AI-009","AI Deal Sourcing","Private Markets",p["CEO"],"Concept",1.4,0.4,0.0,"22%","None","F003+F006","Data sourcing undefined. No infrastructure. CEO sponsorship only.",15.0),
        ("AI-010","Automated KYC / Onboarding","Operations/Compliance",p["CCO"],"Pilot—Stalled",3.2,2.7,0.0,"48%","None","F002+F008","Regulatory AI framework undefined. CDO prerequisite.",11.0),
        ("AI-011","Predictive Cash Flow","Finance/Technology",p["CFO"],"In Development",1.9,1.1,0.0,"64%","None","F003","3-day data lag makes forecasts unreliable.",6.0),
        ("AI-012","AI Trade Surveillance","Compliance/Technology",p["CCO"],"Pilot—Stalled",4.1,3.4,0.0,"58%","None","F002+F006","FCA model validation required. No framework. CDO vacant.",14.0),
        ("AI-013","Dynamic Asset Allocation AI","Investments/Technology",p["CEO"],"Concept",5.6,0.8,0.0,"18%","None","F001+F003+F006","Data readiness 18%. Three prerequisites unmet.",22.0),
        ("AI-014","Automated Proxy Voting","Governance/Investments",p["CEO"],"In Development",1.6,0.9,0.0,"72%","None","F003","Depends on AI-003 ESG scoring.",3.0),
        ("AI-015","Real-Time Perf Attribution","Investments/Technology",p["CIO"],"Pilot—Stalled",3.8,3.2,0.0,"61%","None","F001+F006","AIM data latency. No streaming infrastructure.",8.0),
        ("AI-016","Automated Capital Calc","Finance/Risk",p["CFO"],"Pilot—Stalled",2.9,2.4,0.0,"55%","None","F002","Model validation framework absent. CDO vacant.",12.0),
        ("AI-017","Portfolio Stress Testing AI","Risk/Client",f"({p['CDO']})","Concept",4.2,0.6,0.0,"32%","None","F002+F001","Aladdin integration + CDO prerequisites.",9.0),
        ("AI-018","Vendor Invoice Automation","Operations/Finance",p["CFO"],"In Development",1.2,0.7,0.0,"78%","None","F003","ERP integration incomplete. Data quality sufficient after.",2.0),
        ("AI-019","NL Query Analytics","Technology/Investments",p["CIO"],"Pilot—Stalled",2.6,2.1,0.0,"44%","None","F001+F003","No unified data layer. Queries return inconsistent results.",7.0),
        ("AI-020","Fund Factsheet Automation","Client/Marketing",p["MD_Client"],"In Development",0.9,0.5,0.0,"68%","None","F003","3-day lag produces stale factsheets.",2.0),
        ("AI-021","Predictive Client Churn","Client/Technology",f"({p['CDO']})","Concept",2.2,0.4,0.0,"28%","None","F002+F003","CDO vacant. FSC adoption sparse.",11.0),
        ("AI-022","Trade Matching & Settlement","Operations/Technology",p["CFO"],"Pilot—Stalled",3.4,2.9,0.0,"62%","None","F001+F006","Settlement data across multiple custodians.",9.0),
        ("AI-023","AI Investment Research","Investments/Technology",p["CEO"],"In Development",1.8,1.0,0.0,"71%","Partial","F006","HIGHEST MLOps readiness. Partial Google PSO infra. Priority deployment.",6.0),
        ("AI-024","Dynamic Liquidity Risk","Risk/Treasury",p["CFO"],"Concept",5.8,0.2,0.0,"21%","None","F003+F006","Requires streaming infrastructure. Not begun.",17.0),
        ("AI-025","Compliance Monitoring AI","Compliance/Technology",p["CCO"],"Pilot—Stalled",2.8,2.3,0.0,"58%","None","F001+F003","Mandate data in 14 formats.",10.0),
        ("AI-026","Earnings Call Analysis","Investments/Research",p["CEO"],"In Development",1.4,0.8,0.0,"74%","None","F006","External data only. Near production. Serving layer needed.",4.0),
        ("AI-027","Counterparty Risk Scoring","Risk/Operations",p["CIO"],"Pilot—Stalled",3.6,3.0,0.0,"48%","None","F002+F001","Credit data across 5 providers. CDO vacant.",13.0),
        ("AI-028","Automated Board Reporting","Executive/Technology",p["CFO"],"Concept",1.8,0.3,0.0,"38%","None","F003","3-day lag makes board data unreliable.",3.0),
    ]

    status_fills={"Pilot—Stalled":fills["red"],"BREACH—Overdue":fills["dkred"],"In Development":fills["amb"],"Concept":fills["wht"]}
    for r,row in enumerate(initiatives,3):
        ws.row_dimensions[r].height=36
        id_,name,div,sponsor,status,budget,spent,roi,data_r,mlops,genome,blocker,value=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,id_,sf,align="center",b=bdr); C(ws,r,2,name,sf,bold=True,b=bdr)
        C(ws,r,3,div,sf,b=bdr)
        C(ws,r,4,sponsor,fills["red"] if "VACANT" in sponsor else sf,b=bdr)
        st_f=status_fills.get(status,sf)
        C(ws,r,5,status,st_f,b=bdr)
        C(ws,r,6,budget,sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,7,spent,sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,8,roi,fills["red"],"£#,##0.0",align="center",b=bdr)
        dr=int(data_r.replace("%",""))
        C(ws,r,9,dr/100,fills["red"] if dr<40 else fills["amb"] if dr<65 else fills["grn"],"0%",align="center",b=bdr)
        C(ws,r,10,mlops,fills["red"] if mlops=="None" else fills["amb"] if mlops=="Partial" else fills["grn"],align="center",b=bdr)
        C(ws,r,11,genome,fills["amb"],b=bdr); C(ws,r,12,blocker,sf,b=bdr)
        C(ws,r,13,value,fills["grn"],"£#,##0.0",align="center",bold=True,b=bdr)

    tr=len(initiatives)+3
    C(ws,tr,1,f"TOTALS — {ARC['ai_initiatives']} initiatives · £{ARC['ai_budget_committed_m']:.0f}M committed · £{ARC['ai_verified_roi_m']:.0f}M ROI verified",bold=True,b=bdr)
    ws.cell(tr,6,value=ARC["ai_budget_committed_m"]); ws.cell(tr,6).number_format="£#,##0.0"
    ws.cell(tr,7,value=ARC["ai_spent_to_date_m"]); ws.cell(tr,7).number_format="£#,##0.0"
    ws.cell(tr,8,value=0); ws.cell(tr,8).number_format="£#,##0.0"
    ws.cell(tr,13,value=sum(r[12] for r in initiatives)); ws.cell(tr,13).number_format="£#,##0.0"
    for c in [6,7,8,13]:
        ws.cell(tr,c).font=Font(bold=True,size=10,name="Arial"); ws.cell(tr,c).alignment=Alignment(horizontal="center")

    ws.freeze_panes="C3"
    save(wb, f"{base}/arcturus/pdlc/ARC-P01_AI_Initiative_Inventory.xlsx")
    return "ARC-P01 done"


def arc_p02_data_architecture(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Data Architecture"
    T(ws,12,f"{ARC['name']} — Data Architecture & Quality Assessment",fills)
    ws.row_dimensions[2].height=32
    hdrs=[("Data Domain",22),("Primary System",20),("Secondary Systems",22),
          ("Record Count",13),("Update Freq",14),("Quality Score",12),
          ("Completeness",12),("Accuracy",11),("Lag (Hrs)",11),
          ("Golden Record?",13),("AI Ready?",10),("Blocker",30)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    domains=[
        ("Client Master","Salesforce FSC","Bloomberg AIM, SimCorp, internal CRM",142000,"Daily",38,0.61,0.72,26,"No — 4 conflicting sources","No","4 systems, different values. 44% FSC adoption = 56% clients have stale data."),
        ("Position Data","Bloomberg AIM","Aladdin, SimCorp, Charles River",2800000,"Real-time (15-45min)",52,0.88,0.91,0.25,"No — AIM vs Aladdin conflict","Partial","AIM and Aladdin positions diverge up to 0.3% daily. Manual reconciliation."),
        ("Trade / Order Data","Bloomberg AIM","Charles River, Murex, Broadridge",4200000,"Real-time",71,0.94,0.96,0.1,"No — settlement lag","Partial","Trade data reliable. Settlement confirmation lags 4-8hrs."),
        ("Market Data","Bloomberg Terminal","FactSet, Reuters",1800000000,"Real-time",88,0.99,0.98,0,"No (not needed)","Yes","Best domain. External sourcing. High quality. AI-023, AI-026 rely on this."),
        ("Risk Analytics","BlackRock Aladdin","Bloomberg AIM feed",320000,"Daily",58,0.82,0.88,24,"No","No","Daily metrics available. Stress testing monthly only. Regulatory requires daily."),
        ("Fund Accounting / NAV","SimCorp Dimension","Geneva, SQL Server DW",48000,"Daily T+1",72,0.91,0.94,24,"No — SimCorp vs Geneva diverge","No","Manual reconciliation daily. T+1 lag prevents same-day reporting."),
        ("Client Reporting","SQL Server DW (manual)","SimCorp, Aladdin, FSC, AIM",280000,"Daily T+2 to T+3",28,0.61,0.74,72,"No","No","Manual aggregation from 14 systems. 3-day lag structural."),
        ("Compliance / Mandate","OpenPages GRC","Bloomberg AIM rules engine",18000,"Weekly",44,0.71,0.68,168,"No","No","Mandate rules not synchronised with AIM. Breaches can go undetected 1 week."),
        ("ESG Data","Manual (Excel)","FactSet ESG, Bloomberg ESG",9400,"Monthly",32,0.48,0.61,720,"No","No","Manually compiled monthly. AI-003 blocked by this."),
        ("Alternatives / PE","Geneva","SimCorp (partial)",4200,"Weekly",51,0.78,0.82,168,"No","No","Alternatives not connected to main performance system. 1 week lag."),
        ("Regulatory Reporting","Manual (Excel + OpenPages)","MAS portal, FCA, SEC",82000,"Daily (manual)",21,0.54,0.61,24,"No","No","All regulatory reporting manual. MAS FEAT breach partly caused by this."),
        ("Treasury / Cash","Murex","Bloomberg AIM",380000,"Intraday",68,0.89,0.91,4,"No","No","Treasury reliable. 4-hour lag before available in DW."),
        ("Vendor / Contract","Manual (SharePoint)","Various",2400,"Ad-hoc",18,0.42,0.55,720,"No","No","Contracts in SharePoint. No consistent taxonomy. Renewal tracked in spreadsheets."),
        ("Employee / HR","Workday","N/A",13000,"Real-time",82,0.96,0.97,1,"No (not needed)","Partial","HR data clean. Relevant for Maestro team design."),
    ]

    for r,row in enumerate(domains,3):
        ws.row_dimensions[r].height=32
        domain,primary,secondary,records,freq,quality,comp,acc,lag,golden,ai_r,blocker=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,domain,sf,bold=True,b=bdr); C(ws,r,2,primary,sf,b=bdr); C(ws,r,3,secondary,sf,b=bdr)
        C(ws,r,4,records,sf,"#,##0",align="center",b=bdr)
        C(ws,r,5,freq,sf,align="center",b=bdr)
        qf=fills["red"] if quality<40 else fills["amb"] if quality<65 else fills["grn"]
        C(ws,r,6,quality,qf,"#,##0",align="center",bold=True,b=bdr)
        for col,pct in [(7,comp),(8,acc)]:
            pf=fills["red"] if pct<0.70 else fills["amb"] if pct<0.85 else fills["grn"]
            C(ws,r,col,pct,pf,"0.0%",align="center",b=bdr)
        lag_num=float(lag) if isinstance(lag,(int,float)) else 0
        lf=fills["red"] if lag_num>48 else fills["amb"] if lag_num>4 else fills["grn"]
        C(ws,r,9,lag,lf,align="center",b=bdr)
        gf=fills["red"] if golden.startswith("No") else fills["grn"]
        C(ws,r,10,golden,gf,b=bdr)
        af=fills["grn"] if ai_r=="Yes" else fills["amb"] if ai_r=="Partial" else fills["red"]
        C(ws,r,11,ai_r,af,align="center",b=bdr)
        C(ws,r,12,blocker,sf,b=bdr)

    # Readiness scorecard
    ws2=wb.create_sheet("Readiness Scorecard")
    T(ws2,5,"Data Readiness Scorecard — AI Deployment Prerequisites",fills)
    ws2.row_dimensions[2].height=28
    sc_hdrs=[("Dimension",28),("Score /100",13),("Peer Benchmark",13),("Gap",11),("Blocker",36)]
    for i,(hd,w) in enumerate(sc_hdrs,1): H(ws2,2,i,hd,w,fills,bdr)
    scorecard=[
        ("Golden Record Availability",4,72,-68,"0 of 14 data domains have a golden record."),
        ("Data Pipeline Automation",12,68,-56,"11 of 14 integrations are manual. 3-day lag structural."),
        ("Data Quality (avg)",48,74,-26,"Client, ESG, Regulatory critically low. Market data excellent."),
        ("Real-time Availability",18,61,-43,"Only market data and trade execution are real-time."),
        ("AI Training Data Readiness",11,58,-47,"No feature store. No data versioning. Insufficient labelled data."),
        ("Data Governance & Ownership",15,62,-47,"CDO vacant. No data governance board. Ownership ad-hoc."),
        ("Regulatory Data Compliance",22,71,-49,f"MAS FEAT breach. GDPR partial. Data residency incomplete."),
        ("OVERALL DATA READINESS",12,65,-53,"Not ready for production AI. Foundation programmes required first."),
    ]
    for r,(dim,score,bench,gap,blocker) in enumerate(scorecard,3):
        ws2.row_dimensions[r].height=26
        sf=fills["prp"] if dim.startswith("OVERALL") else fills["alt"] if r%2==0 else fills["wht"]
        C(ws2,r,1,dim,sf,bold=dim.startswith("OVERALL"),b=bdr)
        sf_s=fills["red"] if score<30 else fills["amb"] if score<60 else fills["grn"]
        C(ws2,r,2,score,sf_s,"#,##0",align="center",bold=True,b=bdr)
        C(ws2,r,3,bench,sf,"#,##0",align="center",b=bdr)
        C(ws2,r,4,gap,fills["red"],"#,##0",align="center",bold=True,b=bdr)
        C(ws2,r,5,blocker,sf,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/pdlc/ARC-P02_Data_Architecture.xlsx")
    return "ARC-P02 done"


def arc_p03_mlops(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "MLOps Assessment"
    T(ws,8,f"{ARC['name']} — MLOps Infrastructure Assessment",fills)
    ws.row_dimensions[2].height=32
    hdrs=[("MLOps Capability",26),("Current State",28),("Maturity 0-5",12),
          ("Initiatives Blocked",13),("Effort to Fix",14),
          ("Unblock Sequence",22),("Recommendation",30)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    caps=[
        ("Model Registry","No registry. Models tracked in spreadsheets.",0,28,"Medium — 2mo","Step 1","MLflow on AWS SageMaker. Link to Bedrock model catalog."),
        ("Feature Store","No feature store. Each model rebuilds features.",0,22,"High — 4mo","Step 3","Start with position and client data features — highest demand."),
        ("ML CI/CD Pipeline","No ML CI-CD. Infrastructure pipeline not extended to ML.",0,28,"Medium — 2mo","Step 2","Extend AWS CodePipeline. Add model validation gate."),
        ("Model Serving","No serving. Models run ad-hoc on analyst laptops.",0,28,"Medium — 2mo","Step 2","Deploy SageMaker Endpoints. AI-023 and AI-026 first."),
        ("Model Monitoring","No monitoring. No drift detection.",0,18,"Medium — 2mo","Step 4","CloudWatch + Evidently AI. After first production deployments."),
        ("Experiment Tracking","Partial — 2 scientists use MLflow locally. No shared server.",1,20,"Low — 3wks","Step 1","Deploy shared MLflow server. Quick win."),
        ("Data Versioning","No data versioning. Reproducibility impossible.",0,24,"Medium — 2mo","Step 3","DVC. Version training datasets. Link to model registry."),
        ("Model Validation Framework","No validation. No independent review. CDO sign-off required but vacant.",0,14,"High — requires CDO","CDO prerequisite","Establish model risk framework. Interim model risk officer."),
        ("Automated Retraining","No automated retraining. All models static.",0,18,"Medium — 4mo","Step 5","Implement after serving layer. AI-005 MAS model will drift."),
        ("Data Pipeline for ML","Manual extraction. No automated feature pipeline. 3-day lag.",0,26,"High — 6mo (golden record dep)","Blocked by golden record","Golden record must precede ML pipeline automation."),
        ("Security & Compliance","No model governance. No GDPR-compliant AI data handling.",0,12,"High — requires CDO + Legal","CDO prerequisite","GDPR and MAS AI compliance framework. CDO prerequisite."),
    ]

    for r,row in enumerate(caps,3):
        ws.row_dimensions[r].height=28
        cap,state,maturity,blocked,effort,seq,rec=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,cap,sf,bold=True,b=bdr)
        C(ws,r,2,state,fills["red"] if maturity==0 else sf,b=bdr)
        mf=fills["red"] if maturity<2 else fills["amb"] if maturity<4 else fills["grn"]
        C(ws,r,3,maturity,mf,"#,##0",align="center",bold=True,b=bdr)
        C(ws,r,4,blocked,fills["red"] if blocked>=20 else fills["amb"],"#,##0",align="center",b=bdr)
        ef=fills["red"] if "High" in effort else fills["amb"] if "Medium" in effort else fills["grn"]
        C(ws,r,5,effort,ef,b=bdr)
        C(ws,r,6,seq,fills["teal"] if "Step" in seq else fills["red"],align="center",b=bdr)
        C(ws,r,7,rec,sf,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/pdlc/ARC-P03_MLOps_Assessment.xlsx")
    return "ARC-P03 done"


def arc_p04_bloomberg_aim(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Customisation Register"
    T(ws,7,f"{ARC['name']} — Bloomberg AIM Customisation Register (Migration Risk)",fills)
    ws.row_dimensions[2].height=32
    hdrs=[("ID",10),("Customisation",34),("Business Purpose",28),("Complexity",12),
          ("Migration Risk",14),("Bloomberg Only?",13),("Est Rebuild Cost £M",14),("Notes",30)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    custs=[
        ("AIM-C001","Custom FIX adapter — Broadridge settlement","Route settlement in Arcturus proprietary format","High","Critical","Yes",0.8,"Built by Bloomberg 2008. No documentation. Cannot replicate without Bloomberg."),
        ("AIM-C002","Pre-trade compliance rule engine (93 rules)","93 mandate rules for Arcturus investment mandates","High","Critical","Yes",2.4,"Embedded in AIM core. No export capability. Full rule re-implementation required."),
        ("AIM-C003","Real-time P&L attribution model","Custom attribution per Arcturus performance framework","High","Critical","Yes",1.8,"Arcturus-specific calculation. Bloomberg won't support in standard product."),
        ("AIM-C004","Multi-currency hedging overlay","Automatic hedge ratio calculation for multi-currency mandates","Medium","High","Yes",0.9,"Complex logic. Internal team partial understanding."),
        ("AIM-C005","Portfolio rebalancing workflow","Step-by-step rebalancing with approval gates","Medium","High","Partial",0.6,"Internal team contributed. Some documentation."),
        ("AIM-C006","Singapore MAS regulatory reporting","Automated MAS trade reporting in MASNET format","High","Critical","Yes",1.2,"MAS-specific. Bloomberg built for Arcturus. Critical for Singapore AUM compliance."),
        ("AIM-C007","Aladdin risk feed adapter","Data transformation for BlackRock Aladdin position feed","Medium","High","Yes",0.5,"Bloomberg-built adapter. Internal aware of logic, cannot maintain."),
        ("AIM-C008","Client mandate compliance dashboard","Real-time mandate breach view for client-specific rules","Medium","Medium","Partial",0.4,"Joint build. Internal team understands well."),
        ("AIM-C009","Corporate actions workflow","Custom workflow for complex corporate actions","High","High","Yes",0.7,"Complex Bloomberg workflow. Internal cannot replicate without effort."),
        ("AIM-C010","FX overlay execution logic","Automatic FX hedge execution for overlay mandates","High","Critical","Yes",1.4,"Proprietary strategy embedded in AIM. Cannot extract without Bloomberg."),
        ("AIM-C011","Geneva to AIM PE position feed","Alternatives position connector","Low","Medium","No",0.2,"Internal team built. Well-documented."),
        ("AIM-C012","EMEA trade allocation algorithm","Multi-client trade allocation across EMEA strategies","High","Critical","Yes",1.1,"Regulatory requirement embedded. Bloomberg-only capability."),
        ("AIM-C013","ESG data overlay","Manual ESG override on positions","Low","Low","No",0.1,"Simple internal customisation. Easily migrated."),
        ("AIM-C014","Board reporting extract","Custom extract for board investment committee","Low","Low","Partial",0.1,"Partially documented. Medium migration complexity."),
    ]

    for r,row in enumerate(custs,3):
        ws.row_dimensions[r].height=28
        id_,desc,purpose,comp,risk,blm_only,rebuild_cost,notes=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        rf_map={"Critical":fills["red"],"High":fills["amb"],"Medium":fills["amb"],"Low":fills["grn"]}
        cf_map={"High":fills["red"],"Medium":fills["amb"],"Low":fills["grn"]}
        bf_map={"Yes":fills["red"],"Partial":fills["amb"],"No":fills["grn"]}
        C(ws,r,1,id_,sf,align="center",b=bdr); C(ws,r,2,desc,sf,bold=True,b=bdr)
        C(ws,r,3,purpose,sf,b=bdr); C(ws,r,4,comp,cf_map.get(comp,sf),align="center",b=bdr)
        C(ws,r,5,risk,rf_map.get(risk,sf),align="center",bold=True,b=bdr)
        C(ws,r,6,blm_only,bf_map.get(blm_only,sf),align="center",b=bdr)
        C(ws,r,7,rebuild_cost,sf,"£#,##0.0",align="center",b=bdr); C(ws,r,8,notes,sf,b=bdr)

    tr=len(custs)+3
    C(ws,tr,1,"TOTAL",fills["prp"],bold=True,b=bdr)
    ws.cell(tr,7,value=f"=SUM(G3:G{tr-1})"); ws.cell(tr,7).number_format="£#,##0.0"
    ws.cell(tr,7).font=Font(bold=True,size=10,name="Arial"); ws.cell(tr,7).alignment=Alignment(horizontal="center")

    # Failed migrations sheet
    ws2=wb.create_sheet("Failed Migration Post-Mortems")
    T(ws2,7,f"Bloomberg AIM — 3 Failed Modernisation Attempts ({ARC['bloomberg']['failed_modernisations']} Attempts)",fills)
    ws2.row_dimensions[2].height=28
    pm_hdrs=[("Attempt",10),("Year",9),("Vendor",18),("Target Platform",18),
             ("Duration",11),("Total Cost £M",13),("Reason for Failure",34),("Root Cause Pattern",18)]
    for i,(hd,w) in enumerate(pm_hdrs,1): H(ws2,2,i,hd,w,fills,bdr)
    post_mortems=[
        (1,2009,"TCS + Bloomberg","SimCorp Dimension OMS","18 months",8.2,
         "Bloomberg customisations (14 at that time) could not be replicated in SimCorp. The compliance rule engine (93 rules) had no documented specification. TCS could not reverse-engineer Bloomberg logic. Abandoned after Phase 2 failure.",
         "F001 (72%) — Vendor dependency without internal capability. Internal team had no OMS knowledge independent of Bloomberg."),
        (2,2016,"Accenture + Murex","Murex MX.3 (full OMS replacement)","24 months",14.6,
         "Data migration complexity catastrophically underestimated. 28 years of Bloomberg position history in proprietary format. MAS MASNET reporting adapter could not be rebuilt. Singapore AUM compliance risk halted programme. Board pulled funding at month 24.",
         "F001 (72%) + F003 (68%) — Data readiness below threshold. No data migration assessment before programme start."),
        (3,2021,"Infosys + SS&C Eze","SS&C Eze Investment Suite","14 months",9.8,
         "Internal team insufficient to govern the programme. Infosys embedded consultants who lacked Bloomberg domain knowledge. Three critical customisations (AIM-C002 compliance rules, AIM-C006 MAS reporting, AIM-C010 FX overlay) could not be replicated. CDO appointed to lead programme resigned after 4 months. Programme abandoned.",
         "F002 (84%) — No named executive sponsor. CDO departure = programme collapse. Same pattern as attempts 1 and 2."),
    ]
    for r,(attempt,year,vendor,target,duration,cost,reason,pattern) in enumerate(post_mortems,3):
        ws2.row_dimensions[r].height=80
        sf=fills["red"]
        C(ws2,r,1,attempt,sf,"#,##0",align="center",bold=True,b=bdr)
        C(ws2,r,2,year,sf,"#,##0",align="center",b=bdr)
        C(ws2,r,3,vendor,sf,bold=True,b=bdr); C(ws2,r,4,target,sf,b=bdr)
        C(ws2,r,5,duration,sf,align="center",b=bdr)
        C(ws2,r,6,cost,fills["red"],"£#,##0.0",align="center",bold=True,b=bdr)
        C(ws2,r,7,reason,fills["wht"],b=bdr)
        C(ws2,r,8,pattern,fills["amb"],b=bdr)
    # Total cost
    ws2.cell(len(post_mortems)+3,6,value=sum(r[5] for r in post_mortems))
    ws2.cell(len(post_mortems)+3,6).number_format="£#,##0.0"
    ws2.cell(len(post_mortems)+3,6).font=Font(bold=True,size=11,name="Arial",color="CC0000")
    ws2.cell(len(post_mortems)+3,1,value="TOTAL FAILED MIGRATION COST")
    ws2.cell(len(post_mortems)+3,1).font=Font(bold=True,size=10,name="Arial")

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/pdlc/ARC-P04_Bloomberg_AIM_Customisations.xlsx")
    return "ARC-P04 done"


def arc_p05_engineering_cost(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Engineering Cost"
    T(ws,9,f"{ARC['name']} — Engineering Cost Breakdown {ARC['fy']} (£000s)",fills)
    ws.row_dimensions[2].height=32
    hdrs=[("Squad / Function",26),("Internal FTE Cost",13),("Contractor Cost",13),
          ("Consulting Vendor",13),("Licence/Infra",13),("Total Cost £000s",13),
          ("Story Pts Delivered",13),("Cost Per SP £",12),("vs Peer",14),("Notes",28)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    costs_data=[
        ("OMS Core Engineering",         2100,1680,3360,8400,None,120,None,"+£6.2k/SP","Bloomberg licence dominates. Peers pay £2M for equivalent OMS."),
        ("OMS Integration & APIs",       1680, 840,2240, 400,None, 90,None,"+£4.1k/SP","Integration complexity. Bloomberg dependency inflates."),
        ("Risk Technology",              1680, 840,1120,4200,None,167,None,"Near peer","Aladdin licence reasonable. Internal team cost-efficient."),
        ("Client Data Platform",         1120,1400,2520, 400,None, 75,None,"+£5.8k/SP","Wipro + vacant lead. Zero golden record progress."),
        ("Portfolio Analytics",          1960, 560, 840,2900,None,248,None,"Best performer","Best value. Highest output. Recommend investment."),
        ("Compliance & Regulatory",      1400, 280,1400, 800,None, 54,None,"+£3.2k/SP","Deloitte advisory + waterfall inefficiency."),
        ("Client Portal (FSC)",          1120, 840,1960,3100,None, 96,None,"+£2.4k/SP","Wipro + Salesforce PS. Duplication."),
        ("Data Engineering",             1680,1120,1680, 400,None, 71,None,"+£3.1k/SP","Manual ETL dominates. No automation ROI."),
        ("AI/ML Platform",               1120, 560,2240, 280,None,  0,None,"£1.8M / 0 output","Zero production deployments in 12 months."),
        ("Singapore Technology",         1680, 840,1120, 600,None,108,None,"At benchmark","TCS relationship reasonable for APAC."),
        ("New York Markets",             1400, 560,1400, 600,None,132,None,"Slight above","Bloomberg dependency moderate. Good team."),
        ("Infrastructure & Cloud",       1400, 840, 840,1200,None,180,None,"At benchmark","Solid. ML pipeline absent."),
        ("Enterprise Architecture",       560,1120, 560,   0,None,  0,None,"+£0.8k overhead","Contractor EA. No permanent leadership."),
        ("IT Management & Governance",   1960, 280, 840, 200,None,None,None,"At benchmark",f"CDO vacancy saves £400k in salary. Costs £{ARC['cdo_vacancy']['estimated_value_blocked_m']:.0f}M in blocked AI value."),
    ]

    total_fte_cost=total_contr_cost=total_consult_cost=total_lic_cost=0
    for r,row in enumerate(costs_data,3):
        ws.row_dimensions[r].height=22
        squad,fte,contr,consult,lic,_,sp,__,vs_peer,notes=row
        total=fte+contr+consult+lic
        total_fte_cost+=fte; total_contr_cost+=contr; total_consult_cost+=consult; total_lic_cost+=lic
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,squad,sf,bold=True,b=bdr)
        for col,val in [(2,fte),(3,contr),(4,consult),(5,lic)]:
            C(ws,r,col,val,sf,"#,##0",align="center",b=bdr)
        C(ws,r,6,f"=B{r}+C{r}+D{r}+E{r}",sf,"#,##0",align="center",bold=True,b=bdr)
        ws.cell(r,6).font=Font(size=9,name="Arial",bold=True)
        if sp:
            C(ws,r,7,sp,sf,"#,##0",align="center",b=bdr)
            C(ws,r,8,f"=F{r}/G{r}",sf,"£#,##0",align="center",b=bdr)
        else:
            C(ws,r,7,"N/A",fills["red"],align="center",b=bdr)
            C(ws,r,8,"N/A" if sp is None else "£∞",fills["red"],align="center",b=bdr)
        bench_f=fills["red"] if "+" in vs_peer and "k" in vs_peer else fills["amb"] if "overhead" in vs_peer or "above" in vs_peer.lower() else fills["grn"]
        C(ws,r,9,vs_peer,bench_f,b=bdr); C(ws,r,10,notes,sf,b=bdr)

    tr=len(costs_data)+3
    C(ws,tr,1,"TOTAL ENGINEERING COST",fills["prp"],bold=True,b=bdr)
    total_all=total_fte_cost+total_contr_cost+total_consult_cost+total_lic_cost
    for col,val in [(2,total_fte_cost),(3,total_contr_cost),(4,total_consult_cost),(5,total_lic_cost),(6,total_all)]:
        ws.cell(tr,col,value=val); ws.cell(tr,col).number_format="#,##0"
        ws.cell(tr,col).font=Font(bold=True,size=10,name="Arial")
        ws.cell(tr,col).fill=fills["prp"]; ws.cell(tr,col).alignment=Alignment(horizontal="center"); ws.cell(tr,col).border=bdr
    peer_total=22400
    tr2=tr+2
    for label,val in [("Peer Benchmark (£000s):",peer_total),
                       ("Arcturus Overspend (£000s):",total_all-peer_total),
                       ("Overspend as % of peer:",f"=(F{tr}-{peer_total})/{peer_total}")]:
        ws.cell(tr2,1,value=label); ws.cell(tr2,1).font=Font(bold=True,size=9,name="Arial",color="CC0000" if "Over" in label else "333333")
        ws.cell(tr2,6,value=val); ws.cell(tr2,6).number_format="#,##0" if isinstance(val,(int,float)) else "0.0%"
        ws.cell(tr2,6).font=Font(bold=True,size=10,name="Arial",color="CC0000" if "Over" in label else "333333")
        ws.cell(tr2,6).alignment=Alignment(horizontal="center"); tr2+=1

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/pdlc/ARC-P05_Engineering_Cost.xlsx")
    return "ARC-P05 done"


def arc_d01_consulting_audit(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Consulting Audit"
    T(ws,10,f"{ARC['name']} — Consulting Output vs Promise Audit (All Active Engagements)",fills)
    ws.row_dimensions[2].height=32
    hdrs=[("Vendor",16),("Engagement",26),("Promised",26),("Delivered?",14),
          ("Quality /10",10),("KT Score /100",12),("Annual Cost £M",12),
          ("Value Rating",16),("Recoverable £M pa",13),("Maestro Replacement?",16)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    audit=[
        ("Bloomberg LP","OMS Core Maintenance","24/7 support, 99.9% uptime, quarterly features","Partial — SLA met, features delayed",6,8,ARC["bloomberg"]["annual_cost_m"],"Poor — trapped dependency",6.3,"Yes — API wrapper + internal OMS team (18mo)"),
        ("Infosys","Risk Tech & Analytics","90% sprint velocity, knowledge wiki, upskilling","No — 71% velocity, wiki incomplete",5,22,3.6,"Below expectations",1.8,"Yes — 2 Maestros replace 12 Infosys (Wave 1)"),
        ("Wipro","Client Portal + Data Platform","85% velocity, full KT, golden record","No — 58% velocity, KT 15%, golden record not started",3,15,4.8,"Poor — knowledge hostage",3.2,"Yes — exit plan + 2 Maestros"),
        ("Deloitte","MAS FEAT Compliance","Compliance by Dec 2025, internal capability","No — deadline missed, team not capable",4,45,2.6,"Below expectations",0.8,"Partial — keep advisory, Maestro for execution"),
        ("TCS","APAC Technology Operations","96% SLA, KT, documentation","Yes for SLA, partial KT",6,38,0.9,"Adequate",0.0,"No — retain, value for APAC ops"),
        ("Google PSO","AI/ML Platform Setup","Full MLOps infrastructure delivered","No — 22% delivered, engagement ended",2,5,0.0,"Failed",0.0,"Yes — internal ML engineering with Maestro"),
        ("AWS ProServe","Infrastructure Modernisation","Runbook delivery, team training","Mostly — 88% complete",7,62,1.4,"Good",0.0,"No — continue with internal support"),
        ("Salesforce PS","FSC Customisation","Feature delivery, internal admin capability","Partial — 71% features, no admin capability",5,28,1.6,"Below expectations",0.8,"Yes — internal Salesforce admin team"),
        ("Contractors","Enterprise Architecture","Architecture decisions, documentation","No — no documentation, no KT",1,0,1.8,"Critical risk",1.8,"Yes — permanent EA hire immediately"),
    ]

    total_recoverable=0
    for r,row in enumerate(audit,3):
        ws.row_dimensions[r].height=28
        vendor,eng,promised,delivered,quality,kt,cost,rating,recoverable,replace=row
        total_recoverable+=recoverable
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,vendor,sf,bold=True,b=bdr); C(ws,r,2,eng,sf,b=bdr); C(ws,r,3,promised,sf,b=bdr)
        df=fills["red"] if delivered.startswith("No") else fills["amb"] if "Partial" in delivered or "Mostly" in delivered else fills["grn"]
        C(ws,r,4,delivered,df,b=bdr)
        qf=fills["red"] if quality<5 else fills["amb"] if quality<7 else fills["grn"]
        C(ws,r,5,quality,qf,"#,##0",align="center",b=bdr)
        kf=fills["red"] if kt<20 else fills["amb"] if kt<50 else fills["grn"]
        C(ws,r,6,kt,kf,"#,##0",align="center",b=bdr)
        C(ws,r,7,cost,sf,"£#,##0.0",align="center",b=bdr)
        rf2=fills["red"] if any(w in rating for w in ["Poor","Failed","Critical"]) else fills["amb"] if "Below" in rating or "Adequate" in rating else fills["grn"]
        C(ws,r,8,rating,rf2,b=bdr)
        C(ws,r,9,recoverable,fills["grn"] if recoverable>0 else sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,10,replace,fills["grn"] if "Yes" in replace else fills["amb"] if "Partial" in replace else fills["wht"],b=bdr)

    tr=len(audit)+3
    C(ws,tr,1,"TOTAL RECOVERABLE CONSULTING SPEND",bold=True,b=bdr)
    ws.cell(tr,7,value=sum(r[6] for r in audit)); ws.cell(tr,7).number_format="£#,##0.0"
    ws.cell(tr,9,value=total_recoverable); ws.cell(tr,9).number_format="£#,##0.0"
    for c in [7,9]:
        ws.cell(tr,c).font=Font(bold=True,size=11,name="Arial",color="1A3A5C")
        ws.cell(tr,c).alignment=Alignment(horizontal="center")

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/delivery/ARC-D01_Consulting_Audit.xlsx")
    return "ARC-D01 done"


def arc_d02_knowledge_risk(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Knowledge Risk"
    T(ws,7,f"{ARC['name']} — Knowledge Retention Risk Register",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Knowledge Domain",26),("Current Owner",22),("Type",14),
          ("If Lost — Impact",28),("Retention Plan?",13),("Risk Score",11),("Mitigation",32)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    knowledge=[
        ("Bloomberg AIM customisation logic (14 rules)","Bloomberg LP engineers","Vendor-owned","Cannot maintain or modify OMS. Entire trading operation at vendor mercy.","No",98,"AIM API wrapper + internal capability build. 12-month programme."),
        ("Wipro FSC customisation code","Wipro developers","Vendor-owned","Cannot deploy or maintain Salesforce FSC. Client portal frozen.","No",88,"Code escrow clause + internal Salesforce admin hire immediately."),
        ("Google PSO MLOps design (22% complete)","Google PSO (ended)","Departed vendor","AI/ML platform design knowledge walked out. Starting from zero.","No",95,"Reconstruct from artefacts. Hire ML engineer. Use AbarVa MLOps playbook."),
        ("MAS FEAT compliance process","Deloitte consultants","Vendor-dependent","Regulatory process not documented internally. Next audit will fail.","Partial",72,"Internal compliance team documentation sprint. Deloitte exit plan."),
        ("Enterprise architecture decisions","Contractor (rolling)","Contractor-owned","Architecture logic undocumented. No permanent owner.","No",91,"Permanent EA hire. Documentation retrospective. Architecture decision records."),
        ("Data pipeline logic (14 integrations)","Wipro + contractors","Mixed","Manual processes undocumented. Loss of any operator causes outage.","No",85,"Runbook creation sprint. All manual processes documented."),
        ("Risk model logic (Aladdin config)","Head of Risk + vendor","Internal + vendor","Risk model config not fully documented. Vendor changes break models.","Partial",61,"Model documentation programme. Internal model risk function."),
        ("Client mandate rules (93 Bloomberg rules)","Bloomberg LP","Vendor-owned","Compliance rule logic inaccessible without Bloomberg. Mandate breaches undetected.","No",94,"Compliance rule extraction and documentation. Internal rules engine assessment."),
        ("Portfolio Analytics squad IP","Internal team (Rachel Kim)","Internal","Best internal team. Knowledge concentrated in 3 people. Attrition risk.","Partial",55,"Knowledge sharing programme. Documentation. Succession planning."),
        (f"CDO-sponsored initiative context",f"({ARC['people']['CDO']})","Departed employee","14 initiatives orphaned. Context and rationale lost.","No",89,"Initiative documentation retrospective. CIO assumes interim."),
    ]

    for r,row in enumerate(knowledge,3):
        ws.row_dimensions[r].height=28
        domain,owner,type_,impact,plan,risk,mit=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,domain,sf,bold=True,b=bdr)
        C(ws,r,2,owner,fills["red"] if "VACANT" in owner or "ended" in owner or "Departed" in owner else sf,b=bdr)
        C(ws,r,3,type_,sf,b=bdr); C(ws,r,4,impact,sf,b=bdr)
        pf=fills["grn"] if plan=="Yes" else fills["amb"] if plan=="Partial" else fills["red"]
        C(ws,r,5,plan,pf,align="center",b=bdr)
        rf2=fills["red"] if risk>=80 else fills["amb"] if risk>=60 else fills["grn"]
        C(ws,r,6,risk,rf2,"#,##0",align="center",bold=True,b=bdr)
        C(ws,r,7,mit,sf,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/delivery/ARC-D02_Knowledge_Risk.xlsx")
    return "ARC-D02 done"


def arc_d03_maestro_design(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Maestro Team Design"
    T(ws,8,f"{ARC['name']} — Proposed Maestro Team Design (Replaces 50+ Consultants)",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Maestro Role",24),("Scope",30),("Replaces",24),
          ("Annual Cost Replaced £M",14),("Wave",10),("Duration",14),
          ("KPIs",28),("Success Metric",24)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    maestros=[
        ("Delivery Maestro — OMS & Trading","Govern Bloomberg AIM relationship. Build internal OMS capability. Reduce vendor dependency.","12-14 Bloomberg LP engineers",7.2,"Wave 1","18 months","Bloomberg dependency ratio <40%. Internal team governs AIM changes.","OMS change capability internal. No vendor approval needed for standard changes."),
        ("Delivery Maestro — Data & AI","Build MLOps foundation. Govern AI initiative delivery. Golden record programme.","8-10 Wipro + Google PSO staff",5.6,"Wave 1","24 months","3 AI initiatives in production. Data pipeline automated. Reporting lag <4hrs.","AI-023 and AI-026 in production by month 3. Golden record for 1 domain by month 6."),
        ("Delivery Maestro — Client Platform","Transfer FSC capability internally. Exit Wipro Salesforce. Build internal admin.","7 Wipro Salesforce team",3.2,"Wave 1","12 months","FSC adoption >70%. Internal team owns all deployments.","Wipro exited. Internal Salesforce admin certified. Adoption from 44% to 70%."),
        ("Delivery Maestro — Regulatory Tech","Build MAS FEAT compliance capability internally. Exit Deloitte dependency.","5 Deloitte consultants",2.6,"Wave 1","12 months","MAS FEAT compliant. Internal team capable of next audit.","MAS breach remediated. Compliance team independent of Deloitte."),
        ("Delivery Maestro — Risk Technology","Model risk governance. CDO onboarding support. Unblock AI-006.","4 Infosys risk tech",1.8,"Wave 2","18 months","Model risk framework established. 5 AI initiatives unblocked.","CDO onboarded with full context. Model risk committee operational."),
        ("Architecture Maestro","Replace contractor EA. Permanent architecture governance.","4 contractors (EA function)",1.8,"Wave 1","12 months (active), permanent knowledge transfer","Architecture decisions documented. No contractor in permanent role.","Permanent EA hire in place. All architecture decisions recorded."),
    ]

    for r,row in enumerate(maestros,3):
        ws.row_dimensions[r].height=28
        role,scope,replaces,cost,wave,duration,kpis,success=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        wf=fills["grn"] if wave=="Wave 1" else fills["amb"]
        C(ws,r,1,role,sf,bold=True,b=bdr); C(ws,r,2,scope,sf,b=bdr)
        C(ws,r,3,replaces,sf,b=bdr)
        C(ws,r,4,cost,fills["grn"],"£#,##0.0",align="center",bold=True,b=bdr)
        C(ws,r,5,wave,wf,align="center",b=bdr); C(ws,r,6,duration,sf,b=bdr)
        C(ws,r,7,kpis,sf,b=bdr); C(ws,r,8,success,sf,b=bdr)

    tr=len(maestros)+3
    C(ws,tr,1,f"TOTAL RECOVERABLE — replaces {sum(int(r[3]/0.3) for r in maestros)} consultants",bold=True,b=bdr)
    ws.cell(tr,4,value=sum(r[3] for r in maestros)); ws.cell(tr,4).number_format="£#,##0.0"
    ws.cell(tr,4).font=Font(bold=True,size=11,name="Arial",color="1A3A5C"); ws.cell(tr,4).alignment=Alignment(horizontal="center")

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/delivery/ARC-D03_Maestro_Team_Design.xlsx")
    return "ARC-D03 done"


def arc_m01_pl_detail(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "P&L by Business Unit"
    T(ws,10,f"{ARC['name']} — P&L by Business Unit {ARC['fy']} (£M)",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Business Unit",22),("AUM £B",10),("Revenue £M",12),("Op Costs £M",12),
          ("Op Profit £M",12),("C/I Ratio",11),("Target C/I",11),
          ("Gap pp",9),("Gap £M",11),("AI Opportunity £M",13)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    for r,(unit,data) in enumerate(ARC["revenue_by_bu"].items(),3):
        ws.row_dimensions[r].height=22
        aum=data["aum_b"]; rev=data["revenue_m"]; cost=data["cost_m"]
        ci=data["ci"]; profit=rev-cost if rev else -cost
        sf=fills["prp"] if unit=="Corporate/Technology" else fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,unit,sf,bold=True,b=bdr)
        C(ws,r,2,aum if aum else "N/A",sf,"#,##0" if aum else None,align="center",b=bdr)
        C(ws,r,3,rev if rev else "N/A",sf,"£#,##0" if rev else None,align="center",b=bdr)
        C(ws,r,4,cost,sf,"£#,##0",align="center",b=bdr)
        pf=fills["grn"] if profit>50 else fills["amb"] if profit>0 else fills["red"]
        C(ws,r,5,profit,pf,"£#,##0",align="center",b=bdr)
        if ci:
            cif=fills["red"] if ci>0.80 else fills["amb"] if ci>0.70 else fills["grn"]
            C(ws,r,6,ci,cif,"0.0%",align="center",b=bdr)
            C(ws,r,7,ARC["ci_target"],fills["grn"],"0.0%",align="center",b=bdr)
            gap_pp=round((ci-ARC["ci_target"])*100,1)
            C(ws,r,8,gap_pp,fills["red"],"+0.0;-0.0",align="center",b=bdr)
            gap_m=round((ci-ARC["ci_target"])*rev,1) if rev else None
            C(ws,r,9,gap_m,fills["red"],"£#,##0" if gap_m else None,align="center",b=bdr)
        else:
            for c in range(6,10): C(ws,r,c,"N/A",sf,align="center",b=bdr)
        ai_opp={"Global Equities":18.4,"Fixed Income":12.2,"Multi-Asset":8.4,"Alternatives":6.1,
                "Asia Pacific":4.2,"Client Solutions":3.8,"Corporate/Technology":ARC["ai_budget_committed_m"]}.get(unit,0)
        C(ws,r,10,ai_opp,fills["teal"],"£#,##0.0",align="center",b=bdr)

    tr=len(ARC["revenue_by_bu"])+3
    C(ws,tr,1,f"TOTAL — C/I {ARC['ci_ratio']:.0%} vs {ARC['ci_target']:.0%} target — £{ARC['ci_gap_m']:.0f}M gap",fills["prp"],bold=True,b=bdr)
    C(ws,tr,3,ARC["total_revenue_m"],fills["prp"],"£#,##0",align="center",bold=True,b=bdr)
    C(ws,tr,4,ARC["total_cost_m"],fills["prp"],"£#,##0",align="center",bold=True,b=bdr)
    C(ws,tr,6,ARC["ci_ratio"],fills["red"],"0.0%",align="center",bold=True,b=bdr)
    C(ws,tr,7,ARC["ci_target"],fills["grn"],"0.0%",align="center",bold=True,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/margin/ARC-M01_PL_by_Business_Unit.xlsx")
    return "ARC-M01 done"


def arc_m02_ai_roi(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "AI Spend ROI"
    T(ws,8,f"{ARC['name']} — AI Investment ROI Tracker (£{ARC['ai_budget_committed_m']:.0f}M Committed · £{ARC['ai_verified_roi_m']:.0f}M Verified)",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Initiative",28),("Budget £M",12),("Spent £M",12),("Expected ROI £M pa",14),
          ("Verified ROI £M",13),("ROI Gap £M",12),("Months Active",12),("Root Cause",28)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    # Summary rows from the detailed initiative list
    ai_summary=[
        ("AI trade execution + portfolio rebalancing",7.3,6.4,20.0,0.0,-20.0,25,"No MLOps. No CDO sign-off. Bloomberg data latency."),
        ("ESG + proxy voting automation",4.0,2.7,7.0,0.0,-7.0,19,"ESG data quality. Depends on AI-003 completion."),
        ("Client sentiment + churn prediction",4.0,1.8,17.0,0.0,-17.0,24,"CDO vacant. FSC adoption sparse. Training data insufficient."),
        ("MAS FEAT compliance AI",3.6,2.9,18.0,0.0,-18.0,21,"REGULATORY BREACH. Manual workaround. Emergency priority."),
        ("Risk model automation (AI-006, AI-016, AI-017)",11.9,7.1,30.0,0.0,-30.0,28,"CDO approval required. Vacant 11 months."),
        ("NL reporting + query + factsheets",5.6,3.8,14.0,0.0,-14.0,18,"3-day lag makes outputs unreliable."),
        ("Fraud detection + trade matching",6.2,5.2,16.0,0.0,-16.0,26,"Transaction data fragmented. No unified feed."),
        ("AI research + earnings analysis",3.2,1.8,10.0,0.0,-10.0,15,"Nearest to production. Serving layer only gap."),
        ("Compliance monitoring + trade surveillance",6.9,5.7,24.0,0.0,-24.0,23,"Mandate data in 14 formats. CDO for governance."),
        ("Cash flow + capital calc automation",4.8,3.5,18.0,0.0,-18.0,21,"Data lag + CDO governance."),
        ("Dynamic allocation + liquidity AI",11.4,1.0,39.0,0.0,-39.0,8,"Foundation not in place. Should not have started yet."),
        ("Onboarding + invoice automation + misc",24.5,23.9,5.0,0.0,-5.0,"Various","Various exploratory. Poorly tracked."),
    ]

    for r,row in enumerate(ai_summary,3):
        ws.row_dimensions[r].height=24
        name,budget,spent,expected,verified,gap,months,cause=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,name,sf,bold=True,b=bdr)
        C(ws,r,2,budget,sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,3,spent,sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,4,expected,fills["grn"],"£#,##0.0",align="center",b=bdr)
        C(ws,r,5,verified,fills["red"],"£#,##0.0",align="center",bold=True,b=bdr)
        C(ws,r,6,gap,fills["red"],"£#,##0.0",align="center",b=bdr)
        C(ws,r,7,months,sf,"#,##0" if isinstance(months,int) else None,align="center",b=bdr)
        C(ws,r,8,cause,fills["red"] if "BREACH" in cause else sf,b=bdr)

    tr=len(ai_summary)+3
    C(ws,tr,1,f"TOTALS — {ARC['ai_initiatives']} initiatives — Genome pattern F010 confirmed",fills["prp"],bold=True,b=bdr)
    C(ws,tr,2,ARC["ai_budget_committed_m"],fills["prp"],"£#,##0.0",align="center",bold=True,b=bdr)
    C(ws,tr,3,ARC["ai_spent_to_date_m"],fills["prp"],"£#,##0.0",align="center",bold=True,b=bdr)
    C(ws,tr,5,0,fills["red"],"£#,##0.0",align="center",bold=True,b=bdr)
    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/margin/ARC-M02_AI_Spend_ROI.xlsx")
    return "ARC-M02 done"


def arc_m03_cost_structure(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Cost Structure"
    T(ws,7,f"{ARC['name']} — Cost Structure Detail & Margin Recovery Opportunities",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Cost Category",26),("Actual £M",12),("% of Revenue",13),
          ("Peer Benchmark %",14),("Overspend £M",13),("AI Recovery £M",13),("Programme",26)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    costs=[
        ("Compensation — investment professionals",192,None,0.228,None,0,"Maestro team replaces consulting, not investment staff"),
        ("Compensation — technology staff",72,None,0.082,None,18,"Maestro delivery replaces £18M Infosys/Wipro spend"),
        ("Compensation — operations staff",84,None,0.096,None,0,"Operational efficiency via AI automation (Wave 2)"),
        ("Bloomberg AIM licence + maintenance",ARC["bloomberg"]["annual_cost_m"],None,0.010,None,0,"AIM modernisation reduces this over 3 years"),
        ("Other technology licences",28,None,0.030,None,0,"Review utilization. FSC adoption improvement reduces per-user cost"),
        ("External consulting (non-Bloomberg)",ARC["total_consulting_m"]-ARC["bloomberg"]["annual_cost_m"],None,0.022,None,ARC["total_consulting_m"]-ARC["bloomberg"]["annual_cost_m"]-18,"Maestro delivery programme reduces by £24M (Wave 1+2)"),
        ("Data & market data",21,None,0.023,None,0,"At benchmark. Renegotiate FactSet on renewal."),
        ("Premises & occupancy",26,None,0.030,None,0,"Singapore office cost elevated post-MAS compliance breach"),
        ("Regulatory & compliance",16,None,0.016,None,0,"MAS FEAT breach adds £2-4M in remediation cost"),
        (f"AI portfolio — zero documented ROI",ARC["ai_budget_committed_m"],None,0.008,None,ARC["ai_budget_committed_m"],"Full AI portfolio reorientation. CDO-led governance from Day 1."),
        ("Other operating costs",55,None,0.060,None,0,"Various. Review with Maestro in Margin engagement."),
    ]

    total_rev=ARC["total_revenue_m"]
    for r,row in enumerate(costs,3):
        ws.row_dimensions[r].height=22
        cat,cost,_,bench_pct,__,ai_r,programme=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        bench_cost=round(bench_pct*total_rev,1)
        overspend=round(cost-bench_cost,1)
        C(ws,r,1,cat,sf,bold=True,b=bdr)
        C(ws,r,2,cost,sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,3,cost/total_rev,sf,"0.0%",align="center",b=bdr)
        C(ws,r,4,bench_pct,sf,"0.0%",align="center",b=bdr)
        of=fills["red"] if overspend>3 else fills["amb"] if overspend>0 else fills["grn"]
        C(ws,r,5,overspend,of,"£#,##0.0",align="center",b=bdr)
        C(ws,r,6,ai_r,fills["teal"] if ai_r>0 else sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,7,programme,sf,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/margin/ARC-M03_Cost_Structure.xlsx")
    return "ARC-M03 done"


def arc_t01_tech_modernisation(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Modernisation Options"
    T(ws,9,f"{ARC['name']} — Technology Modernisation Options Scorecard",fills)
    ws.row_dimensions[2].height=32
    hdrs=[("System",20),("Option",22),("Vendor",18),("Total Cost £M",12),
          ("Duration",11),("Risk Level",11),("Internal Capability",14),
          ("Genome Score",12),("Recommended?",12),("Rationale",30)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    options=[
        # Bloomberg AIM options
        ("Bloomberg AIM","Option A: Full OMS replacement","SS&C Eze / Fidessa",28,36,"Critical","Very Low",28,"No","3 previous failures. Same root causes unaddressed. No internal capability to govern."),
        ("Bloomberg AIM","Option B: API wrapper + internal build","Internal (Maestro-led)",8,24,"High","Medium (builds over 24mo)",64,"CONDITIONAL","Reduces dependency without ripping out. Internal capability builds in parallel. Conditional on CDO hire."),
        ("Bloomberg AIM","Option C: Phased modernisation with ION","ION Trading",18,30,"High","Low → Medium",58,"Alternative","ION has better asset management focus than prior attempts. Still requires executive sponsor."),
        ("Bloomberg AIM","Option D: Stay + cost renegotiation","Bloomberg LP",2,3,"Low","High",42,"Short-term","Renegotiate 2027 contract now. Exit penalty £4.2M. Use Options B/C as leverage."),
        # SQL Server DW (EOL)
        ("SQL Server 2017 DW","Option A: Migrate to Azure SQL","Microsoft Azure",1.2,4,"Medium","High",78,"RECOMMENDED","EOL passed. Azure SQL migration straightforward. Internal team capable. 4-month project."),
        ("SQL Server 2017 DW","Option B: Migrate to Snowflake","Snowflake",2.8,6,"Medium","Medium",72,"Alternative","Better analytics capability but longer migration. Azure SQL simpler given existing Microsoft estate."),
        # Salesforce FSC
        ("Salesforce FSC","Option A: Optimise existing deployment","Maestro-led",2.4,12,"Low","Low → High",81,"RECOMMENDED","44% → 70%+ adoption. Wipro exit. Internal admin. Highest ROI option."),
        ("Salesforce FSC","Option B: Replace with Dynamics 365","Microsoft",8,18,"High","Medium",44,"Not recommended","Migration cost and disruption not justified given FSC capability."),
        # Aladdin
        ("BlackRock Aladdin","Option A: Enhance daily risk","BlackRock + internal",1.8,6,"Low","Medium",74,"RECOMMENDED","Daily stress testing regulatory requirement. Internal data feed automation. Low risk."),
        ("BlackRock Aladdin","Option B: Replace with Axioma","Qontigo / Axioma",6,18,"High","Low",48,"Not recommended","Aladdin well-embedded. Replacement cost not justified."),
    ]

    rec_fills={"RECOMMENDED":fills["grn"],"CONDITIONAL":fills["amb"],"Alternative":fills["amb"],"No":fills["red"],"Not recommended":fills["red"],"Short-term":fills["amb"]}
    for r,row in enumerate(options,3):
        ws.row_dimensions[r].height=28
        system,option,vendor,cost,duration,risk,cap,genome,rec,rationale=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,system,sf,bold=True,b=bdr); C(ws,r,2,option,sf,b=bdr); C(ws,r,3,vendor,sf,b=bdr)
        C(ws,r,4,cost,sf,"£#,##0.0",align="center",b=bdr); C(ws,r,5,duration,sf,align="center",b=bdr)
        rf2={"Critical":fills["red"],"High":fills["amb"],"Medium":fills["amb"],"Low":fills["grn"]}.get(risk,sf)
        C(ws,r,6,risk,rf2,align="center",b=bdr)
        cf=fills["red"] if "Very Low" in cap else fills["amb"] if "Low" in cap else fills["grn"]
        C(ws,r,7,cap,cf,b=bdr)
        gf=fills["grn"] if genome>=70 else fills["amb"] if genome>=50 else fills["red"]
        C(ws,r,8,genome,gf,"#,##0",align="center",b=bdr)
        C(ws,r,9,rec,rec_fills.get(rec,sf),align="center",bold=True,b=bdr)
        C(ws,r,10,rationale,sf,b=bdr)

    # MAS FEAT gap analysis
    ws2=wb.create_sheet("MAS FEAT Compliance Gap")
    T(ws2,6,f"MAS FEAT Compliance Assessment — {ARC['mas_feat']['status']}",fills)
    ws2.row_dimensions[2].height=28
    mas_hdrs=[("FEAT Principle",30),("Category",16),("Status",14),("Gap Description",32),("Risk",10),("Remediation",28)]
    for i,(hd,w) in enumerate(mas_hdrs,1): H(ws2,2,i,hd,w,fills,bdr)
    feat_gaps=[
        ("F1 — Fairness","Fairness","Partial","AI decision audit trail incomplete for 14 AI initiatives.",  "High","Implement full audit logging for all AI model decisions."),
        ("F2 — Ethics","Ethics","Not met","AI ethics review board not constituted. CDO vacancy.","Critical","Constitute AI ethics board. CDO appointment prerequisite."),
        ("F3 — Accountability","Accountability","Not met","No named accountable executive for AI decisions. CDO vacant.","Critical","CDO appointment. Model risk committee with clear accountability."),
        ("F4 — Transparency","Transparency","Partial","Model explainability documented for only 3 of 28 initiatives.","High","Explainability documentation programme across all 28 initiatives."),
        ("F1-F4 — Data & Model Governance","Governance","Not met","No model registry. No validation framework. No data governance board.","Critical","MLOps foundation + model risk framework. CDO-led."),
        ("F5 — Risk Management","Risk","Not met","AI risk not integrated into enterprise risk framework.","High","AI risk taxonomy + OpenPages integration."),
        ("F6 — Human Oversight","Human oversight","Partial","Some models have human override. Not systematic.","High","Systematic human-in-the-loop for all high-risk AI decisions."),
        ("Singapore Data Residency","Data residency","Not met",f"£{ARC['mas_feat']['singapore_aum_b']:.1f}B Singapore AUM data not segregated from global data lake.","Critical","Data residency programme. Singapore data lake separation. 6-month programme."),
    ]
    for r,(principle,cat,status,gap,risk,rem) in enumerate(feat_gaps,3):
        ws2.row_dimensions[r].height=28
        sf=fills["alt"] if r%2==0 else fills["wht"]
        sf_s=fills["red"] if status=="Not met" else fills["amb"] if status=="Partial" else fills["grn"]
        rf2={"Critical":fills["red"],"High":fills["amb"],"Medium":fills["amb"],"Low":fills["grn"]}.get(risk,sf)
        C(ws2,r,1,principle,sf,bold=True,b=bdr); C(ws2,r,2,cat,sf,b=bdr)
        C(ws2,r,3,status,sf_s,align="center",bold=True,b=bdr); C(ws2,r,4,gap,sf,b=bdr)
        C(ws2,r,5,risk,rf2,align="center",bold=True,b=bdr); C(ws2,r,6,rem,sf,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/arcturus/tech/ARC-T01_Modernisation_Options.xlsx")
    return "ARC-T01 done"


# ══════════════════════════════════════════════════════════════════════════
# MERIDIAN — CORE FILES (reusing from generate_meridian.py patterns)
# but now with cross-references to MER constants
# ══════════════════════════════════════════════════════════════════════════

def mer_c01_financial_statements(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "P&L Summary"
    T(ws,6,f"{MER['name']} — Financial Statements Summary {MER['fy']} ($M)",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Line Item",32),("FY2023 ($M)",14),("FY2024 ($M)",14),("FY2025 ($M)",14),("Budget ($M)",14),("Variance ($M)",14)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    pl=[
        ("REVENUE","","","","",""),
        ("Patient services — inpatient",       4240,4420,4614, 4680,  -66),
        ("Patient services — outpatient",       2816,2940,3082, 3120,  -38),
        ("Health plan premium revenue",          956, 996,1044, 1040,    4),
        ("Research & grants",                    168, 172, 178,  180,   -2),
        ("Other revenue",                        264, 276, 282,  280,    2),
        ("TOTAL NET REVENUE",                   8444,8804,9200, 9300, -100),
        ("","","","","",""),
        ("OPERATING COSTS","","","","",""),
        ("Salaries & wages — clinical",         2210,2348,2480, 2400,  -80),
        ("Travel nurse / agency labour",          420, 448, 480,  280, -200),
        ("Physician fees",                       1008,1064,1120, 1060,  -60),
        ("Supplies & implants",                  1176,1260,1344, 1300,  -44),
        ("Purchased services (RCM — Ensemble)",  588, 632, 682,  600,  -82),
        ("IT & technology (Epic + other)",        386, 416, 448,  400,  -48),
        ("AI portfolio (zero verified ROI)",        0,  12,  28,   20,   -8),
        ("Depreciation & amortisation",          560, 588, 616,  600,  -16),
        ("Other operating costs",                840, 896, 946,  900,  -46),
        ("TOTAL OPERATING COSTS",               7188,7664,8144, 7560, -584),
        ("","","","","",""),
        ("OPERATING PROFIT / (LOSS)",           1256,1140,1056, 1740, -684),
        ("OPERATING MARGIN %",               0.149,0.130,0.115,0.187,None),
        ("","","","","",""),
        ("KEY RATIOS — PERFORMANCE GAPS","","","","",""),
        ("RCM Denial Rate",               0.142,0.168,0.182,0.120,None),
        ("Days in AR",                       48,   51,   52,   35, None),
        ("Travel Nurse % of Labour",       0.133,0.139,0.148,0.082,None),
        ("Epic Optimization Score /100",      62,   60,   58,   80, None),
        ("MyChart Patient Adoption",        0.28, 0.31, 0.34, 0.60, None),
        ("MA Star Rating",                   4.0,  3.8,  3.5,  4.0, None),
        ("AI Spend with Verified ROI",       0.0,  0.0,  0.0,  0.5, None),
    ]

    section_rows={"REVENUE","OPERATING COSTS","KEY RATIOS — PERFORMANCE GAPS"}
    total_rows={"TOTAL NET REVENUE","TOTAL OPERATING COSTS","OPERATING PROFIT / (LOSS)"}
    for r,row in enumerate(pl,3):
        ws.row_dimensions[r].height=18
        label,fy23,fy24,fy25,budget,var=row
        if not label: continue
        is_section=label in section_rows; is_total=label in total_rows
        sf=fills["prp"] if is_total else fills["blu"] if is_section else fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,label,sf,bold=is_total or is_section,b=bdr)
        for col,val in [(2,fy23),(3,fy24),(4,fy25),(5,budget),(6,var)]:
            if val=="": C(ws,r,col,None,sf,b=bdr); continue
            is_pct=isinstance(val,float) and abs(val)<2
            fmt="0.0%" if is_pct else "$#,##0"
            trend_bad = (col==4 and label in {"RCM Denial Rate","Days in AR","Travel Nurse % of Labour"} and isinstance(fy25,(int,float)) and isinstance(fy24,(int,float)) and fy25>fy24)
            vf=fills["red"] if col==6 and isinstance(val,(int,float)) and val<-50 else fills["grn"] if col==6 and isinstance(val,(int,float)) and val>20 else sf
            if label=="RCM Denial Rate" and col==4: vf=fills["red"]
            if label=="MA Star Rating" and col==4 and fy25<4.0: vf=fills["red"]
            C(ws,r,col,val,vf,fmt,bold=is_total,align="center",b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/meridian/core/MER-C01_Financial_Statements.xlsx")
    return "MER-C01 done"


def mer_c02_consulting_contracts(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Consulting Contracts"
    T(ws,10,f"{MER['name']} — Consulting & Vendor Contract Register",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Vendor",18),("Scope",28),("Annual $M",12),("Contract End",12),
          ("Notice Period",11),("KT Obligation",18),("KT Compliance %",13),
          ("Exit Penalty $M",13),("Performance",16),("Risk",10)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    contracts_ext=[
        ("Ensemble Health Partners","Full RCM outsourcing — billing, denials, collections, AR management",14.2,"2026-06-30","6 months","Full process documentation required","18%",8.0,"UNDERPERFORMING — 18.2% denial rate vs 12% benchmark. $8M SLA penalties enforceable now.","Critical"),
        ("Epic Systems (PS)","Epic implementation support, upgrades, training",4.2,"2026-12-31","3 months","Epic-controlled documentation","62%",1.2,"Good — Epic responsive. Upgrade quality high. Training completion our gap, not vendor.","Low"),
        ("Wipro","Salesforce Health Cloud administration and development",1.8,"2025-12-31","2 months","Documentation and handover sessions required","22%",0.3,"Below expectations — 38% adoption. Internal team capability not built.","High"),
        ("Cohere Health","Prior auth AI platform — pilot 3 payers",0.4,"2025-06-30","1 month","Platform documentation shared","45%",0.0,"Pilot positive — 94/100 Genome vendor score. Ready to expand.","Low"),
        ("3M / Solventum","Clinical coding AI — pilot",0.6,"2025-09-30","1 month","Model documentation","38%",0.0,"In progress — 84% accuracy. Target 96%. Continue.","Low"),
        ("AWS Professional Services","Cloud migration support — 60% complete",1.2,"2025-12-31","1 month","Runbook delivery required","68%",0.1,"Good — AWS ProServe adds value. Cloud migration on track.","Low"),
        ("Deloitte","Regulatory advisory — HIPAA, CMS compliance",0.8,"2026-03-31","2 months","Policy documentation","52%",0.1,"Adequate for advisory. Execution should be internal.","Medium"),
        ("Various contractors","Epic + data analyst contractors (12 FTE)",2.4,"Rolling","2 weeks","None required","12%",0.0,"Risk — contractors in critical roles. Knowledge walks out weekly.","High"),
    ]

    for r,row in enumerate(contracts_ext,3):
        ws.row_dimensions[r].height=24
        vendor,scope,annual,end_date,notice,kt,kt_pct,penalty,perf,risk=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        risk_fills={"Critical":fills["red"],"High":fills["amb"],"Medium":fills["amb"],"Low":fills["grn"]}
        C(ws,r,1,vendor,sf,bold=True,b=bdr); C(ws,r,2,scope,sf,b=bdr)
        C(ws,r,3,annual,sf,"$#,##0.0",align="center",b=bdr)
        C(ws,r,4,end_date,sf,align="center",b=bdr); C(ws,r,5,notice,sf,align="center",b=bdr)
        C(ws,r,6,kt,sf,b=bdr)
        kt_v=int(kt_pct.replace("%",""))/100
        C(ws,r,7,kt_v,fills["red"] if kt_v<0.25 else fills["amb"] if kt_v<0.5 else fills["grn"],"0%",align="center",b=bdr)
        C(ws,r,8,penalty,fills["red"] if penalty>5 else sf,"$#,##0.0",align="center",b=bdr)
        pf=fills["red"] if "UNDER" in perf else fills["amb"] if "Below" in perf or "progress" in perf else fills["grn"]
        C(ws,r,9,perf,pf,b=bdr)
        C(ws,r,10,risk,risk_fills.get(risk,sf),align="center",bold=True,b=bdr)

    tr=len(contracts_ext)+3
    C(ws,tr,1,"TOTAL ANNUAL VENDOR SPEND",fills["prp"],bold=True,b=bdr)
    C(ws,tr,3,sum(r[2] for r in contracts_ext),fills["prp"],"$#,##0.0",align="center",bold=True,b=bdr)
    C(ws,tr,8,sum(r[7] for r in contracts_ext),fills["red"],"$#,##0.0",align="center",bold=True,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/meridian/core/MER-C02_Consulting_Contracts.xlsx")
    return "MER-C02 done"


def mer_c03_workforce_analytics(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Workforce Analytics"
    T(ws,9,f"{MER['name']} — Workforce Analytics & AI Opportunity",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Department",22),("FTE Count",10),("Travel Nurse FTE",13),("TN % of Dept",12),
          ("TN Annual Cost $M",14),("Benchmark TN %",12),("Gap $M",11),
          ("Turnover Rate",12),("AI Opportunity",14),("Initiative",22)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    departments=[
        ("Medical / Surgical",    3840,  248, 0.064, 7.44, 0.020, -4.84, 0.22,"Demand forecasting","Travel nurse demand forecasting AI"),
        ("Emergency Department",  1680,  312, 0.186,12.48, 0.060, -8.28, 0.31,"Surge prediction","ED surge prediction + staffing AI"),
        ("Intensive Care (ICU)",   680,  104, 0.153, 6.24, 0.050, -2.24, 0.28,"Acuity prediction","Nursing acuity AI for ICU staffing"),
        ("Operating Room",         840,   72, 0.086, 2.88, 0.030, -1.68, 0.18,"Scheduling AI","OR scheduling optimisation AI"),
        ("Cardiovascular",         520,   56, 0.108, 2.24, 0.035, -1.12, 0.24,"Demand forecasting","CV demand forecasting"),
        ("Oncology",               480,   48, 0.100, 1.92, 0.035, -0.84, 0.21,"Infusion scheduling","Infusion centre AI scheduling"),
        ("Behavioural Health",     620,  112, 0.181, 4.48, 0.060, -2.88, 0.38,"Crisis prediction","Crisis prediction + staffing"),
        ("Women's Health / OB",    580,   64, 0.110, 2.56, 0.040, -1.12, 0.26,"Labour onset prediction","L&D labour onset prediction"),
        ("Outpatient / Clinics",  2240,   96, 0.043, 2.88, 0.020, -0.48, 0.14,"No-show prediction","Patient no-show prediction (AI-010)"),
        ("Post-Acute / Rehab",     480,   40, 0.083, 1.60, 0.030, -0.40, 0.19,"Discharge planning","Discharge planning AI"),
        ("Radiology",              320,   16, 0.050, 0.64, 0.020,  0.00, 0.12,"Scheduling","Radiology AI scheduling"),
        ("Other clinical",        5220,  432, 0.083,13.44, 0.030, -3.84, 0.20,"Various","Various workforce AI"),
        ("TOTAL — CLINICAL",     18500, 1600, 0.086,48.00, 0.031,-20.00, 0.22,"$8M recoverable","Travel nurse AI (Wave 1)"),
    ]

    for r,row in enumerate(departments,3):
        ws.row_dimensions[r].height=22
        dept,fte,tn_fte,tn_pct,tn_cost,bench_pct,gap,turnover,ai_opp,initiative=row
        sf=fills["prp"] if "TOTAL" in dept else fills["alt"] if r%2==0 else fills["wht"]
        bold="TOTAL" in dept
        C(ws,r,1,dept,sf,bold=bold,b=bdr)
        C(ws,r,2,fte,sf,"#,##0",align="center",b=bdr)
        C(ws,r,3,tn_fte,sf,"#,##0",align="center",b=bdr)
        tnf=fills["red"] if tn_pct>0.12 else fills["amb"] if tn_pct>0.04 else fills["grn"]
        C(ws,r,4,tn_pct,tnf,"0.0%",align="center",b=bdr)
        C(ws,r,5,tn_cost,fills["red"] if tn_cost>5 else fills["amb"] if tn_cost>2 else sf,"$#,##0.00",align="center",b=bdr)
        C(ws,r,6,bench_pct,fills["grn"],"0.0%",align="center",b=bdr)
        C(ws,r,7,gap,fills["red"] if gap<-1 else sf,"$#,##0.00",align="center",b=bdr)
        tf=fills["red"] if turnover>0.25 else fills["amb"] if turnover>0.15 else fills["grn"]
        C(ws,r,8,turnover,tf,"0.0%",align="center",b=bdr)
        C(ws,r,9,ai_opp,fills["teal"],b=bdr); C(ws,r,10,initiative,sf,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/meridian/core/MER-C03_Workforce_Analytics.xlsx")
    return "MER-C03 done"


def mer_d01_consulting_audit(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Consulting Audit"
    T(ws,9,f"{MER['name']} — Consulting Output vs Promise Audit",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Vendor",18),("Engagement",24),("Promised",24),("Delivered?",14),
          ("Quality /10",10),("KT Score",11),("Annual $M",12),
          ("Value Rating",16),("Recoverable $M",13),("Maestro Replacement?",16)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    audit=[
        ("Ensemble Health","Full RCM outsourcing","Denial rate to 12%, SLA compliance, KT","No — denial 18.2% vs 12%. SLA breached. KT 18%",3,18,14.2,"Poor — $8M SLA penalties enforceable",8.0,"Partial — prior auth AI (Cohere) replaces part. Renegotiate core contract."),
        ("Epic PS","Implementation + ongoing support","Epic best practices, training, 80% optimization","Partial — training low (41% completion). Score 58 vs 80.",6,62,4.2,"Good for support. Training gap is Meridian's issue.",0.0,"No — Epic PS value appropriate. Internal training programme needed."),
        ("Wipro","Salesforce Health Cloud","Platform admin, 80% adoption, internal capability","No — 38% adoption. No internal admin built. Wipro owns all customisations.",3,22,1.8,"Poor — no internal capability built after 2 years",0.8,"Yes — internal Salesforce Health Cloud admin with Maestro."),
        ("Various contractors","Epic + data analysts","Project delivery, knowledge transfer","No — knowledge walks out weekly. 12% KT score.",2,12,2.4,"Critical risk","1.8","Yes — permanent Epic analysts. Contractor to FTE programme."),
        ("Deloitte","Regulatory advisory","CMS compliance guidance, HIPAA frameworks","Adequate — advisory quality good. Execution not included.",6,52,0.8,"Adequate — advisory only. Cannot execute.",0.0,"No — keep advisory. Build internal execution capability with Maestro."),
    ]

    for r,row in enumerate(audit,3):
        ws.row_dimensions[r].height=28
        vendor,eng,promised,delivered,quality,kt,cost,rating,recoverable,replace=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,vendor,sf,bold=True,b=bdr); C(ws,r,2,eng,sf,b=bdr); C(ws,r,3,promised,sf,b=bdr)
        df=fills["red"] if delivered.startswith("No") else fills["amb"] if "Partial" in delivered else fills["grn"]
        C(ws,r,4,delivered,df,b=bdr)
        qf=fills["red"] if quality<5 else fills["amb"] if quality<7 else fills["grn"]
        C(ws,r,5,quality,qf,"#,##0",align="center",b=bdr)
        kf=fills["red"] if kt<25 else fills["amb"] if kt<50 else fills["grn"]
        C(ws,r,6,kt,kf,"#,##0",align="center",b=bdr)
        C(ws,r,7,float(str(cost).replace("$","").replace(",","")),sf,"$#,##0.0",align="center",b=bdr)
        rf2=fills["red"] if any(w in rating for w in ["Poor","Critical"]) else fills["amb"] if "Adequate" in rating or "Partial" in rating else fills["grn"]
        C(ws,r,8,rating,rf2,b=bdr)
        rec=float(str(recoverable).replace("$","").replace(",","")) if recoverable!="N/A" else 0
        C(ws,r,9,rec,fills["grn"] if rec>0 else sf,"$#,##0.0",align="center",b=bdr)
        C(ws,r,10,replace,fills["grn"] if "Yes" in replace else fills["amb"] if "Partial" in replace else fills["wht"],b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/meridian/delivery/MER-D01_Consulting_Audit.xlsx")
    return "MER-D01 done"


def mer_d02_maestro_design(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Maestro Team Design"
    T(ws,7,f"{MER['name']} — Proposed Maestro Team Design",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Maestro Role",24),("Scope",30),("Replaces",22),
          ("Annual Cost Replaced $M",14),("Wave",10),("Duration",13),("Success Metric",28)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    maestros=[
        ("Delivery Maestro — RCM & Prior Auth","Govern Ensemble Health relationship. Deploy Cohere Health prior auth AI. Build internal RCM capability.","8-10 Ensemble Health on-site",8.0,"Wave 1","18 months",f"Denial rate to {MER['rcm']['denial_benchmark']:.0%}. Prior auth days to {MER['rcm']['prior_auth_bench']:.1f}. Internal RCM team capable."),
        ("Delivery Maestro — Epic Optimisation","Drive Epic optimization from 58/100 to 80/100. Deploy AI modules. Build internal Epic team capability.","4 contractors + Wipro Epic staff",3.2,"Wave 1","12 months",f"Epic score 80/100. MyChart adoption {MER['epic']['mychart_target']:.0%}. Training completion 80%. All AI modules configured."),
        ("Delivery Maestro — AI & Clinical","Deploy GenAI clinical documentation. Govern AI clinical portfolio. CDO onboarding support.","Various contractors + 6 AI Lab",2.8,"Wave 1","18 months",f"GenAI documentation live for {MER['workforce']['physicians_fte']} physicians. Prior auth AI in production. CDO onboarded."),
        ("Architecture Maestro — Data & Integration","Build data pipeline automation. Reduce 3-day lag. Establish clinical data architecture for AI.","4 contractors (data analytics)",2.0,"Wave 2","24 months","Real-time clinical data pipeline. Reporting lag <4hrs. Feature store for ML."),
    ]

    for r,row in enumerate(maestros,3):
        ws.row_dimensions[r].height=28
        role,scope,replaces,cost,wave,duration,metric=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        wf=fills["grn"] if wave=="Wave 1" else fills["amb"]
        C(ws,r,1,role,sf,bold=True,b=bdr); C(ws,r,2,scope,sf,b=bdr); C(ws,r,3,replaces,sf,b=bdr)
        C(ws,r,4,cost,fills["grn"],"$#,##0.0",align="center",bold=True,b=bdr)
        C(ws,r,5,wave,wf,align="center",b=bdr); C(ws,r,6,duration,sf,b=bdr); C(ws,r,7,metric,sf,b=bdr)

    tr=len(maestros)+3
    C(ws,tr,1,"TOTAL RECOVERABLE",fills["prp"],bold=True,b=bdr)
    C(ws,tr,4,sum(r[3] for r in maestros),fills["prp"],"$#,##0.0",align="center",bold=True,b=bdr)
    ws.freeze_panes="B3"
    save(wb, f"{base}/meridian/delivery/MER-D02_Maestro_Team_Design.xlsx")
    return "MER-D02 done"


def mer_m01_payer_performance(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Payer Contract Performance"
    T(ws,9,f"{MER['name']} — Payer Contract Performance & Prior Auth Complexity",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Payer",20),("Revenue Mix",11),("Net Rev $M",12),("Denial Rate",11),
          ("Denial $M",11),("Days AR",10),("Collect Rate",11),
          ("Prior Auth Electronic %",14),("Cohere Compatible?",14),("Priority",12)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    payers=[
        ("Medicare FFS",         0.28,2576,0.142,366,54,0.871,0.22,"Yes — CMS X12","CRITICAL — CMS mandate Jan 2027"),
        ("Medicare Advantage",   0.18,1656,0.224,372,61,0.812,0.18,"Yes — pilot done","CRITICAL — highest MA denial"),
        ("Medicaid NC",          0.22,2024,0.168,340,58,0.842,0.12,"Yes — MMIS","HIGH — high volume, complex"),
        ("BCBS NC",              0.12,1104,0.088,97, 42,0.924,0.42,"Yes","MEDIUM — near benchmark"),
        ("Aetna",                0.08, 736,0.096,71, 44,0.912,0.31,"Yes","HIGH"),
        ("Cigna",                0.06, 552,0.112,62, 48,0.904,0.28,"In evaluation","MEDIUM"),
        ("Self-Pay / Uninsured", 0.04, 368,0.000,0,  78,0.448,0.00,"N/A","LOW — collection focus"),
        ("Other / Managed Care", 0.02, 184,0.094,17, 45,0.894,0.35,"Varies","LOW"),
        ("TOTAL",                1.00,9200,MER["rcm"]["denial_rate"],MER["rcm"]["denial_annual_m"],MER["rcm"]["days_in_ar"],MER["rcm"]["collection_rate"],0.21,"—","Benchmark denial 12%"),
    ]

    for r,row in enumerate(payers,3):
        ws.row_dimensions[r].height=22
        payer,mix,rev,denial,denial_d,ar,coll,elec_pct,cohere,priority=row
        sf=fills["prp"] if payer=="TOTAL" else fills["alt"] if r%2==0 else fills["wht"]
        bold=payer=="TOTAL"
        C(ws,r,1,payer,sf,bold=bold,b=bdr)
        C(ws,r,2,mix,sf,"0%",align="center",b=bdr)
        C(ws,r,3,rev,sf,"$#,##0",align="center",b=bdr)
        df=fills["red"] if isinstance(denial,(float,int)) and denial>0.15 else fills["amb"] if isinstance(denial,(float,int)) and denial>0.10 else fills["grn"] if isinstance(denial,(float,int)) else sf
        C(ws,r,4,denial,df,"0.0%" if isinstance(denial,float) else None,align="center",bold=bold,b=bdr)
        C(ws,r,5,denial_d,fills["red"] if isinstance(denial_d,(int,float)) and denial_d>100 else sf,"$#,##0",align="center",b=bdr)
        af=fills["red"] if isinstance(ar,(int,float)) and ar>55 else fills["amb"] if isinstance(ar,(int,float)) and ar>40 else fills["grn"] if isinstance(ar,(int,float)) else sf
        C(ws,r,6,ar,af,"#,##0",align="center",b=bdr)
        C(ws,r,7,coll,sf,"0.0%",align="center",b=bdr)
        ef=fills["grn"] if isinstance(elec_pct,(float,int)) and elec_pct>0.35 else fills["amb"] if isinstance(elec_pct,(float,int)) and elec_pct>0.20 else fills["red"] if isinstance(elec_pct,(float,int)) else sf
        C(ws,r,8,elec_pct,ef,"0%" if isinstance(elec_pct,float) else None,align="center",b=bdr)
        cf=fills["grn"] if "Yes" in str(cohere) else fills["amb"] if "evaluat" in str(cohere).lower() else fills["wht"]
        C(ws,r,9,cohere,cf,b=bdr)
        pf=fills["red"] if "CRITICAL" in priority else fills["amb"] if "HIGH" in priority else fills["grn"] if "MEDIUM" in priority else sf
        C(ws,r,10,priority,pf,align="center",b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/meridian/margin/MER-M01_Payer_Performance.xlsx")
    return "MER-M01 done"


def mer_m02_drg_cost(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Cost Per Case by DRG"
    T(ws,8,f"{MER['name']} — Cost Per Case by Top 20 DRGs",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("DRG",10),("Description",28),("Annual Volume",12),("Revenue Per Case $",13),
          ("Cost Per Case $",13),("Margin Per Case $",13),("Margin %",11),
          ("Benchmark Margin %",13),("Gap",10),("AI Opportunity",22)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    drgs=[
        ("470","Major joint replacement — lower extremity",2840,22400,19800,2600,0.116,0.150,-0.034,"Surgical AI scheduling + implant cost reduction"),
        ("291","Heart failure w/ CC",1920,8200,7800,400,0.049,0.080,-0.031,"Readmission prediction AI — $6M annual"),
        ("392","Esophagitis, gastroenteritis",2280,4800,4400,400,0.083,0.100,-0.017,"Clinical documentation AI — coding accuracy"),
        ("690","Kidney & urinary tract infections",1680,4200,4000,200,0.048,0.075,-0.027,"Sepsis prediction AI — reduce LOS"),
        ("378","GI haemorrhage",1440,7400,7100,300,0.041,0.072,-0.031,"Clinical decision support — appropriate intervention"),
        ("871","Septicaemia — no ventilator",1320,12800,13400,-600,-0.047,0.040,-0.087,"Sepsis early warning AI — reduce mortality + cost"),
        ("193","Simple pneumonia w/o MCC",1560,5600,5200,400,0.071,0.090,-0.019,"GenAI documentation — reduce coding denials"),
        ("460","Spinal fusion except cervical",960,38400,32000,6400,0.167,0.200,-0.033,"OR scheduling + implant cost AI"),
        ("247","Percutaneous cardiovascular procedure",1080,24200,20400,3800,0.157,0.180,-0.023,"Cath lab scheduling AI"),
        ("698","Total cholecystectomy",1200,8400,7200,1200,0.143,0.160,-0.017,"Same-day surgery optimisation AI"),
        ("312","Syncope & collapse",1440,3800,3600,200,0.053,0.070,-0.017,"Observation vs admit AI decision support"),
        ("552","Medical back problems",1320,4600,4400,200,0.043,0.068,-0.025,"Pain management AI + early discharge prediction"),
        ("065","Intracranial haemorrhage — stroke",720,18400,19200,-800,-0.043,0.040,-0.083,"Stroke pathway AI — reduce door-to-treatment time"),
        ("795","Normal newborn",2640,2800,2600,200,0.071,0.085,-0.014,"Discharge planning AI"),
        ("775","Vaginal delivery",1920,4200,3800,400,0.095,0.110,-0.015,"OB pathway AI — reduce LOS"),
        ("101","Seizures w/o MCC",840,6400,6200,200,0.031,0.060,-0.029,"Neurology AI — medication optimisation"),
        ("682","Renal failure w/o CC/MCC",1080,5200,5000,200,0.038,0.065,-0.027,"CKD management AI — reduce progression"),
        ("330","Major small/large bowel procedures",480,28400,25200,3200,0.113,0.150,-0.037,"Surgical AI + enhanced recovery protocol"),
        ("287","Circulatory disorders",960,8800,9200,-400,-0.045,0.045,-0.090,"Clinical decision support — appropriate intervention"),
        ("603","Cellulitis",1320,3600,3400,200,0.056,0.075,-0.019,"Discharge planning AI — reduce LOS"),
    ]

    for r,row in enumerate(drgs,3):
        ws.row_dimensions[r].height=22
        drg,desc,vol,rev,cost,margin_d,margin_pct,bench,gap,ai_opp=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,drg,sf,align="center",b=bdr); C(ws,r,2,desc,sf,bold=True,b=bdr)
        C(ws,r,3,vol,sf,"#,##0",align="center",b=bdr)
        C(ws,r,4,rev,sf,"$#,##0",align="center",b=bdr)
        C(ws,r,5,cost,sf,"$#,##0",align="center",b=bdr)
        mf=fills["grn"] if margin_d>2000 else fills["amb"] if margin_d>0 else fills["red"]
        C(ws,r,6,margin_d,mf,"$#,##0",align="center",b=bdr)
        mpf=fills["red"] if margin_pct<0 else fills["amb"] if margin_pct<0.08 else fills["grn"]
        C(ws,r,7,margin_pct,mpf,"0.0%",align="center",b=bdr)
        C(ws,r,8,bench,fills["grn"],"0.0%",align="center",b=bdr)
        C(ws,r,9,gap,fills["red"] if gap<-0.04 else fills["amb"],"0.0%",align="center",b=bdr)
        C(ws,r,10,ai_opp,fills["teal"],b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/meridian/margin/MER-M02_DRG_Cost_Analysis.xlsx")
    return "MER-M02 done"


def mer_t01_epic_integration_map(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Epic Integration Map"
    T(ws,8,f"{MER['name']} — Epic Integration Map (42 HL7/FHIR Connections)",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Source System",20),("Target",20),("Interface Type",14),("Standard",12),
          ("Frequency",14),("Automated?",11),("Critical?",10),("AI Impact",20),("Status",12)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    integrations=[
        ("Epic","Ensemble Health Partners","Claims outbound","HL7 v2.3","Real-time","Yes","Critical","Prior auth workflow — needs enhancement for Cohere integration","Active"),
        ("Epic","Cohere Health (Pilot)","Prior auth requests","FHIR R4","Real-time","Partial","Critical","CMS mandate Jan 2027 — electronic prior auth","Pilot — 60%"),
        ("Ensemble Health","Epic","Denial management","HL7 v2.3","Daily","Yes","Critical","Denial data drives workflow. Integration quality affects AI training data","Active"),
        ("Epic","Radiology (PACS)","Radiology orders/results","HL7 v2.3 ORM/ORU","Real-time","Yes","High","AI radiology interpretation needs this feed","Active"),
        ("Epic","Pharmacy (Omnicell)","Medication administration","HL7 v2.3","Real-time","Yes","High","Medication AI decision support depends on this","Active"),
        ("Epic","Lab (Sunquest)","Lab orders/results","HL7 v2.3","Real-time","Yes","High","Clinical AI sepsis prediction needs real-time lab","Active"),
        ("Epic","Patient monitoring","ADT / vitals","HL7 v2.3 ADT","Real-time","Yes","High","ICU AI early warning systems","Active"),
        ("Epic","Infor SCM","Supply chain orders","HL7 custom","Daily","No — manual","Medium","Supply chain AI needs clean order data","Partial — manual"),
        ("Epic","Workday","Employee / workforce","Custom API","Weekly","No — manual","Low","Travel nurse forecasting AI — Workday data needed","Manual"),
        ("Epic","Press Ganey","Patient satisfaction","SFTP","Monthly","Yes","Medium","MA Star Rating — patient experience component","Active"),
        ("Epic","Salesforce Health Cloud","Patient outreach","SFTP + API","Daily","Partial","Medium","Care gap outreach AI — needs real-time ideally","Partial"),
        ("Epic","State Immunization Registry","Immunization data","HL7 v2.5","Daily","Yes","Medium","Population health AI","Active"),
        ("Epic","Cerner (2 legacy hospitals)","Patient record sharing","HL7 v2.3","Daily","Yes","Critical","Epic migration in progress — data migration complexity high","Active — migration planned"),
        ("Cerner","Epic","ADT notifications","HL7 ADT","Real-time","Yes","Critical","Duplicate patient risk. Critical until migration complete","Active"),
        ("Epic","State HIE","Care Everywhere","FHIR R4","On-demand","Yes","High","Care coordination AI — external record access","Active"),
        ("Epic","Nuance DAX (Pilot)","Ambient documentation","API","Real-time","Yes","High","GenAI documentation — Nuance pilot. Competing with AWS Bedrock option","Pilot"),
        ("Epic","AWS Bedrock (POC)","GenAI documentation POC","API","Real-time","Partial","High","GenAI documentation — AWS route competing with Nuance","POC"),
        ("Epic","3M HIS Coding AI","Clinical coding","FHIR R4","Daily","Partial","Medium","Coding AI pilot — 84% accuracy. Target 96%","Pilot"),
        ("Epic","Medicare portal (CMS)","Value-based care reporting","CMS API","Monthly","Yes","High","MA Star Rating data submission","Active"),
        ("Epic","BCBS NC portal","Prior auth (BCBS)","FHIR R4","Real-time","Partial","High","Electronic prior auth — partial. Full activation opportunity","Active — 42% electronic"),
    ]

    for r,row in enumerate(integrations,3):
        ws.row_dimensions[r].height=22
        src,tgt,type_,std,freq,auto,crit,ai_impact,status=row
        sf=fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,src,sf,bold=True,b=bdr); C(ws,r,2,tgt,sf,b=bdr); C(ws,r,3,type_,sf,b=bdr)
        C(ws,r,4,std,sf,align="center",b=bdr); C(ws,r,5,freq,sf,align="center",b=bdr)
        af=fills["grn"] if auto=="Yes" else fills["amb"] if auto=="Partial" else fills["red"]
        C(ws,r,6,auto,af,align="center",b=bdr)
        cf=fills["red"] if crit=="Critical" else fills["amb"] if crit=="High" else fills["grn"]
        C(ws,r,7,crit,cf,align="center",b=bdr)
        C(ws,r,8,ai_impact,fills["teal"] if "AI" in ai_impact or "GenAI" in ai_impact else sf,b=bdr)
        sf_s=fills["red"] if "migration" in status.lower() or "Pilot" in status else fills["amb"] if "Partial" in status or "manual" in status.lower() else fills["grn"]
        C(ws,r,9,status,sf_s,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/meridian/tech/MER-T01_Epic_Integration_Map.xlsx")
    return "MER-T01 done"


def mer_t02_cerner_migration(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Cerner Migration Assessment"
    T(ws,8,f"{MER['name']} — Cerner to Epic Migration Assessment (2 Hospitals)",fills)
    ws.row_dimensions[2].height=28
    hdrs=[("Dimension",26),("Current State",28),("Target State",22),("Complexity",12),
          ("Risk",10),("Duration",11),("Cost $M",10),("Blocker",24)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    dimensions=[
        ("Hospitals on Cerner","2 community hospitals — St. Andrews (420 beds), Valley Medical (280 beds)","Epic — same instance as main system","High","High","18 months","12.4","Data migration planning not started. Go-live Q4 2026 target."),
        ("Patient data migration","14 years of Cerner patient history. 284k patient records. Complex discrete data.","Full historical data in Epic","Very High","Critical","Concurrent with build","4.2","Cerner data model differs significantly from Epic. Custom mapping required."),
        ("Cerner interface decommission","12 active HL7 interfaces on Cerner. ADT feed to main Epic instance.","All interfaces on Epic","High","High","Last 6 months","1.8","Dual-live period required. Complex ADT reconciliation."),
        ("Staff training","1,400 clinical and administrative staff across 2 hospitals.","Epic-certified staff","Medium","Medium","6 months pre-go-live","2.8","Training schedule not developed. Epic trainer capacity uncertain."),
        ("Downtime procedures","Cerner downtime procedures established. Epic downtime different.","Epic downtime procedures","Medium","Medium","3 months","0.4","Downtime policy rewrite required."),
        ("Clinical workflow redesign","Current Cerner workflows embedded in clinical practice for 14 years.","Epic-optimized workflows","High","High","12 months","1.6","Change management critical. Clinician workflow champions required."),
        ("Pharmacy system migration","Cerner PharmNet currently. Epic Willow target.","Epic Willow","High","High","12 months","1.4","Medication history reconciliation critical for patient safety."),
        ("Revenue cycle migration","Cerner RevCycle in both hospitals. Different denial management process.","Epic revenue cycle","High","Critical","12 months","2.2","Denial rate may temporarily increase during transition. Ensemble Health contract impact."),
        ("TOTAL MIGRATION PROGRAMME","","","Very High","High","Q4 2026 target","26.8","CDO hire supports governance. Maestro recommended for programme delivery."),
    ]

    for r,row in enumerate(dimensions,3):
        ws.row_dimensions[r].height=28
        dim,curr,target,comp,risk,dur,cost,blocker=row
        sf=fills["prp"] if "TOTAL" in dim else fills["alt"] if r%2==0 else fills["wht"]
        bold="TOTAL" in dim
        C(ws,r,1,dim,sf,bold=bold,b=bdr); C(ws,r,2,curr,sf,b=bdr); C(ws,r,3,target,sf,b=bdr)
        cf={"Very High":fills["red"],"High":fills["amb"],"Medium":fills["amb"],"Low":fills["grn"]}.get(comp,sf)
        C(ws,r,4,comp,cf,align="center",b=bdr)
        rf2={"Critical":fills["red"],"High":fills["amb"],"Medium":fills["amb"],"Low":fills["grn"]}.get(risk,sf)
        C(ws,r,5,risk,rf2,align="center",bold=True,b=bdr)
        C(ws,r,6,dur,sf,align="center",b=bdr)
        C(ws,r,7,float(cost.replace("$","").replace(",","")),sf if not bold else fills["prp"],"$#,##0.0",align="center",bold=bold,b=bdr)
        C(ws,r,8,blocker,fills["red"] if "not started" in blocker.lower() else sf,b=bdr)

    ws.freeze_panes="B3"
    save(wb, f"{base}/meridian/tech/MER-T02_Cerner_Migration.xlsx")
    return "MER-T02 done"


# ══════════════════════════════════════════════════════════════════════════
# TASK RUNNER
# ══════════════════════════════════════════════════════════════════════════
def run_task(args):
    func, base = args
    try:
        result = func(base)
        return [result] if isinstance(result, str) else result
    except Exception as e:
        return [f"ERROR in {func.__name__}: {e}\n{traceback.format_exc()}"]


if __name__ == "__main__":
    tasks = [
        # Arcturus core
        (arc_c01_engineering_org,     BASE),
        (arc_c02_financial_statements, BASE),
        (arc_c03_leadership,          BASE),
        (arc_c04_technology_landscape, BASE),
        (arc_c05_sprint_velocity,     BASE),
        # Arcturus PDLC
        (arc_p01_ai_initiatives,      BASE),
        (arc_p02_data_architecture,   BASE),
        (arc_p03_mlops,               BASE),
        (arc_p04_bloomberg_aim,       BASE),
        (arc_p05_engineering_cost,    BASE),
        # Arcturus Delivery
        (arc_d01_consulting_audit,    BASE),
        (arc_d02_knowledge_risk,      BASE),
        (arc_d03_maestro_design,      BASE),
        # Arcturus Margin
        (arc_m01_pl_detail,           BASE),
        (arc_m02_ai_roi,              BASE),
        (arc_m03_cost_structure,      BASE),
        # Arcturus Tech
        (arc_t01_tech_modernisation,  BASE),
        # Meridian core
        (mer_c01_financial_statements, BASE),
        (mer_c02_consulting_contracts, BASE),
        (mer_c03_workforce_analytics,  BASE),
        # Meridian delivery
        (mer_d01_consulting_audit,    BASE),
        (mer_d02_maestro_design,      BASE),
        # Meridian margin
        (mer_m01_payer_performance,   BASE),
        (mer_m02_drg_cost,            BASE),
        # Meridian tech
        (mer_t01_epic_integration_map, BASE),
        (mer_t02_cerner_migration,    BASE),
    ]

    print(f"\n{'='*60}")
    print(f"AbarVa Complete Dataset Generator v2")
    print(f"{len(tasks)} files — Arcturus + Meridian — fully cross-referenced")
    print(f"{'='*60}\n")

    start = time.time()
    with mp.Pool(processes=min(len(tasks), mp.cpu_count())) as pool:
        results = pool.map(run_task, tasks)
    elapsed = time.time() - start

    all_results = [r for group in results for r in group]
    errors = [r for r in all_results if "ERROR" in r]
    successes = [r for r in all_results if "ERROR" not in r]

    print("Results:")
    for r in all_results:
        print(f"  {'OK' if 'ERROR' not in r else 'FAIL'} {r}")

    print(f"\n{'='*60}")
    print(f"Complete: {len(successes)} files, {len(errors)} errors, {elapsed:.1f}s")

    all_files = [f for f in BASE.rglob("*.xlsx") if "scripts" not in str(f)]
    print(f"\nAll files ({len(all_files)}):")
    for f in sorted(all_files):
        print(f"  {str(f.relative_to(BASE))}")

    if errors:
        print("\nErrors:")
        for e in errors: print(e)
        sys.exit(1)
    else:
        print(f"\nDone. Run: git add datasets/ && git commit -m 'datasets: complete v2'")
