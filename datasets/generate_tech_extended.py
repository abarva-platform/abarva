"""
AbarVa — Technology Modernization Extended Datasets
Covers: Technical Debt, Vendor Contracts, Epic Roadmap,
        Integration Enhancement, ERP Genome, Cloud Architecture Advisory
For: Arcturus Financial + Meridian Health
Run: python3 generate_tech_extended.py
"""
import multiprocessing as mp
import sys, time, traceback
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).parent))
from constants import MER, ARC, GENOME

BASE = Path(__file__).parent

def S():
    t = Side(style="thin", color="CCCCCC")
    b = Border(left=t, right=t, top=t, bottom=t)
    f = {
        "hdr":  PatternFill("solid", start_color="1A3A5C"),
        "alt":  PatternFill("solid", start_color="F2F7FC"),
        "wht":  PatternFill("solid", start_color="FFFFFF"),
        "red":  PatternFill("solid", start_color="FDE8E8"),
        "amb":  PatternFill("solid", start_color="FFF4E5"),
        "grn":  PatternFill("solid", start_color="E8F5E9"),
        "blu":  PatternFill("solid", start_color="E8F0F8"),
        "prp":  PatternFill("solid", start_color="F0EEFF"),
        "teal": PatternFill("solid", start_color="E0F7F4"),
        "dkred":PatternFill("solid", start_color="FF6B6B"),
    }
    return b, f

def H(ws, row, col, val, w=16, f=None, b=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    c.fill = f["hdr"]; c.border = b
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(col)].width = w

def C(ws, row, col, val, fill=None, fmt=None, bold=False, align="left", b=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(size=9, name="Arial", bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = b
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    return c

def T(ws, cols, text, f):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    ws["A1"].value = text
    ws["A1"].font = Font(bold=True, size=12, color="1A3A5C", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = f["blu"]; ws.row_dimensions[1].height = 16

def save(wb, path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════
# ARC-T02: Technical Debt Assessment (document-based)
# ══════════════════════════════════════════════════════════════════════════

def arc_t02_technical_debt(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Technical Debt Assessment"
    T(ws, 9, f"{ARC['name']} — Technical Debt Assessment by System", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("System",22),("Age Yrs",9),("Tech Debt Score /100",16),
            ("Documentation Score",16),("Test Coverage %",13),
            ("API Surface",14),("Coupling Risk",14),
            ("Est Remediation Cost £M",18),("Recommendation",24)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    systems = [
        # System, Age, Debt Score, Doc Score, Test%, API, Coupling, Cost £M, Rec
        ("Bloomberg AIM",28,94,"None — vendor black box",0,"Proprietary FIX only","Critical — 14 tightly coupled customisations",8.4,"API wrapper first. Full assessment before any migration commitment."),
        ("SQL Server 2017 DW",7,88,"Partial — some ETL docs",12,"SQL queries only — no REST","High — 14 systems feed this manually",1.2,"IMMEDIATE: Migrate to Azure SQL. EOL passed. Security risk today."),
        ("BlackRock Aladdin",11,42,"Good — BlackRock provides","N/A — SaaS","REST API available","Medium — position feed from AIM creates dependency",0.4,"Enhancement only. Daily risk feed automation. No replacement needed."),
        ("Salesforce FSC",4,38,"Good — Salesforce docs","N/A — SaaS","Full REST/APEX API","Low — API-first platform",0.6,"Adoption programme. Internal admin hire. Exit Wipro dependency."),
        ("SimCorp Dimension",9,48,"Moderate","N/A — vendor","SimCorp API","Medium — fund accounting central",0.8,"Upgrade to latest version. Automate NAV extraction."),
        ("Charles River IMS",6,32,"Good — internal team capable","N/A — SaaS","FIX + REST","Low — well-understood",0.2,"No action needed. Best-managed system."),
        ("Murex",12,56,"Moderate","Limited","Murex API","Medium — FX and derivatives",0.3,"Treasury ops automation. API integration with DW."),
        ("Geneva (SS&C)",8,51,"Moderate","N/A — SaaS","SS&C API","Medium — alternatives data silo",0.3,"API connection to SimCorp. Eliminate manual weekly extract."),
        ("OpenPages (IBM)",5,62,"Poor — IBM docs only","N/A — SaaS","IBM REST API","Low — standalone GRC",0.2,"Internal training. Reduce IBM dependency."),
        ("Bloomberg Terminal",15,28,"N/A — market data","N/A","Bloomberg API (B-PIPE)","Low — data only",0.1,"Renegotiate on renewal. Assess FactSet overlap."),
        ("Linedata Longview",7,71,"Poor","None","Limited","Medium — alternatives OMS",0.4,"Replace or migrate to Charles River. Vendor uncertain."),
        ("Broadridge",11,35,"Good","N/A — SaaS","SWIFT + REST","Low — post-trade standard",0.1,"No action. Well-integrated."),
    ]

    for r, row in enumerate(systems, 3):
        ws.row_dimensions[r].height = 28
        name, age, debt, doc, test, api, coupling, cost, rec = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,name,sf,bold=True,b=bdr)
        af = fills["red"] if age>20 else fills["amb"] if age>10 else sf
        C(ws,r,2,age,af,"#,##0",align="center",b=bdr)
        df = fills["red"] if isinstance(debt,int) and debt>75 else fills["amb"] if isinstance(debt,int) and debt>50 else fills["grn"] if isinstance(debt,int) else sf
        C(ws,r,3,debt,df,"#,##0" if isinstance(debt,int) else None,align="center",bold=True,b=bdr)
        C(ws,r,4,doc,sf,b=bdr)
        tf = fills["red"] if isinstance(test,int) and test<20 else fills["amb"] if isinstance(test,int) and test<50 else fills["grn"] if isinstance(test,int) else sf
        C(ws,r,5,test,tf,"0%" if isinstance(test,int) and test>0 else None,align="center",b=bdr)
        C(ws,r,6,api,sf,b=bdr)
        cf = {"Critical":fills["red"],"High":fills["amb"],"Medium":fills["amb"],"Low":fills["grn"]}.get(coupling.split(" — ")[0] if " — " in coupling else coupling,sf)
        C(ws,r,7,coupling,cf,b=bdr)
        C(ws,r,8,cost,sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,9,rec,sf,b=bdr)

    # API Surface map sheet
    ws2 = wb.create_sheet("API Wrapper Opportunities")
    T(ws2, 6, f"{ARC['name']} — API Wrapper Analysis (Reduce Vendor Lock-In)", fills)
    ws2.row_dimensions[2].height = 28
    api_hdrs = [("System",22),("Current API Access",20),("Wrapper Feasibility",14),
                ("Functions to Wrap",28),("Est Build Cost £M",14),
                ("Risk Reduction",16),("Priority",10)]
    for i,(hd,w) in enumerate(api_hdrs,1): H(ws2,2,i,hd,w,fills,bdr)
    wrappers = [
        ("Bloomberg AIM","Proprietary FIX protocol — limited","Medium","Order routing, position query, compliance rules (6 of 14 customisations)",2.4,"High — reduces dependency for 6 customisations","High"),
        ("Salesforce FSC","Full REST/APEX API — excellent","High","All customisations wrappable. Wipro exit enabled by this.",0.6,"High — enables full internal control","Critical"),
        ("SimCorp Dimension","SimCorp API — good","High","NAV extraction, fund accounting query, position feeds",0.4,"Medium — reduces manual extraction",  "High"),
        ("Geneva","SS&C API — available","Medium","Alternatives NAV, LP data extraction",0.3,"Medium — eliminates weekly manual","Medium"),
        ("Murex","Murex API — moderate","Medium","FX rate feed, treasury P&L extraction",0.3,"Medium","Medium"),
        ("SQL Server DW","SQL only — no API","Low — replace first","Migration to Azure SQL creates REST API layer",1.2,"Critical — eliminates EOL risk","Critical — replace first"),
    ]
    for r, row in enumerate(wrappers, 3):
        ws2.row_dimensions[r].height = 28
        system, api_access, feasibility, functions, cost, risk_red, priority = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        ff = fills["grn"] if "High" in feasibility else fills["amb"] if "Medium" in feasibility else fills["red"]
        pf = fills["red"] if "Critical" in priority else fills["amb"] if "High" in priority else fills["grn"]
        C(ws2,r,1,system,sf,bold=True,b=bdr); C(ws2,r,2,api_access,sf,b=bdr)
        C(ws2,r,3,feasibility,ff,align="center",b=bdr); C(ws2,r,4,functions,sf,b=bdr)
        C(ws2,r,5,cost,sf,"£#,##0.0",align="center",b=bdr)
        C(ws2,r,6,risk_red,sf,b=bdr)
        C(ws2,r,7,priority,pf,align="center",bold=True,b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/arcturus/tech/ARC-T02_Technical_Debt_Assessment.xlsx")
    return "ARC-T02 done"


# ══════════════════════════════════════════════════════════════════════════
# ARC-T03: Vendor Contract Intelligence
# ══════════════════════════════════════════════════════════════════════════

def arc_t03_vendor_contracts(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Vendor Contract Intelligence"
    T(ws, 9, f"{ARC['name']} — Vendor Contract Intelligence & Exit Analysis", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Vendor",18),("Contract End",12),("Notice Period",13),
            ("Exit Penalty £M",13),("Data Portability",16),
            ("Exit Complexity",14),("Leverage Points",28),
            ("Recommended Action",24),("Timeline",12)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    contracts = [
        ("Bloomberg LP (AIM)","2027-03-31","12 months",4.2,
         "Position history in proprietary format. No data export tool. Requires Bloomberg cooperation.",
         "Critical","AIM modernisation alternatives (ION, SS&C) as credible threat. MAS FEAT breach gives regulatory urgency.",
         "Begin renegotiation NOW. Use 2027 renewal as leverage. Commission API wrapper to reduce dependency before negotiation.",
         "Start immediately"),
        ("Bloomberg LP (Terminal)","2026-12-31","6 months",0.9,
         "Market data standard formats. B-PIPE API exportable.","Low",
         "FactSet provides 80% of same data at lower cost. Use as competitive threat.",
         "Renegotiate on renewal. 15-20% reduction achievable. FactSet RFP as leverage.","Q3 2026"),
        ("Wipro","2026-05-31","3 months",1.4,
         "FSC customisations in Wipro-controlled codebase. Code escrow clause absent.","High",
         "Salesforce PS can replace Wipro. Internal admin hire is the enabling step. KT score 15% is breach.",
         "Exit plan in Wave 1. Code escrow clause NOW. Internal Salesforce admin hired within 60 days.","Wave 1 — 60 days"),
        ("Infosys","2025-12-31","3 months",1.0,
         "Standard code handover. Some documentation.","Medium",
         "KT score 22% — contractual KT obligation not met. Basis for penalty or reduced exit fee.",
         "Partial exit. Retain Aladdin/Risk team only. Exit Portfolio Analytics (internal can do). Cite KT breach.","Q4 2025 — urgent"),
        ("Deloitte","2025-08-31","1 month",0.2,
         "Policy documents — Deloitte retains IP. Client gets output only.","Low",
         "MAS FEAT breach occurred on their watch. Grounds for SLA penalty and renegotiation.",
         "Do not renew compliance programme scope. Retain advisory only. Invoke SLA breach clauses.","August 2025 — imminent"),
        ("TCS","2025-12-31","3 months",0.3,
         "Standard code handover.","Low","TCS relationship value-for-money in APAC. No leverage needed.","Renew. APAC ops retain value.","Renew Q4 2025"),
        ("AWS ProServe","2025-05-31","1 month",0.1,
         "Infrastructure runbooks — client owns.","Low","AWS ProServe has runbook delivery obligation outstanding (88% complete).","Enforce runbook completion before renewal. Then renew.","Complete first"),
        ("Salesforce PS","2026-10-31","3 months",0.8,
         "Salesforce-documented configurations.","Medium","FSC adoption 44% — below any reasonable target. Grounds for renegotiation.","Renegotiate scope to exclude delivery. Advisory only. Internal admin takes over.","Q2 2026"),
        ("Contractors (EA)","Rolling","2 weeks",0.0,
         "None — no documentation obligation.","Critical","Zero leverage. No contract protections. Critical dependency with no safety net.",
         "IMMEDIATE: Permanent EA hire. Documentation sprint. Remove contractors from architectural decision authority.","Immediate"),
        ("Google PSO","Ended 2024","Expired",0.0,
         "22% of promised deliverables. Partial design docs.","N/A — expired",
         "Engagement ended. No further leverage. Document what was delivered.",
         "Reconstruct MLOps design from artefacts. Do not re-engage Google PSO without reformed scope and governance.","Complete"),
    ]

    for r, row in enumerate(contracts, 3):
        ws.row_dimensions[r].height = 32
        vendor, end, notice, penalty, portability, complexity, leverage, action, timeline = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        cf = {"Critical":fills["red"],"High":fills["amb"],"Medium":fills["amb"],"Low":fills["grn"],"N/A — expired":fills["wht"]}.get(complexity,sf)
        C(ws,r,1,vendor,sf,bold=True,b=bdr)
        end_f = fills["red"] if "2025" in end or "Ended" in end else fills["amb"] if "2026" in end else sf
        C(ws,r,2,end,end_f,align="center",b=bdr)
        C(ws,r,3,notice,sf,align="center",b=bdr)
        C(ws,r,4,penalty,fills["red"] if penalty>3 else fills["amb"] if penalty>0.5 else sf,"£#,##0.0",align="center",b=bdr)
        C(ws,r,5,portability,fills["red"] if "proprietary" in portability.lower() or "absent" in portability.lower() else sf,b=bdr)
        C(ws,r,6,complexity,cf,align="center",bold=True,b=bdr)
        C(ws,r,7,leverage,sf,b=bdr)
        C(ws,r,8,action,fills["red"] if "IMMEDIATE" in action or "immediate" in action.lower() else fills["amb"] if "Wave 1" in action or "urgent" in timeline.lower() else sf,b=bdr)
        tf = fills["red"] if "immediate" in timeline.lower() or "Immediate" in timeline else fills["amb"] if "2025" in timeline or "60 days" in timeline else sf
        C(ws,r,9,timeline,tf,align="center",b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/arcturus/tech/ARC-T03_Vendor_Contract_Intelligence.xlsx")
    return "ARC-T03 done"


# ══════════════════════════════════════════════════════════════════════════
# MER-T03: Epic Optimization Roadmap
# ══════════════════════════════════════════════════════════════════════════

def mer_t03_epic_roadmap(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Epic Optimization Roadmap"
    T(ws, 9, f"{MER['name']} — Epic Optimization Roadmap (58→80/100)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Module",24),("Current Score",13),("Target Score",13),
            ("Current Adoption %",14),("Target Adoption %",14),
            ("Config Changes Required",26),("Training Required",20),
            ("Wave",8),("Score Impact",12),("Value $M",10)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    modules = [
        # Module, Current, Target, CurrAdopt%, TgtAdopt%, Config, Training, Wave, Score Impact, Value
        ("Epic Prior Auth Automation",8,85,0.08,0.80,"Activate Epic Prior Auth module. Configure payer-specific rules for top 8 payers. Connect to Cohere Health via FHIR R4.","All prior auth staff (142 FTE). Physician workflow for prior auth requests.",1,"+8 points","$37.6M — CMS mandate Jan 2027"),
        ("Epic Healthy Planet (Pop Health)",22,75,0.22,0.65,"Activate care gap workflows. Configure quality measure tracking for MA Star Rating. Enable patient outreach templates.","Care management team (84 FTE). Population health nurses.",1,"+6 points","$24M MA Star Rating revenue"),
        ("Epic Tapestry (Health Plan)",44,80,0.44,0.75,"Configure MA plan benefit structures. Enable quality measure automated tracking. Connect to CMS reporting API.","Health plan operations team (48 FTE).",1,"+5 points","$24M — MA Star Rating composite"),
        ("Epic Cogito (Analytics)",28,70,0.28,0.65,"Activate Cogito dashboards for service line leaders. Configure automated daily operational reports. Replace manual Clarity queries.","Service line directors (42 FTE). Finance team (36 FTE).",1,"+5 points","$1.8M manual analytics cost"),
        ("Epic GPT / Ambient Documentation",12,60,0.12,0.55,"Configure ambient documentation for 3 specialties. Integrate with physician workflow. Set up review and approval process.","All physicians (732 FTE). Physician IT champions (24 FTE).",1,"+4 points","$42M physician productivity"),
        ("Epic Cheers (CRM Patient Outreach)",18,65,0.18,0.55,"Configure care gap outreach campaigns. Enable appointment reminder automation. Set up readmission prevention workflows.","Patient access team (68 FTE). Care managers (42 FTE).",2,"+4 points","$4.2M outreach efficiency"),
        ("MyChart Patient Portal",34,65,0.34,0.60,"Simplify MyChart enrollment at registration. Enable MyChart Bedside for inpatients. Configure self-scheduling for 8 specialties.","Registration staff (284 FTE). Patient experience team (18 FTE).",1,"+4 points","$4.2M patient engagement"),
        ("Epic Beacon (Oncology)",52,85,0.52,0.80,"Complete oncology workflow templates. Configure chemotherapy administration protocols. Enable treatment plan documentation.",  "Oncology nurses and physicians (68 FTE).",2,"+3 points","$1.4M oncology efficiency"),
        ("Epic Welcome (Self Check-In)",42,80,0.42,0.72,"Update kiosk software. Add Spanish language option. Integrate with Epic scheduling for real-time wait time display.","Front desk staff (124 FTE). Patients — communication campaign.",2,"+2 points","$0.6M registration cost"),
        ("Epic Rover (Mobile Nursing)",52,82,0.52,0.78,"Expand device availability (currently 1 device per 3 nurses). Configure nursing-specific workflows per unit.","All nursing staff (8,400 FTE). Nurse managers (284 FTE).",2,"+2 points","$0.7M nursing efficiency"),
        ("Epic Radiant (Radiology)",82,92,0.82,0.90,"Activate AI-assisted report prioritization. Enable critical result auto-notification.","Radiologists (24 FTE). Radiology techs (68 FTE).",3,"+2 points","$0.3M radiology efficiency"),
        ("Epic Willow (Pharmacy)",76,90,0.76,0.88,"Activate clinical decision support alerts tier 2. Enable automated refill management.","Pharmacists (142 FTE). Pharmacy techs (84 FTE).",3,"+2 points","$0.8M pharmacy efficiency"),
        ("Epic Scheduling AI",0,50,0.00,0.40,"License activation. Configure no-show prediction model. Set up automated recall for no-show patients.","Scheduling staff (186 FTE). Department coordinators (48 FTE).",2,"+3 points","$1.1M no-show reduction"),
        ("Epic Care Everywhere (HIE)",61,82,0.61,0.78,"Enable automated external record retrieval. Configure reconciliation workflow for incoming records.","All clinicians who receive referrals (284 FTE).",3,"+2 points","$0.9M care coordination"),
    ]

    for r, row in enumerate(modules, 3):
        ws.row_dimensions[r].height = 32
        module, curr, tgt, curr_adopt, tgt_adopt, config, training, wave, score_impact, value = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        curr_f = fills["red"] if curr<30 else fills["amb"] if curr<60 else fills["grn"]
        adopt_f = fills["red"] if curr_adopt<0.30 else fills["amb"] if curr_adopt<0.60 else fills["grn"]
        wave_f = fills["grn"] if wave==1 else fills["amb"] if wave==2 else fills["wht"]
        C(ws,r,1,module,sf,bold=True,b=bdr)
        C(ws,r,2,curr,curr_f,"#,##0",align="center",bold=True,b=bdr)
        C(ws,r,3,tgt,fills["grn"],"#,##0",align="center",b=bdr)
        C(ws,r,4,curr_adopt,adopt_f,"0%",align="center",b=bdr)
        C(ws,r,5,tgt_adopt,fills["grn"],"0%",align="center",b=bdr)
        C(ws,r,6,config,sf,b=bdr); C(ws,r,7,training,sf,b=bdr)
        C(ws,r,8,wave,wave_f,"#,##0",align="center",bold=True,b=bdr)
        C(ws,r,9,score_impact,fills["teal"],align="center",b=bdr)
        C(ws,r,10,value,fills["teal"],align="center",b=bdr)

    # Score projection
    ws2 = wb.create_sheet("Score Projection")
    T(ws2, 5, "Meridian — Epic Optimization Score Projection by Wave", fills)
    ws2.row_dimensions[2].height = 28
    proj_hdrs = [("Milestone",22),("Epic Score",12),("MyChart %",12),
                 ("Modules Activated",14),("Annual Value Unlocked $M",16),("Timeline",14)]
    for i,(hd,w) in enumerate(proj_hdrs,1): H(ws2,2,i,hd,w,fills,bdr)
    projections = [
        ("Current state",58,0.34,4,0,"Today"),
        ("Wave 1 complete (6 modules activated)",72,0.55,10,114.0,"Month 6"),
        ("Wave 2 complete (4 more modules)",80,0.65,14,124.0,"Month 12"),
        ("Wave 3 complete (full optimization)",88,0.72,22,128.0,"Month 18"),
        ("Benchmark (peer health systems)",80,0.60,18,"N/A","N/A"),
    ]
    for r, row in enumerate(projections, 3):
        ws2.row_dimensions[r].height = 22
        milestone, score, mychart, modules, value, timeline = row
        sf = fills["prp"] if "Benchmark" in milestone else fills["alt"] if r%2==0 else fills["wht"]
        score_f = fills["red"] if isinstance(score,int) and score<65 else fills["amb"] if isinstance(score,int) and score<78 else fills["grn"]
        C(ws2,r,1,milestone,sf,bold=True,b=bdr)
        C(ws2,r,2,score,score_f,"#,##0",align="center",bold=True,b=bdr)
        C(ws2,r,3,mychart,sf,"0%",align="center",b=bdr)
        C(ws2,r,4,modules,sf,"#,##0" if isinstance(modules,int) else None,align="center",b=bdr)
        C(ws2,r,5,value,fills["grn"] if isinstance(value,(int,float)) else sf,"$#,##0.0" if isinstance(value,(int,float)) else None,align="center",b=bdr)
        C(ws2,r,6,timeline,sf,align="center",b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/tech/MER-T03_Epic_Optimization_Roadmap.xlsx")
    return "MER-T03 done"


# ══════════════════════════════════════════════════════════════════════════
# MER-T04: Integration Enhancement Plan (prior auth focus)
# ══════════════════════════════════════════════════════════════════════════

def mer_t04_integration_plan(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Integration Enhancement Plan"
    T(ws, 8, f"{MER['name']} — Integration Enhancement Plan (CMS Mandate Jan 2027)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Integration",24),("Current State",22),("Enhancement Required",24),
            ("Standard",12),("Effort (Weeks)",13),("Cost $M",10),
            ("CMS Mandate?",12),("Priority",10)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    integrations = [
        # Integration, Current, Enhancement, Standard, Weeks, Cost, CMS, Priority
        ("Epic → Cohere Health (Prior Auth AI)","Pilot — 60% complete. 3 payers only.","Complete FHIR R4 integration for all 8 priority payers. Configure automated auth request workflow. Enable real-time decision return.","FHIR R4",8,0.4,"Yes — CMS mandate Jan 2027","Critical"),
        ("Epic → Medicare FFS (CMS Direct)","Manual claim submission via intermediary.","Activate CMS electronic prior auth API (Da Vinci standard). Direct submission for Medicare prior auth.","FHIR R4 / Da Vinci",6,0.3,"Yes — primary mandate","Critical"),
        ("Epic → UnitedHealthcare MA","Manual portal submission. 18% denial rate.","Electronic prior auth via UHC API. Automated denial reason capture.","FHIR R4",4,0.2,"Yes — MA plans included","Critical"),
        ("Epic → Humana MA","Manual. 19.8% denial rate.","Electronic prior auth via Humana API. Real-time decision for 80% of request types.","FHIR R4",4,0.2,"Yes","Critical"),
        ("Epic → Aetna","31% electronic currently.","Expand electronic coverage to 95%. Enable automated appeal submission for denied auths.","FHIR R4",3,0.1,"Yes","High"),
        ("Epic → Cigna","28% electronic currently.","Activate Cigna real-time prior auth API for outpatient procedures.","FHIR R4",3,0.1,"Yes","High"),
        ("Epic → Medicaid NC (MMIS)","12% electronic. Most manual.","Connect to NC MMIS electronic prior auth gateway. High volume — 1,440 requests/month.","NC MMIS API",6,0.3,"Yes — state mandate","High"),
        ("Epic → Ensemble Health Partners","Claims outbound feed. No prior auth feedback loop.","Add prior auth status feed from Ensemble back to Epic. Real-time denial reason capture.","HL7 v2.3 enhanced",4,0.2,"Indirect","High"),
        ("Epic → 3M HIS Coding AI","Daily batch. Partial integration.","Real-time coding suggestion feed. Automated coding accuracy tracking dashboard.","FHIR R4",4,0.2,"No","Medium"),
        ("Epic → Nuance DAX / AWS Bedrock","Pilot only — 10 physicians.","Enterprise rollout configuration. Connect to all physician workflows. Approval and review workflow.","Epic Open API",6,0.3,"No","High"),
        ("Epic Cerner → Epic (2 hospitals)","ADT feed only. Records not shared.","Data migration programme. Patient record merge. Medication history reconciliation.","HL7 ADT + FHIR",52,2.8,"No — migration","Critical — Q4 2026"),
        ("Epic → CMS Quality Reporting","Monthly manual submission.","Automated HEDIS and quality measure reporting. MA Star Rating data submission automated.","CMS API / FHIR",6,0.3,"Yes — quality reporting","High"),
    ]

    for r, row in enumerate(integrations, 3):
        ws.row_dimensions[r].height = 28
        integ, curr, enhancement, standard, weeks, cost, cms, priority = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        pf = fills["red"] if priority=="Critical" or "Critical" in priority else fills["amb"] if priority=="High" else fills["grn"]
        cms_f = fills["red"] if cms.startswith("Yes") else fills["wht"]
        C(ws,r,1,integ,sf,bold=True,b=bdr); C(ws,r,2,curr,sf,b=bdr)
        C(ws,r,3,enhancement,sf,b=bdr); C(ws,r,4,standard,sf,align="center",b=bdr)
        wf = fills["red"] if weeks>12 else fills["amb"] if weeks>6 else fills["grn"]
        C(ws,r,5,weeks,wf,"#,##0",align="center",b=bdr)
        C(ws,r,6,cost,sf,"$#,##0.0",align="center",b=bdr)
        C(ws,r,7,cms,cms_f,b=bdr)
        cp = C(ws,r,8,priority,pf,align="center",bold=True,b=bdr)

    tr = len(integrations)+3
    C(ws,tr,1,"TOTAL INTEGRATION PROGRAMME",fills["prp"],bold=True,b=bdr)
    C(ws,tr,5,sum(r[4] for r in integrations),fills["prp"],"#,##0",align="center",bold=True,b=bdr)
    C(ws,tr,6,sum(r[5] for r in integrations),fills["prp"],"$#,##0.0",align="center",bold=True,b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/meridian/tech/MER-T04_Integration_Enhancement_Plan.xlsx")
    return "MER-T04 done"


# ══════════════════════════════════════════════════════════════════════════
# ERP-G01: ERP Genome — Product + SI Track Record + Readiness Template
# ══════════════════════════════════════════════════════════════════════════

def erp_genome(base):
    wb = Workbook(); bdr, fills = S()

    # Sheet 1: ERP Product Comparison
    ws = wb.active; ws.title = "ERP Product Comparison"
    T(ws, 10, "AbarVa Genome — ERP Product Comparison (Genome-Validated)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("ERP Product",20),("Vendor",14),("Best For",20),
            ("Avg Implementation Months",16),("Avg Overrun %",13),
            ("TCO 5yr vs SAP",14),("Cloud Native?",12),
            ("AI Capability",14),("Genome Score /100",14),("Notes",28)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    erp_products = [
        ("SAP S/4HANA","SAP","Manufacturing, utilities, complex multi-entity, global operations",18,0.42,1.0,"Hybrid — cloud edition available","Strong — SAP Joule AI, embedded analytics",68,"Market leader. High complexity. Genome shows 42% average overrun. Best for complex global ops."),
        ("SAP S/4HANA Cloud (Public)","SAP","Mid-market, standardised processes, cloud-first",12,0.18,0.85,"Yes — public cloud","Strong — full Joule AI capability",74,"Lower overrun than on-premise. Standardised processes required. No customisation."),
        ("Oracle Fusion Cloud ERP","Oracle","Finance-heavy, global, strong reporting needs",16,0.38,0.92,"Yes — cloud native","Strong — Oracle AI apps",71,"Strong financials. Genome shows better TCO than SAP for finance-led deployments."),
        ("Microsoft Dynamics 365","Microsoft","SME to mid-market, Microsoft ecosystem, lower complexity",10,0.24,0.72,"Yes — cloud native","Strong — Copilot embedded",79,"Best Genome score for mid-market. Copilot AI strongest in class. Lower TCO. Less capable for complex mfg."),
        ("Workday Finance + HCM","Workday","Services sector, HCM-first, professional services, healthcare",9,0.16,0.68,"Yes — cloud only","Good — Workday AI", 82,"Highest Genome success rate. Best for services/healthcare where HCM is core. Limited mfg capability."),
        ("NetSuite (Oracle)","Oracle","SME, fast-growing companies, ecommerce, simpler operations",7,0.14,0.58,"Yes — cloud only","Moderate",76,"Best for sub-$500M revenue. Fast implementation. Limited for complex multi-entity."),
        ("Infor CloudSuite","Infor","Manufacturing, healthcare supply chain, distribution",12,0.28,0.78,"Hybrid","Moderate — Coleman AI",64,"Industry-specific depth. Genome shows good healthcare supply chain results. Weaker financials."),
        ("Unit4","Unit4","Professional services, government, non-profit, people-first",9,0.18,0.65,"Yes — cloud native","Good — Wanda AI",73,"Underrated. Strong for services and public sector. Less known but high Genome success rate."),
    ]

    for r, row in enumerate(erp_products, 3):
        ws.row_dimensions[r].height = 28
        product, vendor, best_for, impl, overrun, tco, cloud, ai, score, notes = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        sf_s = fills["grn"] if score>=75 else fills["amb"] if score>=65 else fills["red"]
        over_f = fills["red"] if overrun>0.35 else fills["amb"] if overrun>0.20 else fills["grn"]
        C(ws,r,1,product,sf,bold=True,b=bdr); C(ws,r,2,vendor,sf,b=bdr)
        C(ws,r,3,best_for,sf,b=bdr)
        C(ws,r,4,impl,sf,"#,##0",align="center",b=bdr)
        C(ws,r,5,overrun,over_f,"0%",align="center",b=bdr)
        C(ws,r,6,tco,sf,"0.00x",align="center",b=bdr)
        cloud_f = fills["grn"] if "Yes" in cloud else fills["amb"]
        C(ws,r,7,cloud,cloud_f,align="center",b=bdr)
        C(ws,r,8,ai,sf,b=bdr)
        C(ws,r,9,score,sf_s,"#,##0",align="center",bold=True,b=bdr)
        C(ws,r,10,notes,sf,b=bdr)

    # Sheet 2: SI Track Record
    ws2 = wb.create_sheet("SI Track Record")
    T(ws2, 9, "AbarVa Genome — Systems Integrator Track Record (Verified Delivery Data)", fills)
    ws2.row_dimensions[2].height = 32
    si_hdrs = [("SI Firm",18),("ERP Strength",20),("Industry Strength",18),
               ("Avg Overrun %",13),("Avg Delay Months",14),("KT Score /100",12),
               ("Client Satisfaction",14),("Genome Score /100",14),("Watch Out For",28)]
    for i,(hd,w) in enumerate(si_hdrs,1): H(ws2,2,i,hd,w,fills,bdr)

    si_firms = [
        ("Accenture","SAP S/4HANA, Oracle Fusion","Manufacturing, retail, financial services",0.38,4.2,48,3.6,62,"Alliance fees inflate recommendations. SAP preferred regardless of fit. High day rates. Knowledge dependency created."),
        ("Deloitte","SAP S/4HANA, Oracle","Financial services, government, healthcare",0.32,3.8,52,3.8,65,"Strong methodology. Better KT than Accenture. Watch: subcontracting offshore team quality."),
        ("IBM Consulting","SAP, Oracle, Microsoft","Manufacturing, government, utilities",0.44,5.1,44,3.4,58,"Highest overrun in Genome data. Strong on government. Weak on speed. Best avoided for commercial ERP."),
        ("Capgemini","SAP S/4HANA","Manufacturing, automotive, retail",0.29,3.4,54,3.9,68,"Good SAP delivery. European strength. Watch: North America delivery quality lower than Europe."),
        ("PwC","Oracle Fusion, Workday","Financial services, professional services",0.26,2.8,58,4.1,72,"Better TCO management than Big 4 peers. Workday specialist strength. Good KT discipline."),
        ("KPMG","Microsoft D365, SAP","Mid-market, healthcare, not-for-profit",0.22,2.4,62,4.2,74,"Best Big 4 Genome score. Strong D365. Good healthcare ERP track record. Better value than Accenture/Deloitte."),
        ("Cognizant","SAP, Oracle","Healthcare, manufacturing, financial services",0.28,3.2,56,3.8,68,"Strong offshore model. Good healthcare. Watch: onshore-offshore coordination quality."),
        ("Infosys","SAP, Oracle","Manufacturing, retail, utilities",0.31,3.6,52,3.7,65,"Large delivery capacity. Watch: senior talent on project vs junior offshore. KT discipline variable."),
        ("Wipro","SAP, Microsoft D365","Mid-market, manufacturing",0.34,3.8,48,3.5,61,"Lower rates but Genome shows higher overrun. KT scores lowest in class. Not recommended for knowledge-intensive implementations."),
        ("Syntax (boutique)","SAP S/4HANA","Manufacturing, mid-market",0.18,1.8,74,4.6,84,"Highest Genome score for SAP mid-market. Boutique = senior people stay on project. Strong KT. Recommend for $1-5B manufacturers."),
        ("Rimini Street","SAP, Oracle support","All — third-party support","N/A","N/A",62,4.4,78,"Not an SI — third-party support. 50-90% savings vs SAP/Oracle support. Recommended for systems not being replaced."),
        ("EY","Oracle, Workday","Financial services, professional services",0.24,2.6,62,4.2,73,"Strong Oracle and Workday. Good financial services track record. Better governance discipline than some Big 4 peers."),
    ]

    for r, row in enumerate(si_firms, 3):
        ws2.row_dimensions[r].height = 28
        firm, erp, industry, overrun, delay, kt, sat, score, watch = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        sf_s = fills["grn"] if isinstance(score,int) and score>=75 else fills["amb"] if isinstance(score,int) and score>=65 else fills["red"]
        over_f = fills["red"] if isinstance(overrun,(int,float)) and overrun>0.35 else fills["amb"] if isinstance(overrun,(int,float)) and overrun>0.25 else fills["grn"] if isinstance(overrun,(int,float)) else sf
        kt_f = fills["red"] if isinstance(kt,int) and kt<50 else fills["amb"] if isinstance(kt,int) and kt<65 else fills["grn"] if isinstance(kt,int) else sf
        C(ws2,r,1,firm,sf,bold=True,b=bdr); C(ws2,r,2,erp,sf,b=bdr)
        C(ws2,r,3,industry,sf,b=bdr)
        C(ws2,r,4,overrun,over_f,"0%" if isinstance(overrun,(int,float)) else None,align="center",b=bdr)
        C(ws2,r,5,delay,sf,"#,##0.0" if isinstance(delay,(int,float)) else None,align="center",b=bdr)
        C(ws2,r,6,kt,kt_f,"#,##0" if isinstance(kt,int) else None,align="center",b=bdr)
        C(ws2,r,7,sat,sf,"0.0" if isinstance(sat,(int,float)) else None,align="center",b=bdr)
        C(ws2,r,8,score,sf_s,"#,##0" if isinstance(score,int) else None,align="center",bold=True,b=bdr)
        C(ws2,r,9,watch,fills["amb"],b=bdr)

    # Sheet 3: ERP Readiness Assessment Template
    ws3 = wb.create_sheet("Readiness Assessment Template")
    T(ws3, 7, "AbarVa — ERP Readiness Assessment Template (Client-Facing)", fills)
    ws3.row_dimensions[2].height = 28
    ready_hdrs = [("Dimension",22),("Assessment Questions",32),("Score Weight",11),
                  ("Red Flags",28),("Genome Threshold",14),("Data Source",18)]
    for i,(hd,w) in enumerate(ready_hdrs,1): H(ws3,2,i,hd,w,fills,bdr)

    readiness = [
        ("Data Readiness","How clean is your master data? What % of records have been cleansed? Do you have a data governance process?",0.20,"Data quality <60%. No MDM programme. Multiple conflicting masters.","Score >65 to proceed safely","Upload: data quality assessment or answer 3 numbers"),
        ("Process Standardisation","What % of your core processes are documented? How many process variants exist across business units?",0.18,"<40% documented. >10 variants of core process. No process owners named.","Score >60 to proceed safely","Maestro workshop or process inventory upload"),
        ("Change Management Capacity","Do you have a named change management lead? What is your change fatigue level? Recent major programmes?",0.16,"No dedicated CM lead. >2 major programmes in last 3 years. High attrition in affected teams.","Score >55 required","Leadership interviews — Maestro conducts"),
        ("Executive Sponsorship","Is there a C-suite owner whose performance review is tied to this programme's success?",0.15,"No named sponsor. Sponsor too junior. Sponsor history of disengagement mid-programme. F002 signal.",">1 confirmed sponsor with board accountability","Leadership register — Maestro confirms"),
        ("IT Capability to Govern","Can your IT team manage the implementation vendor without becoming dependent?","0.14","IT team has never run a programme this size. No enterprise architect. IT team is the implementation vendor.","Score >50 — or Maestro engaged to fill gap","IT org chart + capability assessment"),
        ("Budget Realism","Is the approved budget based on independent assessment or vendor estimate?",0.12,"Budget = vendor's estimate only. No contingency. No independent validation. Genome says +30-45% typical.","Contingency >25% of base budget","Budget approval documentation"),
        ("Timeline Realism","Is the go-live date based on scope assessment or a board mandate?",0.05,"Go-live date is fixed regardless of scope. No phasing plan. No pilot/cutover rehearsal plan.","Genome average for this programme type","Project charter — Maestro reviews"),
    ]

    for r, row in enumerate(readiness, 3):
        ws3.row_dimensions[r].height = 32
        dim, questions, weight, red_flags, threshold, source = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        C(ws3,r,1,dim,sf,bold=True,b=bdr); C(ws3,r,2,questions,sf,b=bdr)
        C(ws3,r,3,weight,sf,align="center",b=bdr)
        C(ws3,r,4,red_flags,fills["red"],b=bdr)
        C(ws3,r,5,threshold,fills["amb"],b=bdr)
        C(ws3,r,6,source,fills["teal"],b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/erp_genome/ERP-G01_ERP_Product_SI_Genome.xlsx")
    return "ERP-G01 done"


# ══════════════════════════════════════════════════════════════════════════
# CLOUD ARCHITECTURE ADVISORY BLUEPRINT TEMPLATE
# ══════════════════════════════════════════════════════════════════════════

def cloud_architecture_blueprint(base):
    wb = Workbook(); bdr, fills = S()
    ws = wb.active; ws.title = "Cloud Architecture Patterns"
    T(ws, 8, "AbarVa — Cloud Architecture Advisory Patterns (Governance Blueprint Templates)", fills)
    ws.row_dimensions[2].height = 32
    hdrs = [("Use Case",22),("Architecture Pattern",24),("AWS Services",24),
            ("Azure Services",24),("Est Build Cost $M",14),("Est Run Cost pa $M",14),
            ("SI / Vendor to Build",18),("Governance Recommendation",24)]
    for i,(hd,w) in enumerate(hdrs,1): H(ws,2,i,hd,w,fills,bdr)

    patterns = [
        ("Document Intelligence Platform\n(Replace manual document review)",
         "S3 → Textract/Lambda → Bedrock Claude → Vector DB (OpenSearch) → API → App",
         "S3, Bedrock, Claude 3 Sonnet, OpenSearch Serverless, Lambda, API Gateway, Cognito",
         "Blob Storage, Azure OpenAI, Cognitive Search, Azure Functions, APIM",
         0.8, 0.4,
         "AWS ProServe or specialist AI boutique (e.g. Slalom, Thoughtworks)",
         "Maestro governs: data classification, PII handling, model accuracy SLA, vendor delivery milestones"),
        ("Data Platform & Golden Record\n(Eliminate silos, enable AI)",
         "Source systems → Glue ETL → S3 Data Lake → Redshift/Databricks → dbt → BI + AI",
         "Glue, S3, Redshift, Lake Formation, DMS, Bedrock for data quality AI",
         "Azure Data Factory, ADLS Gen2, Synapse, Purview",
         2.4, 0.8,
         "Databricks PS or Snowflake PS — proven data platform specialists",
         "Maestro governs: golden record definition, data ownership matrix, SLA per domain, KT to internal team"),
        ("MLOps Platform\n(Get AI initiatives to production)",
         "S3 → SageMaker Studio → MLflow → SageMaker Endpoints → CloudWatch → Retrain pipeline",
         "SageMaker, S3, ECR, CodePipeline, CloudWatch, Bedrock for model evaluation",
         "Azure ML, Container Registry, Azure DevOps, Azure Monitor",
         1.2, 0.6,
         "AWS ProServe or Slalom — MLOps specialists",
         "Maestro governs: model registry standards, deployment gates, monitoring thresholds, KT milestones"),
        ("IT Service Desk AI\n(Reduce L1/L2 cost 60%)",
         "ServiceNow/JIRA → Bedrock Claude → Knowledge Base (S3 + OpenSearch) → Chat UI → Human escalation",
         "Bedrock, OpenSearch, Lambda, API Gateway, S3, Connect (if phone)",
         "Azure OpenAI, Cognitive Search, Azure Bot Service, Teams integration",
         0.4, 0.2,
         "Internal build feasible — or ServiceNow AI add-on",
         "Maestro governs: resolution rate SLA, escalation design, knowledge base maintenance, KT to internal team"),
        ("Prior Auth AI Platform\n(Healthcare — CMS mandate)",
         "Epic FHIR R4 → API Gateway → Lambda → Bedrock + clinical rules → Payer API → Epic response",
         "Bedrock, Lambda, API Gateway, S3, DynamoDB, HealthLake",
         "Azure Health Data Services, Azure OpenAI, Azure Functions, APIM",
         0.6, 0.3,
         "Cohere Health (preferred — pre-built) or build internally with Maestro",
         "Maestro governs: payer API SLAs, clinical accuracy validation, HIPAA compliance, Epic integration"),
        ("Trading Analytics AI Platform\n(Asset management — replace manual reporting)",
         "Bloomberg/Aladdin → Kinesis → Lambda → Bedrock → Redshift → QuickSight/Tableau",
         "Kinesis, Lambda, Bedrock, Redshift, S3, Glue, QuickSight",
         "Event Hubs, Azure Functions, Azure OpenAI, Synapse, Power BI",
         1.8, 0.6,
         "AWS FinServ competency partner or Thoughtworks",
         "Maestro governs: data latency SLA, model accuracy, MAS FEAT compliance, vendor delivery"),
        ("ERP AI Layer\n(Add AI on top of existing ERP without replacement)",
         "ERP API → Bedrock → Business logic Lambda → Results → ERP write-back",
         "Bedrock Claude, Lambda, API Gateway, S3, DynamoDB",
         "Azure OpenAI, Azure Functions, Logic Apps, Cosmos DB",
         0.4, 0.2,
         "Internal build — low complexity if ERP has API access",
         "Maestro governs: ERP API access agreement, data write-back approval gates, UAT sign-off"),
    ]

    for r, row in enumerate(patterns, 3):
        ws.row_dimensions[r].height = 40
        use_case, pattern, aws, azure, build, run, vendor, governance = row
        sf = fills["alt"] if r%2==0 else fills["wht"]
        C(ws,r,1,use_case,sf,bold=True,b=bdr); C(ws,r,2,pattern,sf,b=bdr)
        C(ws,r,3,aws,fills["teal"],b=bdr); C(ws,r,4,azure,fills["blu"],b=bdr)
        C(ws,r,5,build,sf,"$#,##0.0",align="center",b=bdr)
        C(ws,r,6,run,sf,"$#,##0.0",align="center",b=bdr)
        C(ws,r,7,vendor,sf,b=bdr)
        C(ws,r,8,governance,fills["grn"],b=bdr)

    ws.freeze_panes = "B3"
    save(wb, f"{base}/tech_advisory/CLOUD-A01_Architecture_Advisory_Patterns.xlsx")
    return "CLOUD-A01 done"


# ══════════════════════════════════════════════════════════════════════════
# RUNNER
# ══════════════════════════════════════════════════════════════════════════

def run_task(args):
    func, base = args
    try:
        result = func(base)
        return [result] if isinstance(result, str) else result
    except Exception as e:
        return [f"ERROR in {func.__name__}: {e}\n{traceback.format_exc()}"]

if __name__ == "__main__":
    tasks = [
        (arc_t02_technical_debt, BASE),
        (arc_t03_vendor_contracts, BASE),
        (mer_t03_epic_roadmap, BASE),
        (mer_t04_integration_plan, BASE),
        (erp_genome, BASE),
        (cloud_architecture_blueprint, BASE),
    ]

    print(f"\n{'='*60}")
    print(f"AbarVa — Tech Modernization Extended Datasets")
    print(f"{len(tasks)} files — Arcturus + Meridian + ERP Genome + Cloud Advisory")
    print(f"{'='*60}\n")

    start = time.time()
    with mp.Pool(processes=min(len(tasks), mp.cpu_count())) as pool:
        results = pool.map(run_task, tasks)
    elapsed = time.time() - start

    all_results = [r for group in results for r in group]
    errors = [r for r in all_results if "ERROR" in r]
    successes = [r for r in all_results if "ERROR" not in r]

    print("Results:")
    for r in all_results:
        print(f"  {'OK' if 'ERROR' not in r else 'FAIL'} {r}")

    print(f"\n{'='*60}")
    print(f"Complete: {len(successes)} files, {len(errors)} errors, {elapsed:.1f}s")

    new_files = list(BASE.rglob("*.xlsx"))
    new_files = [f for f in new_files if "scripts" not in str(f) and
                 any(x in str(f) for x in ["T02","T03","T04","ERP","CLOUD"])]
    print(f"\nNew files ({len(new_files)}):")
    for f in sorted(new_files):
        print(f"  {str(f.relative_to(BASE))}")

    if errors:
        for e in errors: print(e)
        sys.exit(1)
    else:
        print(f"\nRun: git add . && git commit -m 'datasets: tech modernization extended + ERP genome + cloud advisory'")
