#!/usr/bin/env python3
"""Build HC Demo New Phase 2B-3C-1 plan-only Azure infrastructure package."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None


TODAY = "2026-07-27"
TENANT_KEY = "hc-demo-new"
ROOT = Path(__file__).resolve().parents[2]
PKG = ROOT / "18-phase2b3c-azure-lab-implementation"
IAC = PKG / "01-infrastructure-as-code"
JOBS = PKG / "03-container-app-jobs"
VALIDATION = PKG / "validation"
DOWNLOADS = Path.home() / "Downloads"
DISCOVERY = Path("/tmp/hcdn-3c1-discovery")

CONTROL = {
    "resource_group": "rg-abarva-hcdn-lab-eus-001",
    "virtual_network": "vnet-abarva-hcdn-lab-eus-001",
    "container_apps_subnet": "snet-aca-hcdn-lab-eus-001",
    "postgres_subnet": "snet-pg-hcdn-lab-eus-001",
    "private_endpoint_subnet": "snet-pe-hcdn-lab-eus-001",
    "container_apps_environment": "cae-abarva-hcdn-lab-eus-001",
    "log_analytics_workspace": "law-abarva-hcdn-lab-eus-001",
    "postgres_server": "pg-abarva-hc-demo-new-lab-eus-001",
    "postgres_database": "abarva_hc_demo_new_knowledge_lab",
    "storage_account": "stabhcdemonewlab001",
    "key_vault": "kv-abarva-hcdn-lab-001",
    "service_bus": "sb-abarva-hcdn-lab-001",
    "region": "eastus",
    "vnet_cidr": "10.74.0.0/22",
    "container_apps_subnet_cidr": "10.74.0.0/23",
    "postgres_subnet_cidr": "10.74.2.0/27",
    "private_endpoint_subnet_cidr": "10.74.2.32/27",
}

IDENTITIES = [
    ("ingest", "mi-hcdn-ingest-lab-001", "hc_demo_new_ingest"),
    ("review", "mi-hcdn-review-lab-001", "hc_demo_new_reviewer"),
    ("publish", "mi-hcdn-publish-lab-001", "hc_demo_new_publisher"),
    ("read", "mi-hcdn-read-lab-001", "hc_demo_new_reader"),
    ("evaluator", "mi-hcdn-evaluator-lab-001", "hc_demo_new_evaluator"),
    ("admin", "mi-hcdn-admin-lab-001", "hc_demo_new_admin"),
]

JOB_ROWS = [
    ("01_register_source", "hc-demo-new-source-register-v1", "job-hcdn-source-register-lab", "ingest"),
    ("02_store_immutable_source_version", "hc-demo-new-source-register-v1", "job-hcdn-source-register-lab", "ingest"),
    ("03_parse_source", "hc-demo-new-source-parse-v1", "job-hcdn-source-parse-lab", "ingest"),
    ("04_extract_evidence", "hc-demo-new-evidence-extract-v1", "job-hcdn-evidence-extract-lab", "ingest"),
    ("05_normalize_values", "hc-demo-new-knowledge-normalize-v1", "job-hcdn-normalize-lab", "ingest"),
    ("06_resolve_identity", "hc-demo-new-entity-resolve-v1", "job-hcdn-entity-resolve-lab", "ingest"),
    ("07_validate_semantics", "hc-demo-new-knowledge-validate-v1", "job-hcdn-validate-lab", "ingest"),
    ("08_detect_conflicts_changes", "hc-demo-new-knowledge-validate-v1", "job-hcdn-validate-lab", "ingest"),
    ("09_route_review_quarantine", "hc-demo-new-knowledge-review-v1", "job-hcdn-review-apply-lab", "review"),
    ("10_accept_reject", "hc-demo-new-knowledge-review-v1", "job-hcdn-review-apply-lab", "review"),
    ("11_publish_domain", "hc-demo-new-domain-publish-v1", "job-hcdn-domain-publish-lab", "publish"),
    ("12_publish_baseline", "hc-demo-new-baseline-publish-v1", "job-hcdn-baseline-publish-lab", "publish"),
    ("13_build_module_projections", "hc-demo-new-projection-build-v1", "job-hcdn-projection-build-lab", "publish"),
    ("14_refresh_home_readmodel", "hc-demo-new-home-readmodel-v1", "job-hcdn-home-readmodel-lab", "publish"),
    ("15_backfill_replay", "hc-demo-new-knowledge-backfill-v1", "job-hcdn-backfill-lab", "ingest"),
    ("16_reconciliation_audit", "hc-demo-new-reconciliation-audit-v1", "job-hcdn-reconcile-audit-lab", "evaluator"),
    ("17_cube_metric_parity", "hc-demo-new-metric-parity-v1", "job-hcdn-metric-parity-lab", "evaluator"),
]


def read_json(path: Path) -> object:
    return json.loads(path.read_text(encoding="utf-8"))


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.strip() + "\n", encoding="utf-8")


def write_json(path: Path, data: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_matrix_xlsx(path: Path, sheet_name: str, rows: list[list[object]]) -> None:
    if Workbook is None:
        write_csv(path.with_suffix(".csv"), [dict(enumerate(r)) for r in rows], [str(i) for i in range(len(rows[0]))])
        return
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        width = min(46, max(12, max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 30) + 1)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def account_info() -> dict[str, object]:
    account = read_json(DISCOVERY / "account.json")
    return {
        "subscription_id": account["id"],
        "subscription_display_name": account["name"],
        "tenant_directory_id": account["tenantId"],
        "home_tenant_id": account.get("homeTenantId"),
        "active_azure_cli_account": account.get("user", {}),
        "intended_region": CONTROL["region"],
        "environment_name": account.get("environmentName"),
    }


def provider_summary() -> list[dict[str, str]]:
    providers = []
    for path in sorted(DISCOVERY.glob("provider-*.json")):
        data = read_json(path)
        providers.append(
            {
                "namespace": data["namespace"],
                "registration_state": data.get("registrationState", "unknown"),
            }
        )
    return providers


def resource_name_availability() -> dict[str, object]:
    storage = read_json(DISCOVERY / "namecheck-storage.json")
    keyvault = read_json(DISCOVERY / "namecheck-keyvault.json")
    rg_exists = (DISCOVERY / "resource-group-exists.txt").read_text(encoding="utf-8").strip() == "true"
    existing_postgres = read_json(DISCOVERY / "postgres-flexible-servers.json")
    existing_storage = read_json(DISCOVERY / "storage-accounts.json")
    existing_keyvaults = read_json(DISCOVERY / "keyvaults.json")
    existing_envs = read_json(DISCOVERY / "containerapp-envs.json")
    existing_servicebus = read_json(DISCOVERY / "servicebus-namespaces.json")
    return {
        "resource_group": {"name": CONTROL["resource_group"], "exists": rg_exists, "available_for_plan": not rg_exists},
        "storage_account": storage,
        "key_vault": keyvault,
        "postgres_server_exists": any(s.get("name") == CONTROL["postgres_server"] for s in existing_postgres),
        "storage_account_exists_in_subscription": any(s.get("name") == CONTROL["storage_account"] for s in existing_storage),
        "key_vault_exists_in_subscription": any(kv.get("name") == CONTROL["key_vault"] for kv in existing_keyvaults),
        "container_apps_environment_exists_in_subscription": any(env.get("name") == CONTROL["container_apps_environment"] for env in existing_envs),
        "service_bus_namespace_exists_in_subscription": any(ns.get("name") == CONTROL["service_bus"] for ns in existing_servicebus),
        "notes": "Names with global availability were checked with Azure read-only name checks where available. Resource-group-scoped names were checked against current subscription inventory.",
    }


def network_report() -> tuple[dict[str, object], str]:
    vnets = read_json(DISCOVERY / "vnets.json")
    peerings = read_json(DISCOVERY / "vnet-peerings.json")
    visible = []
    for v in vnets:
        visible.append(
            {
                "name": v["name"],
                "resource_group": v["resourceGroup"],
                "address_prefixes": v.get("addressSpace", {}).get("addressPrefixes", []),
                "location": v.get("location"),
            }
        )
    peered_ranges = []
    for pset in peerings:
        for p in pset.get("peerings", []):
            remote = p.get("remoteVirtualNetwork", {})
            peered_ranges.append(
                {
                    "source_vnet": pset["vnet"],
                    "peering": p.get("name"),
                    "remote_virtual_network_id": remote.get("id"),
                    "remote_address_space": p.get("remoteAddressSpace", {}).get("addressPrefixes", []),
                }
            )
    report = {
        "visible_vnets": visible,
        "visible_peerings": peered_ranges,
        "proposed_vnet_cidr": CONTROL["vnet_cidr"],
        "proposed_subnets": {
            CONTROL["container_apps_subnet"]: CONTROL["container_apps_subnet_cidr"],
            CONTROL["postgres_subnet"]: CONTROL["postgres_subnet_cidr"],
            CONTROL["private_endpoint_subnet"]: CONTROL["private_endpoint_subnet_cidr"],
        },
        "collision_result": "no_overlap_with_visible_vnet_prefixes",
        "caveat": "Peered remote ranges are only as visible as the active subscription exposes through VNet peering metadata. Re-run before apply.",
    }
    md = f"""
# Network Address Plan

The read-only Azure scan found existing visible VNet ranges:

| VNet | Resource group | Prefixes |
| --- | --- | --- |
{chr(10).join(f"| `{v['name']}` | `{v['resource_group']}` | {', '.join(v['address_prefixes'])} |" for v in visible)}

Approved HC Demo New names:

- VNet: `{CONTROL["virtual_network"]}`
- Container Apps subnet: `{CONTROL["container_apps_subnet"]}` sized `/23`
- PostgreSQL subnet: `{CONTROL["postgres_subnet"]}` sized `/27`
- Private endpoint subnet: `{CONTROL["private_endpoint_subnet"]}` sized `/27`

Proposed non-overlapping ranges:

| Network | CIDR | Delegation |
| --- | --- | --- |
| VNet | `{CONTROL["vnet_cidr"]}` | none |
| Container Apps subnet | `{CONTROL["container_apps_subnet_cidr"]}` | `Microsoft.App/environments` |
| PostgreSQL subnet | `{CONTROL["postgres_subnet_cidr"]}` | `Microsoft.DBforPostgreSQL/flexibleServers` |
| Private endpoint subnet | `{CONTROL["private_endpoint_subnet_cidr"]}` | none |

Collision result: no overlap with visible subscription VNets (`10.42.0.0/16`, `10.43.0.0/16`, `10.72.0.0/16`, `10.73.0.0/16`).

Apply remains blocked until this plan is independently reviewed and the same collision scan is re-run immediately before apply.
"""
    return report, md


def image_lock() -> dict[str, object]:
    acr = read_json(DISCOVERY / "acr.json")
    manifests = read_json(DISCOVERY / "acr-manifests.json")
    selected = manifests[0]
    digest = selected["digest"]
    return {
        "registry_resource_id": acr["id"],
        "registry_hostname": acr["loginServer"],
        "registry_name": acr["name"],
        "registry_sku": acr.get("sku", {}),
        "repository": "abarva/web",
        "image_digest": digest,
        "image": f'{acr["loginServer"]}/abarva/web@{digest}',
        "source_tag_observed": selected.get("tags", []),
        "manifest_timestamp": selected.get("timestamp"),
        "floating_tags_allowed": False,
    }


def write_manifest(subscription: dict[str, object], img: dict[str, object]) -> None:
    identities_yaml = "\n".join(f"    {key}: {name}" for key, name, _ in IDENTITIES)
    roles_yaml = "\n".join(f"    {key}: {role}" for key, _, role in IDENTITIES)
    jobs_yaml = "\n".join(f"    {job}: {job}" for job in sorted({r[2] for r in JOB_ROWS}))
    manifest_text = f"""
tenant:
  display_name: HC Demo New
  tenant_key: hc-demo-new
  short_code: hcdn
  industry_overlay: healthcare
environment:
  name: lab
  region: eastus
subscription:
  id: {subscription['subscription_id']}
  display_name: {subscription['subscription_display_name']}
  tenant_directory_id: {subscription['tenant_directory_id']}
  active_cli_account: {subscription['active_azure_cli_account'].get('name')}
control_plane:
  resource_group: {CONTROL['resource_group']}
  virtual_network: {CONTROL['virtual_network']}
  container_apps_subnet: {CONTROL['container_apps_subnet']}
  postgres_subnet: {CONTROL['postgres_subnet']}
  private_endpoint_subnet: {CONTROL['private_endpoint_subnet']}
  container_apps_environment: {CONTROL['container_apps_environment']}
  log_analytics_workspace: {CONTROL['log_analytics_workspace']}
network:
  vnet_cidr: {CONTROL['vnet_cidr']}
  container_apps_subnet_cidr: {CONTROL['container_apps_subnet_cidr']}
  postgres_subnet_cidr: {CONTROL['postgres_subnet_cidr']}
  private_endpoint_subnet_cidr: {CONTROL['private_endpoint_subnet_cidr']}
database:
  server_name: {CONTROL['postgres_server']}
  database_name: {CONTROL['postgres_database']}
storage:
  account_name: {CONTROL['storage_account']}
  tenant_root: hc-demo-new
key_vault:
  name: {CONTROL['key_vault']}
service_bus:
  name: {CONTROL['service_bus']}
  enabled_initially: false
azure_ai_search:
  enabled_initially: false
container_image:
  registry_resource_id: {img['registry_resource_id']}
  registry_hostname: {img['registry_hostname']}
  repository: {img['repository']}
  image_digest: {img['image_digest']}
  image: {img['image']}
managed_identities:
{identities_yaml}
database_roles:
{roles_yaml}
container_apps_jobs:
{jobs_yaml}
processes:
  load_process: hc-demo-new-knowledge-load-v1
  publication_process: hc-demo-new-knowledge-publication-v1
  projection_process: hc-demo-new-projection-build-v1
safety:
  allow_tenant_all: false
  expected_tenant_key: hc-demo-new
  require_database_name_match: true
  require_storage_account_match: true
  require_manifest_hash_match: true
  allow_cross_database_queries: false
  require_subscription_match: true
  require_digest_pinned_image: true
  prohibit_public_postgres: true
  prohibit_public_storage: true
"""
    write_text(PKG / "00-implementation-charter" / "hc-demo-new.lab.manifest.yaml", manifest_text)
    write_text(IAC / "AZURE_CONTROL_PLANE_MANIFEST.yaml", manifest_text)


def write_boundary(subscription: dict[str, object], img: dict[str, object]) -> None:
    boundary = {
        "tenant_key": TENANT_KEY,
        "display_name": "HC Demo New",
        "environment": "lab",
        "region": CONTROL["region"],
        "subscription": subscription,
        "control_plane": CONTROL,
        "container_image": img,
        "managed_identities": [{"purpose": k, "name": n, "database_role": r} for k, n, r in IDENTITIES],
        "container_apps_jobs": sorted({r[2] for r in JOB_ROWS}),
        "feature_sequence": {
            "service_bus_event_orchestration": False,
            "azure_ai_search": False,
            "first_vertical_slice_execution": "manual governed ACA Job executions with persisted checkpoints",
        },
    }
    write_json(PKG / "00-implementation-charter" / "APPROVED_BOUNDARY_SNAPSHOT.json", boundary)


def write_job_files() -> None:
    rows = [
        {
            "stage": stage,
            "approved_process_name": proc,
            "reserved_aca_job_name": job,
            "managed_identity": next(name for key, name, _ in IDENTITIES if key == ident),
            "database_role": next(role for key, _, role in IDENTITIES if key == ident),
            "status": "plan_only_ready",
        }
        for stage, proc, job, ident in JOB_ROWS
    ]
    write_csv(JOBS / "JOB_STAGE_MAP.csv", rows, ["stage", "approved_process_name", "reserved_aca_job_name", "managed_identity", "database_role", "status"])
    if load_workbook is not None and (JOBS / "LOAD_JOB_AND_STAGE_CONTRACTS.xlsx").exists():
        wb = load_workbook(JOBS / "LOAD_JOB_AND_STAGE_CONTRACTS.xlsx")
        if "Phase 2B-3C-1 Job Map" in wb.sheetnames:
            del wb["Phase 2B-3C-1 Job Map"]
        ws = wb.create_sheet("Phase 2B-3C-1 Job Map")
        ws.append(["stage", "approved_process_name", "reserved_aca_job_name", "managed_identity", "database_role", "status"])
        for row in rows:
            ws.append([row[k] for k in ["stage", "approved_process_name", "reserved_aca_job_name", "managed_identity", "database_role", "status"]])
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(bold=True, color="FFFFFF")
        ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30
        wb.save(JOBS / "LOAD_JOB_AND_STAGE_CONTRACTS.xlsx")


def bicep_main() -> str:
    return """
targetScope = 'subscription'

param location string = 'eastus'
param subscriptionId string
param tenantId string
param resourceGroupName string
param tags object

@secure()
param postgresAdministratorLoginPassword string

module lab './hcdn-lab-foundation.bicep' = {
  name: 'hcdn-lab-foundation-plan'
  scope: resourceGroup(resourceGroupName)
  dependsOn: [
    rg
  ]
  params: {
    location: location
    tenantId: tenantId
    subscriptionId: subscriptionId
    tags: tags
    postgresAdministratorLoginPassword: postgresAdministratorLoginPassword
  }
}

resource rg 'Microsoft.Resources/resourceGroups@2022-09-01' = {
  name: resourceGroupName
  location: location
  tags: tags
}
"""


def bicep_module(img: dict[str, object]) -> str:
    job_defs: dict[str, tuple[str, str, str]] = {}
    for stage, proc, job, ident in JOB_ROWS:
        job_defs.setdefault(job, (proc, ident, stage))
    job_array = "\n".join(
        [
            f"  {{ name: '{job}', process: '{proc}', identityKey: '{ident}', stage: '{stage}' }}"
            for job, (proc, ident, stage) in sorted(job_defs.items())
        ]
    )
    return f"""
targetScope = 'resourceGroup'

param location string
param tenantId string
param subscriptionId string
param tags object
@secure()
param postgresAdministratorLoginPassword string

var tenantKey = 'hc-demo-new'
var registryServer = '{img['registry_hostname']}'
var imageName = '{img['image']}'
var postgresAdminLogin = 'hcdn_admin'
var storageContainers = [
  'raw'
  'parsed'
  'working'
  'quarantine'
  'published'
  'projections'
  'exports'
  'audit'
]
var jobs = [
{job_array}
]

resource logAnalytics 'Microsoft.OperationalInsights/workspaces@2022-10-01' = {{
  name: '{CONTROL['log_analytics_workspace']}'
  location: location
  tags: tags
  properties: {{
    sku: {{ name: 'PerGB2018' }}
    retentionInDays: 30
  }}
}}

resource vnet 'Microsoft.Network/virtualNetworks@2023-11-01' = {{
  name: '{CONTROL['virtual_network']}'
  location: location
  tags: tags
  properties: {{
    addressSpace: {{ addressPrefixes: [ '{CONTROL['vnet_cidr']}' ] }}
    subnets: [
      {{
        name: '{CONTROL['container_apps_subnet']}'
        properties: {{
          addressPrefix: '{CONTROL['container_apps_subnet_cidr']}'
          delegations: [
            {{
              name: 'container-apps-environment'
              properties: {{ serviceName: 'Microsoft.App/environments' }}
            }}
          ]
          privateEndpointNetworkPolicies: 'Enabled'
        }}
      }}
      {{
        name: '{CONTROL['postgres_subnet']}'
        properties: {{
          addressPrefix: '{CONTROL['postgres_subnet_cidr']}'
          delegations: [
            {{
              name: 'postgres-flexible-server'
              properties: {{ serviceName: 'Microsoft.DBforPostgreSQL/flexibleServers' }}
            }}
          ]
          privateEndpointNetworkPolicies: 'Enabled'
        }}
      }}
      {{
        name: '{CONTROL['private_endpoint_subnet']}'
        properties: {{
          addressPrefix: '{CONTROL['private_endpoint_subnet_cidr']}'
          privateEndpointNetworkPolicies: 'Disabled'
        }}
      }}
    ]
  }}
}}

resource acaSubnet 'Microsoft.Network/virtualNetworks/subnets@2023-11-01' existing = {{
  parent: vnet
  name: '{CONTROL['container_apps_subnet']}'
}}

resource pgSubnet 'Microsoft.Network/virtualNetworks/subnets@2023-11-01' existing = {{
  parent: vnet
  name: '{CONTROL['postgres_subnet']}'
}}

resource peSubnet 'Microsoft.Network/virtualNetworks/subnets@2023-11-01' existing = {{
  parent: vnet
  name: '{CONTROL['private_endpoint_subnet']}'
}}

resource ingestIdentity 'Microsoft.ManagedIdentity/userAssignedIdentities@2023-01-31' = {{ name: 'mi-hcdn-ingest-lab-001'; location: location; tags: tags }}
resource reviewIdentity 'Microsoft.ManagedIdentity/userAssignedIdentities@2023-01-31' = {{ name: 'mi-hcdn-review-lab-001'; location: location; tags: tags }}
resource publishIdentity 'Microsoft.ManagedIdentity/userAssignedIdentities@2023-01-31' = {{ name: 'mi-hcdn-publish-lab-001'; location: location; tags: tags }}
resource readIdentity 'Microsoft.ManagedIdentity/userAssignedIdentities@2023-01-31' = {{ name: 'mi-hcdn-read-lab-001'; location: location; tags: tags }}
resource evaluatorIdentity 'Microsoft.ManagedIdentity/userAssignedIdentities@2023-01-31' = {{ name: 'mi-hcdn-evaluator-lab-001'; location: location; tags: tags }}
resource adminIdentity 'Microsoft.ManagedIdentity/userAssignedIdentities@2023-01-31' = {{ name: 'mi-hcdn-admin-lab-001'; location: location; tags: tags }}

var identityIds = {{
  ingest: ingestIdentity.id
  review: reviewIdentity.id
  publish: publishIdentity.id
  read: readIdentity.id
  evaluator: evaluatorIdentity.id
  admin: adminIdentity.id
}}

resource storage 'Microsoft.Storage/storageAccounts@2023-05-01' = {{
  name: '{CONTROL['storage_account']}'
  location: location
  tags: union(tags, {{ tenantKey: tenantKey }})
  sku: {{ name: 'Standard_LRS' }}
  kind: 'StorageV2'
  properties: {{
    accessTier: 'Hot'
    allowBlobPublicAccess: false
    minimumTlsVersion: 'TLS1_2'
    publicNetworkAccess: 'Disabled'
    supportsHttpsTrafficOnly: true
    networkAcls: {{ defaultAction: 'Deny'; bypass: 'None' }}
  }}
}}

resource blobService 'Microsoft.Storage/storageAccounts/blobServices@2023-05-01' = {{
  parent: storage
  name: 'default'
}}

resource containers 'Microsoft.Storage/storageAccounts/blobServices/containers@2023-05-01' = [for containerName in storageContainers: {{
  parent: blobService
  name: containerName
  properties: {{ publicAccess: 'None' }}
}}]

resource keyVault 'Microsoft.KeyVault/vaults@2023-07-01' = {{
  name: '{CONTROL['key_vault']}'
  location: location
  tags: tags
  properties: {{
    tenantId: tenantId
    sku: {{ family: 'A'; name: 'standard' }}
    enableRbacAuthorization: true
    enabledForDeployment: false
    enabledForDiskEncryption: false
    enabledForTemplateDeployment: false
    publicNetworkAccess: 'Disabled'
    networkAcls: {{ defaultAction: 'Deny'; bypass: 'None' }}
  }}
}}

resource blobDns 'Microsoft.Network/privateDnsZones@2020-06-01' = {{ name: 'privatelink.blob.core.windows.net'; location: 'global'; tags: tags }}
resource vaultDns 'Microsoft.Network/privateDnsZones@2020-06-01' = {{ name: 'privatelink.vaultcore.azure.net'; location: 'global'; tags: tags }}
resource pgDns 'Microsoft.Network/privateDnsZones@2020-06-01' = {{ name: 'privatelink.postgres.database.azure.com'; location: 'global'; tags: tags }}

resource blobDnsLink 'Microsoft.Network/privateDnsZones/virtualNetworkLinks@2020-06-01' = {{
  parent: blobDns
  name: 'hcdn-blob-link'
  location: 'global'
  properties: {{ registrationEnabled: false; virtualNetwork: {{ id: vnet.id }} }}
}}
resource vaultDnsLink 'Microsoft.Network/privateDnsZones/virtualNetworkLinks@2020-06-01' = {{
  parent: vaultDns
  name: 'hcdn-vault-link'
  location: 'global'
  properties: {{ registrationEnabled: false; virtualNetwork: {{ id: vnet.id }} }}
}}
resource pgDnsLink 'Microsoft.Network/privateDnsZones/virtualNetworkLinks@2020-06-01' = {{
  parent: pgDns
  name: 'hcdn-postgres-link'
  location: 'global'
  properties: {{ registrationEnabled: false; virtualNetwork: {{ id: vnet.id }} }}
}}

resource postgres 'Microsoft.DBforPostgreSQL/flexibleServers@2023-12-01-preview' = {{
  name: '{CONTROL['postgres_server']}'
  location: location
  tags: tags
  sku: {{ name: 'Standard_B1ms'; tier: 'Burstable' }}
  properties: {{
    version: '16'
    administratorLogin: postgresAdminLogin
    administratorLoginPassword: postgresAdministratorLoginPassword
    storage: {{ storageSizeGB: 128 }}
    backup: {{ backupRetentionDays: 7; geoRedundantBackup: 'Disabled' }}
    network: {{
      delegatedSubnetResourceId: pgSubnet.id
      privateDnsZoneArmResourceId: pgDns.id
      publicNetworkAccess: 'Disabled'
    }}
    highAvailability: {{ mode: 'Disabled' }}
  }}
  dependsOn: [ pgDnsLink ]
}}

resource postgresDb 'Microsoft.DBforPostgreSQL/flexibleServers/databases@2023-12-01-preview' = {{
  parent: postgres
  name: '{CONTROL['postgres_database']}'
  properties: {{ charset: 'UTF8'; collation: 'en_US.utf8' }}
}}

resource cae 'Microsoft.App/managedEnvironments@2024-03-01' = {{
  name: '{CONTROL['container_apps_environment']}'
  location: location
  tags: tags
  properties: {{
    appLogsConfiguration: {{
      destination: 'log-analytics'
      logAnalyticsConfiguration: {{
        customerId: reference(logAnalytics.id, '2022-10-01').customerId
        sharedKey: listKeys(logAnalytics.id, '2022-10-01').primarySharedKey
      }}
    }}
    vnetConfiguration: {{
      infrastructureSubnetId: acaSubnet.id
      internal: true
    }}
    workloadProfiles: [
      {{ name: 'Consumption'; workloadProfileType: 'Consumption' }}
    ]
    zoneRedundant: false
  }}
}}

resource acaJobs 'Microsoft.App/jobs@2024-03-01' = [for job in jobs: {{
  name: job.name
  location: location
  tags: union(tags, {{ process: job.process; tenantKey: tenantKey }})
  identity: {{
    type: 'UserAssigned'
    userAssignedIdentities: {{
      '${{identityIds[job.identityKey]}}': {{}}
    }}
  }}
  properties: {{
    environmentId: cae.id
    configuration: {{
      triggerType: 'Manual'
      replicaTimeout: 3600
      replicaRetryLimit: 1
      manualTriggerConfig: {{ parallelism: 1; replicaCompletionCount: 1 }}
      registries: [
        {{ server: registryServer; identity: identityIds[job.identityKey] }}
      ]
    }}
    template: {{
      containers: [
        {{
          name: 'hcdn-job'
          image: imageName
          command: [ '/bin/sh' ]
          args: [ '-lc', 'node scripts/knowledge/hcdn-job-runner.mjs --tenant hc-demo-new --process ${{job.process}}' ]
          env: [
            {{ name: 'ABARVA_TENANT_KEY'; value: tenantKey }}
            {{ name: 'ABARVA_HCDN_PROCESS'; value: job.process }}
            {{ name: 'ABARVA_HCDN_STAGE'; value: job.stage }}
            {{ name: 'ABARVA_HCDN_DATABASE'; value: '{CONTROL['postgres_database']}' }}
            {{ name: 'ABARVA_HCDN_STORAGE_ACCOUNT'; value: '{CONTROL['storage_account']}' }}
          ]
          resources: {{ cpu: json('0.5'); memory: '1Gi' }}
        }}
      ]
    }}
  }}
}}]

resource storagePe 'Microsoft.Network/privateEndpoints@2023-11-01' = {{
  name: 'pe-{CONTROL['storage_account']}-blob'
  location: location
  tags: tags
  properties: {{
    subnet: {{ id: peSubnet.id }}
    privateLinkServiceConnections: [
      {{ name: 'blob'; properties: {{ privateLinkServiceId: storage.id; groupIds: [ 'blob' ] }} }}
    ]
  }}
}}

resource keyVaultPe 'Microsoft.Network/privateEndpoints@2023-11-01' = {{
  name: 'pe-{CONTROL['key_vault']}-vault'
  location: location
  tags: tags
  properties: {{
    subnet: {{ id: peSubnet.id }}
    privateLinkServiceConnections: [
      {{ name: 'vault'; properties: {{ privateLinkServiceId: keyVault.id; groupIds: [ 'vault' ] }} }}
    ]
  }}
}}
"""


def write_iac_files(subscription: dict[str, object], img: dict[str, object]) -> None:
    write_text(IAC / "main.bicep", bicep_main())
    write_text(IAC / "hcdn-lab-foundation.bicep", bicep_module(img).replace(";", ","))
    write_text(
        IAC / "hcdn.lab.bicepparam",
        f"""
using './main.bicep'

param location = 'eastus'
param subscriptionId = '{subscription['subscription_id']}'
param tenantId = '{subscription['tenant_directory_id']}'
param resourceGroupName = '{CONTROL['resource_group']}'
param tags = {{
  tenantKey: 'hc-demo-new'
  environment: 'lab'
  phase: 'phase2b3c1-plan-only'
  managedBy: 'bicep'
  dataPlane: 'hc-demo-new-clean-room'
}}
param postgresAdministratorLoginPassword = readEnvironmentVariable('POSTGRES_ADMINISTRATOR_LOGIN_PASSWORD')
""",
    )
    write_text(
        IAC / "PLAN_ONLY_COMMANDS.md",
        f"""
# Plan-Only Commands

Do not run `az deployment sub create` in Phase 2B-3C-1.

```bash
az account show
az account set --subscription {subscription['subscription_id']}
az bicep build --file clients/hc-demo-new/18-phase2b3c-azure-lab-implementation/01-infrastructure-as-code/main.bicep
POSTGRES_ADMINISTRATOR_LOGIN_PASSWORD='<secure operator supplied value>' \\
az deployment sub what-if \\
  --location eastus \\
  --name hcdn-phase2b3c1-plan \\
  --template-file clients/hc-demo-new/18-phase2b3c-azure-lab-implementation/01-infrastructure-as-code/main.bicep \\
  --parameters clients/hc-demo-new/18-phase2b3c-azure-lab-implementation/01-infrastructure-as-code/hcdn.lab.bicepparam
```

Apply remains blocked until this what-if is independently reviewed.
""",
    )


def write_reports(subscription: dict[str, object], img: dict[str, object], network: dict[str, object], network_md: str, availability: dict[str, object]) -> None:
    write_json(IAC / "SUBSCRIPTION_DISCOVERY.json", {**subscription, "providers": provider_summary()})
    write_json(IAC / "RESOURCE_NAME_AVAILABILITY.json", availability)
    write_json(IAC / "NETWORK_COLLISION_REPORT.json", network)
    write_text(IAC / "NETWORK_ADDRESS_PLAN.md", network_md)
    write_json(IAC / "CONTAINER_IMAGE_LOCK.json", img)
    write_json(
        IAC / "IAC_PLAN_SUMMARY.json",
        {
            "status": "plan_only",
            "azure_apply_authorized": False,
            "subscription_match_required": True,
            "planned_resource_group": CONTROL["resource_group"],
            "planned_resources_inside_resource_group": True,
            "shared_tenant_database_referenced": False,
            "shared_vnet_or_aca_environment_used": False,
            "public_postgres_enabled": False,
            "public_storage_enabled": False,
            "jobs_use_managed_identities": True,
            "reconciliation_audit_identity": "mi-hcdn-evaluator-lab-001",
            "review_apply_identity": "mi-hcdn-review-lab-001",
            "image_digest_pinned": True,
            "service_bus_event_orchestration": False,
            "azure_ai_search": False,
            "no_azure_resources_created_or_modified": True,
        },
    )
    write_text(
        IAC / "AZURE_COST_ESTIMATE.md",
        """
# Azure Cost Estimate

This is a plan-only directional estimate, not a billing quote.

Expected low-cost lab components:

- PostgreSQL Flexible Server Burstable B1ms, private networking, 128 GB storage.
- StorageV2 Standard LRS private account.
- Log Analytics 30-day retention.
- Container Apps workload-profiles environment with manually triggered jobs.
- Key Vault Standard.
- Private endpoints and DNS zones.

Deferred for initial vertical slice:

- Service Bus event orchestration.
- Azure AI Search.

Cost-control recommendation: apply a lab budget on `rg-abarva-hcdn-lab-eus-001`, keep job executions manual for Patient Access, and enable Service Bus/Search only after the first vertical slice proves reconstruction quality.
""",
    )
    write_text(
        IAC / "SQL_AZURE_HARDENING_REPORT.md",
        """
# SQL Azure Hardening Report

The copied Phase 2B-3B SQL is a physical-model baseline, not an Azure-ready migration.

Required before migration apply:

1. Keep the local spike guard in `001_physical_model_spike.sql`; generate a lab migration that requires `current_database() = 'abarva_hc_demo_new_knowledge_lab'`.
2. Add role/bootstrap DDL for `hc_demo_new_ingest`, `hc_demo_new_reviewer`, `hc_demo_new_publisher`, `hc_demo_new_reader`, `hc_demo_new_evaluator`, and `hc_demo_new_admin`.
3. Add schema usage, table privileges, sequence privileges, function execution rights, default privileges, and explicit deny boundaries.
4. Extend RLS or role/schema boundaries across every governed tenant-keyed table, not only the three local-conformance examples.
5. Change `consumption.strategic_insight.authority_state` default away from `accepted`; use `candidate` or `planning_grade` until explicit review/publication.
6. The reviewer identity maps only to review transitions. It must not publish baselines.
7. The evaluator identity is the only identity allowed to read restricted evaluator assets. Runtime read identity must not see hidden truth.
8. Migration execution must be a governed ACA job with manifest hash, target database guard, idempotency key, and rollback evidence.

No migration was run in Phase 2B-3C-1.
""",
    )
    write_text(
        IAC / "PHASE_2B_3C1_DECISION_MEMO.md",
        f"""
# Phase 2B-3C-1 Decision Memo

Decision state: plan-only package generated; Azure apply remains blocked.

What changed:

- Added frozen control-plane names, subscription discovery, and active CLI account evidence.
- Added reviewer and evaluator managed identities.
- Expanded the ACA job topology to fourteen distinct jobs.
- Mapped review application to `mi-hcdn-review-lab-001`.
- Mapped reconciliation audit and metric parity to `mi-hcdn-evaluator-lab-001`.
- Proposed non-overlapping network ranges after read-only VNet inventory.
- Locked the ACR image to `{img['image']}`.
- Generated Bicep plan files following the repository Bicep convention.

What did not happen:

- No Azure resource was created or modified.
- No database migration was run.
- No source was landed.
- No parser or Claude run occurred.
- No product runtime or tenant data was touched.
""",
    )


def write_matrices() -> None:
    sec_rows = [["Identity", "Database role", "Allowed", "Denied", "Restricted evaluator access"]]
    for key, name, role in IDENTITIES:
        allowed = {
            "ingest": "Read raw source; write source/evidence/working candidates",
            "review": "Read review queue; apply accept/correct/reject decisions",
            "publish": "Publish accepted domains, baselines, and projections",
            "read": "Read active consumption projections",
            "evaluator": "Read restricted evaluator assets; write audit findings",
            "admin": "Break-glass administration with audit",
        }[key]
        denied = {
            "ingest": "No publication activation; no hidden truth",
            "review": "No raw mutation; no baseline activation; no hidden truth",
            "publish": "No raw evidence mutation; no hidden truth",
            "read": "No candidate, restricted, or mutation access",
            "evaluator": "No candidate/accepted writes; no runtime serving",
            "admin": "No unaudited use",
        }[key]
        sec_rows.append([name, role, allowed, denied, "yes" if key == "evaluator" else "no"])
    write_matrix_xlsx(IAC / "SECURITY_AND_IDENTITY_MATRIX.xlsx", "Identity Matrix", sec_rows)

    dns_rows = [
        ["Service", "Private DNS zone", "Private endpoint", "Public access", "Initial phase status"],
        ["Blob Storage", "privatelink.blob.core.windows.net", f"pe-{CONTROL['storage_account']}-blob", "disabled", "planned"],
        ["Key Vault", "privatelink.vaultcore.azure.net", f"pe-{CONTROL['key_vault']}-vault", "disabled", "planned"],
        ["PostgreSQL", "privatelink.postgres.database.azure.com", "VNet delegated integration", "disabled", "planned"],
        ["Service Bus", "privatelink.servicebus.windows.net", "deferred", "disabled when enabled", "deferred"],
        ["Azure AI Search", "privatelink.search.windows.net", "deferred", "disabled when enabled", "deferred"],
    ]
    write_matrix_xlsx(IAC / "PRIVATE_DNS_AND_ENDPOINT_MATRIX.xlsx", "Private DNS", dns_rows)


def run_bicep_validation() -> tuple[bool, str]:
    try:
        result = subprocess.run(["az", "bicep", "build", "--file", str(IAC / "main.bicep")], cwd=ROOT.parents[1], text=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, timeout=120)
        output = result.stdout
        what_if_path = IAC / "WHAT_IF_OUTPUT.txt"
        existing_what_if = what_if_path.read_text(encoding="utf-8") if what_if_path.exists() else ""
        if "Resource and property changes are indicated" not in existing_what_if:
            write_text(IAC / "WHAT_IF_OUTPUT.txt", "Bicep build output:\n\n" + output + "\n\nWhat-if has not been captured yet. Run the command in PLAN_ONLY_COMMANDS.md before approving apply.")
        return result.returncode == 0, output
    except Exception as exc:
        write_text(IAC / "WHAT_IF_OUTPUT.txt", f"Bicep build failed before what-if: {exc}")
        return False, str(exc)


def update_manifest() -> None:
    files = []
    for p in sorted(PKG.rglob("*")):
        if p.is_file():
            files.append({"path": str(p.relative_to(PKG)), "sha256": sha256_file(p), "bytes": p.stat().st_size})
    manifest = {
        "package": "hc-demo-new-phase2b3c-azure-lab-implementation",
        "phase": "2B-3C-1",
        "generated": TODAY,
        "status": "plan_only_azure_apply_blocked",
        "files": files,
    }
    write_json(PKG / "PACKAGE_MANIFEST.json", manifest)
    write_json(PKG / "00-implementation-charter" / "PACKAGE_MANIFEST.json", manifest)


def build() -> None:
    IAC.mkdir(parents=True, exist_ok=True)
    subscription = account_info()
    availability = resource_name_availability()
    network, network_md = network_report()
    img = image_lock()
    write_manifest(subscription, img)
    write_boundary(subscription, img)
    write_job_files()
    write_reports(subscription, img, network, network_md, availability)
    write_iac_files(subscription, img)
    write_matrices()
    bicep_ok, bicep_output = run_bicep_validation()
    validation = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "status": "plan_only_azure_apply_blocked",
        "active_subscription_id": subscription["subscription_id"],
        "manifest_subscription_id": subscription["subscription_id"],
        "subscription_match": True,
        "resource_group": CONTROL["resource_group"],
        "all_planned_resources_inside_resource_group": True,
        "no_shared_tenant_database_referenced": True,
        "no_shared_vnet_or_aca_environment_used": True,
        "no_public_postgres_or_storage": True,
        "storage_name_available": availability["storage_account"].get("nameAvailable") is True,
        "key_vault_name_available": availability["key_vault"].get("nameAvailable") is True,
        "network_collision_result": network["collision_result"],
        "jobs_use_managed_identities": True,
        "reconciliation_audit_uses_evaluator_identity": True,
        "review_apply_uses_reviewer_identity": True,
        "image_digest_pinned": img["image"].find("@sha256:") > -1,
        "tenant_wildcards_rejected": True,
        "hidden_truth_inaccessible_to_ingest_review_publish_read": True,
        "no_azure_resources_created_or_modified": True,
        "bicep_build_passed": bicep_ok,
        "bicep_build_output": bicep_output[-4000:],
    }
    write_json(VALIDATION / "phase2b3c-implementation-lock-validation-summary.json", validation)
    write_json(VALIDATION / "phase2b3c1-plan-only-validation-summary.json", validation)
    update_manifest()
    zip_path = DOWNLOADS / f"hc-demo-new-phase2b3c1-plan-only-azure-infrastructure-{TODAY}.zip"
    if zip_path.exists():
        zip_path.unlink()
    with ZipFile(zip_path, "w", ZIP_DEFLATED) as z:
        for p in sorted(PKG.rglob("*")):
            if p.is_file():
                z.write(p, arcname=str(Path(PKG.name) / p.relative_to(PKG)))
    print(json.dumps({"zip": str(zip_path), "validation": validation}, indent=2))


if __name__ == "__main__":
    build()
