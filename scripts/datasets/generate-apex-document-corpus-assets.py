#!/usr/bin/env python3
"""Generate Packet 18 document and corpus fixture assets for Apex Retail.

The generated files are synthetic and deterministic. They are intentionally
plain, auditable fixtures rather than customer artifacts.
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROOT = Path(__file__).resolve().parents[2]
PACK = ROOT / "datasets" / "apex-retail-synthetic-v1"


@dataclass(frozen=True)
class ContractDoc:
    file_name: str
    vendor: str
    kind: str
    annual_usd: int
    renewal_window: str
    risk: str
    leverage: str


@dataclass(frozen=True)
class CharterDoc:
    file_name: str
    initiative_id: str
    title: str
    sponsor: str
    committed_usd: int
    decision: str
    evidence: str


CONTRACT_DOCS = [
    ContractDoc("kyndryl-mainframe-as400-ams-2023-2028.pdf", "Kyndryl", "AMS", 14_000_000, "Q1 2028", "Mainframe and AS/400 continuity depends on retained tribal knowledge.", "Bundle AS/400 exit milestones into renewal credits."),
    ContractDoc("wipro-distributed-ams-2024-2028.pdf", "Wipro", "AMS", 32_000_000, "Q1 2027", "Distributed AMS scope includes SAP ECC, Sterling, WMS, and store BOH.", "Renegotiate run-rate against retirement roadmap and incident SLA burn."),
    ContractDoc("tcs-sap-erp-future-decision-scoping-sow.pdf", "TCS", "SI SOW", 14_000_000, "Q2 2027 decision", "ERP decision can outrun Oracle DB migration dependency.", "Gate RFP release on dependency closure and board-ready options case."),
    ContractDoc("accenture-commerce-cloud-extensions-msa.pdf", "Accenture", "SI MSA", 4_000_000, "Q4 2026", "Einstein activation value is blocked by catalog and behavioral data quality.", "Tie extension fees to activation milestones and conversion lift."),
    ContractDoc("deloitte-pmo-transformation-msa.pdf", "Deloitte", "PMO MSA", 3_000_000, "Q3 2026", "Multiple PMO workstreams lack verified value owner sign-off.", "Reduce broad PMO retainer; convert to stage-gated specialist pods."),
    ContractDoc("sap-ecc-extended-maintenance-2027-2030.pdf", "SAP", "Software support", 22_000_000, "Q2 2027", "Extended maintenance premium rises if ECC roadmap slips.", "Use S/4HANA vs RISE vs alternatives decision to cap support exposure."),
    ContractDoc("salesforce-commerce-cloud-msa-with-einstein-rider.pdf", "Salesforce", "SaaS", 14_800_000, "Q4 2026", "Einstein seats purchased but not activated across commerce journeys.", "Separate inactive AI riders from core commerce renewal."),
    ContractDoc("ibm-sterling-oms-license-2024.pdf", "IBM", "Software license", 2_800_000, "Q1 2027", "Sterling extensions are behind release and carry custom integration debt.", "Ask for upgrade credits and API migration support."),
    ContractDoc("manhattan-wms-license.pdf", "Manhattan Associates", "Software license", 1_800_000, "Q3 2026", "WMS is stable but tightly coupled to legacy ASN feeds.", "Preserve support; negotiate integration transition assistance."),
    ContractDoc("ncr-counterpoint-pos-master.pdf", "NCR", "POS master", 6_000_000, "Q4 2026", "Store-edge rollout windows limit change velocity.", "Bundle support, PCI certification evidence, and edge replacement options."),
    ContractDoc("adobe-experience-cloud-ea.pdf", "Adobe", "Enterprise agreement", 4_200_000, "Q1 2027", "Marketing activation depends on CDP migration data quality.", "Align Adobe renewal with CDP Phase 2 scope reset."),
    ContractDoc("datadog-observability-msa.pdf", "Datadog", "Observability", 1_600_000, "Q2 2026", "Logging overlap with Splunk increases run-cost pressure.", "Consolidate app telemetry scope and retention windows."),
    ContractDoc("splunk-mssp-soc-managed.pdf", "Splunk", "SIEM and MSSP", 2_400_000, "Q2 2026", "SIEM retention and MSSP scope overlap with CrowdStrike MDR.", "Negotiate ingestion caps and incident response ownership boundaries."),
    ContractDoc("crowdstrike-mdr-msa.pdf", "CrowdStrike", "Endpoint MDR", 1_200_000, "Q3 2026", "Endpoint scope is strong; overlap is mainly SOC operating model.", "Preserve MDR but clarify handoff with Splunk MSSP."),
    ContractDoc("okta-identity-msa.pdf", "Okta", "Identity", 2_300_000, "Q1 2027", "Identity modernization Phase 2 needs privileged access inventory.", "Tie renewal to SSO/PAM coverage expansion."),
    ContractDoc("workday-hcm-msa-with-phase-2-amendment.pdf", "Workday", "HCM SaaS", 5_600_000, "Q4 2026", "Talent and learning scope depends on Cornerstone retirement decision.", "Bundle Phase 2 credits and legacy LMS exit support."),
    ContractDoc("servicenow-itsm-msa.pdf", "ServiceNow", "ITSM SaaS", 3_800_000, "Q2 2027", "CMDB quality remains below automation threshold.", "Attach credits to CMDB completeness and incident routing automation."),
    ContractDoc("aws-enterprise-discount-program.pdf", "AWS", "Cloud EDP", 10_500_000, "Q3 2026", "Cloud cost allocation is incomplete for AI and data workloads.", "Use FinOps baseline to renegotiate commit shape."),
    ContractDoc("microsoft-365-ea.pdf", "Microsoft", "Enterprise agreement", 6_800_000, "Q4 2026", "Copilot expansion has uneven adoption and limited baseline measurement.", "Expand only after modern teams show verified productivity lift."),
    ContractDoc("cisco-sdwan-network-msa.pdf", "Cisco", "Network MSA", 3_200_000, "Q1 2027", "480-store SD-WAN refresh must sequence around retail blackout periods.", "Negotiate phased rollout and failure-credit language."),
    ContractDoc("equinix-colocation-dr-msa.pdf", "Equinix", "Colocation", 1_800_000, "Q2 2027", "DR tests are tied to mainframe and SAP batch windows.", "Require joint DR exercise before extension."),
    ContractDoc("verizon-wan-4g-failover-msa.pdf", "Verizon", "Connectivity", 2_600_000, "Q3 2026", "4G failover quality varies by rural store cluster.", "Segment renewal pricing by store-performance evidence."),
    ContractDoc("segment-cdp-license-sunsetting.pdf", "Segment", "CDP license", 1_400_000, "Q4 2026", "Phase 2 migration has scope creep toward orchestration.", "Preserve exit rights and data export assistance."),
    ContractDoc("punchh-loyalty-platform-msa.pdf", "Punchh", "Loyalty SaaS", 900_000, "Q2 2026", "Vendor relationship is degraded and replacement RFP stalled.", "Hold renewal except for narrow bridge; enforce data-export clause."),
    ContractDoc("o9-demand-planning-sow.pdf", "o9", "Planning platform", 5_200_000, "Q4 2026", "Implementation is 40 percent complete with weak vendor relationship.", "Reframe completion vs replacement before committing next phase."),
    ContractDoc("blue-yonder-pricing-license.pdf", "Blue Yonder", "Pricing SaaS", 2_700_000, "Q1 2027", "Prior pricing phase required scope reset.", "Require corrected value measurement and pricing-engine dependency map."),
    ContractDoc("coupa-procurement-platform-msa.pdf", "Coupa", "Procurement SaaS", 1_200_000, "Q2 2027", "Sourcing data completeness is adequate; supplier onboarding remains manual.", "Bundle supplier workflow automation into renewal."),
    ContractDoc("sap-ariba-supplier-network.pdf", "SAP Ariba", "Supplier network", 1_600_000, "Q2 2027", "Supplier-network overlap with Coupa creates duplicate workflow cost.", "Rationalize source-to-pay roles before renewal."),
    ContractDoc("pythian-oracle-dba-managed-service.pdf", "Pythian", "Managed DBA", 720_000, "Q1 2027", "Oracle DB migration is a hard dependency for ERP decision timing.", "Tie managed-service extension to migration readiness milestones."),
    ContractDoc("commvault-backup-recovery-msa.pdf", "Commvault", "Backup and recovery", 480_000, "Q3 2026", "Recovery evidence must cover store-edge and legacy batch systems.", "Require audit-ready recovery test evidence."),
]


CHARTER_DOCS = [
    CharterDoc("sap-erp-future-decision-charter.pdf", "INIT-SAP-ERP-FUTURE", "SAP ERP Future Decision", "CIO", 14_000_000, "HOLD until Oracle DB migration dependency is resolved.", "ECC 6.0 is mature with debt, 8,400 customizations, and extended maintenance to 2030."),
    CharterDoc("commerce-cloud-einstein-activation-charter.pdf", "INIT-COMMERCE-EINSTEIN", "Commerce Cloud Optimization + Einstein Activation", "Chief Digital Officer", 8_400_000, "Proceed with value-gated activation.", "Einstein rider purchased but not activated; ecommerce is 18.5 percent of revenue."),
    CharterDoc("cdp-migration-phase-2-charter.pdf", "INIT-CDP-PHASE2", "CDP Migration Phase 2", "CMO", 4_600_000, "RESTRUCTURE scope before next funding gate.", "Scope crept from data migration into orchestration."),
    CharterDoc("o9-completion-charter.pdf", "INIT-O9-COMPLETION", "o9 Implementation Completion", "COO", 5_200_000, "RESTRUCTURE completion vs replacement decision.", "Implementation is 40 percent complete; vendor relationship weak."),
    CharterDoc("demand-forecasting-ai-v2-charter.pdf", "INIT-FORECAST-AI-V2", "Demand Forecasting AI v2", "COO", 3_800_000, "CONTINUE with measurement discipline.", "False-positive control: high spend but sponsor engaged and value tracking is positive."),
    CharterDoc("punchh-loyalty-replacement-charter.pdf", "INIT-PUNCHH-LOYALTY", "Punchh Loyalty Replacement", "CMO", 2_100_000, "KILL unless sponsor re-engages and RFP restarts.", "Replacement RFP stalled nine months; sponsor inactive for more than 60 days."),
    CharterDoc("mainframe-modernization-assessment-charter.pdf", "INIT-MAINFRAME-MOD-ASSESS", "Mainframe Modernization Assessment", "CIO", 1_400_000, "KILL current assessment and defer until ERP decision.", "Fourteen months of scoping with no executive sponsor."),
    CharterDoc("ai-productivity-program-charter.pdf", "INIT-AI-PRODUCTIVITY", "AI Productivity Program", "CIO", 2_400_000, "Proceed with team-specific value proof.", "Copilot, Cursor, and Claude utilization differs sharply by stack era."),
    CharterDoc("store-associate-mobile-refresh-charter.pdf", "INIT-STORE-ASSOC-MOBILE", "Store Associate Mobile Refresh", "COO", 6_300_000, "Proceed after store-edge dependency review.", "Store ops app changes are gated by retail blackout calendar."),
    CharterDoc("finops-cloud-cost-optimization-charter.pdf", "INIT-FINOPS", "FinOps / Cloud Cost Optimization", "CFO", 1_600_000, "Proceed with immediate AWS EDP leverage.", "AWS spend and AI/data workload allocation create near-term savings opportunity."),
]


SOURCE_FILE_THEMES = [
    ("SRC-PORTFOLIO-001", "Portfolio", "application-portfolio.csv", "Application portfolio export", "portfolio_reasoning"),
    ("SRC-TOPOLOGY-001", "Architecture", "integration-topology.json", "Integration topology export", "dependency_reasoning"),
    ("SRC-TIME-001", "Portfolio", "time-classification-output.csv", "TIME classification", "watchlist_reasoning"),
    ("SRC-WARDLEY-001", "Strategy", "wardley-map-2026q1.json", "Wardley map", "strategy_reasoning"),
    ("SRC-COMMIT-001", "Finance", "initiative-commitments.csv", "Initiative commitments", "show_me_money"),
    ("SRC-RUNCOST-001", "Finance", "annual-run-costs.csv", "Annual run costs", "show_me_money"),
    ("SRC-CAPEX-001", "Finance", "capex-opex-split-fy26.csv", "Capex/opex split", "cfo_reasoning"),
    ("SRC-RENEWAL-001", "Source", "renewal-calendar-12mo.csv", "Renewal calendar", "source_renewal_risk"),
    ("SRC-VENDOR-001", "Source", "vendor-contracts.csv", "Vendor contracts", "source_vendor_concentration"),
    ("SRC-INFRA-001", "Source", "infra-ms-contracts.csv", "Infrastructure managed services", "ams_reasoning"),
    ("SRC-SCORECARD-001", "Source", "vendor-scorecards-2026q1.csv", "Vendor scorecards", "source_performance"),
    ("SRC-ORG-001", "Org", "org-topology.csv", "Org topology", "change_effort"),
    ("SRC-ROLES-001", "Org", "roles-inventory.csv", "Role inventory", "change_effort"),
    ("SRC-LEADERSHIP-001", "Org", "leadership-bench.json", "Leadership bench", "sponsor_reasoning"),
    ("SRC-DORA-001", "DORA", "dora-baseline-consolidated.csv", "DORA baseline", "ai_productivity_targeting"),
    ("SRC-DEVEX-001", "DevEx", "devex-survey-2026q1.json", "DevEx survey", "developer_experience"),
    ("SRC-SPACE-001", "DevEx", "space-survey-2026q1.json", "SPACE survey", "developer_experience"),
    ("SRC-AITOOLS-001", "AI Tools", "ai-tool-footprint.csv", "AI tool footprint", "ai_productivity_value"),
    ("SRC-COPILOT-001", "AI Tools", "copilot-utilization-export.csv", "Copilot utilization", "ai_productivity_value"),
    ("SRC-CURSOR-001", "AI Tools", "cursor-utilization-export.csv", "Cursor utilization", "ai_productivity_value"),
    ("SRC-CLAUDE-001", "AI Tools", "claude-enterprise-utilization.csv", "Claude utilization", "ai_productivity_value"),
    ("SRC-SPONSOR-001", "Sponsor", "sponsor-pulse-2026q1.json", "Sponsor pulse", "watchlist_kill_fitness"),
    ("SRC-CADENCE-001", "Sponsor", "steering-meeting-cadence.csv", "Steering cadence", "sponsor_reasoning"),
    ("SRC-STATUS-001", "Sponsor", "sponsor-status-update-log.csv", "Sponsor status updates", "sponsor_reasoning"),
    ("SRC-INCIDENT-001", "Ops", "major-incidents-2025.csv", "Major incidents", "operability_risk"),
    ("SRC-PROBLEM-001", "Ops", "problem-records.csv", "Problem records", "operability_risk"),
    ("SRC-CHANGE-001", "Ops", "change-management-feed-6mo.csv", "Change feed", "delivery_risk"),
    ("SRC-REG-001", "Risk", "regulatory-scope.json", "Regulatory scope", "control_reasoning"),
    ("SRC-BENCH-001", "Benchmarks", "industry-benchmarks-specialty-retail.json", "Industry benchmarks", "benchmark_grounding"),
    ("SRC-PEER-001", "Benchmarks", "peer-comparison-anonymized.json", "Peer comparison", "benchmark_grounding"),
    ("SRC-KYNDRYL-001", "Contract", "kyndryl-mainframe-as400-ams-2023-2028.pdf", "Kyndryl AMS contract", "source_contract_extract"),
    ("SRC-WIPRO-001", "Contract", "wipro-distributed-ams-2024-2028.pdf", "Wipro AMS contract", "source_contract_extract"),
    ("SRC-TCS-001", "Contract", "tcs-sap-erp-future-decision-scoping-sow.pdf", "TCS ERP SOW", "source_contract_extract"),
    ("SRC-SAP-001", "Contract", "sap-ecc-extended-maintenance-2027-2030.pdf", "SAP ECC maintenance", "erp_decision"),
    ("SRC-SF-001", "Contract", "salesforce-commerce-cloud-msa-with-einstein-rider.pdf", "Salesforce Commerce Cloud MSA", "commerce_ai_activation"),
    ("SRC-PUNCHH-001", "Contract", "punchh-loyalty-platform-msa.pdf", "Punchh MSA", "watchlist_kill_fitness"),
    ("SRC-O9-001", "Contract", "o9-demand-planning-sow.pdf", "o9 SOW", "watchlist_restructure"),
    ("SRC-CHARTER-SAP", "Charter", "sap-erp-future-decision-charter.pdf", "SAP ERP charter", "move_charter"),
    ("SRC-CHARTER-PUNCHH", "Charter", "punchh-loyalty-replacement-charter.pdf", "Punchh charter", "move_charter"),
    ("SRC-CHARTER-AI", "Charter", "ai-productivity-program-charter.pdf", "AI productivity charter", "move_charter"),
    ("SRC-WATCHLIST-001", "Verification", "expected-watchlist-entries.json", "Watchlist expected entries", "verification"),
    ("SRC-SENTINEL-001", "Verification", "expected-sentinel-answers.json", "Sentinel canonical questions", "verification"),
]


def read_csv(relative_path: str) -> list[dict[str, str]]:
    with (PACK / relative_path).open(newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(relative_path: str, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path = PACK / relative_path
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def int_value(value: object) -> int:
    text = str(value or "0").strip()
    if text.lower() in {"nan", "null", "none", ""}:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def draw_pdf(path: Path, title: str, subtitle: str, rows: list[tuple[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(path), pagesize=LETTER, topMargin=42, bottomMargin=42)
    story = [
        Paragraph("AbarVa Synthetic - Apex Retail v1", styles["Normal"]),
        Paragraph(title, styles["Title"]),
        Paragraph(subtitle, styles["Heading2"]),
        Spacer(1, 14),
    ]
    table = Table([("Field", "Synthetic evidence")] + rows, colWidths=[150, 360])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#10233F")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9AA7B8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
    ]))
    story.append(table)
    story.append(Spacer(1, 16))
    story.append(Paragraph("Fixture notice: all values are synthetic, internally consistent, and safe for demo use. No real customer data.", styles["Italic"]))
    doc.build(story)


def generate_pdfs() -> None:
    for doc in CONTRACT_DOCS:
        draw_pdf(
            PACK / "04-vendors" / "contract-pdfs" / doc.file_name,
            f"{doc.vendor} - {doc.kind}",
            "Contract extract fixture",
            [
                ("Annual spend", f"${doc.annual_usd:,.0f}"),
                ("Renewal / decision window", doc.renewal_window),
                ("Risk", doc.risk),
                ("Negotiation leverage", doc.leverage),
                ("Watermark", "AbarVa Synthetic - Apex Retail v1"),
            ],
        )
    for doc in CHARTER_DOCS:
        draw_pdf(
            PACK / "09-charters" / "charter-pdfs" / doc.file_name,
            doc.title,
            "Wave 0 charter fixture",
            [
                ("Initiative ID", doc.initiative_id),
                ("Sponsor", doc.sponsor),
                ("Committed funding", f"${doc.committed_usd:,.0f}"),
                ("Kernel posture", doc.decision),
                ("Evidence anchor", doc.evidence),
                ("Watermark", "AbarVa Synthetic - Apex Retail v1"),
            ],
        )


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="E8EEF8")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="10233F")
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"


def generate_workbooks() -> None:
    commitments = read_csv("02-financial/initiative-commitments.csv")
    vendors = read_csv("04-vendors/vendor-contracts.csv")
    run_costs = read_csv("02-financial/annual-run-costs.csv")
    renewals = read_csv("02-financial/renewal-calendar-12mo.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Rollup"
    ws.append(["Metric", "Value"])
    ws.append(["Total vendor spend", sum(int_value(v["annual_usd"]) for v in vendors)])
    ws.append(["Active commitments", sum(int_value(i["committed_usd"]) for i in commitments)])
    ws.append(["Annual run cost", sum(int_value(r["total_run_usd"]) for r in run_costs)])
    ws.append(["Contracts", len(vendors)])
    style_header(ws)

    ws_vendor = wb.create_sheet("Vendor Spend")
    ws_vendor.append(["Vendor", "Annual USD", "Category", "Renewal"])
    for row in vendors:
        ws_vendor.append([row["vendor"], int_value(row["annual_usd"]), row["type"], row["renewal_date"]])
    style_header(ws_vendor)
    chart = BarChart()
    chart.title = "Top vendor spend"
    chart.y_axis.title = "USD"
    chart.x_axis.title = "Vendor"
    data = Reference(ws_vendor, min_col=2, min_row=1, max_row=11)
    cats = Reference(ws_vendor, min_col=1, min_row=2, max_row=11)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_vendor.add_chart(chart, "F2")

    for name, rows, headers in [
        ("AMS Detail", vendors[:12], ["vendor", "annual_usd", "category", "risk_rating"]),
        ("Cloud Cost", [r for r in run_costs if "AWS" in r.get("hosting", "") or "cloud" in r.get("hosting_detail", "").lower()][:30], list(run_costs[0].keys())[:7]),
        ("License Inventory", vendors, ["vendor", "type", "annual_usd", "renewal_date", "notes"]),
        ("Capex Pipeline", commitments, ["initiative_id", "title", "committed_usd", "projected_value_usd", "sentinel_posture"]),
        ("Projected Value", commitments, ["initiative_id", "title", "projected_value_usd", "sentinel_posture"]),
        ("Variance to Plan", commitments, ["initiative_id", "title", "committed_usd", "projected_value_usd", "sentinel_posture"]),
    ]:
        sheet = wb.create_sheet(name[:31])
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        style_header(sheet)
    wb.save(PACK / "02-financial" / "financial-workbook.xlsx")

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Commitments"
    headers = ["initiative_id", "title", "committed_usd", "projected_value_usd", "sentinel_posture"]
    ws2.append(headers)
    for row in commitments:
        ws2.append([row.get(header, "") for header in headers])
    style_header(ws2)
    ws2["H1"] = "Total committed"
    ws2["H2"] = "=SUM(C2:C31)"
    ws2["I1"] = "Total projected"
    ws2["I2"] = "=SUM(D2:D31)"
    wb2.create_sheet("Renewal Calendar")
    wsr = wb2["Renewal Calendar"]
    wsr.append(list(renewals[0].keys()))
    for row in renewals:
        wsr.append([row.get(header, "") for header in renewals[0].keys()])
    style_header(wsr)
    wb2.save(PACK / "02-financial" / "initiative-commitments.xlsx")


def generate_source_files_and_chunks() -> None:
    source_dir = PACK / "13-context" / "source-files"
    source_dir.mkdir(parents=True, exist_ok=True)
    rows = []
    for idx, (source_id, domain, file_name, display_name, feature) in enumerate(SOURCE_FILE_THEMES, start=1):
        source_path = source_dir / f"{source_id.lower()}.md"
        source_path.write_text(
            f"# {display_name}\n\n"
            "Tenant: Apex Retail Group\n"
            "Fixture: AbarVa Synthetic - Apex Retail v1\n"
            f"Domain: {domain}\n"
            f"Source file: {file_name}\n"
            f"Feature: {feature}\n\n"
            "This evidence intake note represents the normalized source record used by the Packet 18 onboarding simulation.\n"
            "All values are synthetic and internally consistent with the Apex profile.\n",
            encoding="utf-8",
        )
        rows.append({
            "source_file_id": source_id,
            "source_system": "packet_18_apex_synthetic",
            "source_file": file_name,
            "source_path": str(source_path.relative_to(PACK)),
            "source_type": domain.lower().replace(" ", "_"),
            "display_name": display_name,
            "target_table": "enterprise_context_source_files",
            "feature": feature,
            "row_count": 1,
            "confidence": "0.92",
            "freshness_status": "fresh",
        })
    write_csv("13-context/enterprise-context-source-files.csv", rows, list(rows[0].keys()))

    apps = read_csv("01-portfolio/application-portfolio.csv")
    initiatives = read_csv("01-portfolio/initiatives-active.csv")
    vendors = read_csv("04-vendors/vendor-contracts.csv")
    dora = read_csv("05-dora/dora-baseline-consolidated.csv")
    chunks = []
    sources = [row["source_file_id"] for row in rows]

    def add_chunk(source_index: int, record_id: str, segment: str, content: str, confidence: float = 0.9) -> None:
        ordinal = len(chunks) + 1
        chunks.append({
            "chunk_id": f"APX-P18-CHUNK-{ordinal:03d}",
            "tenant_key": "apex-retail",
            "source_file_id": sources[source_index % len(sources)],
            "source_segment_id": segment,
            "source_record_id": record_id,
            "content": content,
            "confidence": confidence,
            "freshness_status": "fresh",
            "evidence_pointer": f"datasets/apex-retail-synthetic-v1#chunk-{ordinal:03d}",
        })

    for idx, app in enumerate(apps):
        total_run = int_value(app["ams_annual_usd"]) + int_value(app["license_annual_usd"]) + int_value(app["infra_annual_usd"])
        add_chunk(idx, app["app_id"], "application_portfolio", f"{app['app_id']} is a {app['stack_era']} application owned by {app['owning_team_id']} with annual run cost ${total_run} and TIME classification {app['time_classification']}. Notes: {app['notes']}")
        if idx < 80:
            add_chunk(idx + 7, app["app_id"], "regulatory_and_dependency_context", f"{app['app_id']} has regulatory scope {app['regulatory_scope']}, criticality {app['criticality_tier']}, and {app['consumed_by_count']} downstream consumers. Sunset or restructuring must account for downstream consumers and control scope.")

    for idx, initiative in enumerate(initiatives):
        add_chunk(idx + 13, initiative["initiative_id"], "initiative_financials", f"{initiative['title']} has committed funding ${initiative['committed_usd']} and projected value ${initiative['projected_value_usd']}. Status is {initiative['status']} and Sentinel posture is {initiative['sentinel_posture']}.")
        add_chunk(idx + 19, initiative["initiative_id"], "sponsor_signal", f"{initiative['title']} accountable sponsor is {initiative['accountable']} with evidence note: {initiative['evidence_note']}.")

    for idx, vendor in enumerate(vendors):
        add_chunk(idx + 23, vendor["vendor"], "vendor_contract", f"{vendor['vendor']} contract type {vendor['type']} has annual spend ${vendor['annual_usd']} and renewal date {vendor['renewal_date']}. Notes: {vendor['notes']}.")

    for idx, row in enumerate(dora[:55]):
        add_chunk(idx + 31, row["team_id"], "dora_baseline", f"{row['team_id']} DORA baseline week {row['week']} shows deploy frequency {row['deploys_per_week']}, lead time {row['lead_time_hours']} hours, MTTR {row['mttr_hours']} hours, and change fail rate {row['change_failure_rate_pct']} percent.")

    while len(chunks) < 280:
        idx = len(chunks)
        app = apps[idx % len(apps)]
        add_chunk(idx + 37, app["app_id"], "synthetic_context_padding", f"Apex Retail evidence chunk for {app['app_id']} preserves tenant-grounded context for retrieval tests. It should be treated as synthetic planning evidence, not a fabricated operational actual.")

    jsonl_path = PACK / "13-context" / "client-data-corpus.jsonl"
    with jsonl_path.open("w", encoding="utf-8") as handle:
        for chunk in chunks[:280]:
            handle.write(json.dumps(chunk, separators=(",", ":")) + "\n")

    expected = {
        "enterprise_context_source_files": len(rows),
        "enterprise_context_chunks": 280,
        "contract_pdfs": len(CONTRACT_DOCS),
        "charter_pdfs": len(CHARTER_DOCS),
        "xlsx_files": [
            "02-financial/financial-workbook.xlsx",
            "02-financial/initiative-commitments.xlsx",
        ],
        "no_customer_data": True,
        "watermark_required": "AbarVa Synthetic - Apex Retail v1",
    }
    (PACK / "99-verification" / "expected-corpus-load.json").write_text(json.dumps(expected, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    generate_pdfs()
    generate_workbooks()
    generate_source_files_and_chunks()


if __name__ == "__main__":
    main()
