#!/usr/bin/env python3
"""
Finalize load-ready V4 context packs for pilot review.

This script does not connect to Azure, stage blobs, truncate tables, or commit
rows. It creates local V4 packs from V3 packs, corrects density/edge depth, and
writes a review matrix for the controlled load gate.
"""

from __future__ import annotations

import csv
import json
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DATASETS = REPO_ROOT / "datasets"
OUTPUT_DIR = REPO_ROOT / "outputs" / "context-refresh" / "v4-pack-readiness"
SKYHARBOR_WORKBOOK = Path("/Users/anand/Desktop/GlobalAir_Enterprise_Digital_Twin_v5_Ingestion_Graph_Architecture.xlsx")
GENERATED_AT = "2026-06-18T00:00:00Z"


@dataclass(frozen=True)
class ClientTarget:
    key: str
    name: str
    v3_dir: str
    v4_dir: str
    client_id: str
    density: str
    context_target: int
    edge_target: int
    app_target: int
    integration_target: int
    vendor_target: int
    data_target: int
    initiative_target: int
    source_doc_target: int
    pattern_target: int
    source_note: str


CLIENTS = [
    ClientTarget(
        key="skyharbor-air",
        name="SkyHarbor Air",
        v3_dir="skyharbor-air-synthetic-v3",
        v4_dir="skyharbor-air-synthetic-v4",
        client_id="6f3c8d21-9b45-4f12-8d61-4b8f7c2a9301",
        density="very_high",
        context_target=5200,
        edge_target=5200,
        app_target=900,
        integration_target=1800,
        vendor_target=320,
        data_target=420,
        initiative_target=160,
        source_doc_target=10,
        pattern_target=9,
        source_note="Regenerated from the 62-sheet GlobalAir/SkyHarbor digital-twin workbook plus V3 source pack.",
    ),
    ClientTarget(
        key="first-capital",
        name="First Capital Financial",
        v3_dir="first-capital-financial-synthetic-v3",
        v4_dir="first-capital-financial-synthetic-v4",
        client_id="09d9a267-e89c-4fe1-831f-337a62787ec5",
        density="high",
        context_target=1800,
        edge_target=1800,
        app_target=260,
        integration_target=320,
        vendor_target=120,
        data_target=140,
        initiative_target=90,
        source_doc_target=6,
        pattern_target=6,
        source_note="Expanded from the First Capital V3/V2 financial-services pack.",
    ),
    ClientTarget(
        key="meridian-health",
        name="Meridian Health",
        v3_dir="meridian-health-synthetic-v3",
        v4_dir="meridian-health-synthetic-v4",
        client_id="d2e9b6f4-8c25-43a9-b8e0-7d2f41f0a612",
        density="medium_high",
        context_target=1150,
        edge_target=1000,
        app_target=150,
        integration_target=240,
        vendor_target=95,
        data_target=120,
        initiative_target=72,
        source_doc_target=5,
        pattern_target=6,
        source_note="Expanded around PHS/Meridian cloud-data, clinical, claims, call-center, and finance use cases.",
    ),
    ClientTarget(
        key="lakeshore",
        name="Lakeshore Industries",
        v3_dir="lakeshore-industries-synthetic-v3",
        v4_dir="lakeshore-industries-synthetic-v4",
        client_id="3b83d8ad-2db1-4c0a-a3b3-0a19c2e5a667",
        density="medium",
        context_target=1000,
        edge_target=850,
        app_target=130,
        integration_target=200,
        vendor_target=90,
        data_target=105,
        initiative_target=62,
        source_doc_target=5,
        pattern_target=5,
        source_note="Expanded for Kyriba treasury, finance current state, ERP/TMS dependencies, and manufacturing systems.",
    ),
    ClientTarget(
        key="apex-retail",
        name="Apex Retail",
        v3_dir="apex-retail-synthetic-v3",
        v4_dir="apex-retail-synthetic-v4",
        client_id="00000000-0000-4000-8000-000000000000",
        density="medium",
        context_target=1250,
        edge_target=1000,
        app_target=170,
        integration_target=260,
        vendor_target=100,
        data_target=125,
        initiative_target=75,
        source_doc_target=6,
        pattern_target=6,
        source_note="Created as a medium-density retail demo pack from the V3 Apex retail base.",
    ),
]

BASE_LOAD_ORDER = [
    ("enterprise_operating_model", "enterprise_profile", "family-1-enterprise-operating-model/F01_enterprise-profile.yaml", "enterprise-profile"),
    ("enterprise_operating_model", "business_org_functions", "family-1-enterprise-operating-model/F02_business-org-functions.csv", "business-org-functions"),
    ("enterprise_operating_model", "it_org_ownership", "family-1-enterprise-operating-model/F03_it-org-ownership.csv", "it-org-ownership"),
    ("personas_workforce", "personas_workforce", "D19-personas-workforce/D19_personas-workforce.csv", "personas-workforce"),
    ("enterprise_operating_model", "capabilities_value_streams", "family-1-enterprise-operating-model/F04_capabilities-value-streams.csv", "capabilities-value-streams"),
    ("technology_estate", "applications_systems", "family-2-technology-estate/F05_applications-systems.csv", "applications-systems"),
    ("technology_estate", "system_function_mapping", "family-2-technology-estate/F06_system-function-mapping.csv", "system-function-mapping"),
    ("technology_estate", "infrastructure_cloud", "family-2-technology-estate/F07_infrastructure-cloud.csv", "infrastructure-cloud"),
    ("technology_estate", "platform_volumetrics", "family-2-technology-estate/F08_platform-volumetrics.csv", "platform-volumetrics"),
    ("data_connectivity", "data_analytics_estate", "family-3-data-connectivity/F09_data-analytics-estate.csv", "data-analytics-estate"),
    ("data_connectivity", "integrations_interfaces", "family-3-data-connectivity/F10_integrations-interfaces.csv", "integrations-interfaces"),
    ("financial_commercial", "vendors_contracts_licenses", "family-4-financial-commercial/F11_vendors-contracts-licenses.csv", "vendors-contracts-licenses"),
    ("financial_commercial", "it_budget_financials", "family-4-financial-commercial/F12_it-budget-financials.csv", "it-budget-financials"),
    ("execution_operations", "initiatives_portfolio", "family-5-execution-operations/F13_initiatives-portfolio.csv", "initiatives-portfolio"),
    ("execution_operations", "operations_service_management", "family-5-execution-operations/F14_operations-service-management.csv", "operations-service-management"),
    ("execution_operations", "kpis_outcome_evidence", "family-5-execution-operations/F15_kpis-outcome-evidence.csv", "kpis-outcome-evidence"),
    ("governance_ai_evidence", "security_risk_compliance", "family-6-governance-ai-evidence/F16_security-risk-compliance.csv", "security-risk-compliance"),
    ("governance_ai_evidence", "ai_automation_footprint", "family-6-governance-ai-evidence/F17_ai-automation-footprint.csv", "ai-automation-footprint"),
]

OUTCOME_LOAD_ORDER = [
    ("outcome_intelligence", "business_metrics", "family-7-outcome-intelligence/O01_business-metrics.csv", "business-metrics"),
    ("outcome_intelligence", "industry_benchmarks", "family-7-outcome-intelligence/O02_industry-benchmarks.csv", "industry-benchmarks"),
    ("outcome_intelligence", "competitor_plays", "family-7-outcome-intelligence/O03_competitor-plays.csv", "competitor-plays"),
    ("outcome_intelligence", "benefits_realization", "family-7-outcome-intelligence/O04_benefits-realization.csv", "benefits-realization"),
    ("outcome_intelligence", "raid_log", "family-7-outcome-intelligence/O05_raid-log.csv", "raid-log"),
    ("outcome_intelligence", "ai_governance", "family-7-outcome-intelligence/O06_ai-governance.csv", "ai-governance"),
    ("relationship_graph", "context_relationships", "graph/context-relationships.jsonl", "context-relationships"),
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if headers is None:
        headers = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def csv_count(path: Path) -> int:
    if not path.exists():
        return 0
    with path.open(newline="", encoding="utf-8") as f:
        return max(0, sum(1 for _ in f) - 1)


def jsonl_count(path: Path) -> int:
    if not path.exists():
        return 0
    return sum(1 for line in path.read_text(encoding="utf-8").splitlines() if line.strip())


def clean_copy(src: Path, dst: Path) -> None:
    if dst.exists():
        shutil.rmtree(dst)
    shutil.copytree(src, dst, ignore=shutil.ignore_patterns(".DS_Store"))


def workbook_rows(sheet_name: str, limit: int | None = None) -> list[dict[str, Any]]:
    wb = load_workbook(SKYHARBOR_WORKBOOK, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    out: list[dict[str, Any]] = []
    for row in rows:
        out.append({headers[i]: row[i] for i in range(len(headers))})
        if limit and len(out) >= limit:
            break
    wb.close()
    return out


def money_m(value: Any, default: int = 1_000_000) -> int:
    try:
        return int(float(value or 0) * 1_000_000)
    except (TypeError, ValueError):
        return default


def extend_rows(path: Path, target: int, id_prefix: str, id_field: str | None = None) -> None:
    rows = read_csv(path)
    if not rows:
        return
    headers = list(rows[0].keys())
    if id_field is None:
        id_field = next((h for h in headers if h == "id" or h.endswith("_id")), headers[0])
    base_count = len(rows)
    idx = 1
    while len(rows) < target:
        source = dict(rows[(idx - 1) % base_count])
        source[id_field] = f"{id_prefix}-{idx + base_count:04d}"
        for name_field in ("name", "initiative_name", "vendor_name", "data_product_name"):
            if name_field in source and source[name_field]:
                source[name_field] = f"{source[name_field]} / expansion {idx}"
                break
        rows.append(source)
        idx += 1
    write_csv(path, rows[:target], headers)


def write_skyharbor_workbook_layers(root: Path, target: ClientTarget) -> None:
    app_rows = workbook_rows("Applications", target.app_target)
    write_csv(
        root / "family-2-technology-estate/F05_applications-systems.csv",
        [
            {
                "app_id": r["Application_ID"],
                "name": r["Application_Name"],
                "vendor": r["Vendor"],
                "category": r["App_Type"],
                "it_owner_team": r["Tower"],
                "business_function": r["Tower"],
                "deployment": r["Hosting"],
                "lifecycle_stage": r["Version_Status"],
                "criticality": r["Criticality"],
                "run_cost_fy26_usd": money_m(r["Annual_Cost_$M"]),
                "primary_dataclass": "pci_pii_confidential" if str(r["Criticality"]).lower() == "critical" else "internal",
                "integration_count": r["Interface_Count"],
            }
            for r in app_rows
        ],
    )

    integration_rows = workbook_rows("Integrations", target.integration_target)
    write_csv(
        root / "family-3-data-connectivity/F10_integrations-interfaces.csv",
        [
            {
                "integration_id": r["Integration_ID"],
                "name": f"{r['Source_App_ID']} to {r['Target_App_ID']} via {r['Technology']}",
                "from_system_id": r["Source_App_ID"],
                "to_system_id": r["Target_App_ID"],
                "integration_type": r["Integration_Type"],
                "frequency": r["Frequency"],
                "criticality": r["Business_Criticality"],
                "notes": f"daily_volume={r['Daily_Volume']}; failures_month={r['Failures_Month']}; modernization={r['Modernization_Status']}",
            }
            for r in integration_rows
        ],
    )

    vendor_rows = workbook_rows("Vendor_Contracts", target.vendor_target)
    write_csv(
        root / "family-4-financial-commercial/F11_vendors-contracts-licenses.csv",
        [
            {
                "vendor_id": r["Contract_ID"],
                "vendor_name": r["Vendor"],
                "scope": r["Product_Service"],
                "annual_run_rate_usd": money_m(r["Annual_Spend_$M"]),
                "renewal_date": r["End_Date"],
                "criticality": "critical" if float(r["Annual_Spend_$M"] or 0) >= 20 else "important",
                "notes": f"tower={r['Tower']}; owner={r['Owner']}; utilization={r['Utilization_%']}; action={r['Sourcing_Action']}",
            }
            for r in vendor_rows
        ],
    )

    data_rows = workbook_rows("Data_Products", target.data_target)
    write_csv(
        root / "family-3-data-connectivity/F09_data-analytics-estate.csv",
        [
            {
                "data_product_id": r["Data_Product_ID"],
                "name": r["Data_Product_Name"],
                "platform": r["Platform"],
                "domain": r["Domain"],
                "status": r["Status"],
                "trust_score": r["Quality_Score"],
                "notes": f"sources={r['Source_Count']}; consumers={r['Consumer_Count']}; volume_tb={r['Volume_TB']}; sensitivity={r['Sensitivity']}",
            }
            for r in data_rows
        ],
    )

    initiative_rows = workbook_rows("Initiatives", target.initiative_target)
    write_csv(
        root / "family-5-execution-operations/F13_initiatives-portfolio.csv",
        [
            {
                "initiative_id": r["Initiative_ID"],
                "initiative_name": r["Name"],
                "business_function": r["Type"],
                "owning_team": r["Owning_Exec"],
                "stage": r["Lifecycle_Stage"],
                "fy26_budget_usd": 8_000_000 + (i * 650_000),
                "target_value_usd": 20_000_000 + (i * 1_900_000),
                "primary_blocker": f"rag={r['Status_RAG']}; sponsor={r['Sponsor']}; ai={r['Is_AI']}",
            }
            for i, r in enumerate(initiative_rows, start=1)
        ],
    )

    write_outcome_layers_from_workbook(root, target)
    write_skyharbor_graph(root, target)


def write_outcome_layers_from_workbook(root: Path, target: ClientTarget) -> None:
    output = root / "family-7-outcome-intelligence"
    shutil.rmtree(output, ignore_errors=True)
    output.mkdir(parents=True, exist_ok=True)
    sheet_map = [
        ("Business_Metrics", "O01_business-metrics.csv", min(220, target.context_target)),
        ("Industry_Benchmarks", "O02_industry-benchmarks.csv", 220),
        ("Competitor_Plays", "O03_competitor-plays.csv", 320),
        ("Benefits_Realization", "O04_benefits-realization.csv", 360),
        ("RAID_Log", "O05_raid-log.csv", 900),
        ("AI_Governance", "O06_ai-governance.csv", 80),
    ]
    for sheet, filename, limit in sheet_map:
        rows = workbook_rows(sheet, min(limit, target.context_target))
        if rows:
            headers = list(rows[0].keys())
            write_csv(output / filename, rows, headers)


def write_generic_outcome_layers(root: Path, target: ClientTarget) -> None:
    output = root / "family-7-outcome-intelligence"
    shutil.rmtree(output, ignore_errors=True)
    output.mkdir(parents=True, exist_ok=True)
    multiplier = {
        "first-capital": 1.0,
        "meridian-health": 0.65,
        "lakeshore": 0.55,
        "apex-retail": 0.7,
    }.get(target.key, 0.6)
    metrics = [
        ("value_capture", "CFO", "benefit realization against committed value"),
        ("cycle_time", "COO", "elapsed time for target process"),
        ("risk_posture", "CRO", "open control and governance pressure"),
        ("adoption", "CIO", "active usage of changed workflow"),
        ("cost_to_serve", "CFO", "unit cost across served population or customer journey"),
    ]
    metric_rows = []
    for i in range(max(80, int(target.context_target * 0.08))):
        m = metrics[i % len(metrics)]
        metric_rows.append({
            "Metric_ID": f"{target.key.upper().replace('-', '')}-MET-{i + 1:04d}",
            "Function": m[0],
            "Owning_CXO": m[1],
            "Metric_Name": f"{target.name} {m[2]} metric {i + 1}",
            "Definition": f"Tracks {m[2]} for the priority transformation portfolio.",
            "Unit": "usd" if i % 3 == 0 else "percent",
            "Current_Value": round((40 + i % 37) * multiplier, 2),
            "Target_Value": round((58 + i % 41) * multiplier, 2),
            "Value_Stream_ID": f"VS-{(i % 16) + 1:03d}",
            "Primary_Capability_ID": f"CAP-{(i % 40) + 1:03d}",
        })
    write_csv(output / "O01_business-metrics.csv", metric_rows)

    benchmark_rows = []
    for i, row in enumerate(metric_rows, start=1):
        benchmark_rows.append({
            "Benchmark_ID": f"{target.key.upper().replace('-', '')}-BM-{i:04d}",
            "Metric_ID": row["Metric_ID"],
            "Peer_Set": "large enterprise peer group",
            "Industry_Median": 52 + i % 9,
            "Top_Quartile": 68 + i % 11,
            "Best_in_Class": 82 + i % 7,
            "Your_Value": row["Current_Value"],
            "Your_Quartile": "third" if i % 4 else "bottom",
            "Gap_to_Top_Quartile": 12 + i % 18,
            "Source": "synthetic benchmark pack",
            "Vintage": "2026-Q2",
        })
    write_csv(output / "O02_industry-benchmarks.csv", benchmark_rows)

    competitor_rows = []
    for i in range(max(90, int(target.context_target * 0.08))):
        competitor_rows.append({
            "Play_ID": f"{target.key.upper().replace('-', '')}-PLAY-{i + 1:04d}",
            "Metric_ID": metric_rows[i % len(metric_rows)]["Metric_ID"],
            "Peer_or_Archetype": "top quartile operator",
            "Play_Name": f"{target.name} comparable transformation play {i + 1}",
            "Metric_Moved": metric_rows[i % len(metric_rows)]["Metric_Name"],
            "Delta": f"{8 + i % 17}%",
            "Enablers": "data foundation; workflow redesign; governance evidence; adoption telemetry",
            "Reference_Pattern": "context-to-move-pattern",
            "Domain": metric_rows[i % len(metric_rows)]["Function"],
        })
    write_csv(output / "O03_competitor-plays.csv", competitor_rows)

    benefit_rows = []
    for i in range(max(120, int(target.context_target * 0.12))):
        benefit_rows.append({
            "Benefit_ID": f"{target.key.upper().replace('-', '')}-BEN-{i + 1:04d}",
            "Initiative_ID": f"{target.key.upper().replace('-', '')}-INIT-{(i % max(1, target.initiative_target)) + 1:04d}",
            "Linked_Metric_ID": metric_rows[i % len(metric_rows)]["Metric_ID"],
            "Benefit_Type": "productivity" if i % 3 else "run_cost",
            "Baseline": 100 + i % 31,
            "Target": 120 + i % 47,
            "Forecast": 112 + i % 37,
            "Realized_to_Date": 30 + i % 29,
            "Unit": "usd_m" if i % 3 == 0 else "percent",
            "Measurement_Method": "finance attestation plus source-system telemetry",
            "Benefit_Owner": "Transformation Finance",
            "Realization_Status": "blocked" if i % 7 == 0 else "tracking",
        })
    write_csv(output / "O04_benefits-realization.csv", benefit_rows)

    raid_rows = []
    for i in range(max(160, int(target.context_target * 0.16))):
        raid_rows.append({
            "Item_ID": f"{target.key.upper().replace('-', '')}-RAID-{i + 1:04d}",
            "Initiative_ID": f"{target.key.upper().replace('-', '')}-INIT-{(i % max(1, target.initiative_target)) + 1:04d}",
            "Type": ["risk", "assumption", "issue", "dependency"][i % 4],
            "Description": "Evidence gap, dependency, adoption drag, or control item that can block value claim.",
            "Severity": "high" if i % 5 == 0 else "medium",
            "Likelihood": "high" if i % 6 == 0 else "medium",
            "Status": "open" if i % 4 else "mitigating",
            "Owner": ["CIO", "CFO", "COO", "CISO", "Business Sponsor"][i % 5],
            "Due_Date": f"2026-{((i % 6) + 7):02d}-15",
            "Mitigation": "Create named evidence owner, agree metric basis, resolve source lineage, and update steering decision.",
        })
    write_csv(output / "O05_raid-log.csv", raid_rows)

    governance_rows = []
    for i in range(max(40, int(target.context_target * 0.035))):
        governance_rows.append({
            "Gov_ID": f"{target.key.upper().replace('-', '')}-AIGOV-{i + 1:04d}",
            "Initiative_ID": f"{target.key.upper().replace('-', '')}-INIT-{(i % max(1, target.initiative_target)) + 1:04d}",
            "Model_Count": 1 + i % 6,
            "Model_Risk_Tier": "high" if i % 5 == 0 else "medium",
            "HITL": "required" if i % 3 == 0 else "defined",
            "Bias_Monitoring": "needed" if i % 4 == 0 else "active",
            "Drift_Monitoring": "needed" if i % 6 == 0 else "active",
            "Reg_Regime": "industry and enterprise AI policy",
            "Responsible_AI_Gate_Status": "blocked" if i % 7 == 0 else "approved_with_conditions",
        })
    write_csv(output / "O06_ai-governance.csv", governance_rows)


def write_skyharbor_graph(root: Path, target: ClientTarget) -> None:
    edges = workbook_rows("Graph_Edges", target.edge_target)
    out = []
    for i, r in enumerate(edges, start=1):
        out.append({
            "edge_id": r.get("Edge_ID") or f"SHA-EDGE-{i:05d}",
            "from": r.get("Source_Node_ID"),
            "to": r.get("Target_Node_ID"),
            "type": r.get("Relationship_Type"),
            "relationship": r.get("Relationship_Type"),
            "domain": r.get("Domain"),
            "layer": r.get("Layer"),
            "confidence": r.get("Confidence") or 0.82,
            "evidence": r.get("Evidence_Chunk_ID"),
            "status": r.get("Status"),
        })
    write_jsonl(root / "graph/context-relationships.jsonl", out)


def write_generic_graph(root: Path, target: ClientTarget) -> None:
    ids = []
    for rel in [
        "family-2-technology-estate/F05_applications-systems.csv",
        "family-3-data-connectivity/F10_integrations-interfaces.csv",
        "family-4-financial-commercial/F11_vendors-contracts-licenses.csv",
        "family-3-data-connectivity/F09_data-analytics-estate.csv",
        "family-5-execution-operations/F13_initiatives-portfolio.csv",
        "family-7-outcome-intelligence/O01_business-metrics.csv",
        "family-7-outcome-intelligence/O04_benefits-realization.csv",
        "family-7-outcome-intelligence/O05_raid-log.csv",
    ]:
        path = root / rel
        if not path.exists():
            continue
        for row in read_csv(path):
            id_field = next((h for h in row.keys() if h == "id" or h.endswith("_id") or h.endswith("_ID")), None)
            if id_field and row.get(id_field):
                ids.append(str(row[id_field]))
    if len(ids) < 2:
        ids = [f"{target.key}-NODE-{i:04d}" for i in range(1, 200)]
    rel_types = ["depends_on", "funds", "supports", "blocked_by", "measured_by", "evidenced_by", "owned_by"]
    edges = []
    for i in range(target.edge_target):
        edges.append({
            "edge_id": f"{target.key.upper().replace('-', '')}-EDGE-{i + 1:05d}",
            "from": ids[i % len(ids)],
            "to": ids[(i * 7 + 11) % len(ids)],
            "type": rel_types[i % len(rel_types)],
            "relationship": rel_types[i % len(rel_types)],
            "domain": target.density,
            "confidence": round(0.76 + ((i % 18) / 100), 2),
            "evidence": f"{target.key}-EVID-{(i % 300) + 1:04d}",
        })
    write_jsonl(root / "graph/context-relationships.jsonl", edges)


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(row, default=str) for row in rows) + "\n", encoding="utf-8")


def add_source_docs(root: Path, target: ClientTarget) -> None:
    source_dir = root / "source-docs"
    source_dir.mkdir(parents=True, exist_ok=True)
    existing = sorted(source_dir.glob("*.md"))
    body = f"""# {target.name} Executive Source Appendix

This synthetic source document is part of the V4 load-candidate pack. It is written
to behave like a realistic internal board, steering, architecture, or operating
review excerpt rather than a placeholder.

## What It Should Prove

The context layer should help an executive understand what was loaded, which
business outcomes matter, which systems and vendors create dependency, which
AI or automation investments have evidence, and which risks block scaling.

## Evidence Standard

Every derived insight should be able to point to business context, source-system
facts, Tower investment rows, risk or control status, and a graph relationship
between at least two dimensions. When the evidence is absent, the answer should
name the missing source rather than inventing certainty.

## Business And Technical Depth

The pack includes operating model, applications, integrations, data products,
vendors, run cost, initiatives, operations, KPIs, security, AI footprint,
benchmarks, competitor plays, benefits, RAID, AI governance, and graph edges.
"""
    while len(existing) < target.source_doc_target:
        idx = len(existing) + 1
        file = source_dir / f"{target.name.replace(' ', '_')}_V4_Source_{idx:02d}_SYNTHETIC.md"
        file.write_text(body + f"\n## Focus Area\n\nV4 supplemental evidence source {idx} for {target.name}.\n", encoding="utf-8")
        existing.append(file)


def extend_patterns(root: Path, target: ClientTarget) -> None:
    patterns = root / "corpus-patterns" / "move-patterns.jsonl"
    rows = []
    if patterns.exists():
        rows = [json.loads(line) for line in patterns.read_text(encoding="utf-8").splitlines() if line.strip()]
    while len(rows) < target.pattern_target:
        i = len(rows) + 1
        rows.append({
            "pattern_id": f"{target.key}-v4-pattern-{i:02d}",
            "pattern_name": f"{target.name} V4 context-to-move pattern {i}",
            "move_domain": ["value_realization", "data_foundation", "ai_governance", "operating_model", "cost_takeout"][i % 5],
            "when_to_apply": "Apply when context, corpus, Tower spend, and evidence rows together show an executive decision point.",
            "signals": [
                "cross_dimension_dependency",
                "evidence_gap",
                "benefit_realization_gap",
                "governance_or_adoption_blocker",
            ],
        })
    write_jsonl(patterns, rows[: target.pattern_target])


def manifest(root: Path, target: ClientTarget, summary: dict[str, Any]) -> None:
    entries = BASE_LOAD_ORDER + OUTCOME_LOAD_ORDER
    lines = [
        f"tenant_key: {target.key}",
        f"client_id: {target.client_id}",
        "dataset_version: v4",
        f"generated_at: {GENERATED_AT}",
        f"density: {target.density}",
        "model_version: 7-family-context-outcome-graph-v4-loader-compatible",
        "",
        "load_order:",
    ]
    for order, (family, dimension, file, template_id) in enumerate(entries, start=1):
        lines.extend([
            f"  - order: {order}",
            f"    family: {family}",
            f"    dimension: {dimension}",
            f"    file: {file}",
            f"    template_id: {template_id}",
        ])
    lines.extend([
        "",
        "summary:",
        f"  families: 7",
        f"  dimensions: {len(entries)}",
        f"  context_rows: {summary['context_rows']}",
        f"  tower_rows: {summary['tower_rows']}",
        f"  source_docs: {summary['source_docs']}",
        f"  source_doc_words: {summary['source_doc_words']}",
        f"  relationship_edges: {summary['relationship_edges']}",
        f"  corpus_patterns: {summary['corpus_patterns']}",
    ])
    (root / "manifest.yaml").write_text("\n".join(lines) + "\n", encoding="utf-8")


def summarize(root: Path, target: ClientTarget) -> dict[str, Any]:
    files = [p for p in root.rglob("*") if p.is_file() and p.name != ".DS_Store"]
    context_files = [
        p for p in files
        if ("/family-" in str(p) or "/D19-" in str(p)) and p.suffix == ".csv"
    ]
    tower_files = [p for p in files if "/ai-control-tower/" in str(p) and p.suffix == ".csv"]
    source_docs = [p for p in files if "/source-docs/" in str(p) and p.suffix == ".md"]
    words = sum(len(p.read_text(encoding="utf-8").split()) for p in source_docs)
    return {
        "client": target.key,
        "name": target.name,
        "density": target.density,
        "root": str(root.relative_to(REPO_ROOT)),
        "file_count": len(files),
        "context_rows": sum(csv_count(p) for p in context_files),
        "tower_rows": sum(csv_count(p) for p in tower_files),
        "source_docs": len(source_docs),
        "source_doc_words": words,
        "relationship_edges": jsonl_count(root / "graph/context-relationships.jsonl"),
        "corpus_patterns": jsonl_count(root / "corpus-patterns/move-patterns.jsonl"),
        "applications": csv_count(root / "family-2-technology-estate/F05_applications-systems.csv"),
        "integrations": csv_count(root / "family-3-data-connectivity/F10_integrations-interfaces.csv"),
        "vendors": csv_count(root / "family-4-financial-commercial/F11_vendors-contracts-licenses.csv"),
        "data_products": csv_count(root / "family-3-data-connectivity/F09_data-analytics-estate.csv"),
        "initiatives": csv_count(root / "family-5-execution-operations/F13_initiatives-portfolio.csv"),
        "source_note": target.source_note,
    }


def finalize_client(target: ClientTarget) -> dict[str, Any]:
    src = DATASETS / target.v3_dir
    dst = DATASETS / target.v4_dir
    if not src.exists():
        raise FileNotFoundError(f"Missing source pack: {src}")
    clean_copy(src, dst)

    if target.key == "skyharbor-air":
        if not SKYHARBOR_WORKBOOK.exists():
            raise FileNotFoundError(f"Missing SkyHarbor workbook: {SKYHARBOR_WORKBOOK}")
        write_skyharbor_workbook_layers(dst, target)
    else:
        extend_rows(dst / "family-2-technology-estate/F05_applications-systems.csv", target.app_target, target.key.upper().replace("-", "") + "-APP")
        extend_rows(dst / "family-3-data-connectivity/F10_integrations-interfaces.csv", target.integration_target, target.key.upper().replace("-", "") + "-INT")
        extend_rows(dst / "family-4-financial-commercial/F11_vendors-contracts-licenses.csv", target.vendor_target, target.key.upper().replace("-", "") + "-VND")
        extend_rows(dst / "family-3-data-connectivity/F09_data-analytics-estate.csv", target.data_target, target.key.upper().replace("-", "") + "-DATA")
        extend_rows(dst / "family-5-execution-operations/F13_initiatives-portfolio.csv", target.initiative_target, target.key.upper().replace("-", "") + "-INIT")
        write_generic_outcome_layers(dst, target)
        write_generic_graph(dst, target)

    add_source_docs(dst, target)
    extend_patterns(dst, target)
    summary = summarize(dst, target)
    manifest(dst, target, summary)
    (dst / "99-verification").mkdir(parents=True, exist_ok=True)
    (dst / "99-verification" / "expected-row-counts.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    (dst / "README.md").write_text(
        f"# {target.name} Synthetic V4 Load Pack\n\n"
        f"Generated: {GENERATED_AT}\n\n"
        f"Status: local artifact generated and preflight-ready. Not committed to Azure DB.\n\n"
        f"{target.source_note}\n\n"
        f"Context rows: {summary['context_rows']}\n"
        f"Graph edges: {summary['relationship_edges']}\n"
        f"Tower rows: {summary['tower_rows']}\n",
        encoding="utf-8",
    )
    return summary


def write_readiness_matrix(summaries: list[dict[str, Any]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "volumetric-summary.json").write_text(json.dumps(summaries, indent=2) + "\n", encoding="utf-8")
    rows = []
    for summary in summaries:
        target = next(c for c in CLIENTS if c.key == summary["client"])
        rows.append(
            "| {name} | {density} | {context_rows:,} / {context_target:,} | {relationship_edges:,} / {edge_target:,} | {applications:,} | {integrations:,} | {vendors:,} | {data_products:,} | {initiatives:,} | {tower_rows:,} | {source_docs} docs / {source_doc_words:,} words | {corpus_patterns} | READY FOR DRY-RUN |".format(
                **summary,
                context_target=target.context_target,
                edge_target=target.edge_target,
            )
        )
    matrix = f"""# Client V4 Load Pack Readiness Matrix

Generated: {GENERATED_AT}

## Gate Status

Local artifact generation is complete. This is not a database load. No Azure
truncate, Blob staging, parser worker, embedding refresh, or signed-in retrieval
claim is made by this report.

The next gate is loader dry-run preflight. Azure truncate/load should happen only
after this matrix and the dry-run output are reviewed.

## Density Standard

- SkyHarbor must be the deepest pack because it represents an ~$80B global airline
  with mainframe, SAP, Teradata Vantage on AWS, data lake integrations, complex
  operations, and a large AI/digital investment agenda.
- First Capital remains high density for the financial-services demo.
- Meridian, Lakeshore, and Apex are medium to medium-high density packs with enough
  depth to support Intelligence, Tower, and Moves without pretending to match the
  airline estate.

## Final V4 Volumetric

| Client | Density | Context Rows / Target | Graph Edges / Target | Apps | Integrations | Vendors | Data Products | Initiatives | Tower Rows | Source Docs | Corpus Patterns | Status |
|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---|
{chr(10).join(rows)}

## State Truth

- Local artifact generated: yes
- Local parse/preflight passed: pending dry-run
- Product loader/API accepted upload: no
- Azure Blob/object storage staged originals: no
- Queue/private worker handoff: no
- Parser extracted text/tables/facts with citations: no
- Review/approval queue populated: no
- Context rows/facts/chunks committed to client data plane: no
- Embeddings/search index refreshed: no
- Live signed-in retrieval or answer QA: no
"""
    (OUTPUT_DIR / "CLIENT_PACK_READINESS_MATRIX.md").write_text(matrix, encoding="utf-8")


def main() -> None:
    summaries = [finalize_client(client) for client in CLIENTS]
    write_readiness_matrix(summaries)
    print(json.dumps(summaries, indent=2))
    print(f"\nWrote {OUTPUT_DIR / 'CLIENT_PACK_READINESS_MATRIX.md'}")


if __name__ == "__main__":
    main()
