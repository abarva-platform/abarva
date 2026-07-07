#!/usr/bin/env python3
"""AbarVa Workforce Economics — Workforce_Taxonomy_Master.xlsx
Parametric, traceable substrate. Rates/costs are formula-driven off the Assumptions sheet.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---------- style helpers ----------
FONT = "Arial"
INK = "1A1A1A"; INPUT_BLUE = "0000FF"; LINK_GREEN = "006400"
HDR_FILL = PatternFill("solid", fgColor="102650")
SUB_FILL = PatternFill("solid", fgColor="E8ECF3")
ASSUM_FILL = PatternFill("solid", fgColor="FFF7E0")
TOWER_FILL = PatternFill("solid", fgColor="F1EEE7")
thin = Side(style="thin", color="D5D2CA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def H(cell, text, size=11, color="FFFFFF", fill=HDR_FILL, align="left", wrap=False):
    cell.value = text; cell.font = Font(name=FONT, bold=True, size=size, color=color)
    cell.fill = fill; cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = BORDER

def C(cell, text, bold=False, color=INK, align="left", num=None, wrap=False, size=10, fill=None):
    cell.value = text; cell.font = Font(name=FONT, bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap); cell.border = BORDER
    if num: cell.number_format = num
    if fill: cell.fill = fill

def widths(ws, w):
    for col, width in w.items(): ws.column_dimensions[col].width = width

USD = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
MULT = '0.00x'
PCT = '0.0%'

wb = Workbook()

# ================= 1. COVER / INSTRUCTIONS =================
ws = wb.active; ws.title = "Cover"; ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 40, "C": 70, "D": 18})
C(ws["B2"], "AbarVa", bold=True, size=22, color="102650")
C(ws["B3"], "Enterprise Workforce Economics & AI-Native Delivery Model", bold=True, size=13)
C(ws["B4"], "Workforce Taxonomy Master — v1.0", size=11, color="625F58")
C(ws["B6"], "PURPOSE", bold=True, color="102650")
C(ws["C6"], "Global workforce ontology + parametric cost, rate, agent-economics and pod substrate. "
            "Feeds the AbarVa estimation, business-case and roadmap engines for traditional and AI-native delivery.", wrap=True)
ws.row_dimensions[6].height = 42
C(ws["B8"], "HOW TO USE", bold=True, color="102650")
how = ["1. Edit ONLY blue-font cells (inputs) and the Assumptions tab. All black cells are formulas.",
       "2. Assumptions drives the whole model — change a multiplier and every rate/cost reflows.",
       "3. Rates are parametric: Market Base Rate x Provider Tier x Shore x Scarcity x Delivery. Nothing is a static guess.",
       "4. Internal cost: fully-loaded annual / 2,080 = loaded hourly.",
       "5. Provider names are anonymized archetypes (CONS-T1, SI-T1, SI-T2, ENG-B, AI-B). No real firms.",
       "6. AI-native model estimates every capability twice: People vs People+Agents+Platforms.",
       "7. 'Moves Binding' maps every taxonomy entity to the Move deliverable/estimation engine."]
for i, t in enumerate(how):
    C(ws.cell(row=9+i, column=3), t, wrap=True); ws.row_dimensions[9+i].height = 26
C(ws["B17"], "SHEET INDEX", bold=True, color="102650")
idx = [("Assumptions","Engine inputs: load factors, geo/provider/shore/scarcity multipliers, market base rates"),
       ("Exec Summary","Counts + key model outputs"),
       ("Towers","21 delivery towers"),("Capabilities","100+ capabilities by tower + agent-amenability"),
       ("Career Levels","10-level model with experience and base bands"),
       ("Roles","Role taxonomy by tower/capability/level"),
       ("Geography","17 regions: salary/rate/scarcity/COL multipliers + shore"),
       ("Internal Cost Model","Fully-loaded cost build-up by level"),
       ("Rate Intelligence","Anonymized provider rate cards (parametric)"),
       ("Agent Economics","Agent platform cost + productivity multipliers"),
       ("Delivery Pods","Reusable human+agent pod library"),
       ("Moves Binding","Mapping to the Move estimation/deliverable engine"),
       ("Glossary","Terms")]
H(ws["B18"], "Sheet"); H(ws["C18"], "Contents")
for i,(s,d) in enumerate(idx):
    C(ws.cell(row=19+i, column=2), s, bold=True); C(ws.cell(row=19+i, column=3), d, wrap=True)

# ================= 2. ASSUMPTIONS =================
wa = wb.create_sheet("Assumptions"); wa.sheet_view.showGridLines = False
widths(wa, {"A":3,"B":34,"C":14,"D":14,"E":14,"F":40})
C(wa["B2"], "ASSUMPTIONS — model engine (edit blue cells)", bold=True, size=13, color="102650")
C(wa["B3"], "All downstream cost & rate formulas reference these cells.", color="625F58")
# load components
C(wa["B5"], "FULLY-LOADED COST COMPONENTS (% of base salary)", bold=True, color="102650")
load = [("Bonus",0.15),("Equity",0.08),("Benefits",0.14),("Payroll Taxes",0.0765),("401k Match",0.04),
        ("Recruiting",0.03),("Training",0.025),("Bench / Utilization",0.12),("Facilities",0.05),
        ("Technology Allocation",0.045),("Corporate Allocation",0.06),("Administrative Allocation",0.035),
        ("Travel Allocation",0.02)]
H(wa["B6"],"Component"); H(wa["C6"],"% of base")
r=7
for n,v in load:
    C(wa.cell(row=r,column=2),n); c=wa.cell(row=r,column=3); C(c,v,color=INPUT_BLUE,num=PCT,align="right"); c.fill=ASSUM_FILL; r+=1
C(wa.cell(row=r,column=2),"Total load factor", bold=True)
tot=wa.cell(row=r,column=3); tot.value=f"=SUM(C7:C{r-1})"; tot.font=Font(name=FONT,bold=True); tot.number_format=PCT; tot.alignment=Alignment(horizontal="right"); tot.border=BORDER
LOAD_TOTAL_CELL=f"Assumptions!$C${r}"
C(wa.cell(row=r,column=6),"Loaded annual = base x (1 + total load factor)", color="625F58");
HOURS_ROW=r+2
C(wa.cell(row=HOURS_ROW,column=2),"Billable hours / year", bold=True)
hc=wa.cell(row=HOURS_ROW,column=3); C(hc,2080,color=INPUT_BLUE,align="right"); hc.fill=ASSUM_FILL
HOURS_CELL=f"Assumptions!$C${HOURS_ROW}"

# provider tier multipliers
pr=HOURS_ROW+3
C(wa.cell(row=pr,column=2),"PROVIDER TIER MULTIPLIERS (vs market base bill rate)", bold=True, color="102650")
H(wa.cell(row=pr+1,column=2),"Tier"); H(wa.cell(row=pr+1,column=3),"Multiplier"); H(wa.cell(row=pr+1,column=6),"Archetype")
tiers=[("CONS-T1",1.85,"Premium strategy/advisory"),("SI-T1",1.25,"Global system integrator"),
       ("SI-T2",0.85,"Industrialized offshore"),("ENG-B",1.10,"Engineering boutique"),("AI-B",1.35,"AI-native provider")]
TIER_ROW0=pr+2
for i,(n,v,d) in enumerate(tiers):
    rr=TIER_ROW0+i; C(wa.cell(row=rr,column=2),n,bold=True); c=wa.cell(row=rr,column=3); C(c,v,color=INPUT_BLUE,num=MULT,align="right"); c.fill=ASSUM_FILL
    C(wa.cell(row=rr,column=6),d,color="625F58")
TIER_RANGE=f"Assumptions!$B${TIER_ROW0}:$C${TIER_ROW0+len(tiers)-1}"

# shore multipliers
sr=TIER_ROW0+len(tiers)+2
C(wa.cell(row=sr,column=2),"SHORE MULTIPLIERS", bold=True, color="102650")
H(wa.cell(row=sr+1,column=2),"Shore"); H(wa.cell(row=sr+1,column=3),"Multiplier")
shore=[("Onshore",1.00),("Nearshore",0.72),("Offshore",0.45)]
SHORE_ROW0=sr+2
for i,(n,v) in enumerate(shore):
    rr=SHORE_ROW0+i; C(wa.cell(row=rr,column=2),n,bold=True); c=wa.cell(row=rr,column=3); C(c,v,color=INPUT_BLUE,num=MULT,align="right"); c.fill=ASSUM_FILL

# scarcity multipliers
qr=SHORE_ROW0+len(shore)+2
C(wa.cell(row=qr,column=2),"SCARCITY MULTIPLIERS (by role scarcity tier)", bold=True, color="102650")
H(wa.cell(row=qr+1,column=2),"Scarcity"); H(wa.cell(row=qr+1,column=3),"Multiplier")
scar=[("High",1.30),("Medium",1.10),("Low",1.00)]
SCAR_ROW0=qr+2
for i,(n,v) in enumerate(scar):
    rr=SCAR_ROW0+i; C(wa.cell(row=rr,column=2),n,bold=True); c=wa.cell(row=rr,column=3); C(c,v,color=INPUT_BLUE,num=MULT,align="right"); c.fill=ASSUM_FILL
SCAR_RANGE=f"Assumptions!$B${SCAR_ROW0}:$C${SCAR_ROW0+len(scar)-1}"

# market base bill rate per level (onshore baseline)
mr=SCAR_ROW0+len(scar)+2
C(wa.cell(row=mr,column=2),"MARKET BASE BILL RATE — onshore baseline ($/hr by level)", bold=True, color="102650")
H(wa.cell(row=mr+1,column=2),"Level"); H(wa.cell(row=mr+1,column=3),"Base $/hr"); H(wa.cell(row=mr+1,column=4),"Base salary (US, $)")
LEVELS=[("Partner",625,520000,"25+"),("Managing Director",525,430000,"20+"),("Principal",430,360000,"18+"),
        ("Director",360,300000,"15+"),("Senior Manager",300,245000,"12+"),("Manager",250,200000,"10+"),
        ("Lead",215,175000,"8+"),("Senior",180,150000,"6+"),("Intermediate",140,120000,"3+"),("Junior",105,95000,"0-3")]
MKT_ROW0=mr+2
for i,(n,rate,sal,yoe) in enumerate(LEVELS):
    rr=MKT_ROW0+i; C(wa.cell(row=rr,column=2),n,bold=True)
    c=wa.cell(row=rr,column=3); C(c,rate,color=INPUT_BLUE,num=USD,align="right"); c.fill=ASSUM_FILL
    c2=wa.cell(row=rr,column=4); C(c2,sal,color=INPUT_BLUE,num=USD,align="right"); c2.fill=ASSUM_FILL
MKT_RANGE=f"Assumptions!$B${MKT_ROW0}:$D${MKT_ROW0+len(LEVELS)-1}"
SAL_RANGE=MKT_RANGE
wa.freeze_panes="A5"

# ================= 3. TOWERS =================
TOWERS=["Strategy & Transformation","Industry SMEs","Business Process","Data & Analytics","AI & GenAI",
        "Digital Experience","Marketing Technology","Product Management","Application Engineering","Integration",
        "ERP","Cloud","Infrastructure","Cybersecurity","Quality Engineering","Operations","Managed Services",
        "Legacy & Mainframe","Change Management","Program Management","Enterprise Architecture"]
wt=wb.create_sheet("Towers"); wt.sheet_view.showGridLines=False
widths(wt,{"A":3,"B":10,"C":34,"D":60})
C(wt["B2"],"DELIVERY TOWERS",bold=True,size=13,color="102650")
H(wt["B4"],"ID"); H(wt["C4"],"Tower"); H(wt["D4"],"Scope")
tdesc={"Strategy & Transformation":"Vision, operating model, transformation strategy, value architecture",
 "Industry SMEs":"Vertical domain expertise (FS, healthcare, retail, manufacturing, public)",
 "Business Process":"Process design, reengineering, process mining, automation strategy",
 "Data & Analytics":"Data strategy, governance, engineering, BI, analytics, data products",
 "AI & GenAI":"AI/ML, GenAI, agentic AI, MLOps/LLMOps, responsible AI",
 "Digital Experience":"UX/UI, commerce, content, web/mobile experience platforms",
 "Marketing Technology":"MarTech, Adobe, journey orchestration, customer data",
 "Product Management":"Product strategy, ownership, discovery, lifecycle",
 "Application Engineering":"Custom app dev, microservices, APIs, modernization",
 "Integration":"iPaaS, ESB, event streaming, API management",
 "ERP":"SAP, Oracle, Workday functional & technical delivery",
 "Cloud":"Azure/AWS/GCP architecture, migration, platform engineering",
 "Infrastructure":"Compute, network, storage, identity, end-user",
 "Cybersecurity":"IAM, SOC, GRC, AppSec, threat & vuln management",
 "Quality Engineering":"Test strategy, automation, performance, security testing",
 "Operations":"Run, monitoring, SRE, service operations",
 "Managed Services":"AMS L1/L2/L3, application & platform support",
 "Legacy & Mainframe":"Mainframe, COBOL, DB2, z/OS modernization",
 "Change Management":"OCM, training, communications, adoption",
 "Program Management":"PMO, program/project delivery, agile delivery",
 "Enterprise Architecture":"EA, solution & technical architecture, standards"}
for i,t in enumerate(TOWERS):
    rr=5+i; C(wt.cell(row=rr,column=2),f"TWR-{i+1:02d}",bold=True); C(wt.cell(row=rr,column=3),t,bold=True,fill=TOWER_FILL)
    C(wt.cell(row=rr,column=4),tdesc[t],wrap=True)
wt.freeze_panes="A5"; wt.auto_filter.ref=f"B4:D{4+len(TOWERS)}"

# ================= 4. CAREER LEVELS =================
wl=wb.create_sheet("Career Levels"); wl.sheet_view.showGridLines=False
widths(wl,{"A":3,"B":20,"C":12,"D":18,"E":50})
C(wl["B2"],"CAREER LEVEL MODEL",bold=True,size=13,color="102650")
H(wl["B4"],"Level"); H(wl["C4"],"Years Exp"); H(wl["D4"],"Base Salary (US)"); H(wl["E4"],"Expectation")
exp={"Partner":"Owns client P&L, sells & sponsors transformation",
 "Managing Director":"Owns portfolio, executive relationships, delivery assurance",
 "Principal":"Owns large program design and outcomes",
 "Director":"Owns workstream/program delivery and quality",
 "Senior Manager":"Leads multiple teams / a delivery track",
 "Manager":"Leads a team and a deliverable scope",
 "Lead":"Technical/functional lead for a squad",
 "Senior":"Independent senior contributor",
 "Intermediate":"Contributor with guidance",
 "Junior":"Entry contributor, supervised"}
for i,(n,rate,sal,yoe) in enumerate(LEVELS):
    rr=5+i; C(wl.cell(row=rr,column=2),n,bold=True)
    C(wl.cell(row=rr,column=3),yoe,align="center")
    cc=wl.cell(row=rr,column=4); cc.value=f"=VLOOKUP(B{rr},{SAL_RANGE},3,FALSE)"; cc.font=Font(name=FONT,color=LINK_GREEN); cc.number_format=USD; cc.alignment=Alignment(horizontal="right"); cc.border=BORDER
    C(wl.cell(row=rr,column=5),exp[n],wrap=True)
wl.freeze_panes="A5"
LEVEL_NAMES=[l[0] for l in LEVELS]

# ================= 5. GEOGRAPHY =================
wg=wb.create_sheet("Geography"); wg.sheet_view.showGridLines=False
widths(wg,{"A":3,"B":22,"C":14,"D":14,"E":14,"F":14,"G":14})
C(wg["B2"],"GEOGRAPHY MODEL",bold=True,size=13,color="102650")
for col,t in zip("BCDEFG",["Region","Shore","Salary Mult","Rate Mult","Scarcity Mult","Cost of Living"]):
    H(wg[f"{col}4"],t)
GEOS=[("NYC","Onshore",1.20,1.18,1.10,1.25),("SF / Bay Area","Onshore",1.25,1.22,1.15,1.30),
 ("Seattle","Onshore",1.15,1.12,1.10,1.18),("Chicago","Onshore",1.05,1.05,1.00,1.05),
 ("Houston","Onshore",1.00,1.00,1.00,1.00),("Dallas","Onshore",1.00,1.00,1.00,1.00),
 ("Atlanta","Onshore",0.98,1.00,1.00,0.98),("Canada","Nearshore",0.82,0.85,0.95,0.90),
 ("UK","Onshore",1.05,1.08,1.05,1.10),("Western Europe","Onshore",1.02,1.05,1.05,1.08),
 ("Eastern Europe","Nearshore",0.55,0.60,0.90,0.55),("India Tier 1","Offshore",0.30,0.40,0.85,0.35),
 ("India Tier 2","Offshore",0.24,0.34,0.80,0.28),("Philippines","Offshore",0.26,0.36,0.80,0.30),
 ("LATAM","Nearshore",0.50,0.58,0.90,0.52),("Middle East","Onshore",0.95,1.00,1.05,1.00),
 ("Australia","Onshore",1.08,1.10,1.05,1.12)]
for i,(n,sh,sm,rm,scm,col) in enumerate(GEOS):
    rr=5+i; C(wg.cell(row=rr,column=2),n,bold=True); C(wg.cell(row=rr,column=3),sh,align="center")
    for j,v in enumerate([sm,rm,scm,col]):
        c=wg.cell(row=rr,column=4+j); C(c,v,color=INPUT_BLUE,num=MULT,align="right"); c.fill=ASSUM_FILL
wg.freeze_panes="A5"; wg.auto_filter.ref=f"B4:G{4+len(GEOS)}"

# ================= 6. INTERNAL COST MODEL =================
wc=wb.create_sheet("Internal Cost Model"); wc.sheet_view.showGridLines=False
widths(wc,{"A":3,"B":20,"C":16,"D":16,"E":16,"F":16})
C(wc["B2"],"INTERNAL FULLY-LOADED COST MODEL",bold=True,size=13,color="102650")
C(wc["B3"],"Loaded annual = base salary x (1 + total load factor). Loaded hourly = loaded annual / billable hours.",color="625F58")
for col,t in zip("BCDEF",["Level","Base Salary","Load Factor","Loaded Annual","Loaded Hourly"]):
    H(wc[f"{col}5"],t)
for i,n in enumerate(LEVEL_NAMES):
    rr=6+i; C(wc.cell(row=rr,column=2),n,bold=True)
    b=wc.cell(row=rr,column=3); b.value=f"=VLOOKUP(B{rr},{SAL_RANGE},3,FALSE)"; b.font=Font(name=FONT,color=LINK_GREEN); b.number_format=USD; b.alignment=Alignment(horizontal="right"); b.border=BORDER
    lf=wc.cell(row=rr,column=4); lf.value=f"=1+{LOAD_TOTAL_CELL}"; lf.font=Font(name=FONT,color=LINK_GREEN); lf.number_format=MULT; lf.alignment=Alignment(horizontal="right"); lf.border=BORDER
    la=wc.cell(row=rr,column=5); la.value=f"=C{rr}*D{rr}"; la.font=Font(name=FONT); la.number_format=USD; la.alignment=Alignment(horizontal="right"); la.border=BORDER
    lh=wc.cell(row=rr,column=6); lh.value=f"=E{rr}/{HOURS_CELL}"; lh.font=Font(name=FONT); lh.number_format=USD2; lh.alignment=Alignment(horizontal="right"); lh.border=BORDER
wc.freeze_panes="A6"

# ================= 7. RATE INTELLIGENCE =================
PROVIDERS=[("CONS-T1-A","CONS-T1"),("CONS-T1-B","CONS-T1"),("CONS-T1-C","CONS-T1"),("CONS-T1-D","CONS-T1"),
 ("SI-T1-A","SI-T1"),("SI-T1-B","SI-T1"),("SI-T1-C","SI-T1"),("SI-T1-D","SI-T1"),("SI-T1-E","SI-T1"),
 ("SI-T2-A","SI-T2"),("SI-T2-B","SI-T2"),("SI-T2-C","SI-T2"),("SI-T2-D","SI-T2"),("SI-T2-E","SI-T2"),("SI-T2-F","SI-T2"),
 ("ENG-B1","ENG-B"),("ENG-B2","ENG-B"),("ENG-B3","ENG-B"),("ENG-B4","ENG-B"),
 ("AI-B1","AI-B"),("AI-B2","AI-B"),("AI-B3","AI-B"),("AI-B4","AI-B")]
wr=wb.create_sheet("Rate Intelligence"); wr.sheet_view.showGridLines=False
widths(wr,{"A":3,"B":14,"C":12,"D":20,"E":14,"F":14,"G":14,"H":14})
C(wr["B2"],"GLOBAL RATE INTELLIGENCE — anonymized provider archetypes",bold=True,size=13,color="102650")
C(wr["B3"],"Bill rate = Market Base (onshore, by level) x Provider Tier Mult x Shore Mult. Scarcity applied per role in the estimator. No real firm names.",color="625F58")
for col,t in zip("BCDEFGH",["Provider","Category","Level","Onshore $/hr","Nearshore $/hr","Offshore $/hr","Tier Mult"]):
    H(wr[f"{col}5"],t)
def mkt_rate_ref(level):  # market base $/hr by level
    return f"VLOOKUP(D{{r}},{MKT_RANGE},2,FALSE)"
ON=f"INDEX(Assumptions!$C${SHORE_ROW0}:$C${SHORE_ROW0+2},1)"
NEAR=f"INDEX(Assumptions!$C${SHORE_ROW0}:$C${SHORE_ROW0+2},2)"
OFF=f"INDEX(Assumptions!$C${SHORE_ROW0}:$C${SHORE_ROW0+2},3)"
row=6
for pcode,cat in PROVIDERS:
    for lvl in LEVEL_NAMES:
        C(wr.cell(row=row,column=2),pcode,bold=True); C(wr.cell(row=row,column=3),cat,align="center")
        C(wr.cell(row=row,column=4),lvl)
        tm=wr.cell(row=row,column=8); tm.value=f"=VLOOKUP(C{row},{TIER_RANGE},2,FALSE)"; tm.font=Font(name=FONT,color=LINK_GREEN); tm.number_format=MULT; tm.alignment=Alignment(horizontal="right"); tm.border=BORDER
        base=f"VLOOKUP(D{row},{MKT_RANGE},2,FALSE)"
        for j,sh in zip(range(3),[ON,NEAR,OFF]):
            c=wr.cell(row=row,column=5+j); c.value=f"={base}*H{row}*{sh}"; c.font=Font(name=FONT); c.number_format=USD; c.alignment=Alignment(horizontal="right"); c.border=BORDER
        row+=1
wr.freeze_panes="B6"; wr.auto_filter.ref=f"B5:H{row-1}"

# ================= 8. AGENT ECONOMICS =================
AGENTS=[("Agent Platform A","General copilot",2500),("Agent Platform B","Coding agent",4000),
        ("Agent Platform C","Workflow/RPA agent",3200),("Agent Platform D","Domain reasoning agent",6000),
        ("AbarVa Agents","AbarVa-native delivery agents",3500)]
wae=wb.create_sheet("Agent Economics"); wae.sheet_view.showGridLines=False
widths(wae,{"A":3,"B":20,"C":24,"D":14,"E":14,"F":16,"G":13,"H":13,"I":13,"J":13,"K":13})
C(wae["B2"],"AI AGENT ECONOMICS (provider-neutral)",bold=True,size=13,color="102650")
C(wae["B3"],"Multipliers express capacity gain vs a human-only baseline (1.00 = no gain).",color="625F58")
for col,t in zip("BCDEFGHIJK",["Platform","Type","Monthly $","Annual $","Equiv Eng FTE","Util %","Productivity","Documentation","Testing","Architecture"]):
    H(wae[f"{col}5"],t)
agent_rows=[("Agent Platform A",0.5,0.70,1.35,1.60,1.30,1.10),("Agent Platform B",1.2,0.75,1.80,1.50,1.90,1.25),
 ("Agent Platform C",0.8,0.65,1.50,1.40,1.45,1.10),("Agent Platform D",1.5,0.60,1.70,1.55,1.60,1.45),
 ("AbarVa Agents",1.3,0.72,1.85,1.65,1.80,1.40)]
amap={r[0]:r for r in agent_rows}
for i,(n,ty,mo) in enumerate(AGENTS):
    rr=6+i; _,fte,util,prod,doc,test,arch=amap[n]
    C(wae.cell(row=rr,column=2),n,bold=True); C(wae.cell(row=rr,column=3),ty)
    c=wae.cell(row=rr,column=4); C(c,mo,color=INPUT_BLUE,num=USD,align="right"); c.fill=ASSUM_FILL
    an=wae.cell(row=rr,column=5); an.value=f"=D{rr}*12"; an.font=Font(name=FONT); an.number_format=USD; an.alignment=Alignment(horizontal="right"); an.border=BORDER
    for j,v,fmt in [(6,fte,MULT),(7,util,PCT),(8,prod,MULT),(9,doc,MULT),(10,test,MULT),(11,arch,MULT)]:
        c=wae.cell(row=rr,column=j); C(c,v,color=INPUT_BLUE,num=fmt,align="right"); c.fill=ASSUM_FILL
wae.freeze_panes="A6"

# ================= 9. CAPABILITIES =================
# (tower, capability, scarcity, agent_amenability 1-5)
CAPS=[
("Strategy & Transformation","Transformation Strategy","High",3),("Strategy & Transformation","Operating Model Design","High",3),
("Strategy & Transformation","Value Architecture","Medium",3),("Strategy & Transformation","Business Case & ROI","Medium",4),
("Industry SMEs","Financial Services Domain","High",2),("Industry SMEs","Healthcare Domain","High",2),
("Industry SMEs","Retail & CPG Domain","Medium",2),("Industry SMEs","Manufacturing Domain","Medium",2),("Industry SMEs","Public Sector Domain","Medium",2),
("Business Process","Process Design","Medium",4),("Business Process","Process Mining","Medium",4),("Business Process","Intelligent Automation","High",5),
("Data & Analytics","Data Strategy","High",3),("Data & Analytics","Data Governance","High",4),("Data & Analytics","Data Quality","Medium",4),
("Data & Analytics","Metadata Management","Medium",4),("Data & Analytics","Master Data Management","High",3),("Data & Analytics","Data Products","High",4),
("Data & Analytics","Data Engineering","High",5),("Data & Analytics","Business Intelligence","Medium",4),("Data & Analytics","Reporting","Low",5),
("Data & Analytics","Advanced Analytics","High",4),("Data & Analytics","Snowflake","High",4),("Data & Analytics","Databricks","High",4),
("Data & Analytics","Informatica","Medium",3),("Data & Analytics","Collibra","Medium",3),
("AI & GenAI","Agentic AI","High",5),("AI & GenAI","Generative AI","High",5),("AI & GenAI","Prompt Engineering","High",5),
("AI & GenAI","Machine Learning","High",4),("AI & GenAI","MLOps","High",5),("AI & GenAI","LLMOps","High",5),
("AI & GenAI","Responsible AI","High",3),("AI & GenAI","AI Governance","High",3),("AI & GenAI","Model Validation","High",3),
("Digital Experience","UX Design","Medium",4),("Digital Experience","UI Design","Medium",4),("Digital Experience","Customer Journey","Medium",4),
("Digital Experience","Commerce","Medium",4),("Digital Experience","Content Management","Low",5),
("Marketing Technology","Adobe AEM","High",3),("Marketing Technology","Adobe Analytics","Medium",4),("Marketing Technology","MarTech Strategy","Medium",4),
("Marketing Technology","Customer Data Platform","High",4),("Marketing Technology","Journey Orchestration","Medium",4),
("Product Management","Product Strategy","High",3),("Product Management","Product Ownership","Medium",4),("Product Management","Product Discovery","Medium",4),
("Application Engineering","Microservices","High",5),("Application Engineering","API Engineering","High",5),("Application Engineering","Frontend Engineering","Medium",5),
("Application Engineering","Backend Engineering","High",5),("Application Engineering","Mobile Engineering","Medium",5),("Application Engineering","App Modernization","High",4),
("Integration","iPaaS","Medium",4),("Integration","Event Streaming","High",4),("Integration","API Management","Medium",4),("Integration","ESB / Middleware","Medium",3),
("ERP","SAP Finance","High",3),("ERP","SAP Supply Chain","High",3),("ERP","SAP Basis","Medium",3),("ERP","Oracle Finance","High",3),
("ERP","Oracle SCM","High",3),("ERP","Workday HCM","High",3),("ERP","Workday Finance","High",3),
("Cloud","Azure","High",4),("Cloud","AWS","High",4),("Cloud","GCP","Medium",4),("Cloud","Platform Engineering","High",5),("Cloud","DevOps","High",5),("Cloud","SRE","High",4),
("Infrastructure","Network","Medium",3),("Infrastructure","Storage","Medium",3),("Infrastructure","Compute","Medium",3),("Infrastructure","Identity","High",4),("Infrastructure","End-User Compute","Low",4),
("Cybersecurity","IAM","High",4),("Cybersecurity","SOC","High",4),("Cybersecurity","GRC","High",3),("Cybersecurity","Application Security","High",4),("Cybersecurity","Threat & Vuln Mgmt","High",4),
("Quality Engineering","Test Strategy","Medium",4),("Quality Engineering","Test Automation","High",5),("Quality Engineering","Performance Testing","Medium",4),("Quality Engineering","Security Testing","High",4),
("Operations","Service Operations","Low",5),("Operations","Monitoring & Observability","Medium",5),("Operations","Site Reliability","High",4),
("Managed Services","AMS L1","Low",5),("Managed Services","AMS L2","Medium",5),("Managed Services","AMS L3","High",4),
("Legacy & Mainframe","Mainframe","High",2),("Legacy & Mainframe","COBOL","High",3),("Legacy & Mainframe","DB2","Medium",3),("Legacy & Mainframe","z/OS","High",2),
("Change Management","Organizational Change","Medium",3),("Change Management","Training","Low",4),("Change Management","Communications","Low",4),("Change Management","Adoption","Medium",4),
("Program Management","PMO","Medium",4),("Program Management","Program Delivery","Medium",3),("Program Management","Agile Delivery","Medium",4),
("Enterprise Architecture","Enterprise Architecture","High",3),("Enterprise Architecture","Solution Architecture","High",3),("Enterprise Architecture","Technical Architecture","High",4),
# --- expansion ---
("ERP","SAP S/4HANA","High",3),("ERP","SAP ABAP & Fiori","Medium",4),("ERP","SAP BTP & Integration","High",3),
("ERP","SAP SuccessFactors","Medium",3),("ERP","SAP Ariba","Medium",3),("ERP","SAP Analytics Cloud","Medium",4),
("ERP","Oracle Fusion","High",3),("ERP","Oracle Cloud Infrastructure","High",3),("ERP","Workday Integration","High",3),
("Application Engineering","ServiceNow ITSM","High",4),("Application Engineering","ServiceNow HRSD","Medium",4),
("Application Engineering","ServiceNow CSM","Medium",4),("Application Engineering","ServiceNow ITOM","High",4),
("Application Engineering","ServiceNow SecOps","High",4),("Application Engineering","Salesforce Sales Cloud","High",4),
("Application Engineering","Salesforce Service Cloud","High",4),("Application Engineering","Salesforce Marketing Cloud","Medium",4),
("Application Engineering","Salesforce Platform","High",4),
("Integration","MuleSoft","High",4),("Cloud","Kubernetes","High",4),("Cloud","FinOps","Medium",4),
("Cloud","Infrastructure as Code","High",5),("Cybersecurity","Zero Trust","High",3),("Business Process","Robotic Process Automation","Medium",5),
("Industry SMEs","Banking Domain","High",2),("Industry SMEs","Capital Markets Domain","High",2),("Industry SMEs","Insurance Domain","High",2),
("Industry SMEs","Payments Domain","High",2),("Industry SMEs","Payer Domain","High",2),("Industry SMEs","Provider Domain","High",2),
("Industry SMEs","Life Sciences Domain","High",2),("Industry SMEs","Energy & Utilities Domain","Medium",2),
("Industry SMEs","Telecom Domain","Medium",2),("Industry SMEs","Media Domain","Medium",2),
]
wcap=wb.create_sheet("Capabilities"); wcap.sheet_view.showGridLines=False
widths(wcap,{"A":3,"B":12,"C":26,"D":34,"E":12,"F":18})
C(wcap["B2"],"CAPABILITIES",bold=True,size=13,color="102650")
C(wcap["B3"],f"{len(CAPS)} capabilities. Agent-Amenability (1-5) = how compressible by agents/AI (feeds AI-native model).",color="625F58")
for col,t in zip("BCDEF",["ID","Tower","Capability","Scarcity","Agent-Amenability"]):
    H(wcap[f"{col}5"],t)
for i,(tw,cap,sc,aa) in enumerate(CAPS):
    rr=6+i; C(wcap.cell(row=rr,column=2),f"CAP-{i+1:03d}",bold=True); C(wcap.cell(row=rr,column=3),tw); C(wcap.cell(row=rr,column=4),cap,bold=True)
    C(wcap.cell(row=rr,column=5),sc,align="center"); C(wcap.cell(row=rr,column=6),aa,align="center")
wcap.freeze_panes="A6"; wcap.auto_filter.ref=f"B5:F{5+len(CAPS)}"

# ================= 10. ROLES =================
# role families per tower: (title, capability, scarcity, min_level, max_level)
P,MD,PR,DIR,SM,MGR,LD,SR,IN,JR = LEVEL_NAMES
def lv(a,b): return (a,b)
ROLES=[]
def add(tw,title,cap,sc,lo,hi): ROLES.append((tw,title,cap,sc,lo,hi))
# Strategy & Transformation
add("Strategy & Transformation","Transformation Partner","Transformation Strategy","High",P,P)
add("Strategy & Transformation","Transformation Director","Transformation Strategy","High",DIR,PR)
add("Strategy & Transformation","Operating Model Lead","Operating Model Design","High",SM,DIR)
add("Strategy & Transformation","Value Architect","Value Architecture","Medium",MGR,PR)
add("Strategy & Transformation","Business Case Lead","Business Case & ROI","Medium",MGR,DIR)
add("Strategy & Transformation","Transformation Consultant","Transformation Strategy","Medium",IN,MGR)
# Industry SMEs
for dom in ["Financial Services","Healthcare","Retail & CPG","Manufacturing","Public Sector"]:
    add("Industry SMEs",f"{dom} Principal SME",f"{dom} Domain" if dom in ["Financial Services","Healthcare"] else f"{'Retail & CPG' if dom=='Retail & CPG' else dom} Domain","High",PR,MD)
    add("Industry SMEs",f"{dom} Domain Consultant",f"{dom} Domain" if dom in ["Financial Services","Healthcare"] else f"{dom} Domain","Medium",SR,SM)
# Business Process
add("Business Process","Process Design Lead","Process Design","Medium",MGR,DIR)
add("Business Process","Process Mining Consultant","Process Mining","Medium",SR,MGR)
add("Business Process","Intelligent Automation Architect","Intelligent Automation","High",LD,DIR)
add("Business Process","Automation Engineer","Intelligent Automation","Medium",IN,SR)
# Data & Analytics
add("Data & Analytics","Chief Data Officer (Advisory)","Data Strategy","High",MD,P)
add("Data & Analytics","Enterprise Data Architect","Data Strategy","High",DIR,PR)
add("Data & Analytics","Data Architect","Data Strategy","High",LD,DIR)
add("Data & Analytics","Data Product Manager","Data Products","High",MGR,DIR)
add("Data & Analytics","Data Product Owner","Data Products","Medium",SR,SM)
add("Data & Analytics","Data Governance Lead","Data Governance","High",MGR,DIR)
add("Data & Analytics","Data Steward","Data Governance","Medium",IN,SR)
add("Data & Analytics","Metadata Lead","Metadata Management","Medium",LD,SM)
add("Data & Analytics","MDM Architect","Master Data Management","High",LD,DIR)
add("Data & Analytics","Data Quality Lead","Data Quality","Medium",LD,SM)
add("Data & Analytics","Snowflake Architect","Snowflake","High",LD,DIR)
add("Data & Analytics","Snowflake Engineer","Snowflake","High",IN,SR)
add("Data & Analytics","Databricks Architect","Databricks","High",LD,DIR)
add("Data & Analytics","Databricks Engineer","Databricks","High",IN,SR)
add("Data & Analytics","Lead Data Engineer","Data Engineering","High",LD,SM)
add("Data & Analytics","Senior Data Engineer","Data Engineering","High",SR,SR)
add("Data & Analytics","Data Engineer","Data Engineering","High",IN,SR)
add("Data & Analytics","Informatica Developer","Informatica","Medium",IN,SR)
add("Data & Analytics","Collibra Consultant","Collibra","Medium",SR,MGR)
add("Data & Analytics","BI Architect","Business Intelligence","Medium",LD,DIR)
add("Data & Analytics","BI Developer","Business Intelligence","Medium",IN,SR)
add("Data & Analytics","Analytics Lead","Advanced Analytics","High",LD,SM)
add("Data & Analytics","Analytics Consultant","Advanced Analytics","Medium",IN,SR)
add("Data & Analytics","Reporting Analyst","Reporting","Low",JR,IN)
# AI & GenAI
add("AI & GenAI","Chief AI Officer (Advisory)","AI Governance","High",MD,P)
add("AI & GenAI","Chief AI Architect","Agentic AI","High",DIR,PR)
add("AI & GenAI","AI Strategist","Generative AI","High",MGR,DIR)
add("AI & GenAI","AI Product Manager","Generative AI","High",MGR,DIR)
add("AI & GenAI","Agent Architect","Agentic AI","High",LD,DIR)
add("AI & GenAI","Forward Deployed Engineer","Agentic AI","High",SR,SM)
add("AI & GenAI","Principal AI Engineer","Machine Learning","High",PR,PR)
add("AI & GenAI","Lead AI Engineer","Machine Learning","High",LD,SM)
add("AI & GenAI","AI Engineer","Generative AI","High",IN,SR)
add("AI & GenAI","ML Engineer","Machine Learning","High",IN,SR)
add("AI & GenAI","MLOps Engineer","MLOps","High",SR,LD)
add("AI & GenAI","LLMOps Engineer","LLMOps","High",SR,LD)
add("AI & GenAI","Prompt Engineer","Prompt Engineering","High",IN,SR)
add("AI & GenAI","Responsible AI Lead","Responsible AI","High",MGR,DIR)
add("AI & GenAI","AI Governance Lead","AI Governance","High",MGR,DIR)
add("AI & GenAI","Model Validation Lead","Model Validation","High",LD,SM)
# Digital Experience
add("Digital Experience","Experience Director","UX Design","Medium",DIR,PR)
add("Digital Experience","UX Lead","UX Design","Medium",LD,SM)
add("Digital Experience","UX Designer","UX Design","Medium",IN,SR)
add("Digital Experience","UI Designer","UI Design","Medium",IN,SR)
add("Digital Experience","Journey Architect","Customer Journey","Medium",LD,DIR)
add("Digital Experience","Commerce Architect","Commerce","Medium",LD,DIR)
# Marketing Technology
add("Marketing Technology","Adobe Architect","Adobe AEM","High",LD,DIR)
add("Marketing Technology","AEM Architect","Adobe AEM","High",LD,DIR)
add("Marketing Technology","Adobe Developer","Adobe AEM","Medium",IN,SR)
add("Marketing Technology","Adobe Analytics Consultant","Adobe Analytics","Medium",SR,MGR)
add("Marketing Technology","CDP Architect","Customer Data Platform","High",LD,DIR)
add("Marketing Technology","MarTech Strategist","MarTech Strategy","Medium",MGR,DIR)
# Product Management
add("Product Management","Head of Product (Advisory)","Product Strategy","High",DIR,MD)
add("Product Management","Principal Product Manager","Product Strategy","High",PR,PR)
add("Product Management","Senior Product Manager","Product Ownership","Medium",SM,SM)
add("Product Management","Product Owner","Product Ownership","Medium",SR,MGR)
# Application Engineering
add("Application Engineering","Principal Engineer","Backend Engineering","High",PR,PR)
add("Application Engineering","Engineering Lead","Microservices","High",LD,SM)
add("Application Engineering","Senior Software Engineer","Backend Engineering","High",SR,SR)
add("Application Engineering","Software Engineer","Backend Engineering","Medium",IN,SR)
add("Application Engineering","Frontend Engineer","Frontend Engineering","Medium",IN,SR)
add("Application Engineering","Mobile Engineer","Mobile Engineering","Medium",IN,SR)
add("Application Engineering","API Engineer","API Engineering","High",IN,SR)
add("Application Engineering","Modernization Architect","App Modernization","High",LD,DIR)
# Integration
add("Integration","Integration Architect","API Management","High",LD,DIR)
add("Integration","iPaaS Developer","iPaaS","Medium",IN,SR)
add("Integration","Event Streaming Engineer","Event Streaming","High",SR,LD)
add("Integration","Middleware Engineer","ESB / Middleware","Medium",IN,SR)
# ERP
add("ERP","SAP Enterprise Architect","SAP Finance","High",DIR,PR)
add("ERP","SAP Finance Architect","SAP Finance","High",LD,DIR)
add("ERP","SAP Supply Chain Architect","SAP Supply Chain","High",LD,DIR)
add("ERP","SAP Security Architect","SAP Basis","Medium",LD,SM)
add("ERP","SAP Basis Consultant","SAP Basis","Medium",SR,MGR)
add("ERP","Oracle Finance Architect","Oracle Finance","High",LD,DIR)
add("ERP","Oracle SCM Consultant","Oracle SCM","High",SR,MGR)
add("ERP","Workday Architect","Workday HCM","High",LD,DIR)
add("ERP","Workday Functional Lead","Workday HCM","High",LD,SM)
add("ERP","Workday Technical Lead","Workday Finance","High",LD,SM)
# Cloud
add("Cloud","Cloud Architect","Azure","High",LD,DIR)
add("Cloud","Azure Architect","Azure","High",LD,DIR)
add("Cloud","AWS Architect","AWS","High",LD,DIR)
add("Cloud","GCP Architect","GCP","Medium",LD,DIR)
add("Cloud","Platform Architect","Platform Engineering","High",LD,DIR)
add("Cloud","Platform Engineer","Platform Engineering","High",IN,SR)
add("Cloud","DevOps Architect","DevOps","High",LD,SM)
add("Cloud","DevOps Engineer","DevOps","High",IN,SR)
add("Cloud","SRE Architect","SRE","High",LD,SM)
# Infrastructure
add("Infrastructure","Infrastructure Architect","Compute","Medium",LD,DIR)
add("Infrastructure","Network Architect","Network","Medium",LD,SM)
add("Infrastructure","Storage Engineer","Storage","Medium",IN,SR)
add("Infrastructure","Identity Architect","Identity","High",LD,SM)
# Cybersecurity
add("Cybersecurity","IAM Architect","IAM","High",LD,DIR)
add("Cybersecurity","SOC Lead","SOC","High",LD,SM)
add("Cybersecurity","SOC Analyst","SOC","Medium",JR,SR)
add("Cybersecurity","GRC Lead","GRC","High",MGR,DIR)
add("Cybersecurity","Application Security Engineer","Application Security","High",SR,LD)
add("Cybersecurity","Threat & Vuln Analyst","Threat & Vuln Mgmt","High",IN,SR)
# Quality Engineering
add("Quality Engineering","QA Architect","Test Strategy","Medium",LD,DIR)
add("Quality Engineering","Automation Architect","Test Automation","High",LD,SM)
add("Quality Engineering","Automation Engineer","Test Automation","Medium",IN,SR)
add("Quality Engineering","Performance Architect","Performance Testing","Medium",LD,SM)
add("Quality Engineering","Security Testing Architect","Security Testing","High",LD,SM)
# Operations
add("Operations","Service Operations Lead","Service Operations","Low",LD,SM)
add("Operations","Observability Engineer","Monitoring & Observability","Medium",SR,LD)
add("Operations","Site Reliability Engineer","Site Reliability","High",SR,LD)
# Managed Services
add("Managed Services","AMS Delivery Manager","AMS L3","Medium",MGR,DIR)
add("Managed Services","L3 Support Engineer","AMS L3","High",SR,LD)
add("Managed Services","L2 Support Engineer","AMS L2","Medium",IN,SR)
add("Managed Services","L1 Support Analyst","AMS L1","Low",JR,IN)
# Legacy & Mainframe
add("Legacy & Mainframe","Mainframe Architect","Mainframe","High",LD,DIR)
add("Legacy & Mainframe","COBOL Architect","COBOL","High",LD,SM)
add("Legacy & Mainframe","COBOL Engineer","COBOL","High",IN,SR)
add("Legacy & Mainframe","z/OS Engineer","z/OS","High",SR,LD)
# Change Management
add("Change Management","OCM Lead","Organizational Change","Medium",MGR,DIR)
add("Change Management","Training Lead","Training","Low",LD,SM)
add("Change Management","Communications Lead","Communications","Low",LD,SM)
add("Change Management","Adoption Consultant","Adoption","Medium",SR,MGR)
# Program Management
add("Program Management","Program Director","Program Delivery","Medium",DIR,PR)
add("Program Management","Program Manager","Program Delivery","Medium",SM,DIR)
add("Program Management","Project Manager","PMO","Medium",MGR,SM)
add("Program Management","Scrum Master","Agile Delivery","Medium",SR,MGR)
add("Program Management","PMO Analyst","PMO","Low",JR,IN)
# Enterprise Architecture
add("Enterprise Architecture","Chief Architect","Enterprise Architecture","High",DIR,MD)
add("Enterprise Architecture","Enterprise Architect","Enterprise Architecture","High",DIR,PR)
add("Enterprise Architecture","Solution Architect","Solution Architecture","High",LD,DIR)
add("Enterprise Architecture","Technical Architect","Technical Architecture","High",LD,SM)
# ===== EXPANSION: deeper role specialization =====
# Strategy & Transformation
add("Strategy & Transformation","Sourcing Strategy Lead","Business Case & ROI","Medium",MGR,DIR)
add("Strategy & Transformation","Vendor Management Advisor","Operating Model Design","Medium",SM,DIR)
add("Strategy & Transformation","Cost Optimization Lead","Business Case & ROI","Medium",MGR,DIR)
add("Strategy & Transformation","M&A Integration Lead","Transformation Strategy","High",SM,PR)
add("Strategy & Transformation","Cloud Economics Advisor","Value Architecture","Medium",MGR,DIR)
# Industry SMEs (sub-domains)
for dom,cap in [("Banking","Banking Domain"),("Capital Markets","Capital Markets Domain"),("Insurance","Insurance Domain"),
 ("Payments","Payments Domain"),("Payer","Payer Domain"),("Provider","Provider Domain"),("Life Sciences","Life Sciences Domain"),
 ("Energy & Utilities","Energy & Utilities Domain"),("Telecom","Telecom Domain"),("Media","Media Domain")]:
    add("Industry SMEs",f"{dom} Principal SME",cap,"High",PR,MD)
    add("Industry SMEs",f"{dom} Consultant",cap,"Medium",SR,SM)
# Business Process
add("Business Process","Process Architect","Process Design","Medium",LD,DIR)
add("Business Process","RPA Developer","Robotic Process Automation","Medium",IN,SR)
add("Business Process","Workflow Designer","Intelligent Automation","Medium",IN,SR)
add("Business Process","Process Analyst","Process Mining","Low",JR,SR)
# Data & Analytics
add("Data & Analytics","Analytics Engineer (dbt)","Data Engineering","High",IN,LD)
add("Data & Analytics","Streaming Data Engineer","Data Engineering","High",IN,LD)
add("Data & Analytics","DataOps Engineer","Data Engineering","High",IN,SR)
add("Data & Analytics","Data Modeler","Data Strategy","Medium",IN,LD)
add("Data & Analytics","Data Privacy Lead","Data Governance","High",LD,DIR)
add("Data & Analytics","Data Catalog Analyst","Metadata Management","Medium",JR,SR)
add("Data & Analytics","Power BI Developer","Business Intelligence","Medium",IN,SR)
add("Data & Analytics","Tableau Developer","Business Intelligence","Medium",IN,SR)
add("Data & Analytics","Looker Developer","Business Intelligence","Medium",IN,SR)
add("Data & Analytics","Data Visualization Designer","Reporting","Low",JR,SR)
add("Data & Analytics","Quantitative Analyst","Advanced Analytics","High",SR,SM)
add("Data & Analytics","Decision Scientist","Advanced Analytics","High",SR,SM)
add("Data & Analytics","Data Quality Engineer","Data Quality","Medium",IN,SR)
add("Data & Analytics","Reference Data Analyst","Master Data Management","Medium",IN,SR)
add("Data & Analytics","Snowflake Administrator","Snowflake","Medium",IN,SR)
add("Data & Analytics","Databricks Platform Engineer","Databricks","High",SR,LD)
add("Data & Analytics","Informatica Architect","Informatica","Medium",LD,SM)
# AI & GenAI
add("AI & GenAI","Data Scientist","Machine Learning","High",IN,SM)
add("AI & GenAI","Senior Data Scientist","Machine Learning","High",SR,LD)
add("AI & GenAI","NLP Engineer","Generative AI","High",SR,LD)
add("AI & GenAI","Computer Vision Engineer","Machine Learning","High",SR,LD)
add("AI & GenAI","RAG Engineer","Generative AI","High",IN,SR)
add("AI & GenAI","AI Evaluation Engineer","Model Validation","High",SR,LD)
add("AI & GenAI","AI Safety Engineer","Responsible AI","High",SR,LD)
add("AI & GenAI","Knowledge Engineer","Agentic AI","High",SR,LD)
add("AI & GenAI","Conversational AI Designer","Generative AI","Medium",IN,SR)
add("AI & GenAI","AI Platform Engineer","MLOps","High",SR,LD)
add("AI & GenAI","Retrieval Engineer","LLMOps","High",SR,LD)
add("AI & GenAI","AI Solution Architect","Agentic AI","High",LD,DIR)
# Digital Experience
add("Digital Experience","Service Designer","UX Design","Medium",SR,SM)
add("Digital Experience","Content Strategist","Content Management","Low",IN,SM)
add("Digital Experience","Accessibility Specialist","UX Design","Medium",IN,SR)
add("Digital Experience","Design System Lead","UI Design","Medium",LD,SM)
# Marketing Technology
add("Marketing Technology","Adobe Campaign Consultant","Journey Orchestration","Medium",SR,MGR)
add("Marketing Technology","Adobe Target Specialist","Adobe Analytics","Medium",IN,SR)
add("Marketing Technology","MarTech Engineer","MarTech Strategy","Medium",IN,SR)
add("Marketing Technology","Personalization Lead","Journey Orchestration","Medium",LD,SM)
add("Marketing Technology","CDP Engineer","Customer Data Platform","High",IN,SR)
# Product Management
add("Product Management","Technical Product Manager","Product Strategy","High",SM,DIR)
add("Product Management","Platform Product Manager","Product Strategy","High",SM,DIR)
add("Product Management","Product Analyst","Product Discovery","Medium",JR,SR)
add("Product Management","Product Designer","Product Discovery","Medium",IN,SR)
# Application Engineering (languages + SaaS platforms)
add("Application Engineering","Staff Engineer","Backend Engineering","High",LD,PR)
add("Application Engineering","Full-Stack Engineer","Backend Engineering","Medium",IN,SR)
add("Application Engineering","Java Engineer","Backend Engineering","Medium",IN,SR)
add("Application Engineering",".NET Engineer","Backend Engineering","Medium",IN,SR)
add("Application Engineering","Python Engineer","Backend Engineering","Medium",IN,SR)
add("Application Engineering","React Engineer","Frontend Engineering","Medium",IN,SR)
add("Application Engineering","iOS Engineer","Mobile Engineering","Medium",IN,SR)
add("Application Engineering","Android Engineer","Mobile Engineering","Medium",IN,SR)
add("Application Engineering","GraphQL Engineer","API Engineering","Medium",IN,SR)
add("Application Engineering","ServiceNow Architect","ServiceNow ITSM","High",LD,DIR)
add("Application Engineering","ServiceNow ITSM Developer","ServiceNow ITSM","High",IN,SR)
add("Application Engineering","ServiceNow HRSD Developer","ServiceNow HRSD","Medium",IN,SR)
add("Application Engineering","ServiceNow ITOM Consultant","ServiceNow ITOM","High",SR,SM)
add("Application Engineering","ServiceNow SecOps Consultant","ServiceNow SecOps","High",SR,SM)
add("Application Engineering","Salesforce Architect","Salesforce Platform","High",LD,DIR)
add("Application Engineering","Salesforce Sales Cloud Consultant","Salesforce Sales Cloud","High",SR,MGR)
add("Application Engineering","Salesforce Service Cloud Consultant","Salesforce Service Cloud","High",SR,MGR)
add("Application Engineering","Salesforce Marketing Cloud Consultant","Salesforce Marketing Cloud","Medium",SR,MGR)
add("Application Engineering","Apex Developer","Salesforce Platform","Medium",IN,SR)
add("Application Engineering","LWC Developer","Salesforce Platform","Medium",IN,SR)
# Integration
add("Integration","MuleSoft Architect","MuleSoft","High",LD,DIR)
add("Integration","MuleSoft Developer","MuleSoft","High",IN,SR)
add("Integration","Kafka Engineer","Event Streaming","High",SR,LD)
add("Integration","EDI Specialist","ESB / Middleware","Medium",IN,SR)
add("Integration","API Product Manager","API Management","Medium",SM,DIR)
# ERP — SAP modules
add("ERP","SAP FICO Consultant","SAP Finance","High",SR,SM)
add("ERP","SAP MM Consultant","SAP Supply Chain","High",SR,SM)
add("ERP","SAP SD Consultant","SAP Supply Chain","High",SR,SM)
add("ERP","SAP PP Consultant","SAP Supply Chain","Medium",SR,SM)
add("ERP","SAP EWM Consultant","SAP Supply Chain","Medium",SR,SM)
add("ERP","SAP ABAP Developer","SAP ABAP & Fiori","Medium",IN,SR)
add("ERP","SAP Fiori Developer","SAP ABAP & Fiori","Medium",IN,SR)
add("ERP","SAP BTP Architect","SAP BTP & Integration","High",LD,DIR)
add("ERP","SAP CPI Consultant","SAP BTP & Integration","Medium",SR,SM)
add("ERP","S/4HANA Migration Lead","SAP S/4HANA","High",LD,DIR)
add("ERP","SAP Analytics Cloud Consultant","SAP Analytics Cloud","Medium",SR,SM)
add("ERP","SAP SuccessFactors Consultant","SAP SuccessFactors","Medium",SR,SM)
add("ERP","SAP Ariba Consultant","SAP Ariba","Medium",SR,SM)
# ERP — Oracle / Workday
add("ERP","Oracle Fusion Finance Consultant","Oracle Fusion","High",SR,SM)
add("ERP","Oracle Fusion HCM Consultant","Oracle Fusion","High",SR,SM)
add("ERP","Oracle EBS Consultant","Oracle Finance","Medium",SR,SM)
add("ERP","OCI Architect","Oracle Cloud Infrastructure","High",LD,DIR)
add("ERP","Workday Integration Consultant","Workday Integration","High",SR,SM)
add("ERP","Workday Reporting Analyst","Workday Integration","Medium",IN,SR)
add("ERP","Workday Adaptive Planning Consultant","Workday Finance","Medium",SR,SM)
# Cloud — per platform
add("Cloud","Azure Data Engineer","Azure","High",IN,SR)
add("Cloud","Azure Security Engineer","Azure","High",SR,LD)
add("Cloud","AWS Data Engineer","AWS","High",IN,SR)
add("Cloud","AWS DevOps Engineer","DevOps","High",IN,SR)
add("Cloud","GCP Data Engineer","GCP","Medium",IN,SR)
add("Cloud","Kubernetes Architect","Kubernetes","High",LD,DIR)
add("Cloud","Kubernetes Engineer","Kubernetes","High",IN,SR)
add("Cloud","FinOps Analyst","FinOps","Medium",IN,SR)
add("Cloud","IaC / Terraform Engineer","Infrastructure as Code","High",IN,SR)
add("Cloud","Landing Zone Architect","Platform Engineering","High",LD,DIR)
add("Cloud","Cloud Migration Engineer","Azure","Medium",IN,SR)
add("Cloud","SRE Engineer","SRE","High",IN,SR)
# Infrastructure
add("Infrastructure","Cloud Infrastructure Engineer","Compute","Medium",IN,SR)
add("Infrastructure","Virtualization Engineer","Compute","Medium",IN,SR)
add("Infrastructure","Backup & DR Engineer","Storage","Medium",IN,SR)
add("Infrastructure","Active Directory Engineer","Identity","Medium",IN,SR)
add("Infrastructure","Endpoint Engineer","End-User Compute","Low",JR,SR)
add("Infrastructure","Network Engineer","Network","Medium",IN,SR)
# Cybersecurity
add("Cybersecurity","CISO Advisor","GRC","High",DIR,MD)
add("Cybersecurity","Security Architect","Application Security","High",LD,DIR)
add("Cybersecurity","Cloud Security Engineer","IAM","High",SR,LD)
add("Cybersecurity","Zero Trust Architect","Zero Trust","High",LD,DIR)
add("Cybersecurity","Incident Responder","SOC","High",IN,SR)
add("Cybersecurity","Threat Hunter","Threat & Vuln Mgmt","High",SR,LD)
add("Cybersecurity","Penetration Tester","Threat & Vuln Mgmt","High",SR,LD)
add("Cybersecurity","GRC Analyst","GRC","Medium",JR,SR)
add("Cybersecurity","SIEM Engineer","SOC","High",SR,LD)
add("Cybersecurity","PAM Engineer","IAM","High",SR,LD)
add("Cybersecurity","Privacy Engineer","GRC","High",SR,LD)
# Quality Engineering
add("Quality Engineering","SDET","Test Automation","High",IN,SR)
add("Quality Engineering","API Test Engineer","Test Automation","Medium",IN,SR)
add("Quality Engineering","Mobile Test Engineer","Test Automation","Medium",IN,SR)
add("Quality Engineering","Test Data Engineer","Test Strategy","Medium",IN,SR)
add("Quality Engineering","Chaos Engineer","Performance Testing","High",SR,LD)
# Operations
add("Operations","Incident Manager","Service Operations","Medium",SR,MGR)
add("Operations","Problem Manager","Service Operations","Medium",SR,MGR)
add("Operations","Major Incident Lead","Service Operations","High",LD,SM)
add("Operations","Monitoring Engineer","Monitoring & Observability","Medium",IN,SR)
add("Operations","Capacity Engineer","Site Reliability","Medium",SR,LD)
# Managed Services
add("Managed Services","Service Delivery Manager","AMS L3","Medium",MGR,DIR)
add("Managed Services","Transition Manager","AMS L3","Medium",MGR,SM)
add("Managed Services","AMS Architect","AMS L3","High",LD,DIR)
add("Managed Services","Knowledge Manager","AMS L2","Low",SR,MGR)
# Legacy & Mainframe
add("Legacy & Mainframe","DB2 DBA","DB2","Medium",SR,LD)
add("Legacy & Mainframe","Mainframe Modernization Consultant","Mainframe","High",SR,SM)
add("Legacy & Mainframe","JCL Engineer","z/OS","Medium",IN,SR)
# Change Management
add("Change Management","Change Analyst","Organizational Change","Low",JR,SR)
add("Change Management","Instructional Designer","Training","Low",IN,SR)
add("Change Management","Adoption Analyst","Adoption","Low",JR,SR)
# Program Management
add("Program Management","Portfolio Manager","Program Delivery","Medium",DIR,PR)
add("Program Management","Agile Coach","Agile Delivery","Medium",LD,SM)
add("Program Management","Release Train Engineer","Agile Delivery","Medium",LD,SM)
add("Program Management","Delivery Lead","Program Delivery","Medium",MGR,SM)
add("Program Management","Business Analyst","PMO","Low",JR,SR)
# Enterprise Architecture
add("Enterprise Architecture","Business Architect","Enterprise Architecture","High",LD,DIR)
add("Enterprise Architecture","Domain Architect","Solution Architecture","High",LD,DIR)
add("Enterprise Architecture","Integration Architect (EA)","Technical Architecture","High",LD,SM)
add("Enterprise Architecture","Security Architect (EA)","Technical Architecture","High",LD,SM)
add("Enterprise Architecture","Data Architect (EA)","Technical Architecture","High",LD,SM)

wro=wb.create_sheet("Roles"); wro.sheet_view.showGridLines=False
widths(wro,{"A":3,"B":12,"C":24,"D":34,"E":26,"F":12,"G":16,"H":16,"I":16,"J":16})
C(wro["B2"],"ROLE TAXONOMY",bold=True,size=13,color="102650")
C(wro["B3"],f"{len(ROLES)} roles across {len(TOWERS)} towers. Loaded hourly pulls from Internal Cost Model; scarcity adjusts.",color="625F58")
for col,t in zip("BCDEFGHIJ",["ID","Tower","Role","Capability","Scarcity","Min Level","Max Level","Loaded $/hr (min lvl)","Scarcity-Adj $/hr"]):
    H(wro[f"{col}5"],t)
COST_LH=f"'Internal Cost Model'!$B$6:$F$15"  # B level..F loaded hourly
for i,(tw,title,cap,sc,lo,hi) in enumerate(ROLES):
    rr=6+i
    C(wro.cell(row=rr,column=2),f"ROL-{i+1:03d}",bold=True); C(wro.cell(row=rr,column=3),tw)
    C(wro.cell(row=rr,column=4),title,bold=True); C(wro.cell(row=rr,column=5),cap)
    C(wro.cell(row=rr,column=6),sc,align="center"); C(wro.cell(row=rr,column=7),lo); C(wro.cell(row=rr,column=8),hi)
    lh=wro.cell(row=rr,column=9); lh.value=f"=VLOOKUP(G{rr},{COST_LH},5,FALSE)"; lh.font=Font(name=FONT,color=LINK_GREEN); lh.number_format=USD2; lh.alignment=Alignment(horizontal="right"); lh.border=BORDER
    sa=wro.cell(row=rr,column=10); sa.value=f"=I{rr}*VLOOKUP(F{rr},{SCAR_RANGE},2,FALSE)"; sa.font=Font(name=FONT); sa.number_format=USD2; sa.alignment=Alignment(horizontal="right"); sa.border=BORDER
wro.freeze_panes="B6"; wro.auto_filter.ref=f"B5:J{5+len(ROLES)}"

# ================= 10b. ROLE RATE CARD (role x eligible level) =================
def levels_between(lo, hi):
    i, j = LEVEL_NAMES.index(lo), LEVEL_NAMES.index(hi)
    return LEVEL_NAMES[min(i, j):max(i, j) + 1]
RATECARD = []
for tw, title, cap, sc, lo, hi in ROLES:
    for lvl in levels_between(lo, hi):
        RATECARD.append((tw, title, cap, lvl, sc))
wrc = wb.create_sheet("Role Rate Card"); wrc.sheet_view.showGridLines = False
widths(wrc, {"A":3,"B":12,"C":24,"D":30,"E":26,"F":16,"G":12,"H":16,"I":18,"J":18})
C(wrc["B2"], "ROLE RATE CARD — staffable, priced units (role x career level)", bold=True, size=13, color="102650")
C(wrc["B3"], "Each row is a staffable unit the estimator consumes. Loaded = internal cost; Indicative Bill = SI-T1 onshore market rate. All parametric.", color="625F58")
for col, t in zip("BCDEFGHIJ", ["ID","Tower","Role","Capability","Level","Scarcity","Loaded $/hr","Scarcity-Adj $/hr","Indic. Bill $/hr"]):
    H(wrc[f"{col}5"], t)
for i, (tw, title, cap, lvl, sc) in enumerate(RATECARD):
    rr = 6 + i
    C(wrc.cell(row=rr, column=2), f"RC-{i+1:04d}", bold=True); C(wrc.cell(row=rr, column=3), tw)
    C(wrc.cell(row=rr, column=4), title, bold=True); C(wrc.cell(row=rr, column=5), cap)
    C(wrc.cell(row=rr, column=6), lvl); C(wrc.cell(row=rr, column=7), sc, align="center")
    lh = wrc.cell(row=rr, column=8); lh.value=f"=VLOOKUP(F{rr},{COST_LH},5,FALSE)"; lh.font=Font(name=FONT,color=LINK_GREEN); lh.number_format=USD2; lh.alignment=Alignment(horizontal="right"); lh.border=BORDER
    sa = wrc.cell(row=rr, column=9); sa.value=f"=H{rr}*VLOOKUP(G{rr},{SCAR_RANGE},2,FALSE)"; sa.font=Font(name=FONT); sa.number_format=USD2; sa.alignment=Alignment(horizontal="right"); sa.border=BORDER
    bill = wrc.cell(row=rr, column=10); bill.value=f"=VLOOKUP(F{rr},{MKT_RANGE},2,FALSE)*VLOOKUP(\"SI-T1\",{TIER_RANGE},2,FALSE)*VLOOKUP(G{rr},{SCAR_RANGE},2,FALSE)"; bill.font=Font(name=FONT); bill.number_format=USD; bill.alignment=Alignment(horizontal="right"); bill.border=BORDER
wrc.freeze_panes="B6"; wrc.auto_filter.ref=f"B5:J{5+len(RATECARD)}"

# ================= 11. DELIVERY PODS =================
# (pod, tower, roles summary, headcount, blended level, agent mix, use cases)
PODS=[
("Data Product Pod","Data & Analytics","Data Product Mgr, Data Architect, 3x Data Engineer, BI Dev",6,"Senior","Agent Platform B + AbarVa Agents","Build governed data products end-to-end"),
("Snowflake Pod","Data & Analytics","Snowflake Architect, 3x Snowflake Engineer",4,"Senior","Agent Platform B","Snowflake platform build & migration"),
("Databricks Pod","Data & Analytics","Databricks Architect, 3x Databricks Engineer",4,"Senior","Agent Platform B","Lakehouse & ML pipelines"),
("Data Governance Pod","Data & Analytics","Governance Lead, 2x Steward, Metadata Lead",4,"Manager","Agent Platform A","Stand up governance & catalog"),
("MDM Pod","Data & Analytics","MDM Architect, 2x Data Engineer",3,"Lead","Agent Platform B","Master data domain build"),
("AI Platform Pod","AI & GenAI","AI Architect, 2x AI Engineer, MLOps Engineer",4,"Senior","Agent Platform D + AbarVa Agents","AI platform & MLOps foundation"),
("Agentic AI Pod","AI & GenAI","Agent Architect, FDE, 2x AI Engineer",4,"Senior","AbarVa Agents + Agent Platform D","Design & ship agentic workflows"),
("GenAI Pod","AI & GenAI","AI Engineer, Prompt Engineer, LLMOps",3,"Senior","Agent Platform D","GenAI app build"),
("Responsible AI Pod","AI & GenAI","Responsible AI Lead, Model Validation Lead, AI Governance Lead",3,"Manager","Agent Platform A","RAI controls, validation, governance"),
("Adobe Experience Pod","Marketing Technology","AEM Architect, 2x Adobe Developer, Journey Architect",4,"Senior","Agent Platform A","AEM experience delivery"),
("CDP Pod","Marketing Technology","CDP Architect, 2x Engineer",3,"Lead","Agent Platform A","Customer data platform build"),
("Workday HCM Pod","ERP","Workday Architect, Functional Lead, 2x Consultant",4,"Senior","Agent Platform C","Workday HCM implementation"),
("Workday Finance Pod","ERP","Workday Technical Lead, 2x Consultant",3,"Senior","Agent Platform C","Workday Finance implementation"),
("SAP S/4 Finance Pod","ERP","SAP Finance Architect, 2x Consultant, Basis",4,"Senior","Agent Platform C","SAP Finance transformation"),
("SAP Supply Chain Pod","ERP","SAP SCM Architect, 2x Consultant",3,"Senior","Agent Platform C","SAP supply chain"),
("Oracle Finance Pod","ERP","Oracle Finance Architect, 2x Consultant",3,"Senior","Agent Platform C","Oracle Finance"),
("ServiceNow ITSM Pod","Application Engineering","SN Architect, 2x Developer",3,"Senior","Agent Platform C","ServiceNow ITSM"),
("ServiceNow HRSD Pod","Application Engineering","SN Architect, 2x Developer",3,"Senior","Agent Platform C","ServiceNow HRSD"),
("Salesforce Pod","Application Engineering","SF Architect, 2x Developer",3,"Senior","Agent Platform C","Salesforce delivery"),
("Cloud Migration Pod","Cloud","Cloud Architect, 3x Platform Engineer, SRE",5,"Senior","Agent Platform B","Migrate workloads to cloud"),
("Platform Engineering Pod","Cloud","Platform Architect, 3x Platform Engineer",4,"Senior","Agent Platform B","Internal developer platform"),
("DevOps Pod","Cloud","DevOps Architect, 2x DevOps Engineer",3,"Senior","Agent Platform B","CI/CD & automation"),
("Cyber IAM Pod","Cybersecurity","IAM Architect, 2x Engineer",3,"Lead","Agent Platform A","Identity & access"),
("SOC Pod","Cybersecurity","SOC Lead, 3x Analyst",4,"Senior","Agent Platform C","Security operations"),
("AppSec Pod","Cybersecurity","AppSec Engineer, Security Testing Architect",2,"Lead","Agent Platform B","Application security"),
("Mainframe Modernization Pod","Legacy & Mainframe","Mainframe Architect, COBOL Architect, 2x COBOL Engineer",4,"Senior","AbarVa Agents + Agent Platform B","Modernize legacy estate"),
("AMS Pod","Managed Services","Delivery Mgr, 2x L3, 3x L2, 4x L1",10,"Intermediate","Agent Platform C","Application managed services"),
("Quality Engineering Pod","Quality Engineering","Automation Architect, 3x Automation Engineer",4,"Senior","Agent Platform B","Test automation at scale"),
("Integration Pod","Integration","Integration Architect, 2x iPaaS Dev, Event Engineer",4,"Senior","Agent Platform B","Enterprise integration"),
("Product Engineering Pod","Application Engineering","Eng Lead, 3x Engineer, Product Owner",5,"Senior","Agent Platform B + AbarVa Agents","Full-stack product squad"),
("Transformation Office Pod","Program Management","Program Director, 2x PM, PMO Analyst",4,"Manager","Agent Platform A","Run the transformation office"),
("Change & Adoption Pod","Change Management","OCM Lead, Training Lead, Comms Lead",3,"Manager","Agent Platform A","Drive adoption & change"),
("Enterprise Architecture Pod","Enterprise Architecture","Chief Architect, 2x Solution Architect",3,"Director","Agent Platform A","Architecture runway & standards"),
("Business Process Pod","Business Process","Process Lead, Mining Consultant, Automation Architect",3,"Manager","Agent Platform C","Process redesign & automation"),
# --- expansion ---
("Data Mesh Pod","Data & Analytics","Data Architect, 2x Data Product Mgr, 3x Data Engineer",6,"Senior","Agent Platform B + AbarVa Agents","Federated data mesh enablement"),
("Real-Time Analytics Pod","Data & Analytics","Streaming Engineer x2, Data Engineer, Analytics Engineer",4,"Senior","Agent Platform B","Streaming & real-time analytics"),
("Lakehouse Pod","Data & Analytics","Databricks Architect, 3x Databricks Engineer",4,"Senior","Agent Platform B","Lakehouse foundation"),
("Data Migration Pod","Data & Analytics","Data Architect, 4x Data Engineer",5,"Senior","AbarVa Agents + Agent Platform B","Legacy-to-cloud data migration"),
("Analytics Engineering Pod","Data & Analytics","Analytics Engineer x3, BI Developer",4,"Senior","Agent Platform B","dbt + semantic layer build"),
("BI Modernization Pod","Data & Analytics","BI Architect, 3x BI Developer",4,"Senior","Agent Platform A","Modernize legacy BI estate"),
("Data Privacy Pod","Data & Analytics","Data Privacy Lead, 2x Steward",3,"Manager","Agent Platform A","Privacy controls & DSAR automation"),
("RAG Application Pod","AI & GenAI","AI Architect, 2x RAG Engineer, Retrieval Engineer",4,"Senior","Agent Platform D + AbarVa Agents","Enterprise RAG applications"),
("ML Platform Pod","AI & GenAI","AI Platform Engineer x2, MLOps Engineer",3,"Senior","Agent Platform D","MLOps platform & pipelines"),
("Computer Vision Pod","AI & GenAI","CV Engineer x2, ML Engineer",3,"Senior","Agent Platform D","Vision model delivery"),
("NLP Pod","AI & GenAI","NLP Engineer x2, AI Engineer",3,"Senior","Agent Platform D","NLP & document intelligence"),
("AI Evaluation Pod","AI & GenAI","AI Evaluation Engineer, Model Validation Lead, AI Safety Engineer",3,"Lead","Agent Platform A","Model eval, red-team, validation"),
("AI Center of Excellence Pod","AI & GenAI","Chief AI Architect, AI Governance Lead, 2x AI Engineer",4,"Director","AbarVa Agents","Stand up AI CoE"),
("Forward-Deployed AI Pod","AI & GenAI","2x Forward Deployed Engineer, Agent Architect",3,"Senior","AbarVa Agents + Agent Platform D","Embed with client to ship agents"),
("Web Experience Pod","Digital Experience","UX Lead, 2x Frontend Engineer, UI Designer",4,"Senior","Agent Platform B","Web experience delivery"),
("Mobile App Pod","Digital Experience","iOS Engineer, Android Engineer, UX Designer",3,"Senior","Agent Platform B","Native mobile app build"),
("Design System Pod","Digital Experience","Design System Lead, 2x UI Designer",3,"Lead","Agent Platform A","Enterprise design system"),
("Commerce Build Pod","Digital Experience","Commerce Architect, 3x Engineer",4,"Senior","Agent Platform B","Commerce platform build"),
("Adobe Analytics Pod","Marketing Technology","Adobe Analytics Consultant, Target Specialist",2,"Senior","Agent Platform A","Adobe analytics & optimization"),
("Personalization Pod","Marketing Technology","Personalization Lead, CDP Engineer, MarTech Engineer",3,"Senior","Agent Platform A","Real-time personalization"),
("Product Squad Pod","Product Management","Product Owner, Eng Lead, 3x Engineer, Designer",6,"Senior","Agent Platform B + AbarVa Agents","Cross-functional product squad"),
("Microservices Pod","Application Engineering","Eng Lead, 4x Software Engineer",5,"Senior","Agent Platform B","Microservices build"),
("API Platform Pod","Application Engineering","Integration Architect, 3x API Engineer",4,"Senior","Agent Platform B","API platform & gateway"),
("App Modernization Pod","Application Engineering","Modernization Architect, 4x Engineer",5,"Senior","AbarVa Agents + Agent Platform B","Monolith-to-microservices"),
("ServiceNow CSM Pod","Application Engineering","ServiceNow Architect, 2x Developer",3,"Senior","Agent Platform C","ServiceNow CSM"),
("ServiceNow ITOM Pod","Application Engineering","ITOM Consultant, 2x Developer",3,"Senior","Agent Platform C","ServiceNow ITOM"),
("ServiceNow SecOps Pod","Application Engineering","SecOps Consultant, 2x Developer",3,"Senior","Agent Platform C","ServiceNow SecOps"),
("Salesforce Sales Cloud Pod","Application Engineering","SF Architect, Sales Cloud Consultant, 2x Developer",4,"Senior","Agent Platform C","Salesforce Sales Cloud"),
("Salesforce Service Cloud Pod","Application Engineering","SF Architect, Service Cloud Consultant, 2x Developer",4,"Senior","Agent Platform C","Salesforce Service Cloud"),
("Salesforce Platform Pod","Application Engineering","SF Architect, 2x Apex Developer, LWC Developer",4,"Senior","Agent Platform C","Salesforce custom platform"),
("MuleSoft Pod","Integration","MuleSoft Architect, 3x MuleSoft Developer",4,"Senior","Agent Platform B","MuleSoft integration"),
("Kafka Streaming Pod","Integration","Kafka Engineer x3",3,"Senior","Agent Platform B","Event streaming platform"),
("API Management Pod","Integration","Integration Architect, API Product Manager, 2x Developer",4,"Senior","Agent Platform B","API management & governance"),
("SAP S/4 Supply Chain Pod","ERP","SAP SCM Architect, MM Consultant, SD Consultant",3,"Senior","Agent Platform C","SAP supply chain transformation"),
("SAP BTP Pod","ERP","BTP Architect, 2x CPI Consultant",3,"Senior","Agent Platform C","SAP BTP & integration"),
("SAP ABAP/Fiori Pod","ERP","2x ABAP Developer, 2x Fiori Developer",4,"Senior","Agent Platform B","SAP custom dev"),
("SAP SuccessFactors Pod","ERP","SuccessFactors Consultant x2",2,"Senior","Agent Platform C","SAP HCM SuccessFactors"),
("Oracle Fusion HCM Pod","ERP","Oracle Fusion HCM Consultant x2, Integration",3,"Senior","Agent Platform C","Oracle Fusion HCM"),
("OCI Migration Pod","ERP","OCI Architect, 2x Engineer",3,"Senior","Agent Platform B","Oracle Cloud migration"),
("Workday Integration Pod","ERP","Workday Integration Consultant x2, Reporting Analyst",3,"Senior","Agent Platform C","Workday integrations & reporting"),
("AWS Migration Pod","Cloud","AWS Architect, 3x Platform Engineer, SRE",5,"Senior","Agent Platform B","AWS workload migration"),
("GCP Data Pod","Cloud","GCP Architect, 2x GCP Data Engineer",3,"Senior","Agent Platform B","GCP data platform"),
("Kubernetes Platform Pod","Cloud","Kubernetes Architect, 3x Kubernetes Engineer",4,"Senior","Agent Platform B","K8s platform build"),
("FinOps Pod","Cloud","FinOps Analyst x2, Cloud Architect",3,"Senior","Agent Platform A","Cloud cost optimization"),
("IaC Automation Pod","Cloud","IaC Engineer x3",3,"Senior","Agent Platform B","Infrastructure as code"),
("Landing Zone Pod","Cloud","Landing Zone Architect, 2x Platform Engineer",3,"Lead","Agent Platform B","Cloud landing zone"),
("SRE Pod","Cloud","SRE Architect, 3x SRE Engineer",4,"Senior","Agent Platform B","Reliability engineering"),
("Network Modernization Pod","Infrastructure","Network Architect, 3x Network Engineer",4,"Senior","Agent Platform A","Network modernization"),
("Identity Pod","Infrastructure","Identity Architect, 2x AD Engineer",3,"Lead","Agent Platform A","Identity & directory services"),
("DR & Backup Pod","Infrastructure","2x Backup/DR Engineer, Infra Architect",3,"Senior","Agent Platform A","DR & backup modernization"),
("Zero Trust Pod","Cybersecurity","Zero Trust Architect, 2x Cloud Security Engineer",3,"Lead","Agent Platform A","Zero trust rollout"),
("IR & SOC Surge Pod","Cybersecurity","SOC Lead, 2x Incident Responder, Threat Hunter",4,"Senior","Agent Platform C","Incident response surge"),
("GRC Pod","Cybersecurity","GRC Lead, 2x GRC Analyst",3,"Manager","Agent Platform A","Governance, risk & compliance"),
("Pen Test Pod","Cybersecurity","2x Penetration Tester, Security Architect",3,"Senior","Agent Platform B","Offensive security testing"),
("Privacy Pod","Cybersecurity","Privacy Engineer x2, GRC Analyst",3,"Senior","Agent Platform A","Data protection & privacy"),
("Performance Testing Pod","Quality Engineering","Performance Architect, 2x Engineer",3,"Senior","Agent Platform B","Performance & load testing"),
("SDET Pod","Quality Engineering","4x SDET",4,"Senior","Agent Platform B","Shift-left test automation"),
("Observability Pod","Operations","Observability Engineer x2, SRE",3,"Senior","Agent Platform B","Observability platform"),
("Incident Command Pod","Operations","Major Incident Lead, Incident Manager, Problem Manager",3,"Manager","Agent Platform C","Major incident management"),
("AMS Data Pod","Managed Services","AMS Architect, 2x L3, 3x L2",6,"Intermediate","Agent Platform C","Data platform managed services"),
("AMS Cloud Pod","Managed Services","AMS Architect, 2x L3 (Cloud), 3x L2",6,"Intermediate","Agent Platform C","Cloud managed services"),
("AMS ERP Pod","Managed Services","AMS Architect, 3x L3 (ERP), 4x L2",8,"Intermediate","Agent Platform C","ERP managed services"),
("AMS Salesforce Pod","Managed Services","AMS Lead, 2x L3, 3x L2",6,"Intermediate","Agent Platform C","Salesforce managed services"),
("AMS ServiceNow Pod","Managed Services","AMS Lead, 2x L3, 3x L2",6,"Intermediate","Agent Platform C","ServiceNow managed services"),
("COBOL-to-Java Pod","Legacy & Mainframe","Mainframe Architect, 4x COBOL Engineer",5,"Senior","AbarVa Agents + Agent Platform B","Automated COBOL conversion"),
("DB2 Modernization Pod","Legacy & Mainframe","DB2 DBA x2, Data Engineer",3,"Senior","Agent Platform B","DB2-to-cloud data modernization"),
("SAFe Release Train Pod","Program Management","Release Train Engineer, 3x Scrum Master, Agile Coach",5,"Senior","Agent Platform A","Scaled agile delivery"),
("Delivery Assurance Pod","Program Management","Program Director, 2x Delivery Lead",3,"Director","Agent Platform A","Delivery assurance & QA"),
("EA Runway Pod","Enterprise Architecture","Chief Architect, Business Architect, 2x Solution Architect",4,"Director","Agent Platform A","Architecture runway & standards"),
("Banking Data Pod","Industry SMEs","Banking SME, Data Architect, 2x Data Engineer",4,"Senior","Agent Platform B","Banking data accelerator"),
("Insurance Claims AI Pod","Industry SMEs","Insurance SME, AI Architect, 2x AI Engineer",4,"Senior","Agent Platform D + AbarVa Agents","Claims automation with AI"),
("Healthcare Interop Pod","Industry SMEs","Provider SME, Integration Architect, 2x Engineer",4,"Senior","Agent Platform B","FHIR/HL7 interoperability"),
("Payments Fraud AI Pod","Industry SMEs","Payments SME, AI Architect, 2x ML Engineer",4,"Senior","Agent Platform D","Real-time fraud detection"),
]
wp=wb.create_sheet("Delivery Pods"); wp.sheet_view.showGridLines=False
widths(wp,{"A":3,"B":12,"C":28,"D":22,"E":46,"F":12,"G":14,"H":26,"I":40})
C(wp["B2"],"DELIVERY POD LIBRARY",bold=True,size=13,color="102650")
C(wp["B3"],f"{len(PODS)} reusable human+agent pods. Est. monthly cost is parametric (headcount x blended loaded monthly).",color="625F58")
for col,t in zip("BCDEFGHI",["ID","Pod","Tower","Role Mix","Headcount","Blended Level","Agent Mix","Use Cases / Est. Monthly $"]):
    H(wp[f"{col}5"],t)
# monthly loaded = loaded hourly * hours/12... use loaded annual/12 via cost model VLOOKUP on loaded annual (col 4)
COST_LA=f"'Internal Cost Model'!$B$6:$F$15"
for i,(pod,tw,mix,hc,bl,am,uc) in enumerate(PODS):
    rr=6+i
    C(wp.cell(row=rr,column=2),f"POD-{i+1:03d}",bold=True); C(wp.cell(row=rr,column=3),pod,bold=True,fill=TOWER_FILL)
    C(wp.cell(row=rr,column=4),tw); C(wp.cell(row=rr,column=5),mix,wrap=True)
    C(wp.cell(row=rr,column=6),hc,align="center"); C(wp.cell(row=rr,column=7),bl,align="center")
    C(wp.cell(row=rr,column=8),am,wrap=True)
    # est monthly cost = headcount * (loaded annual for blended level / 12)
    cell=wp.cell(row=rr,column=9); cell.value=f"=F{rr}*VLOOKUP(G{rr},{COST_LA},4,FALSE)/12"
    cell.font=Font(name=FONT); cell.number_format=USD; cell.alignment=Alignment(horizontal="right"); cell.border=BORDER
    wp.row_dimensions[rr].height=30
wp.freeze_panes="B6"; wp.auto_filter.ref=f"B5:I{5+len(PODS)}"

# ================= 12. MOVES BINDING =================
wmb=wb.create_sheet("Moves Binding"); wmb.sheet_view.showGridLines=False
widths(wmb,{"A":3,"B":26,"C":30,"D":60})
C(wmb["B2"],"MOVES ENGINE BINDING MAP",bold=True,size=13,color="102650")
C(wmb["B3"],"How this taxonomy binds to the AbarVa Move estimation/deliverable engine — one economics source of truth.",color="625F58")
for col,t in zip("BCD",["Taxonomy Entity","Move Engine Entity","Binding / Use in business case + roadmap"]):
    H(wmb[f"{col}5"],t)
binds=[("Capability","WorkPackage / Capability scope","Each Move capability scope resolves to one+ taxonomy capabilities; drives WBS"),
 ("Role + Loaded $/hr","Resource line in effort model","Estimator multiplies WBS effort hours x role scarcity-adjusted rate"),
 ("Delivery Pod","Staffing model / squad","Pods seed the resource model with a proven human+agent mix"),
 ("Agent Economics","AI-native scenario","Agent multipliers compress effort hours; cost = agent annual + reduced human FTE"),
 ("Rate Intelligence","Sourcing / vendor benchmark","Provider archetypes benchmark build vs buy and onshore/near/off mix"),
 ("Geography","Delivery model","Geo mix sets blended rate and the onshore/nearshore/offshore split"),
 ("Internal Cost Model","Internal build option","Fully-loaded internal cost = the in-house delivery scenario"),
 ("Career Levels","Seniority mix","Level mix per WorkPackage drives blended cost and quality posture"),
 ("Scarcity / Agent-Amenability","Risk + AI-native compression","Scarcity raises cost/risk; agent-amenability sets traditional vs AI-native delta")]
for i,(a,b,d) in enumerate(binds):
    rr=6+i; C(wmb.cell(row=rr,column=2),a,bold=True); C(wmb.cell(row=rr,column=3),b); C(wmb.cell(row=rr,column=4),d,wrap=True); wmb.row_dimensions[rr].height=28
C(wmb.cell(row=6+len(binds)+1,column=2),"NOTE",bold=True,color="102650")
C(wmb.cell(row=6+len(binds)+1,column=3),"Workbook = editable reference substrate. The estimation/business-case/roadmap COMPUTE runs in-product, consuming these tables. Build that engine phase next.",wrap=True)
wmb.merge_cells(start_row=6+len(binds)+1,start_column=3,end_row=6+len(binds)+1,end_column=4)

# ================= 13. EXEC SUMMARY =================
wx=wb.create_sheet("Exec Summary"); wx.sheet_view.showGridLines=False
widths(wx,{"A":3,"B":36,"C":18,"D":40})
C(wx["B2"],"EXECUTIVE SUMMARY",bold=True,size=14,color="102650")
C(wx["B3"],"AbarVa Workforce Economics — Taxonomy Master v1.0",color="625F58")
H(wx["B5"],"Dimension"); H(wx["C5"],"Count"); H(wx["D5"],"Note")
summ=[("Delivery Towers",len(TOWERS),"Full enterprise transformation coverage"),
 ("Capabilities",len(CAPS),"With scarcity + agent-amenability"),
 ("Roles (distinct families)",len(ROLES),"Across all 21 towers"),
 ("Priced staffable units",len(RATECARD),"Role x eligible career level (Rate Card)"),
 ("Career Levels",len(LEVELS),"Partner → Junior"),
 ("Geographies",len(GEOS),"Onshore / nearshore / offshore"),
 ("Provider Archetypes",len(PROVIDERS),"Anonymized: CONS-T1, SI-T1, SI-T2, ENG-B, AI-B"),
 ("Agent Platforms",len(AGENTS),"Provider-neutral + AbarVa Agents"),
 ("Delivery Pods",len(PODS),"Reusable human+agent squads")]
for i,(n,c,note) in enumerate(summ):
    rr=6+i; C(wx.cell(row=rr,column=2),n,bold=True); C(wx.cell(row=rr,column=3),c,align="center",bold=True); C(wx.cell(row=rr,column=4),note,wrap=True)
r2=6+len(summ)+2
C(wx.cell(row=r2,column=2),"KEY MODEL OUTPUTS (live formulas)",bold=True,color="102650")
C(wx.cell(row=r2+1,column=2),"Loaded hourly — Senior (US baseline)")
o1=wx.cell(row=r2+1,column=3); o1.value="=VLOOKUP(\"Senior\",'Internal Cost Model'!$B$6:$F$15,5,FALSE)"; o1.font=Font(name=FONT,color=LINK_GREEN); o1.number_format=USD2; o1.alignment=Alignment(horizontal="center"); o1.border=BORDER
C(wx.cell(row=r2+2,column=2),"Offshore bill — SI-T2 Senior ($/hr)")
o2=wx.cell(row=r2+2,column=3); o2.value=f"=VLOOKUP(\"Senior\",{MKT_RANGE},2,FALSE)*VLOOKUP(\"SI-T2\",{TIER_RANGE},2,FALSE)*INDEX(Assumptions!$C${SHORE_ROW0}:$C${SHORE_ROW0+2},3)"; o2.font=Font(name=FONT,color=LINK_GREEN); o2.number_format=USD; o2.alignment=Alignment(horizontal="center"); o2.border=BORDER
C(wx.cell(row=r2+3,column=2),"Onshore bill — CONS-T1 Principal ($/hr)")
o3=wx.cell(row=r2+3,column=3); o3.value=f"=VLOOKUP(\"Principal\",{MKT_RANGE},2,FALSE)*VLOOKUP(\"CONS-T1\",{TIER_RANGE},2,FALSE)"; o3.font=Font(name=FONT,color=LINK_GREEN); o3.number_format=USD; o3.alignment=Alignment(horizontal="center"); o3.border=BORDER

# ================= 14. GLOSSARY =================
wgl=wb.create_sheet("Glossary"); wgl.sheet_view.showGridLines=False
widths(wgl,{"A":3,"B":28,"C":74})
C(wgl["B2"],"GLOSSARY",bold=True,size=13,color="102650")
H(wgl["B4"],"Term"); H(wgl["C4"],"Definition")
gl=[("Fully-loaded cost","Base salary plus all employer load (bonus, equity, benefits, taxes, overhead, bench)."),
 ("Loaded hourly","Fully-loaded annual cost / billable hours per year."),
 ("Provider archetype","Anonymized provider category; no real firm names (liability + neutrality)."),
 ("Shore","Onshore / nearshore / offshore delivery location class."),
 ("Scarcity","Talent-market scarcity tier (High/Med/Low) that adjusts rate and risk."),
 ("Agent-amenability","1-5 score of how compressible a capability is by agents/AI; drives AI-native delta."),
 ("Equiv Eng FTE","Human engineering capacity an agent platform approximates."),
 ("Productivity multiplier","Capacity gain vs human-only baseline (1.00 = none)."),
 ("AI-native delivery","People + agents + AI platforms + automation, vs people-only traditional delivery."),
 ("Pod","Reusable squad with a defined human+agent role mix and capacity.")]
for i,(t,d) in enumerate(gl):
    rr=5+i; C(wgl.cell(row=rr,column=2),t,bold=True); C(wgl.cell(row=rr,column=3),d,wrap=True)
wgl.freeze_panes="A5"

# ================= 15. ESTIMATION ENGINE (estimate twice — capacity model) =================
AGT="'Agent Economics'!$B$6:$K$10"   # B..K ; rel idx: Annual=4, EquivFTE=5, Util=6, Productivity=7
wee=wb.create_sheet("Estimation Engine"); wee.sheet_view.showGridLines=False
widths(wee,{"A":3,"B":32,"C":18,"D":18,"E":14})
C(wee["B2"],"ESTIMATION ENGINE — traditional vs AI-native (worked example)",bold=True,size=13,color="102650")
C(wee["B3"],"One scope, estimated twice. Agents add parallel capacity (faster) and are billed by subscription, not human rate (cheaper). No double-counting.",color="625F58",wrap=True)
wee.row_dimensions[3].height=28
def inp(cell,v,num=None): C(cell,v,color=INPUT_BLUE,num=num,align="right"); cell.fill=ASSUM_FILL
C(wee["B5"],"PROGRAM INPUTS",bold=True,color="102650")
C(wee["B6"],"Program"); C(wee["C6"],"Customer Data Product Program",color=INPUT_BLUE); wee["C6"].fill=ASSUM_FILL
C(wee["B7"],"Reference delivery pod"); C(wee["C7"],"Data Product Pod",color=INPUT_BLUE); wee["C7"].fill=ASSUM_FILL
C(wee["B8"],"Onshore mix %"); inp(wee["C8"],0.40,PCT)
C(wee["B9"],"Offshore mix %"); inp(wee["C9"],0.60,PCT)
C(wee["B10"],"Onshore rate (SI-T1 Senior) $/hr")
oc=wee["C10"]; oc.value=f'=VLOOKUP("Senior",{MKT_RANGE},2,FALSE)*VLOOKUP("SI-T1",{TIER_RANGE},2,FALSE)'; oc.font=Font(name=FONT,color=LINK_GREEN); oc.number_format=USD; oc.alignment=Alignment(horizontal="right"); oc.border=BORDER
C(wee["B11"],"Offshore rate (SI-T2 Senior) $/hr")
fc=wee["C11"]; fc.value=f'=VLOOKUP("Senior",{MKT_RANGE},2,FALSE)*VLOOKUP("SI-T2",{TIER_RANGE},2,FALSE)*INDEX(Assumptions!$C${SHORE_ROW0}:$C${SHORE_ROW0+2},3)'; fc.font=Font(name=FONT,color=LINK_GREEN); fc.number_format=USD; fc.alignment=Alignment(horizontal="right"); fc.border=BORDER
C(wee["B12"],"Blended rate $/hr",bold=True)
bc=wee["C12"]; bc.value="=C8*C10+C9*C11"; bc.font=Font(name=FONT,bold=True); bc.number_format=USD; bc.alignment=Alignment(horizontal="right"); bc.border=BORDER
C(wee["B13"],"AI agent platform"); C(wee["C13"],"AbarVa Agents",color=INPUT_BLUE); wee["C13"].fill=ASSUM_FILL
C(wee["B14"],"Agent equiv FTE / agent");
e14=wee["C14"]; e14.value=f"=VLOOKUP(C13,{AGT},5,FALSE)"; e14.font=Font(name=FONT,color=LINK_GREEN); e14.number_format=MULT; e14.alignment=Alignment(horizontal="right"); e14.border=BORDER
C(wee["B15"],"Agent utilization %")
e15=wee["C15"]; e15.value=f"=VLOOKUP(C13,{AGT},6,FALSE)"; e15.font=Font(name=FONT,color=LINK_GREEN); e15.number_format=PCT; e15.alignment=Alignment(horizontal="right"); e15.border=BORDER
C(wee["B16"],"Agent annual cost $")
e16=wee["C16"]; e16.value=f"=VLOOKUP(C13,{AGT},4,FALSE)"; e16.font=Font(name=FONT,color=LINK_GREEN); e16.number_format=USD; e16.alignment=Alignment(horizontal="right"); e16.border=BORDER
C(wee["B17"],"Hours / FTE / month"); inp(wee["C17"],173,'#,##0')
# effort scope table
C(wee["B19"],"EFFORT SCOPE (work breakdown)",bold=True,color="102650")
for col,t in zip("BCD",["Estimate Category","Effort Hours","Agent-Amenable"]):
    H(wee[f"{col}20"],t)
cats=[("Business Analysis",1200,"High"),("Architecture",1400,"Medium"),("Engineering",4500,"High"),
 ("Testing",2200,"High"),("Security",600,"Medium"),("Training",500,"High"),
 ("Change Management",500,"Low"),("Deployment",700,"Medium"),("Governance",400,"Low"),("Program Management",1000,"Low")]
r0=21
for i,(cat,hrs,am) in enumerate(cats):
    rr=r0+i; C(wee.cell(row=rr,column=2),cat); th=wee.cell(row=rr,column=3); inp(th,hrs,'#,##0')
    C(wee.cell(row=rr,column=4),am,align="center")
tr=r0+len(cats)
C(wee.cell(row=tr,column=2),"TOTAL EFFORT (hrs)",bold=True)
ct=wee.cell(row=tr,column=3); ct.value=f"=SUM(C{r0}:C{tr-1})"; ct.font=Font(name=FONT,bold=True); ct.number_format='#,##0'; ct.alignment=Alignment(horizontal="right"); ct.border=BORDER; ct.fill=SUB_FILL
# scenario comparison
sh=tr+2
C(wee.cell(row=sh,column=2),"SCENARIO COMPARISON",bold=True,color="102650")
H(wee.cell(row=sh+1,column=2),"Metric"); H(wee.cell(row=sh+1,column=3),"Traditional"); H(wee.cell(row=sh+1,column=4),"AI-Native")
TR=ct.coordinate  # total effort cell e.g. C31
def scen(off,label,trad,ai,num,bold=True):
    row=sh+2+off; C(wee.cell(row=row,column=2),label,bold=bold)
    for col,val in [(3,trad),(4,ai)]:
        if val is None or val=="-":
            C(wee.cell(row=row,column=col),"-" if val=="-" else "",align="right"); continue
        c=wee.cell(row=row,column=col); c.value=val; c.font=Font(name=FONT,bold=bold); c.number_format=num; c.alignment=Alignment(horizontal="right"); c.border=BORDER
    return row
# fixed layout: humans, agents, capacity, timeline, human hours, agent cost, total cost
R_team=sh+2; R_agents=sh+3; R_cap=sh+4; R_time=sh+5; R_hh=sh+6; R_ac=sh+7; R_cost=sh+8
C(wee.cell(row=R_team,column=2),"Team — humans (FTE)",bold=True)
inp(wee.cell(row=R_team,column=3),18,'#,##0'); inp(wee.cell(row=R_team,column=4),8,'#,##0')
C(wee.cell(row=R_agents,column=2),"AI agents (count)",bold=True)
C(wee.cell(row=R_agents,column=3),"-",align="right"); inp(wee.cell(row=R_agents,column=4),12,'#,##0')
def out2(row,label,ctrad,cai,num):
    C(wee.cell(row=row,column=2),label,bold=True)
    for col,val in [(3,ctrad),(4,cai)]:
        c=wee.cell(row=row,column=col)
        if val=="-": C(c,"-",align="right"); continue
        c.value=val; c.font=Font(name=FONT,bold=True); c.number_format=num; c.alignment=Alignment(horizontal="right"); c.border=BORDER; c.fill=SUB_FILL
out2(R_cap,"Effective capacity (FTE-equiv)",f"=C{R_team}",f"=D{R_team}+D{R_agents}*C14*C15",MULT)
out2(R_time,"Timeline (months)",f"={TR}/(C{R_cap}*C17)",f"={TR}/(D{R_cap}*C17)",'0.0')
out2(R_hh,"Human effort billed (hrs)",f"={TR}",f"=D{R_team}*D{R_time}*C17",'#,##0')
out2(R_ac,"Agent program cost $","-",f"=C16*D{R_time}/12",USD)
out2(R_cost,"TOTAL COST $",f"=C{R_hh}*C12",f"=D{R_hh}*C12+D{R_ac}",USD)
EE={"trad_hrs":f"'Estimation Engine'!$C${R_hh}","ai_hrs":f"'Estimation Engine'!$D${R_hh}",
 "trad_fte":f"'Estimation Engine'!$C${R_team}","ai_fte":f"'Estimation Engine'!$D${R_team}","agents":f"'Estimation Engine'!$D${R_agents}",
 "trad_m":f"'Estimation Engine'!$C${R_time}","ai_m":f"'Estimation Engine'!$D${R_time}",
 "trad_cost":f"'Estimation Engine'!$C${R_cost}","ai_cost":f"'Estimation Engine'!$D${R_cost}"}

# ================= 16. BUSINESS CASE =================
wbc=wb.create_sheet("Business Case"); wbc.sheet_view.showGridLines=False
widths(wbc,{"A":3,"B":34,"C":18,"D":18,"E":18})
C(wbc["B2"],"BUSINESS CASE — Traditional vs AI-Native",bold=True,size=13,color="102650")
C(wbc["B3"],"Pulls live from the Estimation Engine. Edit benefit & discount-rate inputs below.",color="625F58")
C(wbc["B5"],"INPUTS",bold=True,color="102650")
C(wbc["B6"],"Annual business benefit $"); inp(wbc["C6"],2400000,USD)
C(wbc["B7"],"Discount rate"); inp(wbc["C7"],0.10,PCT)
C(wbc["B8"],"Horizon (years)"); inp(wbc["C8"],3,'#,##0')
H(wbc["B10"],"Metric"); H(wbc["C10"],"Traditional"); H(wbc["D10"],"AI-Native"); H(wbc["E10"],"Delta")
def bc_row(row,label,trad,ai,delta,num):
    C(wbc.cell(row=row,column=2),label,bold=True)
    for col,val in [(3,trad),(4,ai),(5,delta)]:
        if val is None: continue
        c=wbc.cell(row=row,column=col); c.value=val; c.font=Font(name=FONT,color=(LINK_GREEN if col<5 else INK),bold=(col==5)); c.number_format=num; c.alignment=Alignment(horizontal="right"); c.border=BORDER
bc_row(11,"Total cost $",f"={EE['trad_cost']}",f"={EE['ai_cost']}",f"={EE['trad_cost']}-{EE['ai_cost']}",USD)
bc_row(12,"Cost reduction %",None,None,f"=({EE['trad_cost']}-{EE['ai_cost']})/{EE['trad_cost']}",PCT)
bc_row(13,"Timeline (months)",f"={EE['trad_m']}",f"={EE['ai_m']}",f"=({EE['trad_m']}-{EE['ai_m']})/{EE['trad_m']}",'0.0')
bc_row(14,"Effort (hours)",f"={EE['trad_hrs']}",f"={EE['ai_hrs']}",f"=({EE['trad_hrs']}-{EE['ai_hrs']})/{EE['trad_hrs']}",'#,##0')
bc_row(15,"Team (FTE)",f"={EE['trad_fte']}",f"={EE['ai_fte']}",f"={EE['trad_fte']}-{EE['ai_fte']}",'#,##0')
bc_row(16,"AI agents",None,f"={EE['agents']}",None,'#,##0')
bc_row(17,"Productivity gain (x)",None,None,f"={EE['trad_hrs']}/{EE['ai_hrs']}",MULT)
# value metrics
C(wbc["B19"],"VALUE",bold=True,color="102650")
def v_row(row,label,trad,ai,num):
    C(wbc.cell(row=row,column=2),label,bold=True)
    for col,val in [(3,trad),(4,ai)]:
        c=wbc.cell(row=row,column=col); c.value=val; c.font=Font(name=FONT); c.number_format=num; c.alignment=Alignment(horizontal="right"); c.border=BORDER
v_row(20,f"Benefit over horizon $","=C6*C8","=C6*C8",USD)
v_row(21,"ROI %",f"=(C6*C8-{EE['trad_cost']})/{EE['trad_cost']}",f"=(C6*C8-{EE['ai_cost']})/{EE['ai_cost']}",PCT)
v_row(22,"Payback (months)",f"={EE['trad_cost']}/(C6/12)",f"={EE['ai_cost']}/(C6/12)",'0.0')
v_row(23,"NPV $","=-"+EE['trad_cost']+"+C6/(1+C7)+C6/(1+C7)^2+C6/(1+C7)^3","=-"+EE['ai_cost']+"+C6/(1+C7)+C6/(1+C7)^2+C6/(1+C7)^3",USD)
C(wbc["B25"],"READ",bold=True,color="102650")
C(wbc["C25"],"AI-native cuts cost and timeline while improving ROI and payback. All figures trace to the Estimation Engine and the taxonomy substrate — editable and auditable.",wrap=True)
wbc.merge_cells("C25:E25"); wbc.row_dimensions[25].height=44

# order sheets
order=["Cover","Exec Summary","Assumptions","Towers","Capabilities","Career Levels","Roles","Role Rate Card","Geography",
       "Internal Cost Model","Rate Intelligence","Agent Economics","Delivery Pods",
       "Estimation Engine","Business Case","Moves Binding","Glossary"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

# tab colors
colors={"Cover":"102650","Exec Summary":"31765B","Assumptions":"A66A1F","Towers":"2F5EA8","Capabilities":"2F5EA8",
 "Career Levels":"2F5EA8","Roles":"2F5EA8","Role Rate Card":"2F5EA8","Geography":"2F5EA8","Internal Cost Model":"31765B",
 "Rate Intelligence":"31765B","Agent Economics":"31765B","Delivery Pods":"5B21B6",
 "Estimation Engine":"A66A1F","Business Case":"31765B","Moves Binding":"9E332E","Glossary":"625F58"}
for s in wb._sheets:
    if s.title in colors: s.sheet_properties.tabColor=colors[s.title]

out="/Users/anand/Downloads/Workforce_Taxonomy_Master.xlsx"
wb.save(out)
print("SAVED", out)
print("Towers",len(TOWERS),"Caps",len(CAPS),"Roles",len(ROLES),"Geos",len(GEOS),"Providers",len(PROVIDERS),"Agents",len(AGENTS),"Pods",len(PODS))
print("Rate rows",len(PROVIDERS)*len(LEVELS))
