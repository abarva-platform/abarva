#!/usr/bin/env python3

"""Land the current owner-facing source Excel workbook package without interpretation.

This is Layer 1 only. It inventories the current 14-workbook intake contract,
captures raw worksheet rows with hashes, inventories source-room extracts, and
reports mapping gaps. It does not mutate Azure, canonical objects, products, or
tenant inputs.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_PACKAGE_ZIP = Path("/Users/anand/Downloads/meridian-v2-2-2b-semantic-mapping-pilot-20260822-092150.zip")
DEFAULT_OUT_DIR = Path("reports/source-excel-raw-landing-2026-08-23")

EXPECTED_WORKBOOKS: dict[str, dict[str, str]] = {
    "00_Getting_Started_and_Review": {
        "workbook": "00_Getting_Started_and_Review.xlsx",
        "family": "review_governance",
        "owner": "AbarVa engagement lead and client PMO",
        "layer2_target": "review/intake_status_adapter",
    },
    "01_Strategy_and_Operating_Model": {
        "workbook": "01_Strategy_and_Operating_Model.xlsx",
        "family": "strategy_operating_model",
        "owner": "Executive sponsors and strategy office",
        "layer2_target": "strategy_operating_model_adapter",
    },
    "02_Operations_Workforce_and_KPIs": {
        "workbook": "02_Operations_Workforce_and_KPIs.xlsx",
        "family": "operations_workforce_kpis",
        "owner": "Operations, HR, and functional KPI owners",
        "layer2_target": "operations_workforce_kpi_adapter",
    },
    "03A_Health_Plan_Applications": {
        "workbook": "03A_Health_Plan_Applications.xlsx",
        "family": "health_plan_applications",
        "owner": "Health plan application and CMDB owners",
        "layer2_target": "application_portfolio_adapter",
    },
    "03B_Clinical_Applications": {
        "workbook": "03B_Clinical_Applications.xlsx",
        "family": "clinical_applications",
        "owner": "Clinical application and CMDB owners",
        "layer2_target": "application_portfolio_adapter",
    },
    "03C_Shared_Applications": {
        "workbook": "03C_Shared_Applications.xlsx",
        "family": "shared_applications",
        "owner": "Enterprise application and CMDB owners",
        "layer2_target": "application_portfolio_adapter",
    },
    "04_Data_Analytics_and_Reporting": {
        "workbook": "04_Data_Analytics_and_Reporting.xlsx",
        "family": "data_analytics_reporting",
        "owner": "Data, analytics, BI, and reporting platform owners",
        "layer2_target": "data_analytics_volumetric_adapter",
    },
    "05_Infrastructure_and_Hosting": {
        "workbook": "05_Infrastructure_and_Hosting.xlsx",
        "family": "infrastructure_hosting",
        "owner": "Infrastructure, cloud, data center, and hosting owners",
        "layer2_target": "infrastructure_hosting_adapter",
    },
    "06_Finance_Budget_and_Programs": {
        "workbook": "06_Finance_Budget_and_Programs.xlsx",
        "family": "finance_budget_programs",
        "owner": "IT finance, FP&A, PMO, and value office",
        "layer2_target": "finance_program_adapter",
    },
    "07_Risk_Controls_and_Compliance": {
        "workbook": "07_Risk_Controls_and_Compliance.xlsx",
        "family": "risk_controls_compliance",
        "owner": "GRC, security, risk, compliance, and audit owners",
        "layer2_target": "risk_control_adapter",
    },
    "08A_Vendors_and_Contracts": {
        "workbook": "08A_Vendors_and_Contracts.xlsx",
        "family": "vendors_contracts",
        "owner": "Vendor management, procurement, CLM, and legal operations",
        "layer2_target": "vendor_contract_adapter",
    },
    "08B_Commercial_Performance_and_Value": {
        "workbook": "08B_Commercial_Performance_and_Value.xlsx",
        "family": "commercial_performance_value",
        "owner": "Procurement, finance, vendor management, and service owners",
        "layer2_target": "commercial_performance_adapter",
    },
    "09A_AI_Portfolio_and_Governance": {
        "workbook": "09A_AI_Portfolio_and_Governance.xlsx",
        "family": "ai_portfolio_governance",
        "owner": "AI governance, architecture, security, and portfolio owners",
        "layer2_target": "ai_portfolio_governance_adapter",
    },
    "09B_AI_Usage_and_Value": {
        "workbook": "09B_AI_Usage_and_Value.xlsx",
        "family": "ai_usage_value",
        "owner": "M365, GitHub, ServiceNow, AI platform admins, finance, and business sponsors",
        "layer2_target": "ai_usage_value_adapter",
    },
}

LEDGER_FILES = {
    "SOURCE_TO_REVIEW_PACK_TRACEABILITY.csv",
    "SYNTHETIC_ID_REGISTRY.csv",
    "SYNTHETIC_PROVENANCE_LEDGER.csv",
    "SYNTHETIC_RELATIONSHIP_LEDGER.csv",
    "PILOT_GOLDEN_ASSERTIONS.csv",
    "SYNTHETIC_SCENARIO_REGISTRY.csv",
}

SOURCE_ROOM_PREFIX = "__synthetic_sources__/"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def normalize(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def read_csv_from_zip(zf: zipfile.ZipFile, name: str) -> list[dict[str, str]]:
    with zf.open(name) as handle:
        text = io.TextIOWrapper(handle, encoding="utf-8-sig", newline="")
        return list(csv.DictReader(text))


def package_root(names: list[str]) -> str:
    roots = [name.split("/", 1)[0] for name in names if "/" in name]
    if not roots:
        return ""
    [(root, _count)] = Counter(roots).most_common(1)
    return f"{root}/"


def find_workbook_members(names: list[str], root: str) -> dict[str, str]:
    found: dict[str, str] = {}
    for folder, spec in EXPECTED_WORKBOOKS.items():
        expected = f"{root}{folder}/{spec['workbook']}"
        if expected in names:
            found[folder] = expected
    return found


def detect_repeated_nonblank_rows(row_values: list[list[str]]) -> int:
    counts: Counter[tuple[str, ...]] = Counter()
    for values in row_values:
        compact = tuple(v for v in values if v)
        if len(compact) >= 4:
            counts[tuple(values)] += 1
    return sum(count - 1 for count in counts.values() if count > 1)


def land_workbook(
    zf: zipfile.ZipFile,
    member: str,
    folder: str,
    package_sha: str,
    workbook_rows: list[dict[str, Any]],
    sheet_rows: list[dict[str, Any]],
    row_jsonl: io.TextIOBase,
) -> dict[str, Any]:
    spec = EXPECTED_WORKBOOKS[folder]
    data = zf.read(member)
    workbook_sha = sha256_bytes(data)
    wb = load_workbook(io.BytesIO(data), data_only=False, read_only=True)

    total_rows = 0
    total_nonblank_rows = 0
    total_cells = 0
    total_nonblank_cells = 0
    duplicate_nonblank_rows = 0
    sheet_count = len(wb.sheetnames)

    for sheet_index, ws in enumerate(wb.worksheets, start=1):
        max_row = ws.max_row or 0
        max_column = ws.max_column or 0
        observed_rows = 0
        sheet_nonblank_rows = 0
        sheet_nonblank_cells = 0
        row_values_for_dupe: list[list[str]] = []

        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            observed_rows = row_number
            values = [normalize(value) for value in row]
            while values and values[-1] == "":
                values.pop()
            nonblank_cells = sum(1 for value in values if value)
            if nonblank_cells:
                sheet_nonblank_rows += 1
                sheet_nonblank_cells += nonblank_cells
                row_values_for_dupe.append(values)
            row_hash = sha256_bytes(
                json.dumps(
                    {
                        "package_sha256": package_sha,
                        "workbook_sha256": workbook_sha,
                        "member": member,
                        "sheet": ws.title,
                        "row_number": row_number,
                        "values": values,
                    },
                    ensure_ascii=False,
                    sort_keys=True,
                ).encode("utf-8")
            )
            row_jsonl.write(
                json.dumps(
                    {
                        "package_sha256": package_sha,
                        "workbook_sha256": workbook_sha,
                        "workbook_folder": folder,
                        "workbook_name": spec["workbook"],
                        "family_id": spec["family"],
                        "source_owner": spec["owner"],
                        "layer2_target": spec["layer2_target"],
                        "sheet_index": sheet_index,
                        "sheet_name": ws.title,
                        "row_number": row_number,
                        "nonblank_cell_count": nonblank_cells,
                        "row_hash": row_hash,
                        "values": values,
                    },
                    ensure_ascii=False,
                    sort_keys=True,
                )
                + "\n"
            )

        if max_row == 0:
            max_row = observed_rows
        sheet_duplicate_rows = detect_repeated_nonblank_rows(row_values_for_dupe)
        duplicate_nonblank_rows += sheet_duplicate_rows
        total_rows += max_row
        total_cells += max_row * max_column
        total_nonblank_rows += sheet_nonblank_rows
        total_nonblank_cells += sheet_nonblank_cells

        sheet_rows.append(
            {
                "workbook_folder": folder,
                "workbook_name": spec["workbook"],
                "family_id": spec["family"],
                "sheet_index": sheet_index,
                "sheet_name": ws.title,
                "max_row": max_row,
                "max_column": max_column,
                "nonblank_rows": sheet_nonblank_rows,
                "nonblank_cells": sheet_nonblank_cells,
                "duplicate_nonblank_rows": sheet_duplicate_rows,
                "layer2_target": spec["layer2_target"],
            }
        )

    workbook_rows.append(
        {
            "workbook_folder": folder,
            "workbook_name": spec["workbook"],
            "zip_member": member,
            "family_id": spec["family"],
            "source_owner": spec["owner"],
            "layer2_target": spec["layer2_target"],
            "sha256": workbook_sha,
            "sheet_count": sheet_count,
            "declared_rows": total_rows,
            "nonblank_rows": total_nonblank_rows,
            "declared_cells": total_cells,
            "nonblank_cells": total_nonblank_cells,
            "duplicate_nonblank_rows": duplicate_nonblank_rows,
            "landing_state": "landed_raw_only",
        }
    )
    return {
        "folder": folder,
        "workbook": spec["workbook"],
        "sha256": workbook_sha,
        "sheet_count": sheet_count,
        "declared_rows": total_rows,
        "nonblank_rows": total_nonblank_rows,
    }


def land_source_room(
    zf: zipfile.ZipFile,
    names: list[str],
    root: str,
    package_sha: str,
    out_dir: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    inventory: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []
    source_members = sorted(
        name
        for name in names
        if name.startswith(f"{root}{SOURCE_ROOM_PREFIX}") and name.lower().endswith(".csv")
    )
    for member in source_members:
        data = zf.read(member)
        source_sha = sha256_bytes(data)
        csv_rows = read_csv_from_zip(zf, member)
        rel = member.removeprefix(root)
        source_room_family = rel.split("/", 2)[1] if "/" in rel else ""
        inventory.append(
            {
                "source_member": rel,
                "source_room_family": source_room_family,
                "sha256": source_sha,
                "row_count": len(csv_rows),
                "column_count": len(csv_rows[0]) if csv_rows else 0,
                "columns": "|".join(csv_rows[0].keys()) if csv_rows else "",
                "landing_state": "landed_raw_extract_only",
            }
        )
        for index, row in enumerate(csv_rows, start=2):
            normalized_row = {key: normalize(value) for key, value in row.items()}
            raw_hash = sha256_bytes(
                json.dumps(
                    {
                        "package_sha256": package_sha,
                        "source_sha256": source_sha,
                        "source_member": rel,
                        "row_number": index,
                        "values": normalized_row,
                    },
                    ensure_ascii=False,
                    sort_keys=True,
                ).encode("utf-8")
            )
            rows.append(
                {
                    "source_member": rel,
                    "source_room_family": source_room_family,
                    "row_number": index,
                    "row_hash": raw_hash,
                    "values_json": json.dumps(normalized_row, ensure_ascii=False, sort_keys=True),
                }
            )

    write_csv(
        out_dir / "source_room_inventory.csv",
        inventory,
        ["source_member", "source_room_family", "sha256", "row_count", "column_count", "columns", "landing_state"],
    )
    write_csv(
        out_dir / "source_room_rows.csv",
        rows,
        ["source_member", "source_room_family", "row_number", "row_hash", "values_json"],
    )
    return inventory, rows


def ledger_inventory(zf: zipfile.ZipFile, names: list[str], root: str) -> tuple[list[dict[str, Any]], dict[str, int]]:
    rows: list[dict[str, Any]] = []
    counts: dict[str, int] = {}
    for filename in sorted(LEDGER_FILES):
        member = f"{root}{filename}"
        if member not in names:
            rows.append({"ledger_file": filename, "present": "no", "row_count": 0, "sha256": "", "columns": ""})
            counts[filename] = 0
            continue
        data = zf.read(member)
        csv_rows = read_csv_from_zip(zf, member)
        rows.append(
            {
                "ledger_file": filename,
                "present": "yes",
                "row_count": len(csv_rows),
                "sha256": sha256_bytes(data),
                "columns": "|".join(csv_rows[0].keys()) if csv_rows else "",
            }
        )
        counts[filename] = len(csv_rows)
    return rows, counts


def add_gap(gaps: list[dict[str, Any]], severity: str, rule_id: str, subject: str, detail: str) -> None:
    gaps.append(
        {
            "severity": severity,
            "rule_id": rule_id,
            "subject": subject,
            "detail": detail,
        }
    )


def build_report(summary: dict[str, Any], workbook_rows: list[dict[str, Any]], gaps: list[dict[str, Any]]) -> str:
    lines = [
        "# Source Excel Raw Landing Report",
        "",
        "This is a Layer 1 raw landing proof for the current 14 owner-facing source workbooks. It does not normalize to canonical objects, load Azure, rebuild cubes, or prove product routes.",
        "",
        "## Summary",
        "",
        f"- Package: `{summary['package_path']}`",
        f"- Package SHA-256: `{summary['package_sha256']}`",
        f"- Expected workbook contract: {summary['expected_workbooks']} workbooks",
        f"- Landed workbooks: {summary['landed_workbooks']}",
        f"- Workbook sheets: {summary['workbook_sheets']}",
        f"- Workbook declared rows: {summary['workbook_declared_rows']}",
        f"- Workbook nonblank rows: {summary['workbook_nonblank_rows']}",
        f"- Source-room extracts: {summary['source_room_extracts']}",
        f"- Source-room rows: {summary['source_room_rows']}",
        f"- Traceability rows: {summary['traceability_rows']}",
        f"- Blocking gaps: {summary['blocking_gaps']}",
        "",
        "## Workbook Contract",
        "",
        "| Workbook | Family | Owner | Rows | Sheets | State |",
        "|---|---|---|---:|---:|---|",
    ]
    for row in workbook_rows:
        lines.append(
            f"| `{row['workbook_name']}` | `{row['family_id']}` | {row['source_owner']} | {row['nonblank_rows']} | {row['sheet_count']} | {row['landing_state']} |"
        )
    lines.extend(
        [
            "",
            "## Gaps",
            "",
        ]
    )
    if not gaps:
        lines.append("No blocking raw-landing gaps detected.")
    else:
        lines.extend(["| Severity | Rule | Subject | Detail |", "|---|---|---|---|"])
        for gap in gaps:
            lines.append(f"| {gap['severity']} | `{gap['rule_id']}` | `{gap['subject']}` | {gap['detail']} |")
    lines.extend(
        [
            "",
            "## What This Proves",
            "",
            "- The current owner-facing workbook contract is visible as raw source.",
            "- Every workbook row is hash-addressed before interpretation.",
            "- Source-room CSV extracts and traceability ledgers are inventoried separately from workbooks.",
            "- Unknowns and thin families remain explicit; this report is not a product proof.",
            "",
            "## Next Layer",
            "",
            "Layer 2 adapters must consume this raw landing and emit canonical objects with upstream workbook/source-room lineage. Product projections and cubes remain downstream only.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--package-zip", type=Path, default=DEFAULT_PACKAGE_ZIP)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    package_zip = args.package_zip.expanduser().resolve()
    out_dir = args.out_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    if not package_zip.exists():
        print(f"Package not found: {package_zip}", file=sys.stderr)
        return 2

    package_bytes = package_zip.read_bytes()
    package_sha = sha256_bytes(package_bytes)
    gaps: list[dict[str, Any]] = []
    workbook_rows: list[dict[str, Any]] = []
    sheet_rows: list[dict[str, Any]] = []

    with zipfile.ZipFile(io.BytesIO(package_bytes)) as zf:
        names = zf.namelist()
        root = package_root(names)
        workbook_members = find_workbook_members(names, root)

        missing = sorted(set(EXPECTED_WORKBOOKS) - set(workbook_members))
        extra_current_like = sorted(
            name.removeprefix(root)
            for name in names
            if name.lower().endswith(".xlsx")
            and not any(name == member for member in workbook_members.values())
            and not name.startswith(f"{root}{SOURCE_ROOM_PREFIX}")
        )
        for folder in missing:
            add_gap(gaps, "blocker", "missing_current_workbook", folder, f"Expected {EXPECTED_WORKBOOKS[folder]['workbook']}")
        for member in extra_current_like:
            add_gap(gaps, "warn", "extra_workbook_in_package", member, "Workbook is outside the current 14-workbook contract.")

        row_landing_path = out_dir / "source_excel_row_landing.jsonl"
        with row_landing_path.open("w", encoding="utf-8") as row_jsonl:
            for folder in sorted(workbook_members):
                land_workbook(zf, workbook_members[folder], folder, package_sha, workbook_rows, sheet_rows, row_jsonl)

        source_inventory, source_rows = land_source_room(zf, names, root, package_sha, out_dir)
        ledger_rows, ledger_counts = ledger_inventory(zf, names, root)

    traceability_columns = ""
    traceability_path = next((row for row in ledger_rows if row["ledger_file"] == "SOURCE_TO_REVIEW_PACK_TRACEABILITY.csv"), None)
    if traceability_path:
        traceability_columns = traceability_path.get("columns", "")
    required_traceability_columns = {"source_field", "target_field_code", "target_business_label", "mapping_rule_id", "raw_value", "normalized_value"}
    actual_traceability_columns = set(traceability_columns.split("|")) if traceability_columns else set()
    if not required_traceability_columns.issubset(actual_traceability_columns):
        add_gap(
            gaps,
            "blocker",
            "traceability_not_field_level",
            "SOURCE_TO_REVIEW_PACK_TRACEABILITY.csv",
            "Missing " + ", ".join(sorted(required_traceability_columns - actual_traceability_columns)),
        )

    for row in workbook_rows:
        if int(row["nonblank_rows"]) == 0:
            add_gap(gaps, "blocker", "empty_workbook", row["workbook_name"], "Workbook has no nonblank rows.")
        if int(row["duplicate_nonblank_rows"]) > 0:
            add_gap(
                gaps,
                "warn",
                "duplicate_nonblank_rows",
                row["workbook_name"],
                f"{row['duplicate_nonblank_rows']} repeated nonblank worksheet rows detected; review for duplicated headers or boilerplate.",
            )

    if len(source_inventory) < 12:
        add_gap(gaps, "warn", "thin_source_room", "__synthetic_sources__", f"{len(source_inventory)} CSV extracts found; expected at least 12 current source-room playbooks.")
    if len(source_rows) < 200:
        add_gap(
            gaps,
            "warn",
            "thin_source_room_rows",
            "__synthetic_sources__",
            f"{len(source_rows)} source-room rows found; enough for mapping proof, not enough for dense current-state simulation.",
        )
    for source in source_inventory:
        if int(source["row_count"]) < 10:
            add_gap(
                gaps,
                "warn",
                "thin_source_extract",
                source["source_member"],
                f"{source['row_count']} rows; needs denser extract simulation before claiming production-depth source coverage.",
            )

    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "package_path": package_zip.as_posix(),
        "package_sha256": package_sha,
        "expected_workbooks": len(EXPECTED_WORKBOOKS),
        "landed_workbooks": len(workbook_rows),
        "workbook_sheets": sum(int(row["sheet_count"]) for row in workbook_rows),
        "workbook_declared_rows": sum(int(row["declared_rows"]) for row in workbook_rows),
        "workbook_nonblank_rows": sum(int(row["nonblank_rows"]) for row in workbook_rows),
        "workbook_nonblank_cells": sum(int(row["nonblank_cells"]) for row in workbook_rows),
        "source_room_extracts": len(source_inventory),
        "source_room_rows": len(source_rows),
        "traceability_rows": ledger_counts.get("SOURCE_TO_REVIEW_PACK_TRACEABILITY.csv", 0),
        "ledger_rows": ledger_counts,
        "blocking_gaps": sum(1 for gap in gaps if gap["severity"] == "blocker"),
        "warning_gaps": sum(1 for gap in gaps if gap["severity"] == "warn"),
        "raw_landing_row_jsonl": row_landing_path.as_posix(),
    }

    write_csv(
        out_dir / "source_excel_workbook_inventory.csv",
        workbook_rows,
        [
            "workbook_folder",
            "workbook_name",
            "zip_member",
            "family_id",
            "source_owner",
            "layer2_target",
            "sha256",
            "sheet_count",
            "declared_rows",
            "nonblank_rows",
            "declared_cells",
            "nonblank_cells",
            "duplicate_nonblank_rows",
            "landing_state",
        ],
    )
    write_csv(
        out_dir / "source_excel_sheet_inventory.csv",
        sheet_rows,
        [
            "workbook_folder",
            "workbook_name",
            "family_id",
            "sheet_index",
            "sheet_name",
            "max_row",
            "max_column",
            "nonblank_rows",
            "nonblank_cells",
            "duplicate_nonblank_rows",
            "layer2_target",
        ],
    )
    write_csv(out_dir / "source_excel_gap_register.csv", gaps, ["severity", "rule_id", "subject", "detail"])
    write_csv(out_dir / "source_excel_ledger_inventory.csv", ledger_rows, ["ledger_file", "present", "row_count", "sha256", "columns"])
    (out_dir / "source_excel_raw_landing_summary.json").write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    (out_dir / "SOURCE_EXCEL_RAW_LANDING_REPORT.md").write_text(build_report(summary, workbook_rows, gaps), encoding="utf-8")

    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0 if summary["blocking_gaps"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
