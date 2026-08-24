#!/usr/bin/env python3

"""Build a local ECL pilot load SQL file from the applications/data/infra workbook.

Design artifact only. This script does not connect to Azure or any shared database.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


TENANT_KEY = "meridian-health"
ASSESSMENT_ID = "assessment-workbook-builder-pilot"
SOURCE_HASH_LABEL = "workbook-builder-pilot-source"
DEFAULT_WORKBOOK = Path(
    "outputs/enterprise-apps-data-infra-intake-rebuild-2026-08-22/"
    "Meridian_Applications_Data_Infrastructure_Intake_SYNTHETIC_v1.xlsx"
)
DEFAULT_OUT_DIR = Path("outputs/ecl-workbook-builder-pilot-2026-08-22")
SOURCE_SHEETS = [
    "01_ENTERPRISE_DA",
    "02_HEALTH_PLAN_APPS",
    "03_PROVIDER_CLINICAL_APPS",
    "04_SHARED_ENTERPRISE_APPS",
    "05_INFRASTRUCTURE_HOSTING",
    "06_MAJOR_INTEGRATIONS",
]
APPLICATION_SHEETS = [
    "02_HEALTH_PLAN_APPS",
    "03_PROVIDER_CLINICAL_APPS",
    "04_SHARED_ENTERPRISE_APPS",
]
ENVIRONMENT_SUFFIX_RE = re.compile(
    r"\s+(?:—|-)\s+(Production|Prod|Test|Testing|Training|Train|Development|Dev|QA|UAT|Staging|Sandbox|DR|Disaster Recovery)$",
    re.IGNORECASE,
)
ENVIRONMENT_LABELS = {
    "prod": "production",
    "production": "production",
    "test": "test",
    "testing": "test",
    "training": "training",
    "train": "training",
    "development": "development",
    "dev": "development",
    "qa": "qa",
    "uat": "uat",
    "staging": "staging",
    "sandbox": "sandbox",
    "dr": "disaster_recovery",
    "disaster recovery": "disaster_recovery",
}


def stable_uuid(*parts: object) -> str:
    digest = bytearray(hashlib.sha256("|".join(str(p) for p in parts).encode("utf-8")).digest()[:16])
    digest[6] = (digest[6] & 0x0F) | 0x40
    digest[8] = (digest[8] & 0x3F) | 0x80
    return str(uuid.UUID(bytes=bytes(digest)))


def clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def slug(value: object, fallback: str) -> str:
    text = re.sub(r"[^A-Z0-9]+", "-", str(value or "").upper().replace("&", " AND "))
    text = text.strip("-")[:72]
    return text or fallback


def as_num(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    matches = re.findall(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    if not matches:
        return None
    return float(matches[-1])


def sql_text(value: object) -> str:
    if value is None or value == "":
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_num(value: object) -> str:
    if value is None:
        return "null"
    return str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)


def sql_json(value: object) -> str:
    return f"{sql_text(json.dumps(value, ensure_ascii=False, default=str))}::jsonb"


def insert(table: str, columns: list[str], rows: list[dict[str, str]]) -> str:
    if not rows:
        return ""
    body = ",\n".join("(" + ", ".join(row[column] for column in columns) + ")" for row in rows)
    return f"insert into {table} ({', '.join(columns)}) values\n{body};\n"


def object_row(
    *,
    object_id: str,
    object_key: str,
    object_type: str,
    display_name: str,
    business_domain: str | None,
    lifecycle_state: str,
    source_record_id: str,
    attributes: dict[str, object],
) -> dict[str, str]:
    return {
        "id": sql_text(object_id),
        "tenant_key": sql_text(TENANT_KEY),
        "assessment_id": sql_text(ASSESSMENT_ID),
        "object_key": sql_text(object_key),
        "object_type": sql_text(object_type),
        "display_name": sql_text(display_name),
        "business_domain": sql_text(business_domain),
        "lifecycle_state": sql_text(lifecycle_state),
        "source_record_id": sql_text(source_record_id),
        "basis": sql_text("source_recorded"),
        "value_state": sql_text("known"),
        "review_state": sql_text("not_reviewed"),
        "confidence": "null",
        "attributes_json": sql_json(attributes),
    }


def measure_row(
    *,
    measure_id: str,
    subject_object_id: str,
    metric_key: str,
    value_number: float | None,
    value_text: str | None,
    unit: str,
    source_record_id: str,
) -> dict[str, str]:
    return {
        "id": sql_text(measure_id),
        "tenant_key": sql_text(TENANT_KEY),
        "assessment_id": sql_text(ASSESSMENT_ID),
        "subject_object_id": sql_text(subject_object_id),
        "metric_key": sql_text(metric_key),
        "value_number": sql_num(value_number),
        "value_text": sql_text(value_text),
        "unit": sql_text(unit),
        "period_start": "null",
        "period_end": "null",
        "scenario": sql_text("current"),
        "source_record_id": sql_text(source_record_id),
        "document_extraction_id": "null",
        "basis": sql_text("source_recorded"),
        "value_state": sql_text("estimated"),
        "quality_state": sql_text("estimated"),
        "review_state": sql_text("not_reviewed"),
        "attributes_json": sql_json({"generated_by": "scripts/ecl/build_workbook_pilot_ecl_sql.py"}),
    }


def parse_workbook(workbook_path: Path) -> dict[str, list[tuple[int, dict[str, object]]]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    parsed: dict[str, list[tuple[int, dict[str, object]]]] = {}
    for sheet_name in SOURCE_SHEETS:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"Missing required worksheet {sheet_name}")
        sheet = workbook[sheet_name]
        header_values = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        headers = [clean(value) for value in header_values]
        rows: list[tuple[int, dict[str, object]]] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            row = {
                header: value
                for header, value in zip(headers, values)
                if header and value not in (None, "")
            }
            if row:
                rows.append((row_number, row))
        parsed[sheet_name] = rows
    return parsed


def row_value(row: dict[str, object], key: str) -> str | None:
    return clean(row.get(key))


def split_environment_variant(app_name: str) -> tuple[str, str | None]:
    match = ENVIRONMENT_SUFFIX_RE.search(app_name)
    if not match:
        return app_name, None
    base_name = app_name[: match.start()].strip()
    environment = ENVIRONMENT_LABELS[match.group(1).lower()]
    return base_name, environment


def app_lifecycle(row: dict[str, object]) -> str:
    lifecycle = (row_value(row, "Lifecycle Status") or "").lower()
    replacement = (row_value(row, "Replacement / Retirement Planned?") or "").lower()
    if "retired" in lifecycle:
        return "retired"
    if "sunset" in lifecycle or replacement == "yes":
        return "candidate"
    return "current"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    workbook_path = args.workbook.resolve()
    out_dir = args.out_dir.resolve()
    workbook_bytes = workbook_path.read_bytes()
    workbook_hash = hashlib.sha256(workbook_bytes).hexdigest()
    parsed = parse_workbook(workbook_path)

    source_file_id = stable_uuid("source_file", workbook_hash)
    source_records: list[dict[str, str]] = []
    source_record_by_sheet_row: dict[tuple[str, int], str] = {}
    for sheet_name, rows in parsed.items():
        for row_number, row in rows:
            native_id = (
                row_value(row, "Application ID")
                or row_value(row, "Infrastructure / Hosting Estate ID")
                or row_value(row, "Integration ID")
                or f"{sheet_name}:{row_number}"
            )
            source_record_id = stable_uuid("source_record", sheet_name, native_id, row_number)
            source_record_by_sheet_row[(sheet_name, row_number)] = source_record_id
            source_records.append(
                {
                    "id": sql_text(source_record_id),
                    "tenant_key": sql_text(TENANT_KEY),
                    "assessment_id": sql_text(ASSESSMENT_ID),
                    "source_file_id": sql_text(source_file_id),
                    "native_id": sql_text(f"{sheet_name}:{native_id}"),
                    "record_type": sql_text(f"workbook_{sheet_name.lower()}"),
                    "row_number": sql_num(row_number),
                    "payload_json": sql_json(row),
                    "parse_state": sql_text("parsed"),
                    "parse_notes": "null",
                }
            )

    objects: dict[str, dict[str, str]] = {}
    relationships: list[dict[str, str]] = []
    relationship_keys: set[tuple[str, str, str]] = set()
    measures: list[dict[str, str]] = []
    app_object_by_name: dict[str, str] = {}
    function_object_by_name: dict[str, str] = {}
    vendor_object_by_name: dict[str, str] = {}
    infra_object_by_id: dict[str, str] = {}
    data_platform_object_by_name: dict[str, str] = {}
    deployment_object_by_name: dict[str, str] = {}
    environment_variant_rows = 0

    def add_relationship(from_id: str | None, rel_type: str, to_id: str | None, source_record_id: str, attrs: dict[str, object]) -> None:
        if not from_id or not to_id or from_id == to_id:
            return
        key = (from_id, rel_type, to_id)
        if key in relationship_keys:
            return
        relationship_keys.add(key)
        relationships.append(
            {
                "id": sql_text(stable_uuid("relationship", *key)),
                "tenant_key": sql_text(TENANT_KEY),
                "assessment_id": sql_text(ASSESSMENT_ID),
                "from_object_id": sql_text(from_id),
                "relationship_type": sql_text(rel_type),
                "to_object_id": sql_text(to_id),
                "direction_label": sql_text(rel_type.lower()),
                "source_record_id": sql_text(source_record_id),
                "basis": sql_text("source_recorded"),
                "value_state": sql_text("known"),
                "review_state": sql_text("not_reviewed"),
                "confidence": "null",
                "attributes_json": sql_json(attrs),
            }
        )

    def add_function(name: str | None, source_record_id: str, business_group: str | None) -> str | None:
        if not name:
            return None
        if name in function_object_by_name:
            return function_object_by_name[name]
        object_id = stable_uuid("object", "business_function", name)
        function_object_by_name[name] = object_id
        objects[object_id] = object_row(
            object_id=object_id,
            object_key=f"FUNC-{slug(name, 'FUNCTION')}",
            object_type="business_function",
            display_name=name,
            business_domain=business_group or name,
            lifecycle_state="current",
            source_record_id=source_record_id,
            attributes={"business_group": business_group, "workbook_source": "applications_data_infrastructure_intake"},
        )
        return object_id

    def add_vendor(name: str | None, source_record_id: str) -> str | None:
        if not name or name.lower() == "unknown":
            return None
        if name in vendor_object_by_name:
            return vendor_object_by_name[name]
        object_id = stable_uuid("object", "vendor", name)
        vendor_object_by_name[name] = object_id
        objects[object_id] = object_row(
            object_id=object_id,
            object_key=f"VEN-{slug(name, 'VENDOR')}",
            object_type="vendor",
            display_name=name,
            business_domain="Vendor / Commercial",
            lifecycle_state="current",
            source_record_id=source_record_id,
            attributes={"workbook_source": "applications_data_infrastructure_intake"},
        )
        return object_id

    def add_data_platform(name: str | None, source_record_id: str, row: dict[str, object], platform_kind: str) -> str | None:
        if not name or name.lower() == "unknown":
            return None
        if name in data_platform_object_by_name:
            return data_platform_object_by_name[name]
        object_id = stable_uuid("object", "data_platform", name)
        data_platform_object_by_name[name] = object_id
        objects[object_id] = object_row(
            object_id=object_id,
            object_key=f"DPLAT-{slug(name, 'DATA-PLATFORM')}",
            object_type="data_platform",
            display_name=name,
            business_domain=row_value(row, "Business Function") or row_value(row, "Primary Business Function") or "Data Analytics",
            lifecycle_state="current",
            source_record_id=source_record_id,
            attributes={
                "platform_kind": platform_kind,
                "technology": row_value(row, "Platform Technology") or row_value(row, "Mart Technology") or row_value(row, "Technology Type"),
                "hosting": row_value(row, "Mart Hosting") or row_value(row, "Hosting Model"),
                "confidence": row_value(row, "Confidence"),
            },
        )
        return object_id

    def add_application(
        *,
        app_id: str,
        app_name: str,
        row: dict[str, object],
        source_record_id: str,
        sheet_name: str,
        function_id: str | None,
        vendor_id: str | None,
    ) -> tuple[str, str]:
        base_name, environment = split_environment_variant(app_name)
        business_group = row_value(row, "Business Group")
        function_name = row_value(row, "Primary Business Function")
        if not environment:
            object_id = stable_uuid("object", "application", app_id)
            app_object_by_name[app_name] = object_id
            objects[object_id] = object_row(
                object_id=object_id,
                object_key=app_id,
                object_type="application",
                display_name=app_name,
                business_domain=function_name or business_group,
                lifecycle_state=app_lifecycle(row),
                source_record_id=source_record_id,
                attributes={
                    "business_group": business_group,
                    "application_category": row_value(row, "Application Category"),
                    "technology_type": row_value(row, "Technology Type"),
                    "product": row_value(row, "Product"),
                    "lifecycle_status": row_value(row, "Lifecycle Status"),
                    "hosting_model": row_value(row, "Hosting Model"),
                    "hosting_estate": row_value(row, "Hosting Estate / Platform"),
                    "infrastructure_id": row_value(row, "Infrastructure / Hosting Estate ID"),
                    "business_criticality": row_value(row, "Business Criticality / Tier"),
                    "source_sheet": sheet_name,
                },
            )
            add_relationship(function_id, "SUPPORTED_BY", object_id, source_record_id, {"source_sheet": sheet_name})
            add_relationship(object_id, "SUPPLIED_BY", vendor_id, source_record_id, {"source_sheet": sheet_name})
            return object_id, object_id

        base_key = f"{row_value(row, 'Vendor') or 'unknown'}::{base_name}"
        base_object_id = stable_uuid("object", "application_base", base_key)
        app_object_by_name[app_name] = base_object_id
        app_object_by_name.setdefault(base_name, base_object_id)
        if base_object_id not in objects:
            objects[base_object_id] = object_row(
                object_id=base_object_id,
                object_key=f"APPBASE-{slug(base_key, 'APPLICATION')}",
                object_type="application",
                display_name=base_name,
                business_domain=function_name or business_group,
                lifecycle_state=app_lifecycle(row),
                source_record_id=source_record_id,
                attributes={
                    "business_group": business_group,
                    "application_category": row_value(row, "Application Category"),
                    "technology_type": row_value(row, "Technology Type"),
                    "product": row_value(row, "Product"),
                    "lifecycle_status": row_value(row, "Lifecycle Status"),
                    "business_criticality": row_value(row, "Business Criticality / Tier"),
                    "environment_variant_source": "application_name_suffix",
                    "deployment_model": "base_application_plus_deployments",
                    "source_sheet": sheet_name,
                },
            )
        deployment_id = stable_uuid("object", "application_deployment", app_id)
        deployment_object_by_name[app_name] = deployment_id
        objects[deployment_id] = object_row(
            object_id=deployment_id,
            object_key=f"DEP-{app_id}",
            object_type="application_deployment",
            display_name=app_name,
            business_domain=function_name or business_group,
            lifecycle_state=app_lifecycle(row),
            source_record_id=source_record_id,
            attributes={
                "base_application_name": base_name,
                "base_application_object_id": base_object_id,
                "environment": environment,
                "deployment_source_application_id": app_id,
                "hosting_model": row_value(row, "Hosting Model"),
                "hosting_estate": row_value(row, "Hosting Estate / Platform"),
                "infrastructure_id": row_value(row, "Infrastructure / Hosting Estate ID"),
                "cloud_provider": row_value(row, "Cloud Provider"),
                "region_or_data_center": row_value(row, "Region / Data Center Location"),
                "source_sheet": sheet_name,
            },
        )
        add_relationship(function_id, "SUPPORTED_BY", base_object_id, source_record_id, {"source_sheet": sheet_name})
        add_relationship(base_object_id, "SUPPLIED_BY", vendor_id, source_record_id, {"source_sheet": sheet_name})
        add_relationship(deployment_id, "DEPLOYMENT_OF", base_object_id, source_record_id, {"source_sheet": sheet_name, "environment": environment})
        return base_object_id, deployment_id

    for sheet_name in APPLICATION_SHEETS:
        for row_number, row in parsed[sheet_name]:
            source_record_id = source_record_by_sheet_row[(sheet_name, row_number)]
            app_id = row_value(row, "Application ID")
            app_name = row_value(row, "Application / System Name")
            if not app_id or not app_name:
                continue
            function_name = row_value(row, "Primary Business Function")
            business_group = row_value(row, "Business Group")
            function_id = add_function(function_name, source_record_id, business_group)
            vendor_id = add_vendor(row_value(row, "Vendor"), source_record_id)
            _, measure_subject_id = add_application(
                app_id=app_id,
                app_name=app_name,
                row=row,
                source_record_id=source_record_id,
                sheet_name=sheet_name,
                function_id=function_id,
                vendor_id=vendor_id,
            )
            if split_environment_variant(app_name)[1]:
                environment_variant_rows += 1
            license_cost = as_num(row.get("Approx. Annual License / Subscription Cost"))
            run_cost = as_num(row.get("Approx. Annual Run / Support Cost"))
            annual_spend = sum(v for v in [license_cost, run_cost] if v is not None)
            if annual_spend:
                measures.append(
                    measure_row(
                        measure_id=stable_uuid("measure", app_id, "annual_spend_usd"),
                        subject_object_id=measure_subject_id,
                        metric_key="annual_spend_usd",
                        value_number=annual_spend,
                        value_text=None,
                        unit="USD",
                        source_record_id=source_record_id,
                    )
                )
            users = as_num(row.get("Approx. # Business Users"))
            if users is not None:
                measures.append(
                    measure_row(
                        measure_id=stable_uuid("measure", app_id, "active_users"),
                        subject_object_id=measure_subject_id,
                        metric_key="active_users",
                        value_number=users,
                        value_text=None,
                        unit="users",
                        source_record_id=source_record_id,
                    )
                )

    for row_number, row in parsed["05_INFRASTRUCTURE_HOSTING"]:
        source_record_id = source_record_by_sheet_row[("05_INFRASTRUCTURE_HOSTING", row_number)]
        infra_key = row_value(row, "Infrastructure / Hosting Estate ID")
        infra_name = row_value(row, "Hosting Estate / Platform Name")
        if not infra_key or not infra_name:
            continue
        object_id = stable_uuid("object", "infrastructure", infra_key)
        infra_object_by_id[infra_key] = object_id
        objects[object_id] = object_row(
            object_id=object_id,
            object_key=infra_key,
            object_type="infrastructure",
            display_name=infra_name,
            business_domain="Infrastructure / Hosting",
            lifecycle_state="current",
            source_record_id=source_record_id,
            attributes={
                "infrastructure_type": row_value(row, "Infrastructure Type"),
                "owner": row_value(row, "Owner"),
                "provider": row_value(row, "Primary Data Center / Cloud Provider"),
                "location": row_value(row, "Location / AWS-Azure Region"),
                "storage_capacity": row_value(row, "Approx. Storage Capacity"),
                "source_sheet": "05_INFRASTRUCTURE_HOSTING",
            },
        )

    for sheet_name in APPLICATION_SHEETS:
        for row_number, row in parsed[sheet_name]:
            source_record_id = source_record_by_sheet_row[(sheet_name, row_number)]
            app_id = row_value(row, "Application ID")
            app_name = row_value(row, "Application / System Name") or ""
            object_id = deployment_object_by_name.get(app_name) or app_object_by_name.get(app_name)
            infra_id = infra_object_by_id.get(row_value(row, "Infrastructure / Hosting Estate ID") or "")
            add_relationship(object_id, "HOSTED_ON", infra_id, source_record_id, {"source_sheet": sheet_name})

    snapshot_id = stable_uuid("snapshot", workbook_hash)
    cube_manifest_id = stable_uuid("cube_manifest", "data_analytics_cube", workbook_hash)
    cube_slice_rows: list[dict[str, str]] = []
    cube_metric_rows: list[dict[str, str]] = []
    cube_measure_rows: list[dict[str, str]] = []
    for row_number, row in parsed["01_ENTERPRISE_DA"]:
        source_record_id = source_record_by_sheet_row[("01_ENTERPRISE_DA", row_number)]
        function_name = row_value(row, "Business Function")
        business_group = row_value(row, "Business Group")
        function_id = add_function(function_name, source_record_id, business_group)
        if not function_id:
            continue
        shared_platform_id = add_data_platform(row_value(row, "Enterprise DW / Lake / Lakehouse Name"), source_record_id, row, "enterprise_shared")
        mart_platform_id = add_data_platform(row_value(row, "Data Mart / Analytical Store Name(s)"), source_record_id, row, "function_mart")
        add_relationship(function_id, "DEPENDS_ON", shared_platform_id, source_record_id, {"source_sheet": "01_ENTERPRISE_DA"})
        add_relationship(function_id, "DEPENDS_ON", mart_platform_id, source_record_id, {"source_sheet": "01_ENTERPRISE_DA"})
        metric_inputs = [
            ("report_count", "Approx. # Production Reports", "reports"),
            ("etl_job_count", "Approx. # ETL Jobs / Pipelines", "jobs"),
            ("stored_procedure_count", "Approx. # Stored Procedures / SQL Scripts", "procedures"),
            ("analytics_user_count", "Approx. # BI / Reporting Users", "users"),
            ("data_volume_tb", "Total Approx. Data Volume", "TB"),
        ]
        measures_json: dict[str, float] = {}
        slice_measure_ids: list[tuple[str, str]] = []
        for metric_key, header, unit in metric_inputs:
            value_number = as_num(row.get(header))
            if value_number is None:
                continue
            measures_json[metric_key] = value_number
            measure_id = stable_uuid("measure", "da", function_name, metric_key)
            slice_measure_ids.append((measure_id, metric_key))
            measures.append(
                measure_row(
                    measure_id=measure_id,
                    subject_object_id=function_id,
                    metric_key=metric_key,
                    value_number=value_number,
                    value_text=None,
                    unit=unit,
                    source_record_id=source_record_id,
                )
            )
        if not measures_json:
            continue
        metric_keys = list(measures_json.keys())
        slice_id = stable_uuid("cube_slice", "data_analytics_cube", function_name)
        cube_slice_rows.append(
            {
                "id": sql_text(slice_id),
                "tenant_key": sql_text(TENANT_KEY),
                "assessment_id": sql_text(ASSESSMENT_ID),
                "snapshot_id": sql_text(snapshot_id),
                "cube_manifest_id": sql_text(cube_manifest_id),
                "cube_key": sql_text("data_analytics_cube"),
                "cube_version": "1",
                "slice_key": sql_text(f"da-{slug(function_name, 'FUNCTION').lower()}"),
                "grain_key": sql_text("function_platform"),
                "primary_object_id": sql_text(function_id),
                "dimensions_json": sql_json(
                    {
                        "business_group": business_group,
                        "function": function_name,
                        "shared_platform": row_value(row, "Enterprise DW / Lake / Lakehouse Name"),
                        "mart": row_value(row, "Data Mart / Analytical Store Name(s)"),
                        "bi_tools": row_value(row, "Primary BI / Reporting Tool(s)"),
                    }
                ),
                "measures_json": sql_json(measures_json),
                "primary_metric_key": sql_text("report_count" if "report_count" in metric_keys else metric_keys[0]),
                "metric_keys_json": sql_json(metric_keys),
                "source_refs_json": sql_json([source_record_id]),
                "basis_summary": sql_text("source_recorded"),
                "value_state": sql_text("estimated"),
                "quality_state": sql_text("warning"),
                "gap_flags_json": sql_json(["synthetic_planning_grade_not_client_attested"]),
                "source_hash": sql_text(SOURCE_HASH_LABEL),
            }
        )
        for index, metric_key in enumerate(metric_keys, start=1):
            unit = "TB" if metric_key == "data_volume_tb" else "users" if "user" in metric_key else "procedures" if "procedure" in metric_key else "jobs" if "etl" in metric_key else "reports"
            cube_metric_rows.append(
                {
                    "tenant_key": sql_text(TENANT_KEY),
                    "assessment_id": sql_text(ASSESSMENT_ID),
                    "cube_slice_id": sql_text(slice_id),
                    "metric_key": sql_text(metric_key),
                    "metric_role": sql_text("primary" if index == 1 else "display"),
                    "unit": sql_text(unit),
                    "sort_order": sql_num(index),
                    "source_hash": sql_text(SOURCE_HASH_LABEL),
                }
            )
        for index, (measure_id, metric_key) in enumerate(slice_measure_ids, start=1):
            cube_measure_rows.append(
                {
                    "tenant_key": sql_text(TENANT_KEY),
                    "assessment_id": sql_text(ASSESSMENT_ID),
                    "cube_slice_id": sql_text(slice_id),
                    "measure_id": sql_text(measure_id),
                    "metric_key": sql_text(metric_key),
                    "measure_role": sql_text("primary" if index == 1 else "display"),
                    "source_hash": sql_text(SOURCE_HASH_LABEL),
                }
            )

    for row_number, row in parsed["06_MAJOR_INTEGRATIONS"]:
        source_record_id = source_record_by_sheet_row[("06_MAJOR_INTEGRATIONS", row_number)]
        source_id = app_object_by_name.get(row_value(row, "Source Application") or "")
        target_id = app_object_by_name.get(row_value(row, "Target Application") or "")
        add_relationship(
            source_id,
            "INTEGRATES_WITH",
            target_id,
            source_record_id,
            {
                "integration_id": row_value(row, "Integration ID"),
                "purpose": row_value(row, "Business Purpose"),
                "protocol": row_value(row, "HL7 / FHIR / API / File / DB / MQ / Event / Other"),
                "frequency": row_value(row, "Frequency"),
            },
        )

    object_rows = list(objects.values())
    app_count = sum(1 for row in object_rows if row["object_type"] == sql_text("application"))
    deployment_count = sum(1 for row in object_rows if row["object_type"] == sql_text("application_deployment"))
    function_count = sum(1 for row in object_rows if row["object_type"] == sql_text("business_function"))
    infra_count = sum(1 for row in object_rows if row["object_type"] == sql_text("infrastructure"))
    data_platform_count = sum(1 for row in object_rows if row["object_type"] == sql_text("data_platform"))
    vendor_count = sum(1 for row in object_rows if row["object_type"] == sql_text("vendor"))
    context_hash = hashlib.sha256(f"{len(object_rows)}|{len(relationships)}|{len(measures)}".encode("utf-8")).hexdigest()
    context_pack_id = stable_uuid("context_pack", "workbook-builder-pilot", workbook_hash)
    home_manifest_id = stable_uuid("projection_manifest", "home_enterprise_landscape", workbook_hash)
    home_rows = [
        {
            "id": sql_text(stable_uuid("home_row", "what_has_been_loaded", workbook_hash)),
            "tenant_key": sql_text(TENANT_KEY),
            "assessment_id": sql_text(ASSESSMENT_ID),
            "snapshot_id": sql_text(snapshot_id),
            "projection_manifest_id": sql_text(home_manifest_id),
            "projection_version": "1",
            "page_key": sql_text("what_has_been_loaded"),
            "row_key": sql_text("workbook-builder-coverage"),
            "section_key": sql_text("coverage"),
            "row_type": sql_text("coverage_summary"),
            "title": sql_text("Applications/data/infrastructure workbook loaded into ECL pilot"),
            "summary": sql_text(
                f"{app_count} applications, {deployment_count} application deployments, {function_count} functions, {infra_count} infrastructure estates, "
                f"{data_platform_count} data platforms, {vendor_count} vendors, {len(relationships)} relationships, "
                f"and {len(measures)} measures generated from workbook rows."
            ),
            "primary_object_id": "null",
            "metric_keys_json": sql_json(["object_count", "relationship_count", "measure_count"]),
            "relationship_ids_json": sql_json([]),
            "source_refs_json": sql_json([source_file_id]),
            "basis_summary": sql_text("source_recorded"),
            "value_state": sql_text("known"),
            "quality_state": sql_text("warning"),
            "admission_status": sql_text("not_applicable"),
            "admission_gate_key": "null",
            "admission_result_json": sql_json({}),
            "gap_flags_json": sql_json(["synthetic_planning_grade_not_client_attested", "commercial_contract_depth_not_in_this_workbook"]),
            "display_payload_json": sql_json(
                {
                    "appCount": app_count,
                    "applicationDeploymentCount": deployment_count,
                    "functionCount": function_count,
                    "infraCount": infra_count,
                    "dataPlatformCount": data_platform_count,
                    "vendorCount": vendor_count,
                    "relationshipCount": len(relationships),
                    "measureCount": len(measures),
                }
            ),
            "source_hash": sql_text(SOURCE_HASH_LABEL),
        },
        {
            "id": sql_text(stable_uuid("home_row", "current_state_architecture", workbook_hash)),
            "tenant_key": sql_text(TENANT_KEY),
            "assessment_id": sql_text(ASSESSMENT_ID),
            "snapshot_id": sql_text(snapshot_id),
            "projection_manifest_id": sql_text(home_manifest_id),
            "projection_version": "1",
            "page_key": sql_text("current_state_architecture"),
            "row_key": sql_text("architecture-workbook-builder"),
            "section_key": sql_text("architecture"),
            "row_type": sql_text("architecture_summary"),
            "title": sql_text("Workbook-derived architecture pilot has FK-backed app/function/vendor/hosting relationships"),
            "summary": sql_text("Builder-generated from intake workbook rows. Architecture fitness/browser proof remains open."),
            "primary_object_id": "null",
            "metric_keys_json": sql_json(["application_count", "vendor_count", "relationship_count"]),
            "relationship_ids_json": sql_json([]),
            "source_refs_json": sql_json([source_file_id]),
            "basis_summary": sql_text("source_recorded"),
            "value_state": sql_text("known"),
            "quality_state": sql_text("warning"),
            "admission_status": sql_text("not_applicable"),
            "admission_gate_key": "null",
            "admission_result_json": sql_json({}),
            "gap_flags_json": sql_json(["architecture_admission_gate_not_run", "browser_proof_not_run"]),
            "display_payload_json": sql_json(
                {
                    "appCount": app_count,
                    "applicationDeploymentCount": deployment_count,
                    "vendorCount": vendor_count,
                    "infraCount": infra_count,
                    "dataPlatformCount": data_platform_count,
                }
            ),
            "source_hash": sql_text(SOURCE_HASH_LABEL),
        },
    ]

    sql_sections = [
        "-- Generated by scripts/ecl/build_workbook_pilot_ecl_sql.py",
        "-- Local design artifact only. Do not run against Azure, lab, preprod, or production without migration authorization.",
        "begin;\n",
        insert(
            "ecl_source.source_file",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "source_type",
                "origin",
                "source_owner",
                "file_name",
                "blob_uri",
                "file_hash",
                "source_date",
                "access_class",
                "quality_state",
                "metadata_json",
            ],
            [
                {
                    "id": sql_text(source_file_id),
                    "tenant_key": sql_text(TENANT_KEY),
                    "assessment_id": sql_text(ASSESSMENT_ID),
                    "source_type": sql_text("manual_workbook"),
                    "origin": sql_text("synthetic_generator"),
                    "source_owner": sql_text("enterprise architecture; CDAO; CMDB owners"),
                    "file_name": sql_text(workbook_path.name),
                    "blob_uri": sql_text(str(workbook_path)),
                    "file_hash": sql_text(workbook_hash),
                    "source_date": "null",
                    "access_class": sql_text("public_demo"),
                    "quality_state": sql_text("accepted"),
                    "metadata_json": sql_json(
                        {
                            "local_only": True,
                            "source_package": "applications_data_infrastructure_intake",
                            "approval_state": "synthetic_planning_grade_not_client_attested",
                        }
                    ),
                }
            ],
        ),
        insert(
            "ecl_source.source_record",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "source_file_id",
                "native_id",
                "record_type",
                "row_number",
                "payload_json",
                "parse_state",
                "parse_notes",
            ],
            source_records,
        ),
        insert(
            "ecl_context.object",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "object_key",
                "object_type",
                "display_name",
                "business_domain",
                "lifecycle_state",
                "source_record_id",
                "basis",
                "value_state",
                "review_state",
                "confidence",
                "attributes_json",
            ],
            object_rows,
        ),
        insert(
            "ecl_context.relationship",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "from_object_id",
                "relationship_type",
                "to_object_id",
                "direction_label",
                "source_record_id",
                "basis",
                "value_state",
                "review_state",
                "confidence",
                "attributes_json",
            ],
            relationships,
        ),
        insert(
            "ecl_context.measure",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "subject_object_id",
                "metric_key",
                "value_number",
                "value_text",
                "unit",
                "period_start",
                "period_end",
                "scenario",
                "source_record_id",
                "document_extraction_id",
                "basis",
                "value_state",
                "quality_state",
                "review_state",
                "attributes_json",
            ],
            measures,
        ),
        insert(
            "ecl_context.snapshot",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "snapshot_key",
                "snapshot_type",
                "source_hash",
                "context_hash",
                "created_by_job",
                "quality_state",
                "proof_uri",
            ],
            [
                {
                    "id": sql_text(snapshot_id),
                    "tenant_key": sql_text(TENANT_KEY),
                    "assessment_id": sql_text(ASSESSMENT_ID),
                    "snapshot_key": sql_text("workbook-builder-pilot-snapshot"),
                    "snapshot_type": sql_text("projection_source"),
                    "source_hash": sql_text(workbook_hash),
                    "context_hash": sql_text(context_hash),
                    "created_by_job": sql_text("scripts/ecl/build_workbook_pilot_ecl_sql.py"),
                    "quality_state": sql_text("warning"),
                    "proof_uri": sql_text("local://outputs/ecl-workbook-builder-pilot-2026-08-22"),
                }
            ],
        ),
        insert(
            "ecl_context.context_pack",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "snapshot_id",
                "pack_key",
                "pack_version",
                "payload_json",
                "payload_hash",
                "retrieval_state",
                "quality_state",
                "proof_uri",
            ],
            [
                {
                    "id": sql_text(context_pack_id),
                    "tenant_key": sql_text(TENANT_KEY),
                    "assessment_id": sql_text(ASSESSMENT_ID),
                    "snapshot_id": sql_text(snapshot_id),
                    "pack_key": sql_text("workbook_builder_orientation"),
                    "pack_version": "1",
                    "payload_json": sql_json(
                        {
                            "appCount": app_count,
                            "applicationDeploymentCount": deployment_count,
                            "functionCount": function_count,
                            "infraCount": infra_count,
                            "dataPlatformCount": data_platform_count,
                            "vendorCount": vendor_count,
                            "relationshipCount": len(relationships),
                            "measureCount": len(measures),
                        }
                    ),
                    "payload_hash": sql_text(hashlib.sha256(f"{workbook_hash}|context-pack".encode("utf-8")).hexdigest()),
                    "retrieval_state": sql_text("not_indexed"),
                    "quality_state": sql_text("warning"),
                    "proof_uri": sql_text("local://outputs/ecl-workbook-builder-pilot-2026-08-22"),
                }
            ],
        ),
        insert(
            "ecl_projection.projection_manifest",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "snapshot_id",
                "projection_key",
                "projection_version",
                "rebuild_command",
                "source_hash",
                "projection_hash",
                "row_count",
                "quality_state",
                "admission_status",
                "admission_gate_results_json",
                "gated_claim_count",
                "proof_uri",
            ],
            [
                {
                    "id": sql_text(home_manifest_id),
                    "tenant_key": sql_text(TENANT_KEY),
                    "assessment_id": sql_text(ASSESSMENT_ID),
                    "snapshot_id": sql_text(snapshot_id),
                    "projection_key": sql_text("home_enterprise_landscape"),
                    "projection_version": "1",
                    "rebuild_command": sql_text("scripts/ecl/build_workbook_pilot_ecl_sql.py --projection home_enterprise_landscape"),
                    "source_hash": sql_text(workbook_hash),
                    "projection_hash": sql_text(hashlib.sha256(f"{workbook_hash}|home".encode("utf-8")).hexdigest()),
                    "row_count": sql_num(len(home_rows)),
                    "quality_state": sql_text("warning"),
                    "admission_status": sql_text("not_applicable"),
                    "admission_gate_results_json": sql_json([]),
                    "gated_claim_count": "0",
                    "proof_uri": sql_text("local://outputs/ecl-workbook-builder-pilot-2026-08-22/home"),
                }
            ],
        ),
        insert(
            "ecl_projection.home_enterprise_landscape",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "snapshot_id",
                "projection_manifest_id",
                "projection_version",
                "page_key",
                "row_key",
                "section_key",
                "row_type",
                "title",
                "summary",
                "primary_object_id",
                "metric_keys_json",
                "relationship_ids_json",
                "source_refs_json",
                "basis_summary",
                "value_state",
                "quality_state",
                "admission_status",
                "admission_gate_key",
                "admission_result_json",
                "gap_flags_json",
                "display_payload_json",
                "source_hash",
            ],
            home_rows,
        ),
        insert(
            "ecl_projection.cube_manifest",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "snapshot_id",
                "cube_key",
                "cube_version",
                "rebuild_command",
                "source_hash",
                "cube_hash",
                "slice_count",
                "quality_state",
                "admission_status",
                "admission_gate_results_json",
                "proof_uri",
            ],
            [
                {
                    "id": sql_text(cube_manifest_id),
                    "tenant_key": sql_text(TENANT_KEY),
                    "assessment_id": sql_text(ASSESSMENT_ID),
                    "snapshot_id": sql_text(snapshot_id),
                    "cube_key": sql_text("data_analytics_cube"),
                    "cube_version": "1",
                    "rebuild_command": sql_text("scripts/ecl/build_workbook_pilot_ecl_sql.py --cube data_analytics_cube"),
                    "source_hash": sql_text(workbook_hash),
                    "cube_hash": sql_text(hashlib.sha256(f"{workbook_hash}|data_analytics_cube".encode("utf-8")).hexdigest()),
                    "slice_count": sql_num(len(cube_slice_rows)),
                    "quality_state": sql_text("warning"),
                    "admission_status": sql_text("not_applicable"),
                    "admission_gate_results_json": sql_json([]),
                    "proof_uri": sql_text("local://outputs/ecl-workbook-builder-pilot-2026-08-22/cubes/data-analytics"),
                }
            ],
        ),
        insert(
            "ecl_projection.cube_slice",
            [
                "id",
                "tenant_key",
                "assessment_id",
                "snapshot_id",
                "cube_manifest_id",
                "cube_key",
                "cube_version",
                "slice_key",
                "grain_key",
                "primary_object_id",
                "dimensions_json",
                "measures_json",
                "primary_metric_key",
                "metric_keys_json",
                "source_refs_json",
                "basis_summary",
                "value_state",
                "quality_state",
                "gap_flags_json",
                "source_hash",
            ],
            cube_slice_rows,
        ),
        insert(
            "ecl_projection.cube_slice_metric",
            [
                "tenant_key",
                "assessment_id",
                "cube_slice_id",
                "metric_key",
                "metric_role",
                "unit",
                "sort_order",
                "source_hash",
            ],
            cube_metric_rows,
        ),
        insert(
            "ecl_projection.cube_slice_measure",
            [
                "tenant_key",
                "assessment_id",
                "cube_slice_id",
                "measure_id",
                "metric_key",
                "measure_role",
                "source_hash",
            ],
            cube_measure_rows,
        ),
        "commit;\n",
    ]

    out_dir.mkdir(parents=True, exist_ok=True)
    sql_path = out_dir / "ecl_workbook_builder_pilot_load.sql"
    manifest_path = out_dir / "ecl_workbook_builder_pilot_manifest.json"
    sql_path.write_text("\n".join(section for section in sql_sections if section), encoding="utf-8")
    manifest = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "generatedBy": "scripts/ecl/build_workbook_pilot_ecl_sql.py",
        "workbookPath": str(workbook_path),
        "workbookHash": workbook_hash,
        "tenantKey": TENANT_KEY,
        "assessmentId": ASSESSMENT_ID,
        "sourceSheets": {sheet: len(rows) for sheet, rows in parsed.items()},
            "output": {
                "sqlPath": str(sql_path),
                "sourceRecords": len(source_records),
                "objects": len(object_rows),
                "applications": app_count,
                "applicationDeployments": deployment_count,
                "environmentVariantRows": environment_variant_rows,
                "relationships": len(relationships),
                "measures": len(measures),
            "homeRows": len(home_rows),
            "dataAnalyticsCubeSlices": len(cube_slice_rows),
            "cubeSliceMetrics": len(cube_metric_rows),
            "cubeSliceMeasures": len(cube_measure_rows),
        },
        "closedGates": [
            "Azure/Postgres load",
            "migration authorization",
            "product route repointing",
            "browser proof",
            "retrieval indexing",
        ],
    }
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps({"sqlPath": str(sql_path), "manifestPath": str(manifest_path), **manifest["output"]}, indent=2))


if __name__ == "__main__":
    main()
