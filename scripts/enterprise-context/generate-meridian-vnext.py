#!/usr/bin/env python3
"""Generate a rich Meridian Health vNext context package.

The output is intentionally file-only. It is designed for Setup/Admin loader
upload and validation, not direct database writes.
"""

from __future__ import annotations

import csv
import json
import random
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_MANIFEST = ROOT / "docs/enterprise-context/templates/meridian/manifest.json"
OUT = ROOT / "docs/enterprise-context/generated/meridian-vnext"
GENERATED_AT = "2026-06-04T00:00:00.000Z"
LAST_VALIDATED = "2026-06-04"
SEED = "meridian-health-vnext-sacramento-integrated-system"
RAND = random.Random(20260604)


PROFILE = {
    "tenantKey": "meridian",
    "tenantSlug": "meridian",
    "displayName": "Meridian Health System",
    "headquarters": "Sacramento, California",
    "businessModel": "Integrated delivery network plus health plan",
    "annualRevenueUsd": 16800000000,
    "workforce": 58000,
    "hospitals": 31,
    "clinics": 280,
    "coveredLives": 1600000,
    "clinicalCore": "Epic",
    "cloudTarget": "100% workloads hosted in AWS by July 2026",
    "modernizationReality": "Most workloads are lift-and-shift legacy VMs hosted in AWS, not cloud-native modernization.",
}


@dataclass(frozen=True)
class WorkbookSpec:
    key: str
    title: str
    csv_name: str
    xlsx_name: str
    columns: list[str]


def common(source_system: str, source_record_id: str, owner: str, confidence: str = "0.86") -> dict[str, str]:
    return {
        "source_system": source_system,
        "source_record_id": source_record_id,
        "source_owner": owner,
        "last_validated_date": LAST_VALIDATED,
        "confidence": confidence,
        "evidence_usable": "true",
        "notes_gaps": "",
    }


def row(columns: list[str], values: dict[str, Any]) -> dict[str, str]:
    return {col: str(values.get(col, "")) for col in columns}


def slug(text: str) -> str:
    return (
        text.lower()
        .replace("&", "and")
        .replace("/", " ")
        .replace("+", " ")
        .replace("(", "")
        .replace(")", "")
        .replace(",", "")
        .replace(".", "")
        .replace(" ", "-")
    )


def as_date(offset: int) -> str:
    return (date(2026, 1, 1) + timedelta(days=offset)).isoformat()


def load_specs() -> list[WorkbookSpec]:
    manifest = json.loads(TEMPLATE_MANIFEST.read_text())
    specs: list[WorkbookSpec] = []
    for index, wb in enumerate(manifest["workbooks"], start=1):
        prefix = f"{index:02d}"
        suffix = wb["path"].split("/")[-1]
        csv_name = suffix.replace(".xlsx", ".csv")
        specs.append(
            WorkbookSpec(
                key=wb["key"],
                title=wb["title"],
                csv_name=csv_name,
                xlsx_name=suffix,
                columns=wb["columns"],
            )
        )
    return specs


EXECUTIVES = [
    ("PERSON-CEO", "Elena Martinez", "President and Chief Executive Officer", "PERSON-BOARD", "Enterprise strategy", "Approve"),
    ("PERSON-COO", "Sarah O'Brien", "Chief Operating Officer", "PERSON-CEO", "Care delivery operations", "Approve"),
    ("PERSON-CFO", "David Park", "Chief Financial Officer", "PERSON-CEO", "Capital allocation and run cost", "Approve"),
    ("PERSON-CDIO", "Anita Krishnamurthy", "Chief Digital and Information Officer", "PERSON-CEO", "Digital, IT, and AI portfolio", "Approve"),
    ("PERSON-CTO", "Marcus Lee", "Chief Technology Officer", "PERSON-CDIO", "Cloud and platform modernization", "Approve"),
    ("PERSON-CDAO", "Renee Walters", "Chief Data and Analytics Officer", "PERSON-CDIO", "Data, analytics, and AI enablement", "Approve"),
    ("PERSON-CISO", "Daniel Reyes", "Chief Information Security Officer", "PERSON-CDIO", "Security and cyber risk", "Approve"),
    ("PERSON-CMO", "Dr. Priya Shah", "Chief Medical Officer", "PERSON-CEO", "Clinical quality and safety", "Approve"),
    ("PERSON-CNE", "Robert Chen", "Chief Nursing Executive", "PERSON-COO", "Nursing operations", "Approve"),
    ("PERSON-CRO", "Patricia Okafor", "Chief Revenue Cycle Officer", "PERSON-CFO", "Revenue cycle performance", "Approve"),
    ("PERSON-HPLAN", "Thomas Hartwell", "President, Meridian Health Plan", "PERSON-CEO", "Health plan operations", "Approve"),
    ("PERSON-COMPLIANCE", "Karen Mercer", "Chief Compliance and Privacy Officer", "PERSON-CEO", "Privacy, compliance, and audit", "Approve"),
    ("PERSON-LEGAL", "Rebecca Hollings", "Chief Legal Officer", "PERSON-CEO", "Legal and contracting", "Approve"),
    ("PERSON-CHRO", "Margaret Liu", "Chief Human Resources Officer", "PERSON-CEO", "Workforce and labor", "Approve"),
    ("PERSON-SUPPLY", "Nora Vasquez", "Chief Supply Chain and Sourcing Officer", "PERSON-CFO", "Supply chain and strategic sourcing", "Approve"),
    ("PERSON-STRATEGY", "Aiden Walsh", "Chief Strategy and Transformation Officer", "PERSON-CEO", "Transformation portfolio", "Approve"),
]


IT_LEADERS = [
    ("PERSON-VP-CLOUD", "Miguel Arroyo", "SVP Infrastructure, Cloud, and Operations", "PERSON-CTO", "Infrastructure modernization"),
    ("PERSON-VP-APPS", "Linda Howard", "VP Enterprise Applications", "PERSON-CDIO", "Enterprise application portfolio"),
    ("PERSON-VP-CLINICAL-APPS", "Maya Ramos", "VP Clinical Applications", "PERSON-CDIO", "Clinical applications"),
    ("PERSON-VP-INTEGRATION", "Leah Benitez", "VP Integration Engineering", "PERSON-CTO", "HL7 and FHIR integration"),
    ("PERSON-VP-ITSM", "Samuel Ito", "VP Service Management", "PERSON-CTO", "ITSM and CMDB"),
    ("PERSON-VP-SECOPS", "Natalie Greer", "VP Cybersecurity Operations", "PERSON-CISO", "Security operations"),
]


CDAO_VPS = [
    ("PERSON-VP-DATA-PLATFORMS", "Isha Raman", "VP Enterprise Data Platforms", "PERSON-CDAO", "Cloud data platform"),
    ("PERSON-VP-ANALYTICS", "Jordan McKenzie", "VP Analytics and BI", "PERSON-CDAO", "Enterprise analytics"),
    ("PERSON-VP-CLIN-ANALYTICS", "Dr. Kavita Patel", "VP Clinical Analytics", "PERSON-CDAO", "Clinical analytics"),
    ("PERSON-VP-PAYER-ANALYTICS", "Owen Li", "VP Health Plan Analytics", "PERSON-CDAO", "Payer analytics"),
    ("PERSON-VP-AI", "Naomi Brooks", "VP AI and Automation", "PERSON-CDAO", "AI enablement"),
    ("PERSON-VP-GOV", "Wei Zhang", "VP Data Governance and Quality", "PERSON-CDAO", "Data governance"),
]


BUSINESS_AREAS = [
    ("ORG-ACUTE", "Acute Care Operations", "PERSON-COO", "Care delivery operations"),
    ("ORG-AMB", "Ambulatory Operations", "PERSON-COO", "Ambulatory access and quality"),
    ("ORG-REV", "Revenue Cycle", "PERSON-CRO", "Revenue cycle"),
    ("ORG-ACCESS", "Access Center", "PERSON-COO", "Patient access"),
    ("ORG-PLAN-CLAIMS", "Claims Operations", "PERSON-HPLAN", "Claims operations"),
    ("ORG-CARE-MGMT", "Care Management", "PERSON-HPLAN", "Care management"),
    ("ORG-QUALITY", "Quality and STAR/HEDIS", "PERSON-CMO", "Quality performance"),
    ("ORG-FINANCE", "Finance Planning", "PERSON-CFO", "Financial planning"),
    ("ORG-SOURCING", "Strategic Sourcing", "PERSON-SUPPLY", "Sourcing decisions"),
    ("ORG-HR", "People Operations", "PERSON-CHRO", "Workforce"),
    ("ORG-LEGAL", "Legal and Privacy", "PERSON-LEGAL", "Legal and privacy"),
    ("ORG-PHARMACY", "Pharmacy Operations", "PERSON-COO", "Pharmacy operations"),
    ("ORG-BH", "Behavioral Health", "PERSON-CMO", "Behavioral health"),
    ("ORG-ONCOLOGY", "Oncology Service Line", "PERSON-CMO", "Oncology"),
    ("ORG-CARDIOLOGY", "Cardiology Service Line", "PERSON-CMO", "Cardiology"),
    ("ORG-PRIMARY", "Primary Care", "PERSON-COO", "Primary care"),
]


def generate_org(columns: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []

    def add(org_unit: str, pid: str, name: str, title: str, reports_to: str, domain: str, right: str = "Recommend"):
        rows.append(
            row(
                columns,
                {
                    "org_unit_id": org_unit,
                    "person_or_group_id": pid,
                    "name": name,
                    "role_title": title,
                    "reports_to_id": reports_to,
                    "decision_domain": domain,
                    "decision_right": right,
                    "approval_authority": f"{right} authority for {domain}",
                    "delegated_to_id": "",
                    "escalation_path": "Executive Steering Committee",
                    **common("Workday", pid, "People Operations", "0.92"),
                },
            )
        )

    for item in EXECUTIVES:
        add("ORG-EXEC", *item)

    for gid, name, title in [
        ("GROUP-AIGOV", "AI Governance Council", "Governance Body"),
        ("GROUP-DATA-GOV", "Data Governance Council", "Governance Body"),
        ("GROUP-CAB", "Clinical Change Advisory Board", "Governance Body"),
        ("GROUP-CLOUD-CAB", "Cloud Migration Control Board", "Governance Body"),
        ("GROUP-AMS", "AMS Transition Governance Forum", "Governance Body"),
        ("GROUP-FINOPS", "Cloud FinOps Council", "Governance Body"),
        ("GROUP-PRIVACY", "Privacy and Compliance Review Board", "Governance Body"),
        ("GROUP-SOURCE", "Strategic Sourcing Council", "Governance Body"),
    ]:
        add("ORG-GOVERNANCE", gid, name, title, "PERSON-CDIO", name, "Approve")

    for item in IT_LEADERS:
        add("ORG-IT-LEADERSHIP", *item, right="Approve")

    it_domains = [
        ("Cloud Migration Factory", "PERSON-VP-CLOUD", "AWS migration wave execution"),
        ("Legacy VM Rehost Operations", "PERSON-VP-CLOUD", "Lift and shift run control"),
        ("Network and Connectivity", "PERSON-VP-CLOUD", "WAN, SD-WAN, and clinic connectivity"),
        ("End User Computing", "PERSON-VP-CLOUD", "Clinical endpoint operations"),
        ("Identity and Access Management", "PERSON-CISO", "Workforce and patient IAM"),
        ("Service Desk", "PERSON-VP-ITSM", "Tier 1 and Tier 2 support"),
        ("CMDB Stewardship", "PERSON-VP-ITSM", "CMDB ownership"),
        ("Epic Platform Operations", "PERSON-VP-CLINICAL-APPS", "Epic operations"),
        ("ERP Platform Operations", "PERSON-VP-APPS", "ERP operations"),
        ("Integration Operations", "PERSON-VP-INTEGRATION", "HL7 and FHIR operations"),
        ("Platform Reliability", "PERSON-CTO", "SRE and service reliability"),
        ("Disaster Recovery", "PERSON-VP-CLOUD", "DR and business continuity"),
    ]
    for idx, (team, leader, domain) in enumerate(it_domains, start=1):
        add(f"ORG-IT-{idx:02d}", f"GROUP-IT-{idx:02d}", team, "IT Operating Group", leader, domain, "Recommend")
        add(f"ORG-IT-{idx:02d}", f"PERSON-DIR-IT-{idx:02d}", f"{team} Director", "Director", leader, domain, "Recommend")
        add(f"ORG-IT-{idx:02d}", f"PERSON-MGR-IT-{idx:02d}", f"{team} Manager", "Manager", f"PERSON-DIR-IT-{idx:02d}", domain, "Consult")

    for item in CDAO_VPS:
        add("ORG-CDAO-LEADERSHIP", *item, right="Approve")

    cdao_domains = [
        ("Cloud Data Platform", "PERSON-VP-DATA-PLATFORMS"),
        ("Epic Analytics", "PERSON-VP-CLIN-ANALYTICS"),
        ("Clinical Data Products", "PERSON-VP-CLIN-ANALYTICS"),
        ("Revenue Cycle Analytics", "PERSON-VP-ANALYTICS"),
        ("Health Plan Analytics", "PERSON-VP-PAYER-ANALYTICS"),
        ("Population Health Analytics", "PERSON-VP-PAYER-ANALYTICS"),
        ("Enterprise BI", "PERSON-VP-ANALYTICS"),
        ("AI Enablement", "PERSON-VP-AI"),
        ("MLOps and Model Monitoring", "PERSON-VP-AI"),
        ("Data Governance", "PERSON-VP-GOV"),
        ("Data Quality Observability", "PERSON-VP-GOV"),
        ("Master Data Management", "PERSON-VP-GOV"),
        ("Interoperability Data Products", "PERSON-VP-DATA-PLATFORMS"),
        ("Research Analytics", "PERSON-VP-CLIN-ANALYTICS"),
        ("Finance Analytics", "PERSON-VP-ANALYTICS"),
        ("Supply Chain Analytics", "PERSON-VP-ANALYTICS"),
        ("Patient Access Analytics", "PERSON-VP-ANALYTICS"),
        ("Data Privacy Engineering", "PERSON-VP-GOV"),
    ]
    cdao_resource_rows = 0
    for idx, (domain, leader) in enumerate(cdao_domains, start=1):
        org_id = f"ORG-CDAO-{idx:02d}"
        add(org_id, f"PERSON-DIR-CDAO-{idx:02d}", f"{domain} Director", "Director", leader, domain, "Recommend")
        cdao_resource_rows += 1
        for mgr in range(1, 3):
            add(org_id, f"PERSON-MGR-CDAO-{idx:02d}-{mgr}", f"{domain} Manager {mgr}", "Manager", f"PERSON-DIR-CDAO-{idx:02d}", domain, "Consult")
            cdao_resource_rows += 1
        squads = 4 if idx <= 12 else 3
        for squad in range(1, squads + 1):
            add(org_id, f"GROUP-CDAO-{idx:02d}-{squad}", f"{domain} Squad {squad}", "Data / Analytics Team", f"PERSON-MGR-CDAO-{idx:02d}-{1 + squad % 2}", domain, "Consult")
            cdao_resource_rows += 1

    for org_id, name, leader, domain in BUSINESS_AREAS:
        add(org_id, f"GROUP-{org_id}", name, "Business Operating Group", leader, domain, "Recommend")
        add(org_id, f"PERSON-OWNER-{org_id}", f"{name} Technology Owner", "Business Technology Owner", leader, domain, "Consult")

    assert cdao_resource_rows + len(CDAO_VPS) + 1 >= 120, "CDAO resource coverage below target"
    return rows


def generate_facilities(columns: list[str]) -> list[dict[str, str]]:
    rows = []
    regions = ["Sacramento Metro", "Central Valley", "Bay Area", "Northern California", "Sierra Foothills"]
    service_lines = ["Acute Care", "Ambulatory", "Behavioral Health", "Oncology", "Cardiology", "Primary Care", "Post-Acute"]
    for i in range(1, PROFILE["hospitals"] + 1):
        rows.append(
            row(
                columns,
                {
                    "facility_id": f"FAC-HOSP-{i:03d}",
                    "business_unit_id": "BU-ACUTE",
                    "facility_name": f"Meridian {regions[i % len(regions)]} Hospital {i:02d}",
                    "facility_type": "Hospital",
                    "region": regions[i % len(regions)],
                    "service_line": service_lines[i % len(service_lines)],
                    "bed_count": 180 + (i * 17) % 260,
                    "annual_encounters": 42000 + i * 3100,
                    "revenue_owner": "Regional CFO",
                    "it_site_owner": "Clinical Applications Operations",
                    "critical_services": "Epic|Identity|Network|ServiceNow|AWS Connectivity",
                    **common("Facilities master", f"FAC-HOSP-{i:03d}", "Facilities Operations", "0.88"),
                },
            )
        )
    for i in range(1, PROFILE["clinics"] + 1):
        region = regions[i % len(regions)]
        rows.append(
            row(
                columns,
                {
                    "facility_id": f"FAC-CLINIC-{i:03d}",
                    "business_unit_id": "BU-AMB",
                    "facility_name": f"Meridian {region} Clinic {i:03d}",
                    "facility_type": "Clinic",
                    "region": region,
                    "service_line": service_lines[(i + 1) % len(service_lines)],
                    "bed_count": 0,
                    "annual_encounters": 9000 + i * 120,
                    "revenue_owner": "Ambulatory Finance",
                    "it_site_owner": "Ambulatory Technology Operations",
                    "critical_services": "Epic|MyChart|Network|Identity",
                    **common("Facilities master", f"FAC-CLINIC-{i:03d}", "Facilities Operations", "0.84"),
                },
            )
        )
    for i, name in enumerate(["Sacramento Primary Data Center", "Elk Grove DR Data Center", "Roseville Network Hub", "Folsom Print and Batch Center"], start=1):
        rows.append(
            row(
                columns,
                {
                    "facility_id": f"FAC-DC-{i:02d}",
                    "business_unit_id": "BU-IT",
                    "facility_name": name,
                    "facility_type": "Data Center",
                    "region": "Sacramento Metro",
                    "service_line": "Infrastructure",
                    "bed_count": 0,
                    "annual_encounters": 0,
                    "revenue_owner": "IT Finance",
                    "it_site_owner": "Infrastructure and Cloud Operations",
                    "critical_services": "Network|Storage|VMware|Epic Batch|AWS Direct Connect",
                    **common("Facilities master", f"FAC-DC-{i:02d}", "Infrastructure Operations", "0.9"),
                },
            )
        )
    for i in range(1, 15):
        rows.append(
            row(
                columns,
                {
                    "facility_id": f"FAC-OPS-{i:03d}",
                    "business_unit_id": "BU-SHARED",
                    "facility_name": f"Meridian Shared Services Site {i:02d}",
                    "facility_type": "Shared Services",
                    "region": regions[i % len(regions)],
                    "service_line": ["Contact Center", "Revenue Cycle", "Health Plan", "Supply Chain"][i % 4],
                    "bed_count": 0,
                    "annual_encounters": 0,
                    "revenue_owner": "Shared Services Finance",
                    "it_site_owner": "Enterprise Applications",
                    "critical_services": "ERP|Contact Center|ServiceNow|M365",
                    **common("Facilities master", f"FAC-OPS-{i:03d}", "Facilities Operations", "0.84"),
                },
            )
        )
    return rows


CORE_CIS = [
    ("CI-APP-EPIC-HYPERSPACE", "Epic Hyperspace", "Application", "Clinical Documentation", "Epic Platform Operations", "Clinical Applications Operations", "VEN-EPIC", "CON-EPIC-CORE-2027"),
    ("CI-APP-EPIC-CABOODLE", "Epic Caboodle", "Data Platform", "Clinical Analytics", "Epic Analytics", "Cloud Data Platform", "VEN-EPIC", "CON-EPIC-CORE-2027"),
    ("CI-APP-EPIC-CLARITY", "Epic Clarity", "Database", "Clinical Analytics", "Epic Analytics", "Database Operations", "VEN-EPIC", "CON-EPIC-CORE-2027"),
    ("CI-APP-EPIC-MYCHART", "Epic MyChart", "Application", "Patient Digital Experience", "Digital Patient Experience", "Clinical Applications Operations", "VEN-EPIC", "CON-EPIC-CORE-2027"),
    ("CI-APP-EPIC-RESOLUTE-HB", "Epic Resolute Hospital Billing", "Application", "Revenue Cycle", "Revenue Cycle Technology", "Clinical Applications Operations", "VEN-EPIC", "CON-EPIC-CORE-2027"),
    ("CI-APP-EPIC-RESOLUTE-PB", "Epic Resolute Professional Billing", "Application", "Revenue Cycle", "Revenue Cycle Technology", "Clinical Applications Operations", "VEN-EPIC", "CON-EPIC-CORE-2027"),
    ("CI-INT-RHAPSODY", "Rhapsody Integration Engine", "Integration", "Clinical Integration", "Integration Operations", "Integration Engineering", "VEN-RHAPSODY", "CON-RHAPSODY-2027"),
    ("CI-API-FHIR-GATEWAY", "FHIR API Gateway", "Integration", "Interoperability", "Integration Operations", "Cloud Platform Engineering", "VEN-AWS", "CON-AWS-EA-2027"),
    ("CI-ERP-ORACLE", "Oracle ERP Cloud", "Application", "Finance and Supply Chain", "ERP Platform Operations", "Enterprise Applications", "VEN-ORACLE", "CON-ORACLE-ERP-2028"),
    ("CI-HCM-WORKDAY", "Workday HCM", "Application", "People Operations", "HR Technology", "Enterprise Applications", "VEN-WORKDAY", "CON-WORKDAY-2028"),
    ("CI-ITSM-SERVICENOW", "ServiceNow ITSM and CMDB", "Application", "IT Service Management", "Service Management", "ITSM and CMDB", "VEN-SNOW", "CON-SNOW-2028"),
    ("CI-DATA-DATABRICKS", "Databricks Lakehouse", "Data Platform", "Enterprise Analytics", "Cloud Data Platform", "Data Engineering", "VEN-DATABRICKS", "CON-DATABRICKS-2028"),
    ("CI-DATA-COLLIBRA", "Collibra Data Catalog", "Data Governance", "Data Governance", "Data Governance and Quality", "Data Governance", "VEN-COLLIBRA", "CON-COLLIBRA-2027"),
    ("CI-IAM-ENTRA", "Microsoft Entra ID", "Identity", "Identity and Access", "Identity and Access Management", "Cybersecurity Operations", "VEN-MICROSOFT", "CON-M365-2028"),
    ("CI-SEC-CROWDSTRIKE", "CrowdStrike Falcon", "Security", "Endpoint Security", "Cybersecurity Operations", "Security Operations", "VEN-CROWDSTRIKE", "CON-CROWDSTRIKE-2027"),
    ("CI-AWS-LANDING-ZONE", "AWS Landing Zone", "Cloud Platform", "Cloud Hosting", "Cloud Migration Factory", "Cloud Platform Engineering", "VEN-AWS", "CON-AWS-EA-2027"),
]


def generate_cis(columns: list[str]) -> list[dict[str, str]]:
    rows = []
    for ci_id, name, ci_type, service, app_owner, tech_owner, vendor, contract in CORE_CIS:
        rows.append(
            row(
                columns,
                {
                    "ci_id": ci_id,
                    "ci_name": name,
                    "ci_type": ci_type,
                    "business_service": service,
                    "application_owner": app_owner,
                    "technical_owner": tech_owner,
                    "support_group": tech_owner,
                    "criticality": "Tier 1",
                    "hosting_model": "AWS-hosted" if ci_id.startswith("CI-AWS") or "DATABRICKS" in ci_id else "Hybrid",
                    "environment": "Production",
                    "vendor_id": vendor,
                    "contract_id": contract,
                    "data_classification": "Restricted",
                    "service_tier": "Gold",
                    **common("ServiceNow CMDB", ci_id, "CMDB Stewardship", "0.9"),
                },
            )
        )
    services = ["Clinical Documentation", "Revenue Cycle", "Patient Access", "Claims", "Care Management", "Population Health", "Finance", "Supply Chain", "Security", "Data Platform"]
    vendors = ["VEN-AWS", "VEN-EPIC", "VEN-MICROSOFT", "VEN-ORACLE", "VEN-SNOW", "VEN-DATABRICKS", "VEN-AMS-EPIC", "VEN-AMS-DATA"]
    contracts = ["CON-AWS-EA-2027", "CON-EPIC-CORE-2027", "CON-M365-2028", "CON-ORACLE-ERP-2028", "CON-SNOW-2028", "CON-DATABRICKS-2028", "CON-AMS-EPIC-2026", "CON-AMS-DATA-2026"]
    for i in range(1, 225):
        legacy = i <= 130
        ci_id = f"CI-LEGACY-VM-{i:03d}" if legacy else f"CI-SVC-{i:03d}"
        service = services[i % len(services)]
        vendor = "VEN-AWS" if legacy else vendors[i % len(vendors)]
        contract = "CON-AWS-EA-2027" if legacy else contracts[i % len(contracts)]
        rows.append(
            row(
                columns,
                {
                    "ci_id": ci_id,
                    "ci_name": f"{service} {'Legacy VM' if legacy else 'Service'} {i:03d}",
                    "ci_type": "Virtual Machine" if legacy else "Application Service",
                    "business_service": service,
                    "application_owner": f"{service} Owner",
                    "technical_owner": "Cloud Migration Factory" if legacy else "Platform Reliability",
                    "support_group": "Legacy VM Rehost Operations" if legacy else "Service Management",
                    "criticality": "Tier 1" if i % 9 == 0 else "Tier 2",
                    "hosting_model": "AWS-hosted lift-and-shift VM" if legacy else ["SaaS", "AWS managed service", "Hybrid"][i % 3],
                    "environment": "Production",
                    "vendor_id": vendor,
                    "contract_id": contract,
                    "data_classification": "Restricted" if service in ["Clinical Documentation", "Claims", "Care Management"] else "Internal",
                    "service_tier": "Gold" if i % 4 == 0 else "Silver",
                    **common("ServiceNow CMDB", ci_id, "CMDB Stewardship", "0.84"),
                },
            )
        )
    return rows


def generate_vendors(columns: list[str]) -> list[dict[str, str]]:
    named = [
        ("VEN-EPIC", "Epic Systems", "CON-EPIC-CORE-2027", "Epic Enterprise License and Hosting", "Clinical Systems", 28600000, "false"),
        ("VEN-AMS-EPIC", "Redwood Clinical AMS", "CON-AMS-EPIC-2026", "Epic AMS Transition and Run Services", "AMS - Epic", 18400000, "true"),
        ("VEN-AMS-DATA", "Sierra Data Operations", "CON-AMS-DATA-2026", "Data and Analytics AMS", "AMS - Data Analytics", 12600000, "true"),
        ("VEN-AWS", "Amazon Web Services", "CON-AWS-EA-2027", "AWS Enterprise Agreement", "Cloud Infrastructure", 48200000, "true"),
        ("VEN-ORACLE", "Oracle", "CON-ORACLE-ERP-2028", "Oracle ERP Cloud", "ERP and Finance", 14300000, "true"),
        ("VEN-WORKDAY", "Workday", "CON-WORKDAY-2028", "Workday HCM", "HCM", 9200000, "true"),
        ("VEN-SNOW", "ServiceNow", "CON-SNOW-2028", "ServiceNow ITSM and CMDB", "ITSM / CMDB", 7100000, "true"),
        ("VEN-DATABRICKS", "Databricks", "CON-DATABRICKS-2028", "Lakehouse Platform", "Data Platform", 9800000, "true"),
        ("VEN-MICROSOFT", "Microsoft", "CON-M365-2028", "Microsoft 365 and Entra", "Productivity and Identity", 22400000, "true"),
        ("VEN-RHAPSODY", "Rhapsody", "CON-RHAPSODY-2027", "Integration Engine", "Integration", 3600000, "true"),
        ("VEN-COLLIBRA", "Collibra", "CON-COLLIBRA-2027", "Data Governance Catalog", "Data Governance", 2900000, "false"),
        ("VEN-CROWDSTRIKE", "CrowdStrike", "CON-CROWDSTRIKE-2027", "Endpoint Security", "Security", 5100000, "false"),
    ]
    rows = []
    for vendor_id, vendor, contract_id, contract, category, spend, baa in named:
        rows.append(
            row(
                columns,
                {
                    "vendor_id": vendor_id,
                    "vendor_name": vendor,
                    "contract_id": contract_id,
                    "contract_name": contract,
                    "category": category,
                    "contract_owner": "Strategic Sourcing",
                    "relationship_owner": "Vendor and AMS Governance",
                    "start_date": "2026-01-01",
                    "end_date": "2028-12-31",
                    "annual_spend_usd": spend,
                    "termination_notice_days": 120,
                    "baa_required": baa,
                    "security_review_status": "Current" if baa == "true" else "Due",
                    **common("Contract repository", contract_id, "Strategic Sourcing", "0.89"),
                },
            )
        )
    categories = ["Clinical Systems", "Cloud Infrastructure", "Data Platform", "Security", "Integration", "Health Plan Operations", "Revenue Cycle", "Contact Center", "Device Management", "Professional Services"]
    for i in range(13, 111):
        vendor_id = f"VEN-MH-{i:03d}"
        contract_id = f"CON-MH-{i:03d}-2027"
        category = categories[i % len(categories)]
        rows.append(
            row(
                columns,
                {
                    "vendor_id": vendor_id,
                    "vendor_name": f"Meridian {category} Vendor {i:03d}",
                    "contract_id": contract_id,
                    "contract_name": f"{category} Services Agreement {i:03d}",
                    "category": category,
                    "contract_owner": "Strategic Sourcing",
                    "relationship_owner": ["Epic Platform Operations", "Cloud Migration Factory", "Data Governance", "Revenue Cycle Technology"][i % 4],
                    "start_date": as_date(i % 180),
                    "end_date": as_date(520 + i * 3),
                    "annual_spend_usd": 250000 + (i * 173000) % 4800000,
                    "termination_notice_days": [60, 90, 120, 180][i % 4],
                    "baa_required": "true" if i % 3 != 0 else "false",
                    "security_review_status": ["Current", "Due", "Conditional"][i % 3],
                    **common("Contract repository", contract_id, "Strategic Sourcing", "0.82"),
                },
            )
        )
    return rows


def generate_relationships(columns: list[str], ci_ids: list[str]) -> list[dict[str, str]]:
    rows = []
    rel_types = ["depends_on", "feeds", "authenticates_with", "hosted_on", "monitored_by", "integrates_with"]
    flows = ["PHI batch", "FHIR API", "HL7 ADT", "Claims EDI", "Financial batch", "Telemetry", "Identity token"]
    for i in range(1, 821):
        from_ci = ci_ids[i % len(ci_ids)]
        to_ci = ci_ids[(i * 7 + 13) % len(ci_ids)]
        rows.append(
            row(
                columns,
                {
                    "relationship_id": f"REL-MH-{i:04d}",
                    "from_ci_id": from_ci,
                    "to_ci_id": to_ci,
                    "relationship_type": rel_types[i % len(rel_types)],
                    "dependency_direction": "from_to",
                    "integration_pattern": ["API", "Batch", "Message queue", "Database replication", "File transfer"][i % 5],
                    "data_flow": flows[i % len(flows)],
                    "criticality": "High" if i % 8 == 0 else "Medium",
                    "recovery_dependency": "true" if i % 11 == 0 else "false",
                    "known_gap": "AWS-hosted VM dependency not yet modernized" if i % 17 == 0 else "",
                    **common("ServiceNow CMDB", f"REL-MH-{i:04d}", "CMDB Stewardship", "0.82"),
                },
            )
        )
    return rows


def generate_renewals(columns: list[str], vendors: list[dict[str, str]]) -> list[dict[str, str]]:
    rows = []
    for i, v in enumerate(vendors[:90], start=1):
        rows.append(
            row(
                columns,
                {
                    "renewal_id": f"REN-MH-{i:03d}",
                    "contract_id": v["contract_id"],
                    "vendor_id": v["vendor_id"],
                    "renewal_date": as_date(180 + i * 5),
                    "notice_date": as_date(120 + i * 5),
                    "renewal_type": ["Auto-renew", "Negotiated renewal", "Competitive sourcing"][i % 3],
                    "estimated_value_usd": v["annual_spend_usd"],
                    "sourcing_required": "true" if i % 4 == 0 or "AMS" in v["category"] else "false",
                    "decision_owner": "Strategic Sourcing",
                    "renewal_risk": "High" if "AMS" in v["category"] or i % 13 == 0 else "Medium",
                    "status": ["Open", "Planning", "In sourcing"][i % 3],
                    **common("Contract repository", f"REN-MH-{i:03d}", "Strategic Sourcing", "0.84"),
                },
            )
        )
    return rows


def generate_spend(columns: list[str], vendors: list[dict[str, str]]) -> list[dict[str, str]]:
    rows = []
    periods = [(date(2026, m, 1), date(2026, m, 28)) for m in range(1, 13)]
    for idx in range(1, 361):
        v = vendors[idx % len(vendors)]
        start, end = periods[idx % len(periods)]
        rows.append(
            row(
                columns,
                {
                    "spend_id": f"SPEND-MH-{idx:04d}",
                    "period_start": start.isoformat(),
                    "period_end": end.isoformat(),
                    "business_unit_id": ["BU-ACUTE", "BU-AMB", "BU-PLAN", "BU-IT", "BU-SHARED"][idx % 5],
                    "cost_center": f"CC-{1000 + idx % 90}",
                    "vendor_id": v["vendor_id"],
                    "contract_id": v["contract_id"],
                    "category": v["category"],
                    "actual_spend_usd": int(v["annual_spend_usd"]) // 12 + (idx * 1100) % 50000,
                    "run_rate_usd": int(v["annual_spend_usd"]),
                    "capex_opex": "Opex" if idx % 4 else "Capex",
                    "budget_owner": ["CFO", "CDAO", "CTO", "Revenue Cycle", "Health Plan"][idx % 5],
                    **common("ERP spend export", f"SPEND-MH-{idx:04d}", "IT Finance", "0.86"),
                },
            )
        )
    return rows


def generate_policies(columns: list[str], ci_ids: list[str]) -> list[dict[str, str]]:
    names = [
        "HIPAA Minimum Necessary",
        "California Privacy Controls",
        "AI Governance Intake",
        "Model Risk Evaluation",
        "Cloud PHI Hosting",
        "Third Party BAA Review",
        "Clinical Safety Gate",
        "FHIR API Access",
        "Data Retention",
        "FinOps Cost Allocation",
    ]
    rows = []
    for i in range(1, 76):
        name = names[i % len(names)]
        rows.append(
            row(
                columns,
                {
                    "policy_id": f"POL-MH-{i:03d}",
                    "policy_name": f"{name} Policy {i:03d}",
                    "policy_type": ["Privacy", "Security", "Clinical Safety", "AI Governance", "Cloud Operations"][i % 5],
                    "effective_date": as_date(i),
                    "version": f"v{1 + i % 4}.0",
                    "policy_owner": ["Chief Compliance and Privacy Officer", "CISO", "AI Governance Council", "Cloud FinOps Council"][i % 4],
                    "applies_to_systems": "|".join([ci_ids[i % len(ci_ids)], ci_ids[(i + 9) % len(ci_ids)]]),
                    "data_domains": ["Clinical", "Claims", "Revenue Cycle", "Identity", "Finance"][i % 5],
                    "control_requirement": "Documented evidence required before production use",
                    "ai_constraint": "AI use requires governance review" if i % 3 == 0 else "No autonomous action without approval",
                    "review_cycle_days": [90, 180, 365][i % 3],
                    "next_review_date": as_date(240 + i),
                    **common("Policy register", f"POL-MH-{i:03d}", "Compliance Office", "0.88"),
                },
            )
        )
    return rows


def generate_problems_incidents_changes_slas(columns_by_key: dict[str, list[str]], ci_ids: list[str]) -> dict[str, list[dict[str, str]]]:
    problems = []
    incidents = []
    changes = []
    slas = []
    services = ["Clinical Documentation", "Revenue Cycle", "Patient Access", "Claims", "Care Management", "Cloud Hosting", "Enterprise Analytics"]
    root_causes = ["AWS rehost performance", "Epic interface backlog", "Data quality", "Identity sync", "Vendor AMS transition", "Legacy batch dependency"]
    for i in range(1, 151):
        ci = ci_ids[i % len(ci_ids)]
        problems.append(
            row(
                columns_by_key["problems"],
                {
                    "problem_id": f"PRB-MH-{i:04d}",
                    "opened_at": as_date(20 + i),
                    "closed_at": "" if i % 5 else as_date(80 + i),
                    "ci_id": ci,
                    "business_service": services[i % len(services)],
                    "priority": ["P1", "P2", "P3"][i % 3],
                    "root_cause_category": root_causes[i % len(root_causes)],
                    "known_error": "true" if i % 4 == 0 else "false",
                    "status": "Open" if i % 5 else "Resolved",
                    "owner": ["Integration Operations", "Cloud Migration Factory", "Epic Platform Operations", "Data Quality Observability"][i % 4],
                    "linked_incident_count": 2 + i % 17,
                    "workaround_available": "true" if i % 3 else "false",
                    **common("ServiceNow ITSM", f"PRB-MH-{i:04d}", "ITSM Stewardship", "0.83"),
                },
            )
        )
    for i in range(1, 321):
        ci = ci_ids[i % len(ci_ids)]
        problem = f"PRB-MH-{1 + i % 150:04d}" if i % 2 == 0 else ""
        incidents.append(
            row(
                columns_by_key["incidents"],
                {
                    "incident_id": f"INC-MH-{i:05d}",
                    "opened_at": as_date(30 + i % 120),
                    "closed_at": "" if i % 7 == 0 else as_date(31 + i % 120),
                    "ci_id": ci,
                    "business_service": services[i % len(services)],
                    "priority": ["P1", "P2", "P3", "P4"][i % 4],
                    "severity": ["Critical", "High", "Medium", "Low"][i % 4],
                    "assignment_group": ["Epic Platform Operations", "Integration Operations", "Cloud Migration Factory", "Service Desk"][i % 4],
                    "short_description": f"{services[i % len(services)]} disruption during AWS-hosted transition wave {i % 8}",
                    "resolution_code": "Monitoring restored" if i % 7 else "",
                    "breach_sla": "true" if i % 11 == 0 else "false",
                    "related_problem_id": problem,
                    **common("ServiceNow ITSM", f"INC-MH-{i:05d}", "ITSM Stewardship", "0.82"),
                },
            )
        )
    for i in range(1, 221):
        ci = ci_ids[i % len(ci_ids)]
        caused = f"INC-MH-{i:05d}" if i % 29 == 0 else ""
        changes.append(
            row(
                columns_by_key["changes"],
                {
                    "change_id": f"CHG-MH-{i:05d}",
                    "planned_start": as_date(70 + i),
                    "planned_end": as_date(71 + i),
                    "ci_id": ci,
                    "business_service": services[i % len(services)],
                    "change_type": ["Standard", "Normal", "Emergency"][i % 3],
                    "risk_level": ["Low", "Medium", "High"][i % 3],
                    "approval_group": ["Clinical Change Advisory Board", "Cloud Migration Control Board", "AI Governance Council"][i % 3],
                    "implementation_owner": ["Cloud Migration Factory", "Epic Platform Operations", "Integration Operations"][i % 3],
                    "status": ["Scheduled", "Approved", "Implemented"][i % 3],
                    "backout_plan_tested": "true" if i % 4 else "false",
                    "caused_incident_id": caused,
                    **common("ServiceNow Change", f"CHG-MH-{i:05d}", "Change Management", "0.84"),
                },
            )
        )
    for i in range(1, 151):
        ci = ci_ids[i % len(ci_ids)]
        slas.append(
            row(
                columns_by_key["slas"],
                {
                    "sla_id": f"SLA-MH-{i:04d}",
                    "business_service": services[i % len(services)],
                    "ci_id": ci,
                    "sla_name": f"{services[i % len(services)]} availability SLA",
                    "metric_name": ["Availability", "P1 response", "Batch completion", "API latency"][i % 4],
                    "target_value": [99.9, 30, 4, 250][i % 4],
                    "unit": ["percent", "minutes", "hours", "milliseconds"][i % 4],
                    "measurement_source": "ServiceNow and observability platform",
                    "owner": ["Platform Reliability", "Epic Platform Operations", "Integration Operations"][i % 3],
                    "breach_count_90d": i % 9,
                    "trending_status": ["Improving", "Flat", "Worsening"][i % 3],
                    "penalty_or_impact": "Clinical continuity risk" if i % 7 == 0 else "Operational escalation",
                    **common("SLA repository", f"SLA-MH-{i:04d}", "Service Management", "0.83"),
                },
            )
        )
    return {"problems": problems, "incidents": incidents, "changes": changes, "slas": slas}


def generate_initiatives(columns: list[str], ci_ids: list[str], contracts: list[str], policies: list[str]) -> list[dict[str, str]]:
    themes = [
        "Epic AMS Transition",
        "Data Analytics AMS Transition",
        "AWS Data Center Exit",
        "Ambient Documentation Rollout",
        "Revenue Cycle Denials Automation",
        "Patient Access Contact Center AI",
        "Caboodle to Lakehouse Integration",
        "Claims Automation",
        "Care Management AI Assist",
        "Cloud FinOps Control Tower",
        "Identity Modernization",
        "Integration Reliability",
        "ERP Controls Modernization",
        "Data Governance Remediation",
        "Clinical Quality Analytics",
    ]
    rows = []
    for i in range(1, 96):
        theme = themes[i % len(themes)]
        rows.append(
            row(
                columns,
                {
                    "initiative_id": f"INIT-MH-{i:03d}",
                    "initiative_name": f"{theme} Wave {1 + i % 6}",
                    "sponsor": ["CIO/CDIO", "CDAO", "COO", "CFO", "CMO", "Health Plan President"][i % 6],
                    "initiative_owner": ["Cloud Migration Factory", "Epic Platform Operations", "Data Governance", "Revenue Cycle Technology", "AI Enablement"][i % 5],
                    "status": ["Proposed", "In Design", "Approved", "In Flight", "At Risk"][i % 5],
                    "stage": ["P1 Charter", "P2 Diagnose", "P3 Design", "P4 Business Case", "P5 Mobilize"][i % 5],
                    "start_date": as_date(i * 2),
                    "target_date": as_date(160 + i * 3),
                    "value_hypothesis": f"Improve {theme.lower()} outcomes while protecting clinical continuity, PHI controls, and run-cost evidence.",
                    "dependent_ci_ids": "|".join([ci_ids[i % len(ci_ids)], ci_ids[(i * 5) % len(ci_ids)], ci_ids[(i * 11) % len(ci_ids)]]),
                    "dependent_contract_ids": "|".join([contracts[i % len(contracts)], contracts[(i * 3) % len(contracts)]]),
                    "policy_constraints": "|".join([policies[i % len(policies)], policies[(i * 7) % len(policies)]]),
                    **common("Enterprise PMO", f"INIT-MH-{i:03d}", "Enterprise PMO", "0.84"),
                },
            )
        )
    return rows


def generate_data_domains(columns: list[str], ci_ids: list[str]) -> list[dict[str, str]]:
    domains = ["Clinical Encounter", "Orders", "Medication", "Revenue Cycle", "Claims", "Member", "Provider", "Access", "Care Management", "Quality", "Finance", "Supply Chain", "Identity", "Security", "Cloud FinOps", "Workforce", "Patient Experience", "Population Health"]
    rows = []
    for i in range(1, 111):
        name = domains[i % len(domains)]
        rows.append(
            row(
                columns,
                {
                    "data_domain_id": f"DATA-MH-{i:03d}",
                    "domain_name": f"{name} Domain {1 + i // len(domains)}",
                    "data_owner": ["CDAO", "CMO", "CRO", "Health Plan President", "CFO"][i % 5],
                    "data_steward": ["Data Governance", "Epic Analytics", "Revenue Cycle Analytics", "Health Plan Analytics", "Cloud Data Platform"][i % 5],
                    "source_systems": "|".join([ci_ids[i % len(ci_ids)], ci_ids[(i * 13) % len(ci_ids)]]),
                    "critical_data_elements": "Patient ID|Encounter ID|Owner|Timestamp|Status",
                    "classification": "Restricted" if i % 2 else "Internal",
                    "quality_score": round(0.68 + (i % 25) / 100, 2),
                    "retention_policy": "Enterprise healthcare retention schedule",
                    "ai_use_allowed": "Allowed with governance" if i % 3 else "Restricted",
                    "last_quality_review": as_date(40 + i),
                    **common("Data catalog", f"DATA-MH-{i:03d}", "Data Governance", "0.82"),
                },
            )
        )
    return rows


def generate_risks(columns: list[str], ci_ids: list[str], vendors: list[str], policies: list[str]) -> list[dict[str, str]]:
    titles = [
        "AWS-hosted workload not modernized",
        "Epic AMS transition knowledge gap",
        "Data analytics AMS scope ambiguity",
        "PHI control evidence missing after migration",
        "Cloud run cost exceeds baseline",
        "Integration backlog affects clinical continuity",
        "CDAO stewardship capacity constraint",
        "BAA chain incomplete for AI workflow",
    ]
    rows = []
    for i in range(1, 211):
        rows.append(
            row(
                columns,
                {
                    "risk_id": f"RISK-MH-{i:04d}",
                    "risk_title": f"{titles[i % len(titles)]} {i:03d}",
                    "risk_type": ["Operational", "Compliance", "Financial", "Clinical Safety", "Sourcing"][i % 5],
                    "control_id": f"CTRL-MH-{i % 75 + 1:03d}",
                    "policy_id": policies[i % len(policies)],
                    "owner": ["CISO", "CDAO", "CTO", "CFO", "AI Governance Council"][i % 5],
                    "severity": ["High", "Medium", "Critical"][i % 3],
                    "likelihood": ["Likely", "Possible", "Unlikely"][i % 3],
                    "status": ["Open", "Mitigating", "Accepted with controls"][i % 3],
                    "linked_ci_ids": "|".join([ci_ids[i % len(ci_ids)], ci_ids[(i * 9) % len(ci_ids)]]),
                    "linked_vendor_ids": "|".join([vendors[i % len(vendors)], vendors[(i * 5) % len(vendors)]]),
                    "remediation_due_date": as_date(120 + i),
                    "evidence_required": "Control evidence, owner attestation, run-cost and continuity proof",
                    **common("Risk register", f"RISK-MH-{i:04d}", "Risk and Compliance", "0.83"),
                },
            )
        )
    return rows


def write_csv(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, title: str, columns: list[str], rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    fixed_timestamp = datetime(2026, 6, 4, 0, 0, 0)
    wb.properties.creator = "AbarVa"
    wb.properties.created = fixed_timestamp
    wb.properties.modified = fixed_timestamp
    ws = wb.active
    safe_title = "".join("_" if char in r'[]:*?/\\' else char for char in title)
    ws.title = safe_title[:31]
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="111827")
    for record in rows:
        ws.append([record.get(col, "") for col in columns])
    ws.freeze_panes = "A2"
    for idx, col in enumerate(columns, start=1):
        width = max(12, min(42, len(col) + 4))
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A"].width = width
    wb.save(path)
    normalize_xlsx_zip(path)


def normalize_xlsx_zip(path: Path) -> None:
    fixed_date = (2026, 6, 4, 0, 0, 0)
    with zipfile.ZipFile(path, "r") as source:
        entries = []
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "docProps/core.xml":
                text = payload.decode("utf-8")
                text = re.sub(
                    r"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    '<dcterms:modified xsi:type="dcterms:W3CDTF">2026-06-04T00:00:00Z</dcterms:modified>',
                    text,
                )
                payload = text.encode("utf-8")
            entries.append((info.filename, payload))
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as target:
        for filename, payload in entries:
            info = zipfile.ZipInfo(filename, fixed_date)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            target.writestr(info, payload)


def validate(datasets: dict[str, list[dict[str, str]]]) -> dict[str, Any]:
    ci_ids = {r["ci_id"] for r in datasets["cmdb_applications_services"]}
    vendor_ids = {r["vendor_id"] for r in datasets["vendors_contract_inventory"]}
    contract_ids = {r["contract_id"] for r in datasets["vendors_contract_inventory"]}
    policy_ids = {r["policy_id"] for r in datasets["policies_procedures"]}
    unresolved: list[dict[str, str]] = []

    def split(value: str) -> list[str]:
        return [part for part in value.split("|") if part]

    for r in datasets["ci_relationships_dependencies"]:
        for field in ["from_ci_id", "to_ci_id"]:
            if r[field] not in ci_ids:
                unresolved.append({"dataset": "ci_relationships_dependencies", "record": r["relationship_id"], "field": field, "value": r[field]})
    for r in datasets["cmdb_applications_services"]:
        if r["vendor_id"] not in vendor_ids:
            unresolved.append({"dataset": "cmdb_applications_services", "record": r["ci_id"], "field": "vendor_id", "value": r["vendor_id"]})
        if r["contract_id"] not in contract_ids:
            unresolved.append({"dataset": "cmdb_applications_services", "record": r["ci_id"], "field": "contract_id", "value": r["contract_id"]})
    for r in datasets["initiative_portfolio"]:
        for value in split(r["dependent_ci_ids"]):
            if value not in ci_ids:
                unresolved.append({"dataset": "initiative_portfolio", "record": r["initiative_id"], "field": "dependent_ci_ids", "value": value})
        for value in split(r["dependent_contract_ids"]):
            if value not in contract_ids:
                unresolved.append({"dataset": "initiative_portfolio", "record": r["initiative_id"], "field": "dependent_contract_ids", "value": value})
        for value in split(r["policy_constraints"]):
            if value not in policy_ids:
                unresolved.append({"dataset": "initiative_portfolio", "record": r["initiative_id"], "field": "policy_constraints", "value": value})
    for r in datasets["risk_compliance_register"]:
        for value in split(r["linked_ci_ids"]):
            if value not in ci_ids:
                unresolved.append({"dataset": "risk_compliance_register", "record": r["risk_id"], "field": "linked_ci_ids", "value": value})
        for value in split(r["linked_vendor_ids"]):
            if value not in vendor_ids:
                unresolved.append({"dataset": "risk_compliance_register", "record": r["risk_id"], "field": "linked_vendor_ids", "value": value})

    cdao_rows = [
        r for r in datasets["org_decision_rights"]
        if r["person_or_group_id"] == "PERSON-CDAO" or r["reports_to_id"].startswith("PERSON-CDAO") or r["org_unit_id"].startswith("ORG-CDAO")
    ]
    hospital_rows = [r for r in datasets["facilities_business_units"] if r["facility_type"] == "Hospital"]
    clinic_rows = [r for r in datasets["facilities_business_units"] if r["facility_type"] == "Clinic"]
    aws_hosted = [r for r in datasets["cmdb_applications_services"] if "AWS" in r["hosting_model"]]

    return {
        "generatedAt": GENERATED_AT,
        "profile": PROFILE,
        "rowCounts": {key: len(value) for key, value in datasets.items()},
        "assertions": {
            "sacramentoProfile": PROFILE["headquarters"] == "Sacramento, California",
            "hospitalCountAtLeast30": len(hospital_rows) >= 30,
            "clinicCountAtLeast280": len(clinic_rows) >= 280,
            "cdaoResourceRowsAtLeast120": len(cdao_rows) >= 120,
            "awsHostedSignalsPresent": len(aws_hosted) >= 100,
            "unresolvedReferences": len(unresolved),
        },
        "unresolvedReferences": unresolved,
    }


def write_supporting_docs(datasets: dict[str, list[dict[str, str]]], validation: dict[str, Any]) -> None:
    profile_md = f"""# Meridian Health System vNext Profile Fact Card

- Sacramento-based integrated delivery system plus health plan
- $16.8B annual revenue
- 31 hospitals and 280 clinics represented in the package
- Roughly 1.6M covered lives
- Roughly 58,000 workforce
- Epic-centered clinical estate with ERP, ServiceNow, Microsoft, AWS, and Databricks-style analytics stack
- CDAO resource model: {sum(1 for r in datasets['org_decision_rights'] if r['person_or_group_id'] == 'PERSON-CDAO' or r['reports_to_id'].startswith('PERSON-CDAO') or r['org_unit_id'].startswith('ORG-CDAO'))} CDAO rows/resources represented
- Cloud truth: 100% workloads targeted to be AWS-hosted by July 2026, mostly via lift-and-shift/rehosted VMs
- Sourcing truth: new AMS pressure for Epic and data analytics operations

Stale facts to reject: Charlotte headquarters, 23 hospitals, or a small regional-system profile.
"""
    (OUT / "profile-fact-card.md").write_text(profile_md)

    loader_md = """# Meridian vNext Loader Checklist

Use Setup/Admin loader only. Do not insert these records directly into the database.

1. Upload all 15 CSV or XLSX files as a single Meridian vNext package.
2. Confirm row counts match `manifest.json`.
3. Confirm unresolved references remain 0 in `validation-report.json`.
4. Confirm loader ledger records source basis, source owner, validation date, confidence, and evidence usability.
5. Confirm Admin shows all 15 dimensions loaded for Meridian.
6. Confirm chunks, graph nodes, graph edges, and quality issues are visible in the loader run report.
7. If a field is dropped, enhance the loader rather than simplifying this package.
"""
    (OUT / "loader-checklist.md").write_text(loader_md)

    qa_md = """# Meridian vNext Agent Grounding Checklist

Run these after Admin loader ingestion:

- Describe Meridian in one paragraph.
- Who owns data and analytics?
- How many resources are under the CDAO?
- What are the biggest risks in moving all workloads to AWS by July?
- Is hosted on AWS the same as modernized?
- What systems are most critical to the Epic AMS contract?
- Which initiatives depend on Epic Caboodle or Clarity?
- Which vendors have BAA or AMS risk?
- Which facilities or regions are most exposed by the migration?
- What would you tell the CFO about AWS run cost risk?
- What would you tell the COO about clinical continuity risk?
- What would you tell the CDAO about data-platform readiness?
- What would Source need before drafting the Epic AMS sourcing event?

Pass criteria: uses Sacramento, 30+ hospitals, 280 clinics, provider plus plan, Epic, AWS-hosted lift-and-shift, 120+ CDAO resources, AMS pressure, and plain CXO language. Fails if it uses Charlotte, 23 hospitals, unexplained raw IDs, or direct table names.
"""
    (OUT / "agent-grounding-checklist.md").write_text(qa_md)


def main() -> None:
    specs = load_specs()
    columns_by_key = {spec.key: spec.columns for spec in specs}
    OUT.mkdir(parents=True, exist_ok=True)

    datasets: dict[str, list[dict[str, str]]] = {}
    datasets["org_decision_rights"] = generate_org(columns_by_key["org_decision_rights"])
    datasets["facilities_business_units"] = generate_facilities(columns_by_key["facilities_business_units"])
    datasets["vendors_contract_inventory"] = generate_vendors(columns_by_key["vendors_contract_inventory"])
    datasets["cmdb_applications_services"] = generate_cis(columns_by_key["cmdb_applications_services"])
    ci_ids = [r["ci_id"] for r in datasets["cmdb_applications_services"]]
    vendor_ids = [r["vendor_id"] for r in datasets["vendors_contract_inventory"]]
    contract_ids = [r["contract_id"] for r in datasets["vendors_contract_inventory"]]
    datasets["ci_relationships_dependencies"] = generate_relationships(columns_by_key["ci_relationships_dependencies"], ci_ids)
    datasets["renewal_calendar"] = generate_renewals(columns_by_key["renewal_calendar"], datasets["vendors_contract_inventory"])
    datasets["spend_baseline"] = generate_spend(columns_by_key["spend_baseline"], datasets["vendors_contract_inventory"])
    datasets["policies_procedures"] = generate_policies(columns_by_key["policies_procedures"], ci_ids)
    policy_ids = [r["policy_id"] for r in datasets["policies_procedures"]]
    datasets.update(generate_problems_incidents_changes_slas(columns_by_key, ci_ids))
    datasets["initiative_portfolio"] = generate_initiatives(columns_by_key["initiative_portfolio"], ci_ids, contract_ids, policy_ids)
    datasets["data_domains_stewardship"] = generate_data_domains(columns_by_key["data_domains_stewardship"], ci_ids)
    datasets["risk_compliance_register"] = generate_risks(columns_by_key["risk_compliance_register"], ci_ids, vendor_ids, policy_ids)

    validation = validate(datasets)

    for spec in specs:
        records = datasets[spec.key]
        write_csv(OUT / spec.csv_name, spec.columns, records)
        write_xlsx(OUT / spec.xlsx_name, spec.title, spec.columns, records)

    manifest = {
        "tenantKey": "meridian",
        "tenantSlug": "meridian",
        "displayName": PROFILE["displayName"],
        "generatedAt": GENERATED_AT,
        "fictional": True,
        "noPhi": True,
        "seed": SEED,
        "templateVersion": "2026.05.day-one.v1",
        "profile": PROFILE,
        "workbookCount": len(specs),
        "totalRows": sum(len(v) for v in datasets.values()),
        "validation": {
            "unresolvedReferences": len(validation["unresolvedReferences"]),
            "assertions": validation["assertions"],
        },
        "datasets": [
            {
                "key": spec.key,
                "title": spec.title,
                "csv": spec.csv_name,
                "xlsx": spec.xlsx_name,
                "rows": len(datasets[spec.key]),
                "columns": spec.columns,
            }
            for spec in specs
        ],
    }
    (OUT / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n")
    (OUT / "validation-report.json").write_text(json.dumps(validation, indent=2) + "\n")
    write_supporting_docs(datasets, validation)

    summary = {
        "out": str(OUT.relative_to(ROOT)),
        "totalRows": manifest["totalRows"],
        "rowCounts": validation["rowCounts"],
        "assertions": validation["assertions"],
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
